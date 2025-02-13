
import pandas as pd
import tempfile
import pandas as pd
import openpyxl
import os
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import pixels_to_points, points_to_pixels
from openpyxl.styles import numbers
from openpyxl.styles import NamedStyle
from openpyxl.drawing.image import Image
from io import BytesIO
from PIL import Image as PILImage
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
# Input Excel file
from openpyxl.utils import column_index_from_string 
import hashlib
import xlwings as xw
from openpyxl.formatting.rule import FormulaRule
import shutil
from  data import *
# Input Excel sheets
METAL_COMPONENTS_SHEET = 'METAL & COMPONENTS'
LABOR_SHEET = 'LABOR'
DIAMOND_SHAPE_QUALITY_SHEET = 'DIAMOND SHAPE & QLTY'
IMAGE_SHEET = 'IMAGE'

# Output directory for extracted images
OUTPUT_IMAGE_DIR = 'img_all'

current_directory = os.getcwd()

set_df=pd.read_excel("SETTING CODE DESCP.xlsx")
SETTING_DESC=dict(zip(set_df['Code'],set_df['Description']))
# Output Excel file

name_mapping = {}
# Function to hash the Itemno to generate a safe filename
def hash_item_name(item_name):
    return hashlib.md5(item_name.encode()).hexdigest()  # Generate a unique hash for the item name

def apply_formatting(ws, cell_range, headers, values, header_font, value_font, alignment, border,dollar=False):
    # Split the range to get the start and end cells
    start_cell, end_cell = cell_range.split(':')
    
    # Determine the row and column of the start and end cells
    start_row = int(start_cell[2:])
    start_col = start_cell[:2]

    # Apply the headers
    for i, header in enumerate(headers):
        cell = f"{start_col}{start_row + i}"
        ws[cell] = header
        ws[cell].font = header_font
        ws[cell].alignment = alignment
        ws[cell].border = border
    
    # Apply the values
    for i, value in enumerate(values):
        cell = f"{start_col}{start_row + 1 + i}"
        # print(cell)
        ws[cell] = value
        if dollar:
            ws[cell].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
            
        ws[cell].font = value_font
        ws[cell].alignment = alignment
        ws[cell].border = border



def index_to_column(index):
    column_str = ""
    while index > 0:
        index, remainder = divmod(index - 1, 26)
        column_str = chr(65 + remainder) + column_str
    return column_str




def get_single_or_longest_string(strings):
    if len(strings) == 1:
        return strings[0]
    else:
        return max(strings, key=len)



def get_column_width(cell):
    if not cell.value:
        return 9  # Default minimum width
    s=get_single_or_longest_string(str(cell.value).split("\n"))

    # Approximate character widths (these are rough estimates and may need adjustment)
    char_widths = {
        'default': 0.09,
        'digit': 0.13,
        'uppercase': 0.14,
        'punctuation': 0.09,
        
    }
    total_width = 0
    for char in s:
        if char.isdigit():
            total_width += char_widths['digit']
        elif char.isupper():
            total_width += char_widths['uppercase']
        elif char in '.,;:!?-':
            total_width += char_widths['punctuation']
        else:
            total_width += char_widths['default']

    return total_width

def download_images(INPUT_EXCEL_FILE):
    # Load the source workbook and select the 'IMAGE' sheet
    source_wb = openpyxl.load_workbook(INPUT_EXCEL_FILE)
    source_sheet = source_wb[IMAGE_SHEET]
    
    # Create output directory if it doesn't exist
    shutil.rmtree(OUTPUT_IMAGE_DIR)
    os.makedirs(OUTPUT_IMAGE_DIR)
    
    # Iterate through rows in the source sheet to find images and corresponding item names
    for row_idx, row in enumerate(source_sheet.iter_rows(min_row=2, min_col=1, max_col=2), start=2):
        item_cell, image_cell = row
        item_name = item_cell.value
        
        # Check if item_name is None, and skip this row if it is
        if item_name is None:
            print(f"Item name in row {row_idx} is None, skipping.")
            continue
        
        # Hash the item name to get a safe filename
        hashed_filename = hash_item_name(item_name)
        image_file_path = os.path.join(OUTPUT_IMAGE_DIR, f"{hashed_filename}.png")
        
        # Skip saving if the file already exists
        if os.path.exists(image_file_path):
            print(f"Image for item {item_name} already exists as {hashed_filename}.png, skipping download.")
            continue
        
        # Save the image using the hashed filename if it doesn't already exist
        for drawing in source_sheet._images:
            if drawing.anchor._from.row + 1 == row_idx:
                image_data = drawing._data()
                image_stream = BytesIO(image_data)
                pil_image = PILImage.open(image_stream)
                pil_image.save(image_file_path)
                name_mapping[item_name] = hashed_filename  # Store mapping for retrieval
                print(f"Saved image for item {item_name} as {hashed_filename}.png at {image_file_path}")
                break
        else:
            print(f"No image found for item {item_name} in row {row_idx}")
    

    
    source_wb.close()
    
   
    
    
    
def add_data_to_sheet(INPUT_EXCEL_FILE, OUTPUT_EXCEL_FILE):

    metal_df1 = pd.read_excel(INPUT_EXCEL_FILE, sheet_name=METAL_COMPONENTS_SHEET)
    labour_df = pd.read_excel(INPUT_EXCEL_FILE, sheet_name=LABOR_SHEET)
    shape_qty_df = pd.read_excel(INPUT_EXCEL_FILE, sheet_name=DIAMOND_SHAPE_QUALITY_SHEET)
    # Extract unique TAB categories
    download_images(INPUT_EXCEL_FILE)


    # Create a new workbook
    wb = openpyxl.Workbook()
    # Remove the default sheet
    wb.remove(wb.active)

    # Check if 'TAB' column is in metal_df1 before accessing it
    if 'TAB' in metal_df1.columns:
        # Extract unique TAB categories
        tab_categories = metal_df1['TAB'].unique()
        
        # Loop through each TAB category and create a corresponding sheet
        for tab in tab_categories:
            # Filter metal_df by the current TAB
            filtered_df = metal_df1[metal_df1['TAB'] == tab]
            print(f"Number of items in {tab}: {len(filtered_df)}")
            
            # Create a new sheet for the current TAB
            ws = wb.create_sheet(title=str(tab))
            
            # Call function to add data to the current sheet
            preprocess(ws, filtered_df, labour_df, shape_qty_df)
    else:
        # If 'TAB' column is not present, create a single sheet called "Replicated Data"
        ws = wb.create_sheet(title="Replicated Data")
        print("TAB column not found. Saving all data in 'Replicated Data' sheet.")
        
        # Call function to add the full metal_df1 to the "Replicated Data" sheet
        preprocess(ws, metal_df1, labour_df, shape_qty_df)
    # Save the workbook to the output file

    # Save the workbook to the output file
    wb.save(OUTPUT_EXCEL_FILE)
    wb.close()
    
    
def preprocess(ws, metal_df,labour_df,shape_qty_df):

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  
    light_blue_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")  

    column_names = [
        "IMAGE", "STYLE #", "DESCRIPTION", "Mnf.\nPolicy", "Item\nStatus\nCode",
        "Manf.\nCode", "COO", "METAL", "Wt.", "Metal\nCost\n /Gm", "METAL\nCOST", 
        "STONE\nTYPE", "QTY", "SHAPE", "Color\nMM\nSize", "wt.", "WT.\n Ext.", "DIA CTTW", 
        "CODE", "Quality", "$ PER \nCT/PC", "COST","IH\n (YES/NO)", "TOTAL\nCS\nCOST", "CS\nIH\nCOST",
        "TOTAL\nDIA\nCOST", "DIA\nIH\nCOST","TOTAL\nSTONE\nCOST","STONE\nIH\nCOST", "SET\nCOST", "EXT\nSET\nCOST", 
         "SETTING\nDESCRIPTION","LABOR TYPE", "LABOR\nQTY", "LABOR\nDTL\nCOST","Labour IH COST\nOR NOT\n(YES/NO)", 
        "LABOR\nTTL\nCOST", "LABOR\nIH\nCOST", "FINDING\nTYPE","FINDING\n(YES/NO)", "FINDING\nQTY", "FINDING\nDTL\nCOST", "FINDING IH COST\nOR NOT\n(YES/NO)",
        "FINDING\nTTL\nCOST","FINDING\nIH COST", "STYLE\nTOTAL","STYLE\nTOTAL\nEXCL.IH", "S/H", "Duty $", "Duty %", 
        "LANDED\nTOTAL", "Box", "Cert", "Target\nGM%", "Sell_1", 
        "Gross\nMargin_1", "Net\nSell_1", "Net\nMargin_1", "MSRP_1", "IMU_1", "AUR_1", "AUR\nDiscount_1", 
        "AUR\nGM%_1", "1st MKD_1", "MKD NM%_1", "Sell_2", "Gross\nMargin_2", "Net\nSell_2", "Net Margin_2", 
        "MSRP_2", "IMU_2", "AUR_2", "AUR\nDiscount_2", "AUR\nGM%_2", "1st MKD_2", "MKD NM%_2","Support Comment"
    ]
    
    op_li=SETTING_CODE
    
    column_dict = {name: index_to_column(i+1) for i, name in enumerate(column_names)}
    if 'Manufacturing Policy' in metal_df.columns and 'AvgOfDuty %' in metal_df.columns:
        single_df=metal_df[[
        "Itemno",
        'Manufacturing Policy',
        'AvgOfDuty %',
        "dbo_NAV_Item_Master.Description",
        "Item Status Code",
        "Vendor No_",
        "Country Of Origin Code","Metal Quality Code","Fin Metal weight"
    ]].drop_duplicates().reset_index(drop=True)

    else:

        single_df=metal_df[[
            "Itemno",
            "dbo_NAV_Item_Master.Description",
            "Item Status Code",
            "Vendor No_",
            "Country Of Origin Code","Metal Quality Code","Fin Metal weight"
        ]].drop_duplicates().reset_index(drop=True)
    
    # single_df['inhouse_or_not'] = ['YES' if i == 'IH Assembly' else 'NO' for i in single_df['Supply Policy']]

    # Define header formatting
    header_font = Font(name='Calibri', size=12, bold=True)
    subheader_font = Font(name='Calibri', size=11)
    subheader_bold = Font(name='Calibri', size=11, bold=True)
    
    alignment_center = Alignment(horizontal='center', vertical='center')
    alignment_left = Alignment(horizontal='left', vertical='center')
    border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))
    


    # Add subheaders
    ws[f"{column_dict['IMAGE']}7"] = "Customer Name:"
    ws[f"{column_dict['IMAGE']}7"].font = subheader_bold
    ws[f"{column_dict['IMAGE']}7"].alignment = alignment_left
    
    ws[f"{column_dict['STYLE #']}7"] = "JCP"
    ws[f"{column_dict['STYLE #']}7"].font = subheader_font
    ws[f"{column_dict['STYLE #']}7"].alignment = alignment_left
    
    ws[f"{column_dict['DESCRIPTION']}7"] = "Sales Quote #"
    ws[f"{column_dict['DESCRIPTION']}7"].font = subheader_bold
    ws[f"{column_dict['DESCRIPTION']}7"].alignment = alignment_left
    
    # Add data rows and format them as needed
    ws[f"{column_dict['IMAGE']}8"] = "Date of Extract:"
    ws[f"{column_dict['IMAGE']}8"].font = subheader_bold
    ws[f"{column_dict['IMAGE']}8"].alignment = alignment_left
    
    # Add data rows and format them as needed
    ws[f"{column_dict['IMAGE']}9"] = "Meeting Date:"
    ws[f"{column_dict['IMAGE']}9"].font = subheader_bold
    ws[f"{column_dict['IMAGE']}9"].alignment = alignment_left
    
    # Add specific data from O1 to P1 and O3 to P3
    ws[column_dict['Metal\nCost\n /Gm'] + "1"] = "S"
    ws[column_dict['Metal\nCost\n /Gm'] + "2"] = "GOS"

    ws[column_dict['Metal\nCost\n /Gm'] + "1"].font = subheader_bold
    ws[column_dict['Metal\nCost\n /Gm'] + "2"].font = subheader_bold
    ws[column_dict['Metal\nCost\n /Gm'] + "1"].alignment = alignment_left
    ws[column_dict['Metal\nCost\n /Gm'] + "2"].alignment = alignment_left
    ws[column_dict['Metal\nCost\n /Gm'] + "1"].border = border
    ws[column_dict['Metal\nCost\n /Gm'] + "2"].border = border

    ws[column_dict['Metal\nCost\n /Gm'] + "3"] = "10"
    ws[column_dict['Metal\nCost\n /Gm'] + "4"] = "14"

    ws[column_dict['Metal\nCost\n /Gm'] + "3"].font = subheader_bold
    ws[column_dict['Metal\nCost\n /Gm'] + "4"].font = subheader_bold
    ws[column_dict['Metal\nCost\n /Gm'] + "3"].alignment = alignment_left
    ws[column_dict['Metal\nCost\n /Gm'] + "4"].alignment = alignment_left
    ws[column_dict['Metal\nCost\n /Gm'] + "3"].border = border
    ws[column_dict['Metal\nCost\n /Gm'] + "4"].border = border

    ws[column_dict['Metal\nCost\n /Gm'] + "7"] = "S"
    ws[column_dict['Metal\nCost\n /Gm'] + "8"] = "GOLD"

    ws[column_dict['Metal\nCost\n /Gm'] + "7"].font = subheader_bold
    ws[column_dict['Metal\nCost\n /Gm'] + "8"].font = subheader_bold
    ws[column_dict['Metal\nCost\n /Gm'] + "7"].alignment = alignment_left
    ws[column_dict['Metal\nCost\n /Gm'] + "8"].alignment = alignment_left
    ws[column_dict['Metal\nCost\n /Gm'] + "7"].border = border
    ws[column_dict['Metal\nCost\n /Gm'] + "8"].border = border

    # Update your formulas using string concatenation
    ws[column_dict['METAL\nCOST'] + "1"] = '=ROUND($' + column_dict['METAL\nCOST'] + '$7/31.1035*0.925*1.2, 2)'
    ws[column_dict['METAL\nCOST'] + "2"] = '=ROUND($' + column_dict['METAL\nCOST'] + '$7/31.1035*0.925*1.2, 2)'
    ws[column_dict['METAL\nCOST'] + "3"] = '=ROUND($' + column_dict['METAL\nCOST'] + '$8/31.1035*0.417*1.088, 2)'
    ws[column_dict['METAL\nCOST'] + "4"] = '=ROUND($' + column_dict['METAL\nCOST'] + '$8/31.1035*0.585*1.088, 2)'


    ws[column_dict['METAL\nCOST'] + "7"] = 40.00
    ws[column_dict['METAL\nCOST'] + "7"].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

    ws[column_dict['METAL\nCOST'] + "8"] = 2600.00
    ws[column_dict['METAL\nCOST'] + "8"].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

    ws[column_dict['METAL\nCOST'] + "7"].font = subheader_font
    ws[column_dict['METAL\nCOST'] + "8"].font = subheader_font
    ws[column_dict['METAL\nCOST'] + "7"].alignment = alignment_left
    ws[column_dict['METAL\nCOST'] + "8"].alignment = alignment_left
    ws[column_dict['METAL\nCOST'] + "7"].border = border
    ws[column_dict['METAL\nCOST'] + "8"].border = border

    ws[column_dict['Target\nGM%'] + "7"] = "Base GM%"
    ws[column_dict['Target\nGM%'] + "8"] = "33.6%"

    ws[column_dict['Target\nGM%'] + "7"].font = subheader_bold
    ws[column_dict['Target\nGM%'] + "7"].alignment = alignment_left
    ws[column_dict['Target\nGM%'] + "7"].border = border

    ws[column_dict['Target\nGM%'] + "8"].font = subheader_font
    ws[column_dict['Target\nGM%'] + "8"].alignment = alignment_left
    ws[column_dict['Target\nGM%'] + "8"].border = border

    
    
    
    
    apply_formatting(ws, column_dict['Target\nGM%'] + "7:" + column_dict['Target\nGM%'] + "8", 
                    ["Base GM%"], [0.336], subheader_bold, subheader_font, alignment_left, border)
    ws[column_dict['Target\nGM%'] + "8"].number_format = '0.0%'
    apply_formatting(ws, column_dict['Net\nSell_2'] + "7:" + column_dict['Net\nSell_2'] + "8", 
                    ["Dilution"], [ 0.14], subheader_bold, subheader_font, alignment_left, border)
    ws[column_dict['Net\nSell_2'] + "8"].number_format = '0%'

    apply_formatting(ws, column_dict['IMU_2'] + "7:" + column_dict['IMU_2'] + "8", 
                    ["Cust IMU%"], [0.5], subheader_bold, subheader_font, alignment_left, border)
    ws[column_dict['IMU_2'] + "8"].number_format = '0%'

    apply_formatting(ws, column_dict['AUR\nDiscount_2'] + "7:" + column_dict['AUR\nDiscount_2'] + "8", 
                    ["Tkt % Off"], [0.25], subheader_bold, subheader_font, alignment_left, border)
    ws[column_dict['AUR\nDiscount_2'] + "8"].number_format = '0%'
    apply_formatting(ws, column_dict['1st MKD_2'] + "7:" + column_dict['1st MKD_2'] + "8", 
                    ["MKD %"], [0.2], subheader_bold, subheader_font, alignment_left, border)
    ws[column_dict['1st MKD_2'] + "8"].number_format = '0%'
    apply_formatting(ws, column_dict['Sell_2'] + "7:" + column_dict['Sell_2'] + "8", 
                    ["Sell Int."], [0.25], subheader_bold, subheader_font, alignment_left, border, dollar=True)

    apply_formatting(ws, column_dict['MSRP_2'] + "7:" + column_dict['MSRP_2'] + "8", 
                    ["MSRP Int."], [25.00], subheader_bold, subheader_font, alignment_left, border, dollar=True)

    apply_formatting(ws, column_dict['Box'] + "7:" + column_dict['Box'] + "8", 
                    ["Box"], [0], subheader_bold, subheader_font, alignment_left, border, dollar=True)

    apply_formatting(ws, column_dict['Cert'] + "7:" + column_dict['Cert'] + "8", 
                    ["Cert"], [0], subheader_bold, subheader_font, alignment_left, border, dollar=True)


    # Place each column name in the appropriate cell in row 13
    for i, column_name in enumerate(column_names, start=1):  # start=1 corresponds to column 'A'
        cell=ws.cell(row=13, column=i, value=column_name)
        cell.font = header_font
        cell.alignment = alignment_left
        cell.border = border
    
    
    ws.row_dimensions[13].height = 45
    
    
    
    
    start_row = 14  # Adjust this to the row where you want to start inserting images
    
    r = 14
    for i in range(len(single_df)):
        # Get the original item name from single_df
        original_item_name = single_df['Itemno'].iloc[i]
        
        # Retrieve the hashed filename from the mapping dictionary
        hashed_filename = name_mapping.get(original_item_name)
        
        if hashed_filename:
            # Construct the image file path using the hashed filename
            image_path = os.path.join(OUTPUT_IMAGE_DIR, f"{hashed_filename}.png")
            
            # Check if the image file exists
            if os.path.exists(image_path):
                img_pil = PILImage.open(image_path)
                original_width, original_height = img_pil.size

                # Desired maximum width and height to fit in the cell
                max_width, max_height = 150, 150

                # Calculate new dimensions while preserving the aspect ratio
                aspect_ratio = original_width / original_height

                if aspect_ratio > 1:  # Wider image (landscape)
                    new_width = min(max_width, original_width)
                    new_height = int(new_width / aspect_ratio)
                else:  # Taller image (portrait) or square
                    new_height = min(max_height, original_height)
                    new_width = int(new_height * aspect_ratio)

                # Load image and apply the new dimensions
                img = Image(image_path)
                img.width, img.height = new_width, new_height

                # Add image using the column for "IMAGE" from the column_dict
                ws.add_image(img, column_dict["IMAGE"] + str(r))
                print(f"Added image for item {original_item_name} at {column_dict['IMAGE']}{r}")
            else:
                print(f"Image file {hashed_filename}.png not found in directory.")
        else:
            print(f"No mapping found for item {original_item_name}")
        # Mapping the data to the worksheet using column_dict
        ws[column_dict["STYLE #"] + str(r)] = single_df["Itemno"].iloc[i]
        ws[column_dict["DESCRIPTION"] + str(r)] = single_df["dbo_NAV_Item_Master.Description"].iloc[i]
        ws[column_dict["Item\nStatus\nCode"] + str(r)] = single_df["Item Status Code"].iloc[i]
        ws[column_dict["Manf.\nCode"] + str(r)] = single_df["Vendor No_"].iloc[i]
        ws[column_dict["COO"] + str(r)] = single_df["Country Of Origin Code"].iloc[i]
        ws[column_dict["METAL"] + str(r)] = str(single_df["Metal Quality Code"].iloc[i])
        ws[column_dict["Wt."] + str(r)] = single_df["Fin Metal weight"].iloc[i]
        # if single_df
        if "Manufacturing Policy" in single_df.columns:
            if single_df["Manufacturing Policy"].iloc[i]=='Make-to-Stock':
                ws[column_dict["Mnf.\nPolicy"] + str(r)] = 'MTS'
            elif single_df["Manufacturing Policy"].iloc[i]=='Assembly-to-order':
                ws[column_dict["Mnf.\nPolicy"] + str(r)] = 'ATO'
            else:
                ws[column_dict["Mnf.\nPolicy"] + str(r)] = single_df["Manufacturing Policy"].iloc[i]

        else:
            ws[column_dict["Mnf.\nPolicy"] + str(r)] = None
        if "AvgOfDuty %" in single_df.columns:
            ws[column_dict["Duty %"] + str(r)] = single_df["AvgOfDuty %"].iloc[i]
            ws[column_dict["Duty %"] + str(r)] = f"{single_df['AvgOfDuty %'].iloc[i]}%"


        else:
            ws[column_dict["Duty %"] + str(r)] = None
            # ws[column_dict["Duty %"] + str(r)] = f"{single_df['AvgOfDuty %'].iloc[i]}%"


        # Formula for Metal Cost Per Gram using dynamic column and row reference
        ws[column_dict["Metal\nCost\n /Gm"]+str(r)]='=IFERROR(VLOOKUP(' + column_dict["METAL"] + str(r) + '&"", ' + column_dict["Metal\nCost\n /Gm"] + '$1:' + column_dict["METAL\nCOST"] + '$4, 2, 0), "")'
        # Formula for Metal Cost (wt. * Metal Cost Per Gram)
        ws[column_dict["METAL\nCOST"] + str(r)] = "=ROUND(" + column_dict["Wt."] + str(r) + "*" + column_dict["Metal\nCost\n /Gm"] + str(r) + ", 2)"

        
        # Set the current row for further reference
        c_r = r
        fin_c = r
        # Filter the metal_df for the current item and add 'inhouse_or_not' based on the 'Supply Policy'
        df_i = metal_df[metal_df.Itemno == single_df["Itemno"].iloc[i]]
        df_i['inhouse_or_not'] = ['YES' if policy == 'IH Assembly' else 'NO' for policy in df_i['Supply Policy']]
        s_q = []
        quantity_column = 'AvgOfQuantity' if 'AvgOfQuantity' in df_i.columns else 'Quantity' 
        print("df_i",quantity_column)
        for index, row in df_i.iterrows():

            if row['Item Category Code'] == "DIAM": 

                ws[column_dict["STONE\nTYPE"] + str(c_r)] = row['Item Category Code']
                ws[column_dict["QTY"] + str(c_r)] = row[quantity_column]
                s_q.append(row[quantity_column])
                ws[column_dict["wt."] + str(c_r)] = row['AvgOfWeight']
                ws[column_dict["CODE"] + str(c_r)] = row['Item No_']   
                ws[column_dict["$ PER \nCT/PC"] + str(c_r)] = row['AvgOfBase Unit Cost (LCY)']

                try:
                    ws[column_dict["SHAPE"] + str(c_r)] = shape_qty_df[shape_qty_df["Item No_"] == row['Item No_']]["SHAPE"].iloc[0]
                    ws[column_dict["Quality"] + str(c_r)] = shape_qty_df[shape_qty_df["Item No_"] == row['Item No_']]["QUALITY"].iloc[0]
                    ws[column_dict["Color\nMM\nSize"] + str(c_r)] = shape_qty_df[shape_qty_df["Item No_"] == row['Item No_']]["MM SIZE"].iloc[0]
                except:
                    pass

                ws[column_dict["WT.\n Ext."] + str(c_r)] = "=" + column_dict["wt."] + str(c_r) + "*" + column_dict["QTY"] + str(c_r)
                ws[column_dict["COST"] + str(c_r)] = "=" + column_dict["$ PER \nCT/PC"] + str(c_r) + "*" + column_dict["WT.\n Ext."] + str(c_r)
                ws[column_dict["IH\n (YES/NO)"] + str(c_r)] = row["inhouse_or_not"]
                if SETTING_DESC.get(row['Setting Type Code']):
                    ws[column_dict["SETTING\nDESCRIPTION"]+str(c_r)]=SETTING_DESC.get(row['Setting Type Code'])
                else:
                    reversed_dict = dict(zip(SETTING_TYPE_DIC.values(), SETTING_TYPE_DIC.keys()))
                    ws[column_dict["SETTING\nDESCRIPTION"]+str(c_r)]=SETTING_DESC.get(reversed_dict.get(row['Setting Type Code']))
                    

                c_r += 1

            if row['Item Category Code'] in ["FIN", "OTHER", "SUBC"]: 
                ws[column_dict["FINDING IH COST\nOR NOT\n(YES/NO)"] + str(fin_c)] = row["inhouse_or_not"]
                ws[column_dict["FINDING\nTYPE"] + str(fin_c)] = row['Item No_']
                
                ws[column_dict["FINDING\n(YES/NO)"] + str(fin_c)]='=IF(ISNUMBER(SEARCH("CERT", ' + column_dict["FINDING\nTYPE"] + str(fin_c) + ')), "YES", "NO")'
                ws[column_dict["FINDING\nQTY"] + str(fin_c)] = row[quantity_column]
            
# ( "=IF("+ column_dict["METAL"] + str(r) + '="S", '  # Check if the Metal column value is "S"
#     + str(row['AvgOfBase Unit Cost (LCY)']) + "/40*40, "  # If "S", divide by 40 and multiply by 40
#     + str(row['AvgOfBase Unit Cost (LCY)']) + "/2200*2600)" 
 
                ws[column_dict["FINDING\nDTL\nCOST"] + str(fin_c)] = ( "=IF("+ column_dict["METAL"] + str(r) + '="S", '  # Check if the Metal column value is "S"
    + str(row['AvgOfCost Amount (LCY)']) + "/40*$"+column_dict["METAL\nCOST"]+"$7"+", "  # If "S", divide by 40 and multiply by 40
    + str(row['AvgOfCost Amount (LCY)']) + "/2200*$"+column_dict["METAL\nCOST"]+"$8)"  )
                
                fin_c += 1

            if row['Item Category Code'] in ["COLOR", "CZ"]:  
                ws[column_dict["STONE\nTYPE"] + str(c_r)] = row['Item Category Code']
                ws[column_dict["QTY"] + str(c_r)] = row[quantity_column]
                s_q.append(row[quantity_column])
                ws[column_dict["wt."] + str(c_r)] = row['AvgOfWeight']
                ws[column_dict["CODE"] + str(c_r)] = row['Item No_']
                ws[column_dict["$ PER \nCT/PC"] + str(c_r)] = row['AvgOfBase Unit Cost (LCY)']

                try:
                    ws[column_dict["SHAPE"] + str(c_r)] = shape_qty_df[shape_qty_df["Item No_"] == row['Item No_']]["SHAPE"].iloc[0]
                    ws[column_dict["Quality"] + str(c_r)] = shape_qty_df[shape_qty_df["Item No_"] == row['Item No_']]["QUALITY"].iloc[0]
                    ws[column_dict["Color\nMM\nSize"] + str(c_r)] = shape_qty_df[shape_qty_df["Item No_"] == row['Item No_']]["MM SIZE"].iloc[0]
                except:
                    pass
                
                ws[column_dict["WT.\n Ext."] + str(c_r)] = "=" + column_dict["wt."] + str(c_r) + "*" + column_dict["QTY"] + str(c_r)
                ws[column_dict["COST"] + str(c_r)] = row['AvgOfCost Amount (LCY)']
                ws[column_dict["IH\n (YES/NO)"] + str(c_r)] = row["inhouse_or_not"]
                if SETTING_DESC.get(row['Setting Type Code']):
                    
                    ws[column_dict["SETTING\nDESCRIPTION"]+str(c_r)]=SETTING_DESC.get(row['Setting Type Code'])
                else:
                    reversed_dict = dict(zip(SETTING_TYPE_DIC.values(), SETTING_TYPE_DIC.keys()))
               
                    ws[column_dict["SETTING\nDESCRIPTION"]+str(c_r)]=SETTING_DESC.get(reversed_dict.get(row['Setting Type Code']))
                c_r += 1

        labour_df['visited'] = 0

        # DIA CTTW
        ws[column_dict["DIA CTTW"] + str(r)] = (
    "=SUMIF("
    + column_dict["STONE\nTYPE"] + str(r) + ":" + column_dict["STONE\nTYPE"] + str(c_r)  # Condition range
    + ',"DIAM",'  # Criteria: "diam"
    + column_dict["WT.\n Ext."] + str(r) + ":" + column_dict["WT.\n Ext."] + str(c_r)  # Sum range
    + ")"
)
        
        # TOTAL CS COST
        ws[column_dict["TOTAL\nCS\nCOST"] + str(r)] = (
            "=SUMIF(" + column_dict["STONE\nTYPE"] + str(r) + ":" + column_dict["STONE\nTYPE"] + str(c_r) + ',"COLOR",' + column_dict["COST"] + str(r) + ":" + column_dict["COST"] + str(c_r) + ")"
            " + SUMIF(" + column_dict["STONE\nTYPE"] + str(r) + ":" + column_dict["STONE\nTYPE"] + str(c_r) + ',"CZ",' + column_dict["COST"] + str(r) + ":" + column_dict["COST"] + str(c_r) + ")"
        )

        # CS IH COST
        ws[column_dict["CS\nIH\nCOST"] + str(r)] = (
            "=SUMIFS(" + column_dict["COST"] + str(r) + ":" + column_dict["COST"] + str(c_r) + ", " + column_dict["STONE\nTYPE"] + str(r) + ":" + column_dict["STONE\nTYPE"] + str(c_r) + ', "COLOR", ' + column_dict["IH\n (YES/NO)"] + str(r) + ":" + column_dict["IH\n (YES/NO)"] + str(c_r) + ', "YES") '
            "+ SUMIFS(" + column_dict["COST"] + str(r) + ":" + column_dict["COST"] + str(c_r) + ", " + column_dict["STONE\nTYPE"] + str(r) + ":" + column_dict["STONE\nTYPE"] + str(c_r) + ', "CZ", ' + column_dict["IH\n (YES/NO)"] + str(r) + ":" + column_dict["IH\n (YES/NO)"] + str(c_r) + ', "YES")'
        )

        # TOTAL DIA COST
        ws[column_dict["TOTAL\nDIA\nCOST"] + str(r)] = "=SUMIF(" + column_dict["STONE\nTYPE"] + str(r) + ":" + column_dict["STONE\nTYPE"] + str(c_r) + ',"DIAM",' + column_dict["COST"] + str(r) + ":" + column_dict["COST"] + str(c_r) + ")"

        # DIA IH COST
        ws[column_dict["DIA\nIH\nCOST"] + str(r)] = (
            "=SUMIFS(" + column_dict["COST"] + str(r) + ":" + column_dict["COST"] + str(c_r) + ", " + column_dict["STONE\nTYPE"] + str(r) + ":" + column_dict["STONE\nTYPE"] + str(c_r) + ', "DIAM", ' + column_dict["IH\n (YES/NO)"] + str(r) + ":" + column_dict["IH\n (YES/NO)"] + str(c_r) + ', "YES")'
        )

        # TTL STONE COST
        ws[column_dict["TOTAL\nSTONE\nCOST"] + str(r)] = "=SUM(" + column_dict["TOTAL\nCS\nCOST"] + str(r) + "," + column_dict["TOTAL\nDIA\nCOST"] + str(r) + ")"

        # STONE IH COST
        ws[column_dict["STONE\nIH\nCOST"] + str(r)] = "=SUM(" + column_dict["CS\nIH\nCOST"] + str(r) + "," + column_dict["DIA\nIH\nCOST"] + str(r) + ")"

        
    
        
        ll=labour_df[(labour_df.Quantity!=0)&(labour_df.Itemno==single_df["Itemno"].iloc[i])&(labour_df["Unit Cost (LCY)"]!=0)&(labour_df.visited==0)]
        
        # ll[(ll["Quantity"]==1) & (ll["Operation Code"]=="SET")][:1]
        # ws[["SET\nCOST"]+str(r)]
        l_c=r
        count_s=0
    
        for q in s_q:       
            # print(ll[(ll["Quantity"]==q) & (ll["Operation Code"]=="SET")&(ll.visited!=1)])
            pick_df=ll[(ll["Quantity"]==q) & ((ll["Operation Code"]=="SET")|(ll["Operation Code"].isin(op_li)) )&(ll.visited!=1)][:1]
            next_df=ll[(ll["Quantity"]==q) & (ll["Operation Code"].isin(op_li)) &(ll.visited!=1)][:1] 
            ll.loc[ll[(ll["Quantity"]==q) &((ll["Operation Code"]=="SET")|(ll["Operation Code"].isin(op_li)) ) &(ll.visited!=1)][:1].index, 'visited']=1
            pick_df=pick_df[pick_df["Unit Cost (LCY)"]>0]
            if len(pick_df)>0:
                ws[column_dict["SET\nCOST"]+str(l_c)]=pick_df["Unit Cost (LCY)"].iloc[0]
                ws[column_dict["SET\nCOST"]+str(l_c)].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
                ws[column_dict["EXT\nSET\nCOST"] + str(l_c)] = '=' + column_dict["SET\nCOST"] + str(l_c) + '*' + column_dict["QTY"] + str(l_c)

                ws[column_dict["EXT\nSET\nCOST"]+str(l_c)].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
                count_s=count_s+1
                if len(next_df)>0:
                    if SETTING_DESC.get(SETTING_TYPE_DIC[next_df['Operation Code'].iloc[0]]):
                        ws[column_dict["SETTING\nDESCRIPTION"]+str(l_c)]=SETTING_DESC.get(SETTING_TYPE_DIC[next_df['Operation Code'].iloc[0]])
                    else:
                        ws[column_dict["SETTING\nDESCRIPTION"]+str(l_c)]=SETTING_DESC.get(next_df['Operation Code'].iloc[0])
                                        
                
            else:
                ws[column_dict["SETTING\nDESCRIPTION"]+str(l_c)]=None
                
            
        
            l_c=l_c+1

        # print( count_s)
        labour_c=r
        if count_s!=0:
            ws[column_dict["LABOR TYPE"]+str(labour_c)]="Setting Cost"
            ws[column_dict["Labour IH COST\nOR NOT\n(YES/NO)"]+str(labour_c)]="NO"
            # LABOR QTY (summing the QTY column)
            ws[column_dict["LABOR\nQTY"] + str(labour_c)] = '=SUM(' + column_dict["QTY"] + str(r) + ':' + column_dict["QTY"] + str(l_c) + ')'

            # LABOR DTL COST (summing the EXT SETTING COST column)
            ws[column_dict["LABOR\nDTL\nCOST"] + str(labour_c)] = '=SUM(' + column_dict["EXT\nSET\nCOST"] + str(r) + ':' + column_dict["EXT\nSET\nCOST"] + str(l_c) + ')'

            
            labour_c=labour_c+1
        ll=ll[ll['visited']==0]
        SumOfTotal_Cost_LCY = 'SumOfTotal Cost (LCY)' if 'SumOfTotal Cost (LCY)' in ll.columns else 'Total Cost (LCY)'
        for Description,Quantity,UC in zip(list(ll['Description']),ll['Quantity'],ll[SumOfTotal_Cost_LCY]):
            
            ws[column_dict["LABOR TYPE"]+str(labour_c)]=Description
            ws[column_dict["Labour IH COST\nOR NOT\n(YES/NO)"]+str(labour_c)]='=IF(ISNUMBER(SEARCH("NY", ' + column_dict["LABOR TYPE"] + str(labour_c) + ')), "YES", "NO")'

            ws[column_dict["LABOR\nQTY"]+str(labour_c)]=Quantity
            ws[column_dict['LABOR\nDTL\nCOST']+str(labour_c)]=UC
            
            labour_c=labour_c+1
    
        ws[column_dict['LABOR\nTTL\nCOST'] + str(r)] = '=SUM(' + column_dict["LABOR\nDTL\nCOST"] + str(r) + ':' + column_dict["LABOR\nDTL\nCOST"] + str(labour_c) + ')'
        ws[column_dict["LABOR\nIH\nCOST"] + str(r)]="=SUMIF(" + column_dict["Labour IH COST\nOR NOT\n(YES/NO)"] + str(r) + ":" + column_dict["Labour IH COST\nOR NOT\n(YES/NO)"] + str(labour_c) + ',"YES",' + column_dict['LABOR\nDTL\nCOST'] + str(r) + ":" + column_dict['LABOR\nDTL\nCOST'] + str(labour_c) + ")"
        ws[column_dict["FINDING\nTTL\nCOST"] + str(r)] = '=SUM(' + column_dict["FINDING\nDTL\nCOST"] + str(r) + ':' + column_dict["FINDING\nDTL\nCOST"] + str(fin_c) + ')'
        ws[column_dict["FINDING\nIH COST"]+ str(r)]="=SUMIF(" + column_dict["FINDING IH COST\nOR NOT\n(YES/NO)"] + str(r) + ":" + column_dict["FINDING IH COST\nOR NOT\n(YES/NO)"] + str(fin_c) + ',"YES",' + column_dict["FINDING\nDTL\nCOST"] + str(r) + ":" + column_dict["FINDING\nDTL\nCOST"] + str(fin_c) + ")"
        # ws[column_dict["FINDING\nCERT COST"]+str(r)]="=SUMIF(" + column_dict["FINDING\n(YES/NO)"] + str(r) + ":" + column_dict["FINDING\n(YES/NO)"] + str(fin_c) + ',"YES",' + column_dict["FINDING\nDTL\nCOST"] + str(r) + ":" + column_dict["FINDING\nDTL\nCOST"] + str(fin_c) + ")"
        # ws[column_dict["STYLE\nTOTAL"] + str(r)] = ('=SUM(' + column_dict["METAL\nCOST"] + str(r) + ':' + column_dict["METAL\nCOST"] + str(labour_c) + ',' +column_dict["TOTAL\nCS\nCOST"] + str(r) + ',' + column_dict["CS\nIH\nCOST"] + str(r) + ',' + column_dict["TOTAL\nDIA\nCOST"] + str(r) + ',' + column_dict["LABOR\nTTL\nCOST"] + str(r) + ',' + column_dict["FINDING\nTTL\nCOST"] + str(r) + ')')+"-"+column_dict["FINDING\nCERT COST"] + str(r)
        # ws[column_dict["STYLE\nTOTAL\nEXCL.IH"]+ str(r)]="="+column_dict["STYLE\nTOTAL"] + str(r)+"-"+column_dict["FINDING\nIH COST"]+ str(r)+"-"+column_dict["LABOR\nIH\nCOST"]+ str(r)+"-"+column_dict["DIA\nIH\nCOST"]+ str(r)+"-"+column_dict["CS\nIH\nCOST"]+ str(r)
        

        # ws[column_dict["S/H"] + str(r)] = '=' + column_dict["STYLE\nTOTAL\nEXCL.IH"] + str(r) + '*0.02'
        # ws[column_dict["Duty $"] + str(r)] = '=' + column_dict["STYLE\nTOTAL\nEXCL.IH"] + str(r) + '*0.06'

                # STYLE TOTAL
        ws[column_dict["STYLE\nTOTAL"] + str(r)] = (
            '=ROUND(SUM(' + column_dict["METAL\nCOST"] + str(r) + ':' + column_dict["METAL\nCOST"] + str(labour_c) + ',' 
            + column_dict["TOTAL\nCS\nCOST"] + str(r) + ',' + column_dict["CS\nIH\nCOST"] + str(r) + ',' 
            + column_dict["TOTAL\nDIA\nCOST"] + str(r) + ',' + column_dict["LABOR\nTTL\nCOST"] + str(r) + ',' 
            + column_dict["FINDING\nTTL\nCOST"] + str(r) + '), 2)'
        )
        
        # STYLE TOTAL EXCL. IH
        ws[column_dict["STYLE\nTOTAL\nEXCL.IH"]+ str(r)] = (
            '=ROUND(' + column_dict["STYLE\nTOTAL"] + str(r) + ', 2) - ROUND(' + column_dict["FINDING\nIH COST"] + str(r) + ', 2) - '
            'ROUND(' + column_dict["LABOR\nIH\nCOST"] + str(r) + ', 2) - ROUND(' + column_dict["DIA\nIH\nCOST"] + str(r) + ', 2) - '
            'ROUND(' + column_dict["CS\nIH\nCOST"] + str(r) + ', 2)'
        )
        
        # S/H
        ws[column_dict["S/H"] + str(r)] = (
            '=ROUND(' + column_dict["STYLE\nTOTAL\nEXCL.IH"] + str(r) + '*0.02, 2)'
        )
        # Duty $
        ws[column_dict["Duty $"] + str(r)] = (
            '=ROUND(' + column_dict["STYLE\nTOTAL\nEXCL.IH"] + str(r) + '*' + column_dict["Duty %"] + str(r) + ', 2)'
        )



    

        
       
        ws[column_dict["LANDED\nTOTAL"] + str(r)] = ('=SUM(' + column_dict["STYLE\nTOTAL"] + str(r) + ',' +    column_dict["S/H"] + str(r) + ',' +   column_dict["Duty $"] + str(r)  + ')')

        # ws[column_dict["LANDED\nTOTAL"] + str(r)].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

        ws[column_dict["Box"] + str(r)] = '=$' + column_dict["Box"] + '$8'
        ws[column_dict["Cert"] + str(r)] = '=$' + column_dict["Cert"] + '$8'


        ws[column_dict["Target\nGM%"] + str(r)] = '=$' + column_dict["Target\nGM%"] + '$8'
        ws[column_dict["Target\nGM%"] +  str(r)].number_format = '0.0%'
        ws[column_dict["Sell_1"] + str(r)] = '=ROUNDUP(' + column_dict["LANDED\nTOTAL"] + str(r) + '/(1-' + column_dict["Target\nGM%"] + str(r) + ')/$' + column_dict["Sell_2"] + '$8,0)*$' + column_dict["Sell_2"] + '$8'
        ws[column_dict["Gross\nMargin_1"] + str(r)] = '=(' + column_dict["Sell_1"] + str(r) + '-' + column_dict["LANDED\nTOTAL"] + str(r) + ')/' + column_dict["Sell_1"] + str(r)
        ws[column_dict["Gross\nMargin_1"] + str(r)].number_format = '0%'

        ws[column_dict["Net\nSell_1"] + str(r)] = '=' + column_dict["Sell_1"] + str(r) + '*(1-$' + column_dict["Net\nSell_2"] + '$8)'
        ws[column_dict["Net\nMargin_1"] + str(r)] = '=(' + column_dict["Net\nSell_1"] + str(r) + '-' + column_dict["LANDED\nTOTAL"] + str(r) + ')/' + column_dict["Net\nSell_1"] + str(r)
        ws[column_dict["Net\nMargin_1"] + str(r)].number_format = '0.0%'

        ws[column_dict["MSRP_1"] + str(r)] = '=ROUNDUP(' + column_dict["Sell_1"] + str(r) + '/(1-$' + column_dict["IMU_2"] + '$8)/$' + column_dict["MSRP_2"] + '$8,0)*$' + column_dict["MSRP_2"] + '$8'

        ws[column_dict["IMU_1"] + str(r)] = '=((' + column_dict["MSRP_1"] + str(r) + '-' + column_dict["Sell_1"] + str(r) + ')/' + column_dict["MSRP_1"] + str(r) + ')'
        ws[column_dict["IMU_1"] + str(r)].number_format = '0.0%'

        ws[column_dict["AUR_1"] + str(r)] = '=(' + column_dict["MSRP_1"] + str(r) + '*(1-$' + column_dict["AUR\nDiscount_2"] + '$8))'
        ws[column_dict["AUR\nDiscount_1"] + str(r)] = '=(' + column_dict["MSRP_1"] + str(r) + '-' + column_dict["AUR_1"] + str(r) + ')/' + column_dict["MSRP_1"] + str(r)
        ws[column_dict["AUR\nDiscount_1"] + str(r)].number_format = '0%'

        ws[column_dict["AUR\nGM%_1"] + str(r)] = '=(' + column_dict["AUR_1"] + str(r) + '-' + column_dict["Sell_1"] + str(r) + ')/' + column_dict["AUR_1"] + str(r)
        ws[column_dict["AUR\nGM%_1"] + str(r)].number_format = '0%'

        ws[column_dict["1st MKD_1"] + str(r)] = '=' + column_dict["Sell_1"] + str(r) + '*(1-$' + column_dict["1st MKD_2"] + '$8)'
        ws[column_dict["MKD NM%_1"] + str(r)] = '=((' + column_dict["1st MKD_1"] + str(r) + '*(1-$' + column_dict["Net\nSell_2"] + '$8))- ' + column_dict["LANDED\nTOTAL"] + str(r) + ')/(' + column_dict["1st MKD_1"] + str(r) + '*(1-$' + column_dict["Net\nSell_2"] + '$8))'
        ws[column_dict["MKD NM%_1"] + str(r)].number_format = '0%'

        ws[column_dict["Sell_2"] + str(r)] = '=' + column_dict["LANDED\nTOTAL"] + str(r) + '/(1-$' + column_dict["Target\nGM%"] + '$8)'
        ws[column_dict["Gross\nMargin_2"] + str(r)] = '=(' + column_dict["Sell_2"] + str(r) + '-' + column_dict["LANDED\nTOTAL"] + str(r) + ')/' + column_dict["Sell_2"] + str(r)
        ws[column_dict["Gross\nMargin_2"] + str(r)].number_format = '0%'

        ws[column_dict["Net\nSell_2"] + str(r)] = '=' + column_dict["Sell_2"] + str(r) + '*(1-$' + column_dict["Net\nSell_2"] + '$8)'
        ws[column_dict["Net Margin_2"] + str(r)] = '=(' + column_dict["Net\nSell_2"] + str(r) + '-' + column_dict["LANDED\nTOTAL"] + str(r) + ')/' + column_dict["Net\nSell_2"] + str(r)
        ws[column_dict["Net Margin_2"] + str(r)].number_format = '0.0%'

        ws[column_dict["MSRP_2"] + str(r)] = '=' + column_dict["Sell_2"] + str(r) + '/(1-$' + column_dict["IMU_2"] + '$8)'
        ws[column_dict["IMU_2"] + str(r)] = '=(' + column_dict["MSRP_2"] + str(r) + '-' + column_dict["Sell_2"] + str(r) + ')/' + column_dict["MSRP_2"] + str(r)
        ws[column_dict["IMU_2"] + str(r)].number_format = '0.0%'

        ws[column_dict["AUR_2"] + str(r)] = '=' + column_dict["MSRP_2"] + str(r) + '*(1-$' + column_dict["AUR\nDiscount_2"] + '$8)'
        ws[column_dict["AUR\nDiscount_2"] + str(r)] = '=(' + column_dict["MSRP_2"] + str(r) + '-' + column_dict["AUR_2"] + str(r) + ')/' + column_dict["MSRP_2"] + str(r)
        ws[column_dict["AUR\nDiscount_2"] + str(r)].number_format = '0%'

        ws[column_dict["AUR\nGM%_2"] + str(r)] = '=(' + column_dict["AUR_2"] + str(r) + '-' + column_dict["Sell_2"] + str(r) + ')/' + column_dict["AUR_2"] + str(r)
        ws[column_dict["AUR\nGM%_2"] + str(r)].number_format = '0%'

        ws[column_dict["1st MKD_2"] + str(r)] = '=' + column_dict["Sell_2"] + str(r) + '*(1-$' + column_dict["1st MKD_2"] + '$8)'
        ws[column_dict["MKD NM%_2"] + str(r)] = '=((' + column_dict["1st MKD_2"] + str(r) + '*(1-$' + column_dict["Net\nSell_2"] + '$8))- ' + column_dict["LANDED\nTOTAL"] + str(r) + ')/(' + column_dict["1st MKD_2"] + str(r) + '*(1-$' + column_dict["Net\nSell_2"] + '$8))'
        ws[column_dict["MKD NM%_2"] + str(r)].number_format = '0%'
                
    
        non_percentage_columns = [
        "Sell_1",
        "Net\nSell_1",
        "MSRP_1",
        "AUR_1",
        "1st MKD_1",
        "Sell_2",
        "Net\nSell_2",
        "MSRP_2",
        "AUR_2",
        "1st MKD_2"
    ]
        for n_p in non_percentage_columns:
            ws[column_dict[n_p] + str(r)].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
        
        thick_border = Border(bottom=Side(style='thick'))
        
        
    # Specify the row where you want to draw the line
        row_to_draw_line = max(labour_c,l_c,r+8)  # Replace with the row number where you want the line
        # Specify the last column (e.g., "BU")
        last_column = column_dict["MKD NM%_2"] 
        
        # Apply the border to all cells from the first column to the specified last column
        for col in range(1, ws[last_column + '1'].column + 1):
            cell = ws.cell(row=row_to_draw_line, column=col)
            cell.border = thick_border
    
        r= max(labour_c+1,l_c+1,r+9)
            
    
    for column in ws.columns:
        
        column_letter = get_column_letter(column[0].column)
        max_width = 9
        for cell in column:
            font = cell.font
            font_size = font.size or ws.sheet_format.baseColWidth
    
            if cell.value:
                if cell.font.bold:
                    cell_width = get_column_width(cell)*1.15*font_size
                else:
                    cell_width = (get_column_width(cell)-0.1)*font_size
                    


                
                if cell_width > max_width:
                    max_width = cell_width
        
        ws.column_dimensions[column_letter].width = max_width

    # Save the file
    ws.column_dimensions[column_dict["METAL\nCOST"]].width = 10 + 11 * 0.08 * 2
    ws.merge_cells('A6:C6')#NTC
    # Set value for merged cells
    ws[column_dict["IMAGE"] + '6'] = "RICHLINE JEWELERY GROUP"

    # Apply font and alignment to the merged cell
    ws[column_dict["IMAGE"] + '6'].font = header_font
    ws[column_dict["IMAGE"] + '6'].alignment = alignment_left
    
    gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    orange_fill =PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    blue_fill = PatternFill(start_color="99CCFF", end_color="99CCFF", fill_type="solid")
    gold_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")  # Gold color
    light_yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    desired_width = 20  # Adjust this value as needed
    # Iterate over the columns from AZ to BU and set the width
    for col in range(ws['BJ1'].column, ws['BY1'].column + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = min(ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width,10+11*2)#NTC
    
    thick_side = Side(border_style="thick", color="000000")

    top_border = Border(top=thick_side)

    # Get the range of cells you want to apply the fill to
    start_col = column_dict["STONE\nTYPE"]
    end_col = column_dict["SETTING\nDESCRIPTION"]

    # Assuming row 12 is to be formatted
    start_cell = f"{start_col}12"
    end_cell = f"{end_col}12"

    # Apply the white fill to the specified range
    for col in ws[start_cell:end_cell]:
        for cell in col:
            cell.fill = gray_fill
    # Dynamically assign value, font, alignment, and fill to the "STONE\nTYPE" column (which corresponds to Q)
    ws[column_dict["Quality"] + '12'] = "STONE BREAKDOWN"
    ws[column_dict["Quality"] + '12'].font = header_font
    ws[column_dict["Quality"] + '12'].alignment = alignment_center

    ws[column_dict["EXT\nSET\nCOST"] + '12'] = "SETTING BREAKDOWN"
    ws[column_dict["EXT\nSET\nCOST"] + '12'].font = header_font
    ws[column_dict["EXT\nSET\nCOST"] + '12'].alignment = alignment_center
    for row in ws[column_dict["STONE\nTYPE"] + '12:' + column_dict["SETTING\nDESCRIPTION"] + '12']:
        for cell in row:
            cell.border = top_border
    ws[column_dict["STONE\nTYPE"] + '12'].border = Border(left=thick_side,top=thick_side)
    ws[column_dict["SET\nCOST"] + '12'].border = Border(left=thick_side,top=thick_side)
    ws[column_dict["SETTING\nDESCRIPTION"] + '12'].border = Border(right=thick_side,top=thick_side)

    start_col = column_dict["Sell_1"]
    end_col = column_dict["MKD NM%_1"]

    # Assuming row 12 is to be formatted
    start_cell = f"{start_col}12"
    end_cell = f"{end_col}12"

    # Apply the white fill to the specified range
    for col in ws[start_cell:end_cell]:
        for cell in col:
            cell.fill = gold_fill

    ws[column_dict["IMU_1"] + '12'] = "DEFAULT PRICING - GROUP IT CLOSE"
    ws[column_dict["IMU_1"] + '12'].font = header_font
    ws[column_dict["IMU_1"] + '12'].alignment = alignment_center
    for row in ws[column_dict["Sell_1"] + '12:' + column_dict["MKD NM%_1"] + '12']:
        for cell in row:
            cell.border = top_border
    ws[column_dict["Sell_1"] + '12'].border = Border(left=thick_side,top=thick_side)
    ws[column_dict["MKD NM%_1"] + '12'].border = Border(right=thick_side,top=thick_side)
    start_col = column_dict["Sell_2"]
    end_col = column_dict["MKD NM%_2"]

    # Assuming row 12 is to be formatted
    start_cell = f"{start_col}12"
    end_cell = f"{end_col}12"

    # Apply the white fill to the specified range
    for col in ws[start_cell:end_cell]:
        for cell in col:
            cell.fill = green_fill

    # Dynamically assign value, font, alignment, and fill to the "Sell_2" column (which corresponds to BU)
    ws[column_dict["IMU_2"] + '12'] = "PRICING REVIEW"
    ws[column_dict["IMU_2"] + '12'].font = header_font
    ws[column_dict["IMU_2"] + '12'].alignment = alignment_center

    for row in ws[column_dict["Sell_2"] + '12:' + column_dict["MKD NM%_2"] + '12']:
        for cell in row:
            cell.border = top_border

    ws[column_dict["Sell_2"] + '12'].border = Border(left=thick_side,top=thick_side)
    ws[column_dict["MKD NM%_2"] + '12'].border = Border(right=thick_side,top=thick_side)
    # Dynamically create formulas using column_dict
    # ws[column_dict["LANDED\nTOTAL"] + '11'] = '=SUBTOTAL(9,' + column_dict["LANDED\nTOTAL"] + '14:' + column_dict["LANDED\nTOTAL"] + '1379)'

    ws[column_dict["LANDED\nTOTAL"] + '11'] = (
    '=ROUND(SUBTOTAL(9,' + column_dict["LANDED\nTOTAL"] + '14:' + column_dict["LANDED\nTOTAL"] + '1379),0)'
)
    ws[column_dict["LANDED\nTOTAL"] + '11'].number_format = '"$"#,##0'

    
    # ws[column_dict["Sell_1"] + '11'] = '=SUBTOTAL(9,' + column_dict["Sell_1"] + '14:' + column_dict["Sell_1"] + '1379)'
    ws[column_dict["Sell_1"] + '11'] = (
    '=INT(SUBTOTAL(9,' + column_dict["Sell_1"] + '14:' + column_dict["Sell_1"] + '1379))'
)

    ws[column_dict["Gross\nMargin_1"] + '11'] = '=(' + column_dict["Sell_1"] + '11-' + column_dict["LANDED\nTOTAL"] + '11)/' + column_dict["Sell_1"] + '11'
    # ws[column_dict["Net\nSell_1"] + '11'] = '=SUBTOTAL(9,' + column_dict["Net\nSell_1"] + '14:' + column_dict["Net\nSell_1"] + '1379)'
    ws[column_dict["Net\nSell_1"] + '11'] = (
    '=INT(SUBTOTAL(9,' + column_dict["Net\nSell_1"] + '14:' + column_dict["Net\nSell_1"] + '1379))'
)
    
    ws[column_dict["Net\nMargin_1"] + '11'] = '=(' + column_dict["Net\nSell_1"] + '11-' + column_dict["LANDED\nTOTAL"] + '11)/' + column_dict["Net\nSell_1"] + '11'

    # ws[column_dict["MSRP_1"] + '11'] = '=SUBTOTAL(9,' + column_dict["MSRP_1"] + '14:' + column_dict["MSRP_1"] + '1379)'

    ws[column_dict["MSRP_1"] + '11'] = (
    '=INT(SUBTOTAL(9,' + column_dict["MSRP_1"] + '14:' + column_dict["MSRP_1"] + '1379))'
)

    ws[column_dict["IMU_1"] + '11'] = '=((' + column_dict["MSRP_1"] + '11-' + column_dict["Sell_1"] + '11)/' + column_dict["MSRP_1"] + '11)'
    # ws[column_dict["AUR_1"] + '11'] = '=SUBTOTAL(9,' + column_dict["AUR_1"] + '14:' + column_dict["AUR_1"] + '1379)'
    ws[column_dict["AUR\nDiscount_1"] + '11'] = '=(' + column_dict["MSRP_1"] + '11-' + column_dict["AUR_1"] + '11)/' + column_dict["MSRP_1"] + '11'
    ws[column_dict["AUR\nGM%_1"] + '11'] = '=(' + column_dict["AUR_1"] + '11-' + column_dict["Sell_1"] + '11)/' + column_dict["AUR_1"] + '11'

    # ws[column_dict["Sell_2"] + '11'] = '=SUBTOTAL(9,' + column_dict["Sell_2"] + '14:' + column_dict["Sell_2"] + '1379)'
    ws[column_dict["Gross\nMargin_2"] + '11'] = '=(' + column_dict["Sell_2"] + '11-' + column_dict["LANDED\nTOTAL"] + '11)/' + column_dict["Sell_2"] + '11'
    # ws[column_dict["Net\nSell_2"] + '11'] = '=SUBTOTAL(9,' + column_dict["Net\nSell_2"] + '14:' + column_dict["Net\nSell_2"] + '1379)'
    ws[column_dict["Net Margin_2"] + '11'] = '=(' + column_dict["Net\nSell_2"] + '11-' + column_dict["LANDED\nTOTAL"] + '11)/' + column_dict["Net\nSell_2"] + '11'

    # ws[column_dict["MSRP_2"] + '11'] = '=SUBTOTAL(9,' + column_dict["MSRP_2"] + '14:' + column_dict["MSRP_2"] + '1379)'
    ws[column_dict["IMU_2"] + '11'] = '=((' + column_dict["MSRP_2"] + '11-' + column_dict["Sell_2"] + '11)/' + column_dict["MSRP_2"] + '11)'
    # ws[column_dict["AUR_2"] + '11'] = '=SUBTOTAL(9,' + column_dict["AUR_2"] + '14:' + column_dict["AUR_2"] + '1379)'
    ws[column_dict["AUR\nDiscount_2"] + '11'] = '=(' + column_dict["MSRP_2"] + '11-' + column_dict["AUR_2"] + '11)/' + column_dict["MSRP_2"] + '11'
    ws[column_dict["AUR\nGM%_2"] + '11'] = '=(' + column_dict["AUR_2"] + '11-' + column_dict["Sell_2"] + '11)/' + column_dict["AUR_2"] + '11'
    # AUR_1
    ws[column_dict["AUR_1"] + '11'] = (
        '=INT(SUBTOTAL(9,' + column_dict["AUR_1"] + '14:' + column_dict["AUR_1"] + '1379))'
    )
    
    # Sell_2
    ws[column_dict["Sell_2"] + '11'] = (
        '=INT(SUBTOTAL(9,' + column_dict["Sell_2"] + '14:' + column_dict["Sell_2"] + '1379))'
    )
    
    # Net Sell_2
    ws[column_dict["Net\nSell_2"] + '11'] = (
        '=INT(SUBTOTAL(9,' + column_dict["Net\nSell_2"] + '14:' + column_dict["Net\nSell_2"] + '1379))'
    )
    
    # MSRP_2
    ws[column_dict["MSRP_2"] + '11'] = (
        '=INT(SUBTOTAL(9,' + column_dict["MSRP_2"] + '14:' + column_dict["MSRP_2"] + '1379))'
    )
    
    # AUR_2
    ws[column_dict["AUR_2"] + '11'] = (
        '=INT(SUBTOTAL(9,' + column_dict["AUR_2"] + '14:' + column_dict["AUR_2"] + '1379))'
    )

    
    # List of columns where you want to apply the percentage format using column_dict
    percent_columns = [
        "Gross\nMargin_1", "Net\nMargin_1", "IMU_1", "AUR\nDiscount_1", "AUR\nGM%_1", 
        "Gross\nMargin_2", "Net Margin_2", "IMU_2", "AUR\nDiscount_2", "AUR\nGM%_2"
    ]

    # Loop through the columns and apply the percentage format to each corresponding cell in row 11
    for col in percent_columns:
        ws[column_dict[col] + '11'].number_format = '0.00%'

    # List of columns where you want to apply the dollar format using column_dict
    dollar_columns = [
        "LANDED\nTOTAL", "Sell_1", "Net\nSell_1", "MSRP_1", "AUR_1", "Sell_2", "Net\nSell_2", "MSRP_2", "AUR_2"
    ]


    for col in dollar_columns:
        ws[column_dict[col] + '11'].number_format = '"$"#,##0'

    # Apply fill to A13, which corresponds to "IMAGE"
    ws[column_dict["IMAGE"] + '13'].fill = yellow_fill

    # Fill colors for B13, C13, D13, and others as per your request
    ws[column_dict["STYLE #"] + '13'].fill = light_blue_fill
    ws[column_dict["DESCRIPTION"] + '13'].fill = light_blue_fill
    ws[column_dict["Mnf.\nPolicy"] + '13'].fill = gray_fill

    # Additional mappings for BJ13, BK13, etc.
    ws[column_dict["Sell_1"] + '13'].fill = light_blue_fill
    ws[column_dict["Gross\nMargin_1"] + '13'].fill = gold_fill
    ws[column_dict["Net\nSell_1"] + '13'].fill = gold_fill
    ws[column_dict["Net\nMargin_1"] + '13'].fill = gold_fill
    ws[column_dict["MSRP_1"] + '13'].fill = light_blue_fill
    ws[column_dict["IMU_1"] + '13'].fill = light_blue_fill

    # Continue with the rest of the cells
    ws[column_dict["AUR_1"] + '13'].fill = light_yellow_fill
    ws[column_dict["AUR\nDiscount_1"] + '13'].fill = light_yellow_fill
    ws[column_dict["AUR\nGM%_1"] + '13'].fill = light_yellow_fill

    # Manually apply gray color to the rest of the cells
    ws[column_dict["1st MKD_1"] + '13'].fill = gold_fill
    ws[column_dict["MKD NM%_1"] + '13'].fill = gold_fill

    ws[column_dict["Sell_2"] + '13'].fill = light_blue_fill
    ws[column_dict["Gross\nMargin_2"] + '13'].fill = gold_fill
    ws[column_dict["Net\nSell_2"] + '13'].fill = gold_fill
    ws[column_dict["Net Margin_2"] + '13'].fill = gold_fill

    ws[column_dict["MSRP_2"] + '13'].fill = light_blue_fill
    ws[column_dict["IMU_2"] + '13'].fill = light_blue_fill

    ws[column_dict["AUR_2"] + '13'].fill = light_yellow_fill
    ws[column_dict["AUR\nDiscount_2"] + '13'].fill = light_yellow_fill
    ws[column_dict["AUR\nGM%_2"] + '13'].fill = light_yellow_fill
    
    ws[column_dict["1st MKD_2"] + '13'].fill = gold_fill
    ws[column_dict["MKD NM%_2"] + '13'].fill = gold_fill
    ws[column_dict["Support Comment"] + '13'].fill = gold_fill
    # Correctly mapped cells for rows 7, 8, 9
    ws[column_dict["STYLE #"] + '7'].fill = yellow_fill
    ws[column_dict["STYLE #"] + '8'].fill = yellow_fill
    ws[column_dict["STYLE #"] + '9'].fill = yellow_fill

    # Correctly mapped for D7
    ws[column_dict["Mnf.\nPolicy"] + '7'].fill = yellow_fill

    # Correctly mapped for P7, P8 (METAL COST)
    ws[column_dict["METAL\nCOST"] + '7'].fill = yellow_fill
    ws[column_dict["METAL\nCOST"] + '8'].fill = yellow_fill

    # Correctly mapped for BZ8 (IMU_2), CB8 (AUR Discount 2), CD8 (1st MKD 2)
    ws[column_dict["IMU_2"] + '8'].fill = yellow_fill
    ws[column_dict["AUR\nDiscount_2"] + '8'].fill = yellow_fill
    ws[column_dict["1st MKD_2"] + '8'].fill = yellow_fill


    # Set freeze panes based on the "D14" equivalent in column_dict
    ws.freeze_panes = ws[column_dict["Mnf.\nPolicy"] + '14']
   

    # Dynamically apply fill for the range J13:BI13 based on specific columns
    for col in range(ws[column_dict["Item\nStatus\nCode"] + '13'].column, ws[column_dict["Target\nGM%"] + '13'].column + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        if col_letter in [column_dict["METAL"], column_dict["STONE\nTYPE"], column_dict["QTY"], column_dict["SHAPE"], 
                        column_dict["Color\nMM\nSize"], column_dict["wt."], column_dict["DIA CTTW"], column_dict["Quality"]]:
            cell = ws.cell(row=13, column=col)
            cell.fill = light_blue_fill
        else:
            cell = ws.cell(row=13, column=col)
            cell.fill = gray_fill

    # Dynamically apply borders for the range A7:B9
    for row in ws[column_dict["IMAGE"] + '7:' + column_dict["STYLE #"] + '9']:
        for cell in row:
            cell.border = border

    # Dynamically apply borders for the range P1:P4
    for row in ws[column_dict["METAL\nCOST"] + '1:' + column_dict["METAL\nCOST"] + '4']:
        for cell in row:
            cell.border = border
    # Define thick border sides
    thick_side = Side(border_style="thick", color="000000")
    top_border = Border(top=thick_side)
    bottom_border = Border(bottom=thick_side)
    side_border = Border(left=thick_side, right=thick_side)

    # Apply thick top border to the range A7:CE7
    for row in ws[column_dict["IMAGE"] + '7:' + column_dict["MKD NM%_2"] + '7']:
        for cell in row:
            cell.border = top_border

    # Apply thick top border to the range A10:CE10
    for row in ws[column_dict["IMAGE"] + '10:' + column_dict["MKD NM%_2"] + '10']:
        for cell in row:
            cell.border = top_border

    # Apply thick top border to the range A10:CE10


    thick_side = Side(border_style="thick", color="000000")
    thin_side = Side(border_style="thin", color="000000")

    # Create a custom border with thick top and thin for other sides
    header_border = Border(top=thick_side, left=thick_side, right=thick_side, bottom=thin_side)
    normal_border = Border(top=thin_side, left=thick_side, right=thick_side, bottom=thin_side)


    # Apply borders to the range A7:C9
    for row in ws[column_dict["IMAGE"] + '7:' + column_dict["STYLE #"] + '9']:#NTC
        for cell in row:
            # Apply thick top border only for the first row (header)
            if cell.row == 7:
                cell.border = header_border
            else:
                cell.border = normal_border

    # Function to apply thick borders and thin inner horizontal line
    def apply_borders_dynamic(sheet, start_row, end_row, start_column, end_column):
        # Define a thick border for the outer and vertical borders and thin for the inner horizontal line
        thick_side = Side(border_style="thick", color="000000")
        thin_side = Side(border_style="thin", color="000000")

        # Convert column letters to numbers
        start_col_idx = column_index_from_string(start_column)
        end_col_idx = column_index_from_string(end_column)

        # Apply borders
        for row in sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col_idx, max_col=end_col_idx):
            for cell in row:
                # Apply thick borders to all edges except inner horizontal
                top = thick_side if cell.row == start_row else None
                bottom = thick_side if cell.row == end_row else thin_side if cell.row == end_row - 1 else thick_side
                left = thick_side if cell.column == start_col_idx else thick_side
                right = thick_side if cell.column == end_col_idx else thick_side

                # Apply the combined thick borders and thin inner horizontal border
                cell.border = Border(top=top, left=left, right=right, bottom=bottom)
    # Dynamically ap        ply borders using column_dict
    apply_borders_dynamic(ws, 7, 8, column_dict["Box"], column_dict["Cert"])
    apply_borders_dynamic(ws, 7, 8, column_dict["Target\nGM%"], column_dict["Target\nGM%"])
    apply_borders_dynamic(ws, 7, 8, column_dict["Sell_2"], column_dict["Sell_2"])
    apply_borders_dynamic(ws, 7, 8, column_dict["Net\nSell_2"], column_dict["Net\nSell_2"])
    apply_borders_dynamic(ws, 7, 8, column_dict["MSRP_2"], column_dict["MSRP_2"])
    apply_borders_dynamic(ws, 7, 8, column_dict["IMU_2"], column_dict["IMU_2"])
    apply_borders_dynamic(ws, 7, 8, column_dict["AUR\nDiscount_2"], column_dict["AUR\nDiscount_2"])
    apply_borders_dynamic(ws, 7, 8, column_dict["1st MKD_2"], column_dict["1st MKD_2"])
    apply_borders_dynamic(ws, 7, 8, column_dict["Metal\nCost\n /Gm"], column_dict["METAL\nCOST"])


# # Dynamically apply borders for the range Q12:AL12/
    # for row in ws[column_dict["STONE\nTYPE"] + '12:' + column_dict["SETTING\nDESCRIPTION"] + '12']:
    #     for cell in row:
    #         cell.border = Border(top=thick_side, left=thick_side, right=thick_side, bottom=thin_side)

    # Dynamically apply borders for the range BJ12:CE13
    # for row in ws[column_dict["Sell_1"] + '12:' + column_dict["MKD NM%_2"] + '13']:
    #     for cell in row:
    #         cell.border = Border(top=thick_side, left=thick_side, right=thick_side, bottom=thin_side)

    ############ Color Section ############

    # Dynamically apply gray fill for the range A7:CE9
    for row in ws[column_dict["IMAGE"] + '7:' + column_dict["MKD NM%_2"] + '9']:
        for cell in row:
            cell.fill = gray_fill

    # Dynamically apply yellow fill for the range B7:B9
    for row in ws[column_dict["STYLE #"] + '7:' + column_dict["STYLE #"] + '9']:
        for cell in row:
            cell.fill = yellow_fill




    # Define thin and thick borders
    thin_side = Side(style='thin')
    thick_side = Side(style='thick')
    thin_border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)

    # Define white background fill
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    # Define top alignment
    top_alignment = Alignment(vertical='top')

    #WHERE WE DONOT NEED MERGING
    skip_columns = [
    column_dict["STONE\nTYPE"],
    column_dict["QTY"],
    column_dict["SHAPE"],
    column_dict["Color\nMM\nSize"],
    column_dict["wt."],
    column_dict["WT.\n Ext."],
    column_dict["CODE"],
    column_dict["Quality"],
    column_dict["$ PER \nCT/PC"],
    column_dict["COST"],
    column_dict["IH\n (YES/NO)"],
    column_dict["SET\nCOST"],
    column_dict["EXT\nSET\nCOST"],
    column_dict["SETTING\nDESCRIPTION"]
]

    # Variable to track the first row with a thick border
    first_thick_border_row = None
    # Iterate over the rows to find where "Style #" (assumed to be in column B) has a value and apply a thick top border
    for row in ws.iter_rows(min_row=13, max_row=ws.max_row, min_col=2, max_col=2):  # Start from row 13, column B
        for cell in row:
            if cell.value:  # If there is a value in the "Style #" column
                # Apply a thick top border to the entire row, including columns in skip_columns
                for col in ws.iter_cols(min_row=cell.row, max_row=cell.row, min_col=1, max_col=ws.max_column):
                    for c in col:
                        col_letter = get_column_letter(c.column)
                        # Apply thick border to all columns, including skip_columns
                        c.border = Border(top=thick_side, left=thin_side, right=thin_side)

                # If this is the first thick border row, mark it as the first thick row
                if first_thick_border_row is None:
                    first_thick_border_row = cell.row

                # If there was already a first thick border row, apply thin lines for skip_columns and remove borders for other columns
                else:
                    for row_num in range(first_thick_border_row + 1, cell.row):
                        for col_num in range(1, ws.max_column + 1):
                            col_letter = get_column_letter(col_num)
                            if col_letter in skip_columns:
                                # Apply thin border for skip_columns
                                ws.cell(row=row_num, column=col_num).border = thin_border
                                ws.cell(row=row_num, column=col_num).fill = white_fill
                            else:
                                # Remove the borders for non-skip_columns and add white background
                                ws.cell(row=row_num, column=col_num).border = Border(top=None, left=thin_side, right=thin_side, bottom=None)
                                ws.cell(row=row_num, column=col_num).fill = white_fill

                    # Reset the first thick border row for the next set of rows
                    first_thick_border_row = cell.row



    # After the loop, handle the case where the last row needs border and background adjustment
    if first_thick_border_row and first_thick_border_row <= ws.max_row:
        # Remove borders and apply white background between the last thick border row and the end
        for row_num in range(first_thick_border_row + 1, ws.max_row + 1):
            for col_num in range(1, ws.max_column + 1):
                col_letter = get_column_letter(col_num)
                if col_letter not in skip_columns:
                    ws.cell(row=row_num, column=col_num).border = Border(top=None, left=thin_side, right=thin_side, bottom=None)
                    ws.cell(row=row_num, column=col_num).fill = white_fill
                else:
                    # Ensure thin borders and white background for skip_columns
                    ws.cell(row=row_num, column=col_num).border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)
                    ws.cell(row=row_num, column=col_num).fill = white_fill
    # Find the max row and apply thick borders to the entire row at the bottom
    for col_num in range(1, ws.max_column + 1):
        for row_num in range(ws.max_row, ws.max_row + 1):  # You can adjust to how many rows you want to apply the thick border
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = Border(top=None, left=thin_side, right=thin_side, bottom=thick_side)

    fill_columns = [
        column_dict["STONE\nTYPE"],
        column_dict["QTY"],
        column_dict["SHAPE"],
        column_dict["Color\nMM\nSize"],
        column_dict["wt."],
        column_dict["DIA CTTW"],
        column_dict["Quality"]
    ]

    # Iterate over the range Q13:AG725 (outer loop)
    # Determine the last row dynamically using ws.max_row
    last_row = ws.max_row

    # Iterate over the range Q13:last_row (outer loop)
    for row_idx, row in enumerate(ws.iter_rows(min_row=13, max_row=last_row, min_col=17, max_col=33), start=13):
        for cell in row:
            column_letter = cell.column_letter  # Get the column letter (e.g., 'Q', 'R', etc.)

        # Apply fill to specific columns
        if column_letter in fill_columns:
            cell.fill = light_blue_fill



    # List of column letters where the fill should be applied
    fill_columns1 = [
    column_dict["Sell_1"],
    column_dict["MSRP_1"],
    column_dict["IMU_1"],
    column_dict["AUR_1"],
    column_dict["AUR\nDiscount_1"],
    column_dict["AUR\nGM%_1"],
    column_dict["1st MKD_1"],
    column_dict["MKD NM%_1"],
    column_dict["Sell_2"],
    column_dict["MSRP_2"],
    column_dict["IMU_2"],
    column_dict["AUR_2"],
    column_dict["AUR\nDiscount_2"],
    column_dict["AUR\nGM%_2"],
    column_dict["1st MKD_2"],
    column_dict["MKD NM%_2"],
    column_dict["Support Comment"]
]

    # Get the last row in the worksheet
    last_row = ws.max_row

    # Iterate over each row starting from row 14 to the last row
    for row_idx, row in enumerate(ws.iter_rows(min_row=14, max_row=last_row), start=14):
        for cell in row:
            column_letter = cell.column_letter  # Get the column letter (e.g., 'AZ', 'BD', etc.)

            # Apply fill only to specific columns
            if column_letter in fill_columns1:
                cell.fill = light_blue_fill  # Apply the light blue fill  
        # List of column letters where the fill should be applied
    fill_columns1 = [
    column_dict["Gross\nMargin_1"],
    column_dict["Net\nSell_1"],
    column_dict["Net\nMargin_1"],
    column_dict["Gross\nMargin_2"],
    column_dict["Net\nSell_2"],
    column_dict["Net Margin_2"]]

    # Get the last row in the worksheet
    last_row = ws.max_row

    # Iterate over each row starting from row 14 to the last row
    for row_idx, row in enumerate(ws.iter_rows(min_row=14, max_row=last_row), start=14):
        for cell in row:
            column_letter = cell.column_letter  # Get the column letter (e.g., 'AZ', 'BD', etc.)

            # Apply fill only to specific columns
            if column_letter in fill_columns1:
                cell.fill = gray_fill  # Apply the light blue fill  

    	
    fill_columns = [
        column_dict["STONE\nTYPE"],
        column_dict["QTY"],
        column_dict["SHAPE"],
        column_dict["Color\nMM\nSize"],
        column_dict["wt."],
        column_dict["DIA CTTW"],
        column_dict["Quality"]
    ]


    # Get the last row dynamically
    last_row = ws.max_row
    
    # Iterate over the range Q13:last_row (columns Q to AG, or 17 to 33 in numeric form)
    for row in ws.iter_rows(min_row=13, max_row=last_row, min_col=17, max_col=33):
        for cell in row:
            # Get the column letter (e.g., 'Q', 'R', etc.)
            column_letter = cell.column_letter
    
            # Apply fill to specific columns
            if column_letter in fill_columns:
                cell.fill = light_blue_fill

    
    
    #CONINUE===========
    
    # Correctly mapped cells for row 8
    ws[column_dict["Box"] + '8'].fill = light_blue_fill
    ws[column_dict["Cert"] + '8'].fill = light_blue_fill
    ws[column_dict["Target\nGM%"] + '8'].fill = orange_fill
    ws[column_dict["Sell_2"] + '8'].fill = light_blue_fill
    ws[column_dict["Net\nSell_2"] + '8'].fill = orange_fill
    ws[column_dict["MSRP_2"] + '8'].fill = light_blue_fill
    ws[column_dict["IMU_2"] + '8'].fill = yellow_fill
    ws[column_dict["AUR\nDiscount_2"] + '8'].fill = yellow_fill
    ws[column_dict["1st MKD_2"] + '8'].fill = yellow_fill

    # Correctly mapped for P7, P8 (METAL COST)
    ws[column_dict["METAL\nCOST"] + '7'].fill = yellow_fill
    ws[column_dict["METAL\nCOST"] + '8'].fill = yellow_fill

    # Correctly mapped for M13 (METAL)
    ws[column_dict["METAL"] + '13'].fill = light_blue_fill



    #alignment

    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Apply wrap text and center alignment to the header row (row 13)
    for cell in ws[13]:  # Row 13
        cell.alignment = header_alignment
    #width

    price_column_indices = [column_index_from_string(column_dict[col]) for col in [
        "METAL\nCOST",
        "Metal\nCost\n /Gm",
        "$ PER \nCT/PC",
        "COST",
        "TOTAL\nCS\nCOST",
        "CS\nIH\nCOST",
        "TOTAL\nDIA\nCOST",
        "DIA\nIH\nCOST",
        "TOTAL\nSTONE\nCOST",
        "STONE\nIH\nCOST",
        "SET\nCOST",
        "EXT\nSET\nCOST",
        "LABOR\nDTL\nCOST",
        "LABOR\nTTL\nCOST",
        "LABOR\nIH\nCOST",
        "FINDING\nDTL\nCOST",
        "FINDING\nTTL\nCOST",
        "FINDING\nIH COST",
        "STYLE\nTOTAL",
        "STYLE\nTOTAL\nEXCL.IH",
        "S/H",
        "Duty $",
        "LANDED\nTOTAL",

    ]]
    # Iterate through rows and apply dollar format
    for row_idx, row in enumerate(ws.iter_rows(min_row=14, max_row=last_row), start=14):
        for col_idx in price_column_indices:
            cell = ws.cell(row=row_idx, column=col_idx)
            # Apply dollar format if the value is not zero
            cell.number_format = '"$"#,##0.00'

    three_deicimal = [column_index_from_string(column_dict[col]) for col in ["wt."]]
    for row_idx, row in enumerate(ws.iter_rows(min_row=14, max_row=last_row), start=14):
        for col_idx in three_deicimal:
            cell = ws.cell(row=row_idx, column=col_idx)
            # Apply dollar format to the cell
            cell.number_format = '#,##0.000'

    two_deicimal = [column_index_from_string(column_dict[col]) for col in ["WT.\n Ext."]]
    for row_idx, row in enumerate(ws.iter_rows(min_row=14, max_row=last_row), start=14):

        for col_idx in two_deicimal:
            cell = ws.cell(row=row_idx, column=col_idx)
            # Apply dollar format to the cell
            cell.number_format = '#,##0.00'
    # Set row height for row 13
    ws.row_dimensions[13].height = 45  # Adjust row height for row 13

    image_width_in_pixels = 150
    column_width = image_width_in_pixels / 7  # Convert pixels to Excel width units
    row_height = image_width_in_pixels / 7     # Add 2 extra units for padding
    #hiide
    #  "Sell_1", "Gross\nMargin_1", "Net\nSell_1", "Net\nMargin_1", "MSRP_1", "IMU_1", "AUR_1", "AUR\nDiscount_1",  "AUR\nGM%_1", "1st MKD_1", "MKD NM%_1"

    ws.column_dimensions[column_dict['IMAGE']].width = column_width
    columns_to_hide=["CS\nIH\nCOST", "DIA\nIH\nCOST","STONE\nIH\nCOST"]
    ws.row_dimensions.group(1, 4, outline_level=1, hidden=True)
    for col in columns_to_hide:
        ws.column_dimensions[column_dict[col]].hidden = True
    ws.column_dimensions[column_dict["DESCRIPTION"]].width = 50
    ws.column_dimensions[column_dict["Metal\nCost\n /Gm"]].width = 10
    ws.column_dimensions[column_dict["DIA CTTW"]].width = 10
    ws.column_dimensions[column_dict["wt."]].width = 10
    for col in ["Sell_1", "Gross\nMargin_1", "Net\nSell_1", "Net\nMargin_1", "MSRP_1", "IMU_1", "AUR_1", "AUR\nDiscount_1", 
        "AUR\nGM%_1", "1st MKD_1", "MKD NM%_1", "Sell_2", "Gross\nMargin_2", "Net\nSell_2", "Net Margin_2", 
        "MSRP_2", "IMU_2", "AUR_2", "AUR\nDiscount_2", "AUR\nGM%_2", "1st MKD_2", "MKD NM%_2"]:
         ws.column_dimensions[column_dict[col]].width = 11

    for col in [ "LABOR\nDTL\nCOST", "LABOR\nTTL\nCOST", "LABOR\nIH\nCOST","FINDING\nQTY", "FINDING\nDTL\nCOST", "FINDING IH COST\nOR NOT\n(YES/NO)",
        "FINDING\nTTL\nCOST","FINDING\nIH COST","FINDING\n(YES/NO)"]:
        ws.column_dimensions[column_dict[col]].width = 15
    for col in ["TOTAL\nCS\nCOST", "CS\nIH\nCOST","TOTAL\nDIA\nCOST", "DIA\nIH\nCOST","TOTAL\nSTONE\nCOST","STONE\nIH\nCOST"]:
        ws.column_dimensions[column_dict[col]].width = 10   
    ws.column_dimensions[column_dict["SETTING\nDESCRIPTION"]].width = 35
    ws.column_dimensions[column_dict["LABOR TYPE"]].width = 20
    ws.column_dimensions[column_dict["FINDING\nTYPE"]].width = 20
    ws.column_dimensions[column_dict["STYLE\nTOTAL"]].width = 20
    ws.column_dimensions[column_dict["STYLE\nTOTAL\nEXCL.IH"]].width = 20
    ws.column_dimensions[column_dict["LANDED\nTOTAL"]].width = 20
    ws.column_dimensions[column_dict["LABOR\nQTY"]].width = 8
    ws.column_dimensions[column_dict["LABOR\nDTL\nCOST"]].width = 8
    ws.column_dimensions[column_dict["Labour IH COST\nOR NOT\n(YES/NO)"]].width = 9
    ws.column_dimensions[column_dict["LABOR\nTTL\nCOST"]].width = 8
    ws.column_dimensions[column_dict["LABOR\nTTL\nCOST"]].width = 8
    ws.column_dimensions[column_dict["LABOR\nTTL\nCOST"]].width = 8
    ws.column_dimensions[column_dict["LABOR\nTTL\nCOST"]].width = 8
    ws.column_dimensions[column_dict["LABOR\nTTL\nCOST"]].width = 8
    ws.column_dimensions[column_dict["S/H"]].width = 8
    ws.column_dimensions[column_dict["Duty $"]].width = 8
    ws.column_dimensions[column_dict["Duty %"]].width = 8

    ws.column_dimensions[column_dict["Support Comment"]].width = 14
    #Rename
    # ws.column_dimensions[column_dict["LABOR\nQTY"]] = "Q"
    ws[column_dict["LABOR\nQTY"] + '13']="QTY"
    ws[column_dict["LABOR\nDTL\nCOST"] + '13']="LABOR\nCOST"
    ws[column_dict["Labour IH COST\nOR NOT\n(YES/NO)"]+'13'] ="IH\n(YES/NO)"
    ws[column_dict["LABOR\nTTL\nCOST"] + '13']="LABOR\nTOTAL\nCOST"
    red_font = Font(color="FF0000",bold=True)  # Hex code for red

    # Cell reference for "Manf.\nCode" in row 13
    cell_key = column_dict["Manf.\nCode"]  # Extract column from the dictionary
    cell_reference = f"{cell_key}13"       # Combine column with row number

    # Apply red font to the cell

    #hiiden
    ws.column_dimensions.group(
    start=column_dict["Sell_1"], 
    end=column_dict["MKD NM%_1"], 
    hidden=True, 
    outline_level=1
)
    ws.column_dimensions[column_dict["LABOR\nIH\nCOST"]].hidden = True
    ws.column_dimensions[column_dict["FINDING\nIH COST"]].hidden = True
    # Detect the last row and last column with data
    last_row = ws.max_row  # Detect the last row with data
    last_col = ws.max_column  # Detect the last column with data

    # Define the range for applying the filter (starting from A13)
    filter_range = f"A13:{openpyxl.utils.get_column_letter(last_col)}{last_row}"

    # Apply the AutoFilter to the range
    ws.auto_filter.ref = filter_range    




def fill_style_column(sheet, start_row=13, style_column=2):
    """Fill down product names or descriptions in the specified column to ensure proper filtering."""
    current_value = None  # Track the current value

    for row in range(start_row, sheet.max_row + 1):
        cell_value = sheet.cell(row, style_column).value

        if cell_value:  # If a value exists, update the tracker
            current_value = cell_value
        elif current_value:  # If the cell is empty, fill with the current value
            sheet.cell(row, style_column).value = current_value

def apply_conditional_formatting(sheet, column_letter, start_row=13):
    """Apply conditional formatting to hide duplicate values in the specified column."""
    last_row = sheet.max_row

    # Define the formula to detect duplicate values
    formula = f'={column_letter}{start_row}={column_letter}{start_row - 1}'

    # Create a conditional formatting rule to hide duplicate values (white font)
    rule = FormulaRule(formula=[formula], font=Font(color="FFFFFF"))  # White font color

    # Apply the rule to the specified column from the start row to the last row
    sheet.conditional_formatting.add(f"{column_letter}{start_row}:{column_letter}{last_row}", rule)

def adjust_images_with_xlwings(input_file, output_file):
    """Adjust image properties using xlwings and apply formatting across all sheets."""
    try:
        # Open workbook with openpyxl to fill values and apply formatting
        wb_openpyxl = openpyxl.load_workbook(input_file)

        # Iterate through all sheets to fill down and apply conditional formatting
        for ws_openpyxl in wb_openpyxl.worksheets:
            # Fill down 'STYLE #' column values (column 2)
            fill_style_column(ws_openpyxl, start_row=13, style_column=2)

            # Fill down 'DESCRIPTION' column values (column 3)
            fill_style_column(ws_openpyxl, start_row=13, style_column=3)

            # Apply conditional formatting to 'STYLE #' column (column B)
            apply_conditional_formatting(ws_openpyxl, column_letter='B', start_row=13)

            # Apply conditional formatting to 'DESCRIPTION' column (column C)
            apply_conditional_formatting(ws_openpyxl, column_letter='C', start_row=13)

        # Save intermediate result
        wb_openpyxl.save(output_file)
        wb_openpyxl.close()  # Close the openpyxl workbook

        # Open the adjusted workbook with xlwings to adjust image placement
        app = xw.App(visible=False)  # Ensure Excel app runs in the background
        wb_xlwings = app.books.open(output_file)

        # Iterate over all sheets to adjust image placement
        for ws_xlwings in wb_xlwings.sheets:
            # Adjust image placement: 'Move and size with cells' 
            for picture in ws_xlwings.api.Shapes:
                picture.Placement = 1  # 1 = 'Move and size with cells'
                print(f"Adjusted {picture.Name} to 'Move and size with cells' in sheet {ws_xlwings.name}.")

        # Save the final workbook and properly close everything
        wb_xlwings.save(output_file)
        wb_xlwings.close()
        app.quit()  # Quit the Excel application
        print(f"Excel saved with adjusted images at: {output_file}")



    except Exception as e:
        print(f"An error occurred: {e}")
