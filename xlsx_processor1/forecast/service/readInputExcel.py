import pandas as pd
from openpyxl import load_workbook
from multiprocessing import Pool, cpu_count
from . import config


def read_single_sheet(args):
    input_path, sheet_name = args
    return sheet_name, pd.read_excel(input_path, sheet_name=sheet_name, engine="openpyxl")


def readInputExcel(input_path):
    # Load sheet names using openpyxl (faster than pandas for just metadata)
    wb = load_workbook(input_path, read_only=True)
    sheet_names = wb.sheetnames

    # Parallel load sheets using Pool
    args = [(input_path, sheet) for sheet in sheet_names]
    with Pool(processes=min(cpu_count(), len(sheet_names))) as pool:
        results = pool.map(read_single_sheet, args)

    # Store in shared config
    config.sheets = {sheet: df for sheet, df in results}
    config.return_QA_df = pd.read_excel(
        r"C:\Users\emman\OneDrive\Desktop\final_excel_deploy\xlsx_processor1\forecast\service\Macys returns 2025.xlsx",
        engine='openpyxl'
    )

    print("Sheets in the Excel file:")
    for sheet in config.sheets.keys():
        print(sheet)
    print("Successfully read the input Excel file.")
