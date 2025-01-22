import os
import pandas as pd
import tempfile
from django.http import HttpResponse, JsonResponse
from django.views.decorators.csrf import csrf_exempt
import pandas as pd
import openpyxl
import os
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side,GradientFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import pixels_to_points, points_to_pixels
from openpyxl.styles import numbers
from openpyxl.styles import NamedStyle
from openpyxl.drawing.image import Image
from io import BytesIO
from PIL import Image as PILImage
from openpyxl import load_workbook
# Input Excel file
from django.conf import settings
from openpyxl.utils import column_index_from_string 
import hashlib
from openpyxl.formatting.rule import FormulaRule
import shutil
from .config import *
import os
import zipfile
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.worksheet.datavalidation import DataValidation
import time
from openpyxl.workbook.defined_name import DefinedName
from datetime import datetime, timedelta
import os
import zipfile
import shutil
import json 
import math
import multiprocessing as mp
from concurrent.futures import ThreadPoolExecutor
import pandas as pd
from datetime import datetime, timedelta
from calendar import monthrange
from multiprocessing import Manager, Pool


def make_zip_and_delete(folder_path):
    folder_path = os.path.normpath(folder_path)
    zip_file_path = os.path.normpath(f'{folder_path}.zip')
    
    try:
        # Create a ZIP file
        with zipfile.ZipFile(zip_file_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for root, dirs, files in os.walk(folder_path):
                for file in files:
                    file_path = os.path.join(root, file)
                    arcname = os.path.relpath(file_path, folder_path)  # Preserve folder structure
                    zipf.write(file_path, arcname)
        
        print(f"Folder '{folder_path}' has been compressed into '{zip_file_path}'")

        # Delete the folder after zipping
        shutil.rmtree(folder_path)
        print(f"Folder '{folder_path}' has been deleted successfully.")
    
    except PermissionError:
        print(f"Permission denied: Cannot access '{folder_path}'. Please check folder permissions.")
    except FileNotFoundError:
        print(f"File not found: '{folder_path}' does not exist.")
    except Exception as e:
        print(f"An error occurred: {e}")


def get_previous_retail_week():
    """
    Get the previous week's month, year of the previous month, 
    last year's occurrence of that month, last month before the previous month in numeric format,
    determine SP (Spring) or FA (Fall) based on the previous month,
    and calculate the number of retail weeks for each month individually.
    """
    # Use the current date as input
    current_date = datetime.now()
    print(f"Current date: {current_date}")

    # Find the current week's Sunday
    current_sunday = current_date - timedelta(days=current_date.weekday() + 1)

    # Calculate the previous week's Sunday
    previous_week_sunday = current_sunday - timedelta(days=7)

    # Determine the previous week number
    previous_week_number = (previous_week_sunday.day - 1) // 7 + 1

    # Get the month and year of the previous week
    previous_month = previous_week_sunday.strftime('%b').upper()
    year_of_previous_month = previous_week_sunday.year

    # Determine last year's occurrence of the same month
    last_year_of_previous_month = year_of_previous_month - 1

    # Determine the last month before the previous month
    last_month_of_previous_month_date = previous_week_sunday.replace(day=1) - timedelta(days=1)
    last_month = last_month_of_previous_month_date.strftime('%b').upper()

    # Custom mapping for months
    month_mapping = {
        'FEB': 1, 'MAR': 2, 'APR': 3, 'MAY': 4,
        'JUN': 5, 'JUL': 6, 'AUG': 7, 'SEP': 8,
        'OCT': 9, 'NOV': 10, 'DEC': 11, 'JAN': 12
    }
    last_month_of_previous_month_numeric = month_mapping[last_month]

    # Determine SP (Spring) or FA (Fall/Winter) based on the previous month
    spring_months = ['FEB', 'MAR', 'APR', 'MAY', 'JUN', 'JUL']
    fall_months = ['AUG', 'SEP', 'OCT', 'NOV', 'DEC', 'JAN']

    season = "SP" if previous_month in spring_months else "FA"

    # Calculate the number of retail weeks for each month of the current year
    current_year = year_of_previous_month
    print(current_year)
    # Individual variables for retail weeks of each month
 

    def get_retail_weeks(year, month):
        """
        Calculate the number of retail weeks in a given month.
        Retail weeks follow the Sunday-to-Saturday structure, 
        and all days in a week belong to the month in which the week starts.
        
        Args:
            year (int): The year of the month.
            month (int): The month (1 for January, 12 for December).

        Returns:
            int: Number of retail weeks in the month.
        """
        # Get the first day and last day of the month
        first_day = datetime(year, month, 1)
        last_day = datetime(year, month, monthrange(year, month)[1])

        # Find the first Sunday of the month
        first_sunday = first_day + timedelta(days=(6 - first_day.weekday()) % 7)

        # Find the last Saturday of the month
        last_saturday = last_day - timedelta(days=last_day.weekday() + 1)

        # Count retail weeks
        current_week_start = first_sunday
        week_count = 0

        while current_week_start <= last_saturday:
            week_count += 1
            current_week_start += timedelta(days=7)  # Move to the next Sunday

        # Check if the final week starts in the current month (partial week rule)
        if current_week_start <= last_day:
            week_count += 1

        return week_count

    feb_weeks = get_retail_weeks(current_year,2)
    mar_weeks = get_retail_weeks(current_year,3)
    apr_weeks = get_retail_weeks(current_year,4)
    may_weeks = get_retail_weeks(current_year,5)
    jun_weeks = get_retail_weeks(current_year,6)
    jul_weeks = get_retail_weeks(current_year,7)
    aug_weeks = get_retail_weeks(current_year,8)
    sep_weeks = get_retail_weeks(current_year,9)
    oct_weeks = get_retail_weeks(current_year,10)
    nov_weeks = get_retail_weeks(current_year,11)
    dec_weeks = get_retail_weeks(current_year,12)
    jan_weeks = get_retail_weeks(current_year + 1, 1)  # January belongs to the next year

    return previous_month, previous_week_number, year_of_previous_month,last_year_of_previous_month, last_month_of_previous_month_numeric,season, feb_weeks, mar_weeks, apr_weeks, may_weeks,jun_weeks, jul_weeks, aug_weeks, sep_weeks, oct_weeks,nov_weeks, dec_weeks, jan_weeks
    # return previous_month, previous_week_number, year_of_previous_month, last_year_of_previous_month, last_month_of_previous_month_numeric, season

# Call the function and print the results
previous_month, previous_week_number, year_of_previous_month,last_year_of_previous_month, last_month_of_previous_month_numeric,season, feb_weeks, mar_weeks, apr_weeks, may_weeks,jun_weeks, jul_weeks, aug_weeks, sep_weeks, oct_weeks,nov_weeks, dec_weeks, jan_weeks = get_previous_retail_week()
@csrf_exempt
def upload_xlsx(request):
    try:
        if request.method == 'POST' and 'file' in request.FILES:
            uploaded_file = request.FILES['file']
            output_folder= request.POST['output_filename']
            month_from = request.POST.get('month_from')
            month_to = request.POST.get('month_to')
            percentage = request.POST.get('percentage')
            categories = request.POST.get('categories')

            print("categories--> ",categories)
            print("month_from", month_from)
            print("month_to", month_to)
            print("percentage", percentage)

            input_tuple = [(item['name'], item['value']) for item in categories]


            # Ensure the directory exists within MEDIA_ROOT
            processed_dir = os.path.join(settings.MEDIA_ROOT, 'processed_files')
            os.makedirs(processed_dir, exist_ok=True)
            output_folder_path = os.path.join(processed_dir, output_folder)
            os.makedirs(output_folder_path, exist_ok=True)
            file_path = os.path.join(processed_dir,output_folder_path)
            #preprocess(uploaded_file,file_path)

            input_path = uploaded_file.temporary_file_path() 
            INPUT_EXCEL_FILE = input_path
            # OUTPUT_EXCEL_FILE = file_path
            
            try:
                # processed_data(INPUT_EXCEL_FILE, OUTPUT_EXCEL_FILE)  # Ensure this function is defined elsewhere
                start_time = time.time()
                process_data(INPUT_EXCEL_FILE,file_path,month_from,month_to,percentage,input_tuple) 
                end_time = time.time() 
                elapsed_time = end_time - start_time
                print(f"Function executed in {elapsed_time:.6f} seconds")
                make_zip_and_delete(file_path)
            except Exception as e:
                print(f"An error occurred during execution: {e}")

            return JsonResponse({'file_path': f'{settings.MEDIA_URL}processed_files/{output_folder}.zip'}, status=200)


        return JsonResponse({'error': 'No file uploaded'}, status=400)
    except Exception as e:
        return JsonResponse({'error': f'Server error: {str(e)}'}, status=500)
    

@csrf_exempt
def download_file(request):
    file_path = request.GET.get('file_path')

    if os.path.exists(file_path):
        with open(file_path, 'rb') as fh:
            response = HttpResponse(fh.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{os.path.basename(file_path)}"'
            print(response)
            return response
    return JsonResponse({'error': 'File not found'}, status=404)


@csrf_exempt
# Function to add a dropdown to a specific cell
def add_dropdown(ws, cell, options):
    # Create a data validation object with the list of options
    dropdown = DataValidation(type="list", formula1=f'"{",".join(options)}"', allow_blank=True)
    dropdown.prompt = "Please select an option"
    dropdown.promptTitle = "Dropdown List"
    # Apply the data validation dropdown to the specified cell
    ws.add_data_validation(dropdown)
    dropdown.add(ws[cell])


def apply_round_format(ws, cell_ranges, decimal_places):
    for cell_range in cell_ranges:
        # Check if the cell range is a single cell
        if ":" in cell_range:
            # This is a range, so we can iterate through it
            for row in ws[cell_range]:
                for cell in row:
                    # Check if cell contains a formula by verifying if the value starts with "="
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                        # Wrap formula in ROUND with the specified decimal places
                        cell.value = f"=ROUND({cell.value[1:]}, {decimal_places})"
        else:
            # This is a single cell reference
            cell = ws[cell_range]
            if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                # Wrap formula in ROUND with the specified decimal places
                cell.value = f"=ROUND({cell.value[1:]}, {decimal_places})"


def apply_format(ws, cell_ranges, number_format):
    for cell_range in cell_ranges:
        # Check if the cell range is a single cell or a range
        if ":" in cell_range:
            # This is a range, so we can iterate through it
            for row in ws[cell_range]:
                for cell in row:
                    cell.number_format = number_format
        else:
            # This is a single cell
            cell = ws[cell_range]
            cell.number_format = number_format


def read_excel_sheet(file_details):
    file_path, sheet_name, kwargs = file_details
    return sheet_name, pd.read_excel(file_path, sheet_name=sheet_name, **kwargs)


def process_sheet(sheet_name, config, input_path):
    """
    Function to process a single sheet with specified configuration.
    """
    print(f"Processing sheet: {sheet_name}")
    try:
        excel_file = pd.ExcelFile(input_path)
        data = excel_file.parse(sheet_name=sheet_name, **config)
        print(f"Finished processing sheet: {sheet_name}")
        return sheet_name, data
    except Exception as e:
        print(f"Error processing sheet {sheet_name}: {e}")
        return sheet_name, None


def process_category(args):
    """
    Process a single category with its associated data.
    """
    category, (code, num_products), dynamic_categories, file_path, other_params = args

    # Extract required variables from other_params
    (
        year_of_previous_month,
        last_year_of_previous_month,
        season,
        current_month,
        current_month_number,
        previous_week_number,
        last_month_of_previous_month_numeric,
        rolling_method,
        feb_weeks,
        mar_weeks,
        apr_weeks,
        may_weeks,
        jun_weeks,
        jul_weeks,
        aug_weeks,
        sep_weeks,
        oct_weeks,
        nov_weeks,
        dec_weeks,
        jan_weeks,
        index_df,
        report_grouping_df,
        planning_df,
        TBL_Planning_VerticalReport__3,
        Macys_Recpts,
        All_DATA,
        MCOM_Data,
        percentage,
        month_from,
        month_to,
        master_sheet,
        vendor_sheet,
        birthstone_sheet

    ) = other_params

    # Placeholder for the processing logic of each category
    # Use the original code logic for processing categories here.
    print(f"Processing Category: {category} with Code: {code} and Num Products: {num_products}")
    data = [
            ["", "TY", year_of_previous_month, "LY", last_year_of_previous_month, "Season", season, "Current Year", "Month", current_month, current_month_number, "Week", previous_week_number, "", "MAY-SEP", "", "", "Last Completed Month", last_month_of_previous_month_numeric, "", "Use EOM Actual?", rolling_method],
            ["", "Count of Items", "", 8215, "", "", "", "Last SP / FA Months", "Month", "Jul", "", "Jan", 12, "Sorted by:", "Dept Grouping >Class ID", "", "", ""],
            ["", "", "", "", "", "", "", "# of Wks in Mth", feb_weeks, mar_weeks, apr_weeks, may_weeks, jun_weeks, jul_weeks, aug_weeks, sep_weeks, oct_weeks, nov_weeks, dec_weeks, jan_weeks],
            ["", category.upper(), code, "", "Avg Sales 1st & last Mth", 8, 11, "Month #", 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, "", "", ""]
        ]

    output_file = f"{category}{code}"
    output_file_path = os.path.join(file_path, f'{output_file}.xlsx')
        # Initialize workbook and create sheets
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = category
    ws_index = wb.create_sheet(title="Index")
    ws_month = wb.create_sheet(title="Month")
    ws_dropdown = wb.create_sheet(title="DropdownData")
    store_products=[]
    birthstone_list=[]
    notify_macys=[]
    pids_below_door_count = []


    # Loop through products to generate PIDs

    # Step 4: Populate Worksheet with Data
    for row_num, row_data in enumerate(data, 1):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")
    # Step 4: Freeze Top 4 Rows
    ws.freeze_panes = ws["A5"]
    # Define options for dropdowns
    lookup_key = f"{category.upper()}{code}"
# Perform a lookup in 'report grouping'
    lookup_value_c2 = report_grouping_df.loc[
    report_grouping_df[0].str.upper() == lookup_key.upper(), 3].iloc[0]
    ws['C2'] = lookup_value_c2
    ws['D2'] = "=C2*51+4"
    ws['K1'] = "=VLOOKUP(J1,Month!A:B,2,0)"
    ws['K2'] = "=VLOOKUP(J2,Month!A:B,2,0)"
    ws['M2'] = "=VLOOKUP(L2,Month!A:B,2,0)"
    # Define additional dropdown options

    # Define your dropdown options (38 items)
    dropdown_options = [
        "BT", "Citrine", "Cross", "CZ", "Dia", "Ear", "EMER", "Garnet", "Gem",
        "GEM EAR", "Gold Chain", "GOLD EAR", "Amy", "Anklet", "Aqua", "Bridal",
        "Heart", "Heavy Gold Chain", "Jade", "KIDS", "Locket", "Mens Gold Bracelet",
        "Mens Misc", "Mens Silver chain", "Mom", "MOP", "Neck", "Onyx", "Opal",
        "Pearl", "Peridot", "Religious", "Ring", "Ruby", "Saph", "Womens Silver Chain",
        "Wrist", "Grand Total"
    ]

    # Write the dropdown options to the "DropdownData" sheet in column A
    for i, option in enumerate(dropdown_options, start=1):
        ws_dropdown.cell(row=i, column=1, value=option)

    # Define the named range for the dropdown options
    # Make sure the range covers exactly the 38 items (A1 to A38)
    named_range = DefinedName(name="DropdownOptions", attr_text="DropdownData!$A$1:$A$38")
    wb.defined_names.add(named_range)

    # Create a data validation that references the named range
    dropdown = DataValidation(type="list", formula1="DropdownOptions", allow_blank=True)
    dropdown.prompt = "Please select an option"
    dropdown.promptTitle = "Dropdown List"



    for loop in range(num_products):
        start_row = 5 + (51 * loop)  # Adjust the range as needed for your loop
        ws.add_data_validation(dropdown)
        dropdown.add(ws[f"F{start_row + 1}"])  # Apply to column F as an example

    forecast_method_options = ["FC by Index", "FC by Trend", "Average", "Current Year", "Last Year"]


    season_option = ["FA", "SP"]
    month_option = ["Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Jan"]
    year_option = ["Current MTH", "YTD", "SPRING", "FALL", "LY FALL"]


    # Function to add a dropdown to a specific cell
    def add_dropdown(ws, cell, options):
        # Create a data validation object with the list of options
        dropdown = DataValidation(type="list", formula1=f'"{",".join(options)}"', allow_blank=True)
        dropdown.prompt = "Please select an option"
        dropdown.promptTitle = "Dropdown List"
        # Apply the data validation dropdown to the specified cell
        ws.add_data_validation(dropdown)
        dropdown.add(ws[cell])
    # Add dropdowns to specified cells
    add_dropdown(ws, "G1", season_option)     # G1 for Season
    add_dropdown(ws, "V1", year_option)       # V1 for Year options

    # Add dropdowns to multiple cells for Months
    for cell in ["J1", "J2", "L2"]:
        add_dropdown(ws, cell, month_option)




    ALL_VALUES = [
        "PID/BLU/MKST", "Current FC Index", "(TY/LY) STD Sales Index/12M FC", "STD Trend / 12M FC",
        "Item Status/Forecasting Method/Safe", "Current Str Cnt/Last Str Cnt/Last Updated",
        "Store Model/Com Model/TTL Model", "MA Proj/Proj Ball/Holiday Bd FC",
        "MCY/OH/OH in Transit/MTD Ships/LW Ships", "Planned WOS/WOP/Wkly Avg Sale/Last 4 Wks Ships",
        "Actual WOS/WOP/Real OOS Loc", "Excess Proj Qty & $ / Recall Wks of Poj & Qty",
        "Proj Qty & $ to Release / Note", "Vendor/Min order", "RL TTL/Net Proj/ORD Unalloc/+/− to Model",
        "MA Bin/FLDC/WIP QTY/REPLN HOLD DATE", "WIP Demand", "MD Status/Store & Mcom Repl",
        "TTL Last Repl/Age/Mths Active", "Last Cost/Owned/TKT Ret/GM/Actual GM", "Metal Lock/MFG Policy",
        "KPI DATA", "Last KPI Door count", "Diff to Current Door",
        "Out of Stock Locations", "Suspended Location count",
        "Click to View online", "DEPT #", "Sub Class", "Masterstyle", "PID Desc",
        "COM 1st Live/Live Site/V2C/WebID/STD Rtn", "Web ID Description",
        "Last Reviewed Date/Code/Qty to Enter", "Current Review Comments"
    ]

    H_VALUES = [
        'ROLLING 12M FC', 'Index', 'FC by Index', 'FC by Trend', 'Recommended FC', 'Planned FC',
        'Planned Shipments', 'Planned EOH (Cal)', 'Gross Projection (Nav)', 'Macys Proj Receipts',
        'Planned Sell thru %', "TOTAL 2023",'Total Sales Units', 'Store Sales Units', 'Com Sales Units',
        'COM % to TTL (Sales)', 'TOTAL EOM OH', 'Store EOM OH', 'COM EOM OH',
        'COM % to TTL (EOH)', 'Omni Sales $', 'COM Sales $', 'Omni AUR/% Diff Own',
        'Omni Sell Thru %', 'Store SellThru %', 'Omni Turn', 'Store turn',
        'TY Store Sales U vs LY', 'TY COM sales U vs LY', 'TY Store EOH vs LY',
        'Omni OO Units', 'COM OO Units', 'Omni Receipts',    "TOTAL 2023",
        "Total Sales Units",
        "Store Sales Units",
        "Com Sales Units",
        "COM % to TTL (Sales)",
        "TOTAL EOM OH",
        "Store EOM OH",
        "COM EOM OH",
        "COM % to TTL (EOH)",
        "Omni Receipts",
        "Omni Sell Thru %",
        "Store SellThru %",
        "Omni Turn",
        "Store Turn",
        "Omni Sales $",
        "COM Sales $",
        "Omni AUR/% Diff Own"
    ]
    MONTHLY_VALUES = ['FEB', 'MAR', 'APR', 'MAY', 'JUN', 'JUL', 'AUG', 'SEP', 'OCT',
                    'NOV', 'DEC', 'JAN', 'ANNUAL', 'SPRING', 'FALL']
    def apply_round_format(ws, cell_ranges, decimal_places):
        for cell_range in cell_ranges:
            # Check if the cell range is a single cell
            if ":" in cell_range:
                # This is a range, so we can iterate through it
                for row in ws[cell_range]:
                    for cell in row:
                        # Check if cell contains a formula by verifying if the value starts with "="
                        if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                            # Wrap formula in ROUND with the specified decimal places
                            cell.value = f"=ROUND({cell.value[1:]}, {decimal_places})"
            else:
                # This is a single cell reference
                cell = ws[cell_range]
                if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                    # Wrap formula in ROUND with the specified decimal places
                    cell.value = f"=ROUND({cell.value[1:]}, {decimal_places})"

    def apply_format(ws, cell_ranges, number_format):
        for cell_range in cell_ranges:
            # Check if the cell range is a single cell or a range
            if ":" in cell_range:
                # This is a range, so we can iterate through it
                for row in ws[cell_range]:
                    for cell in row:
                        cell.number_format = number_format
            else:
                # This is a single cell
                cell = ws[cell_range]
                cell.number_format = number_format


    # Open the source workbook
    
    for row_num, row_data in enumerate(index_df.values, start=1):
        for col_num, value in enumerate(row_data, start=1):
            ws_index.cell(row=row_num, column=col_num, value=value)

    # Step 5: Apply Percentage Formatting to B4:P41
    for row in ws_index.iter_rows(min_row=4, max_row=41, min_col=2, max_col=16):
        for cell in row:
            if isinstance(cell.value, (int, float)) and 0 <= cell.value <= 1:
                cell.number_format = '0.00%'  # Format as percentage


    # Step 6: Apply Filters to A3:P3
    ws_index.auto_filter.ref = "A3:P3"


    month_data = [
        ["All", ""],
        ["Feb", 1],
        ["Mar", 2],
        ["Apr", 3],
        ["May", 4],
        ["Jun", 5],
        ["Jul", 6],
        ["Aug", 7],
        ["Sep", 8],
        ["Oct", 9],
        ["Nov", 10],
        ["Dec", 11],
        ["Jan", 12],
    ]

    # Write the data to the "Month" sheet
    for row_num, (month, number) in enumerate(month_data, start=1):
        ws_month.cell(row=row_num, column=1, value=month)  # Column A
        ws_month.cell(row=row_num, column=2, value=number)  # Column B

    border_color = "D9D9D9"  # Gridline color
    gridline1 = Border(
        left=Side(style="thin", color=border_color),
        right=Side(style="thin", color=border_color),
        top=Side(style="thin", color=border_color),
        bottom=Side(style="thin", color=border_color)
    )

    # Create PatternFill objects for the colors
    light_gray = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    light_yellow = PatternFill(start_color="FFEB9C" , end_color="FFEB9C" , fill_type="solid")
    light_pink = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    # Apply the fills to specific cells
    ws['C1'].fill = light_gray
    ws['E1'].fill = light_gray
    ws['G1'].fill = light_gray
    ws['J1'].fill = light_gray
    ws['K1'].fill = light_gray
    ws['M1'].fill = light_gray
    ws['S1'].fill = light_gray
    ws['J2'].fill = light_gray
    ws['L2'].fill = light_gray

    ws['K2'].fill = light_yellow
    ws['O1'].fill = light_yellow
    ws['M2'].fill = light_yellow

    ws['B2'].fill = yellow_fill
    ws['C2'].fill = yellow_fill
    ws['V1'].fill = light_pink

    #no boader
    # Define white border style
    white_border = Border(
        left=Side(style='thin', color="FFFFFF"),
        right=Side(style='thin', color="FFFFFF"),
        top=Side(style='thin', color="FFFFFF"),
        bottom=Side(style='thin', color="FFFFFF")
    )

    # Apply no border to the range B1:W2
    for row in range(1, 3):  # Rows 1 to 2
        for col in range(2, 24):  # Columns B (2) to W (23)
            ws.cell(row=row, column=col).border = white_border

    # Remove borders for specific cells to show gridlines
    # Define the border style
    border = Border(
        left=Side(style='thin', color='7F7F7F'),  # Thin black border on the left
        right=Side(style='thin', color='7F7F7F'),  # Thin black border on the right
        top=Side(style='thin', color='7F7F7F'),  # Thin black border on the top
        bottom=Side(style='thin', color='7F7F7F')  # Thin black border on the bottom
    )
    #no_boader list
    no_boader_list=['C1','E1','G1','J1','K1','L1','M1','O1','S1','V1','J2','K2','L2','M2']
    for i in no_boader_list:
        ws[i].border = border

    #font color
    red_font_italic= Font(color="FF0000", italic=True) 
    yellow_font = Font(color="9C5700") 
    blue_bold_font = Font(color="0563C1",bold=True) 
    ws['B4'].font = blue_bold_font
    ws['C4'].font = blue_bold_font
    ws['B4'].font = Font(bold=True)
    ws['C4'].font = Font(bold=True)
    ws['K2'].font = yellow_font
    ws['O1'].font = yellow_font
    ws['M2'].font = yellow_font

    ws['H1'].font = red_font_italic
    ws['H2'].font = red_font_italic
    ws['O2'].font = red_font_italic
    # Define the alignment (right-aligned)
    right_alignment = Alignment(horizontal='right')
    ws['H1'].alignment = right_alignment
    ws['H2'].alignment = right_alignment
    # Apply no border to the cells in the no_boader list
    bold_font_list=['C1','E1','G1','J1','K1','M1','S1','V1','L2','J2','C2','E4','F4','G4','H4',]
    for i in bold_font_list:
        ws[i].font = Font(bold=True)
    for row in ws["H3:W4"]:  # Use the dynamically calculated range
        for cell in row:
            cell.border = gridline1 

    ws.column_dimensions['c'].width = 20
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['H'].width = 25
    
    for loop in range(num_products):
        start_row = 5 + (51 * loop)
        g_value = loop + 1
        cross_ref = f"{g_value}{category.upper()}{code}"  # Ensure no spaces in category
        # Find the matching row
        matching_row = planning_df.loc[planning_df['Cross ref'].str.upper() == cross_ref]

        pid_value = matching_row['PID'].iloc[0]
        #print(pid_value)
        RLJ = matching_row['Adjusted RLJ Item'].iloc[0] 
        MKST = matching_row['Mkst'].iloc[0] # Get the first matching PID
        Current_FC_Index = matching_row['FC Index'].iloc[0] # Get the first matching PID
        Safe_Non_Safe=matching_row['Safe/Non-Safe'].iloc[0]
        Item_Code=matching_row['Item Code'].iloc[0]
        Item_Status=f"{Safe_Non_Safe}/{Item_Code}"
        Door_Count=matching_row['Door Count'].iloc[0]
        Last_Str_Cnt=matching_row['Old Door count'].iloc[0]
        Door_count_Updated=matching_row['Door count Updated'].iloc[0]
        Store_Model=matching_row['Model'].iloc[0]
        Com_Model=matching_row['Com Model'].iloc[0]
        Holiday_Bld_FC=matching_row['HolidayBuildFC'].iloc[0]
        MCYOH=matching_row['OH Units'].iloc[0]
        OO=matching_row['OO Units'].iloc[0]
        nav_OO=matching_row['nav OO'].iloc[0]
        MTD_SHIPMENTS=matching_row['MTD SHIPMENTS'].iloc[0]
        LW_Shipments=matching_row['LW Shipments'].iloc[0]
        Wks_of_Stock_OH=matching_row['Wks of Stock OH'].iloc[0]
        Wks_of_on_Proj=matching_row['Wks of on Proj'].iloc[0]
        Last_3Wks_Ships=matching_row['Last 3Wks Ships'].iloc[0]
        Vendor_Name=matching_row['Vendor Name'].iloc[0]
        Min_order=matching_row['Min order'].iloc[0]
        Proj=matching_row['Proj'].iloc[0]
        Net_Proj=matching_row['Net Proj'].iloc[0]
        Unalloc_Orders=matching_row['Unalloc Orders'].iloc[0]
        RLJ_OH=matching_row['RLJ OH'].iloc[0]
        FLDC=matching_row['FLDC'].iloc[0]
        WIP=matching_row['WIP'].iloc[0]
        MD_Status_MZ1=matching_row['MD Status MZ1'].iloc[0]
        Repl_Flag=matching_row['Repl Flag'].iloc[0]
        MCOM_RPL=matching_row['MCOM RPL'].iloc[0]
        Pool_stock=matching_row['Pool stock'].iloc[0]
        st_Rec_Date=matching_row['1st Rec Date'].iloc[0]
        Last_Rec_Date=matching_row['Last Rec Date'].iloc[0]
        Item_Age=matching_row['Item Age'].iloc[0]
        TY_Last_Cost=matching_row['TY Last Cost'].iloc[0]
        Own_Retail=matching_row['Own Retail'].iloc[0]
        AWR_1st_Tkt_Ret=matching_row['AWR 1st Tkt Ret'].iloc[0]
        Metal_Lock=matching_row['Metal Lock'].iloc[0]
        MFG_Policy=matching_row['MFG Policy'].iloc[0]
        KPI_Data_Updated=matching_row['KPI Data Updated'].iloc[0]
        KPI_Door_count=matching_row['KPI Door count'].iloc[0]
        TBL_Planning_VerticalReport__3_matching_row=TBL_Planning_VerticalReport__3.loc[TBL_Planning_VerticalReport__3['PID'].str.upper() == pid_value]
        STD_Sales = TBL_Planning_VerticalReport__3_matching_row['STD SALES'].iloc[0]
        LY_STD_SALES = TBL_Planning_VerticalReport__3_matching_row['LY STD SALES'].iloc[0] 
        OOS_Locs=matching_row['OOS Locs'].iloc[0]
        Suspended_Loc_Count=matching_row['Suspended Loc Count'].iloc[0]
        Masterstyle_Desc=matching_row['Masterstyle Desc'].iloc[0]
        Dpt_ID=matching_row['Dpt ID'].iloc[0]  
        Dpt_Desc=matching_row['Dpt Desc'].iloc[0]
        SC_ID=matching_row['SC ID'].iloc[0]
        SC_Desc=matching_row['SC Desc'].iloc[0]
        MstrSt_ID=matching_row['MstrSt ID'].iloc[0]
        Masterstyle_Desc=matching_row['Masterstyle Desc'].iloc[0]
        PID_Desc=matching_row['PID Desc'].iloc[0]
        st_Live=matching_row['1st Live'].iloc[0]
        Live_Site=matching_row['Live Site'].iloc[0]
        V2C=matching_row['V2C'].iloc[0]
        Mktg_ID=matching_row['Mktg ID'].iloc[0]
        STD_Store_Rtn=matching_row['STD Store Rtn %'].iloc[0]
        Prod_Desc=matching_row['Prod Desc'].iloc[0]
        Last_Proj_Review_Date=matching_row['Last Proj Review Date'].iloc[0]
        Macys_Recpts_matching_row=Macys_Recpts.loc[Macys_Recpts['PID'].str.upper() == pid_value]
        Macys_Spring_Proj_Notes =  f"Macy's Spring Proj Notes: {Macys_Recpts_matching_row['ACTION'].iloc[0]}" if not Macys_Recpts_matching_row.empty else "Macy's Spring Proj Notes: "
        Planner_Response=matching_row['Planner Response'].iloc[0] 

        Nav_Feb=matching_row['Feb'].iloc[0]
        Nav_Mar=matching_row['Mar'].iloc[0]
        Nav_Apr=matching_row['Apr'].iloc[0]
        Nav_May=matching_row['May'].iloc[0]
        Nav_Jun=matching_row['Jun'].iloc[0]
        Nav_Jul=matching_row['Jul'].iloc[0]
        Nav_Aug=matching_row['Aug'].iloc[0]
        Nav_Sep=matching_row['Sep'].iloc[0]
        Nav_Oct=matching_row['Oct'].iloc[0]
        Nav_Nov=matching_row['Nov'].iloc[0]
        Nav_Dec=matching_row['Dec'].iloc[0]
        Nav_Jan=matching_row['Jan'].iloc[0]

        Macys_Proj_Receipts_Feb=matching_row['FEB RECPT'].iloc[0]
        Macys_Proj_Receipts_Mar=matching_row['MAR RECPT'].iloc[0]
        Macys_Proj_Receipts_Apr=matching_row['APR RECPT'].iloc[0]
        Macys_Proj_Receipts_May=matching_row['May RECPT'].iloc[0]
        Macys_Proj_Receipts_Jun=matching_row['JUN RECPT'].iloc[0]
        Macys_Proj_Receipts_Jul=matching_row['JUL RECPT'].iloc[0]
        Macys_Proj_Receipts_Aug=matching_row['AUG RECPT'].iloc[0]
        Macys_Proj_Receipts_Sep=matching_row['SEP RECPT'].iloc[0]
        Macys_Proj_Receipts_oct=matching_row['OCT RECPT'].iloc[0]
        Macys_Proj_Receipts_Nov=matching_row['NOV RECPT'].iloc[0]
        Macys_Proj_Receipts_Dec=matching_row['DEC RECPT'].iloc[0]
        Macys_Proj_Receipts_Jan=matching_row['JAN RECPT'].iloc[0]
        MCOM_Data_matching_row=MCOM_Data.loc[MCOM_Data['PID'].str.upper() == pid_value]
        this_year_value=2024
        last_year_value=2023
        this_year_data = All_DATA.loc[(All_DATA['PID'] == pid_value) & (All_DATA['Year'] == this_year_value)]
        last_year_data = All_DATA.loc[(All_DATA['PID'] == pid_value) & (All_DATA['Year'] == last_year_value)]
        this_year_MCOM=MCOM_Data.loc[(MCOM_Data['PID'] == pid_value) & (MCOM_Data['Year'] == this_year_value)]
        last_year_MCOM=MCOM_Data.loc[(MCOM_Data['PID'] == pid_value) & (MCOM_Data['Year'] == last_year_value)]
        # Define months in order
        months = ['JAN', 'FEB', 'MAR', 'APR', 'MAY', 'JUN', 'JUL', 'AUG', 'SEP', 'OCT', 'NOV', 'DEC']
        # Initialize dictionaries to store results
        TY_Unit_Sales = {month: 0 for month in months}
        LY_Unit_Sales = {month: 0 for month in months}
        LY_OH_Units = {month: 0 for month in months}
        TY_OH_Units = {month: 0 for month in months}
        TY_Receipts = {month: 0 for month in months}
        TY_MCOM_Unit_Sales = {month: 0 for month in months}
        TY_OH_MCOM_Units={month: 0 for month in months}
        PTD_TY_Sales={month: 0 for month in months}
        MCOM_PTD_TY_Sales={month: 0 for month in months}
        LY_MCOM_Unit_Sales={month: 0 for month in months}
        LY_MCOM_OH_Units = {month: 0 for month in months}
        OO_Total_Units={month: 0 for month in months}
        OO_MCOM_Total_Units={month: 0 for month in months}
        LY_Receipts={month: 0 for month in months}
        LY_PTD_Sales={month: 0 for month in months}
        MCOM_PTD_LY_Sales={month: 0 for month in months}
        # Sum data for each month
        for month in months:
            TY_Unit_Sales[month] = this_year_data.loc[this_year_data['Month'].str.upper() == month, 'PTD TY Unit Sales'].sum()
            LY_Unit_Sales[month] = last_year_data.loc[last_year_data['Month'].str.upper() == month, 'PTD TY Unit Sales'].sum()
            LY_OH_Units[month] = last_year_data.loc[last_year_data['Month'].str.upper() == month, 'OH TY Units'].sum()
            TY_OH_Units[month] = this_year_data.loc[this_year_data['Month'].str.upper() == month, 'OH TY Units'].sum()
            TY_Receipts[month] = this_year_data.loc[this_year_data['Month'].str.upper() == month, 'PTD TY RCVD Unit'].sum()
            TY_MCOM_Unit_Sales[month] = this_year_MCOM.loc[this_year_MCOM['Month'].str.upper() == month, 'PTD TY Unit Sales'].sum()
            LY_MCOM_Unit_Sales[month] = last_year_MCOM.loc[last_year_MCOM['Month'].str.upper() == month, 'PTD TY Unit Sales'].sum()
            TY_OH_MCOM_Units[month] = this_year_MCOM.loc[this_year_MCOM['Month'].str.upper() == month, 'OH TY Units'].sum()
            PTD_TY_Sales[month] = this_year_data.loc[this_year_data['Month'].str.upper() == month, 'PTD TY $ Sales'].sum()
            LY_PTD_Sales[month] = last_year_data.loc[last_year_data['Month'].str.upper() == month, 'PTD TY $ Sales'].sum()
            MCOM_PTD_TY_Sales[month] = this_year_MCOM.loc[this_year_MCOM['Month'].str.upper() == month, 'PTD TY $ Sales'].sum()
            MCOM_PTD_LY_Sales[month] = last_year_MCOM.loc[last_year_MCOM['Month'].str.upper() == month, 'PTD TY $ Sales'].sum()
            LY_MCOM_OH_Units[month] = last_year_MCOM.loc[last_year_MCOM['Month'].str.upper() == month, 'OH TY Units'].sum()
            OO_Total_Units[month] = this_year_data.loc[this_year_data['Month'].str.upper() == month, 'OO Total Units'].sum()
            OO_MCOM_Total_Units[month] = this_year_MCOM.loc[this_year_MCOM['Month'].str.upper() == month, 'OO Total Units'].sum()
            LY_Receipts[month] = last_year_data.loc[last_year_data['Month'].str.upper() == month, 'PTD TY RCVD Unit'].sum()
        index_df.columns = index_df.columns.str.strip().str.upper()
        #print("Current_FC_Index",Current_FC_Index)
        if pd.isna(Current_FC_Index):
            Current_FC_Index = "Dia"

        index_row_data = index_df.loc[index_df['INDEX'].astype(str).str.lower() == Current_FC_Index.lower()]

        index_value = {}
        # Loop through each month and fetch its value
        for month in months:
            index_value[month] = index_row_data[month].iloc[0] if not index_row_data.empty else 0


                        #algorithm
        #######step1####################
        # Function to get Lead Time by PID
        def get_lead_time_by_pid(pid_value):
            # Find the Vendor for the given PID
            vendor = master_sheet.loc[master_sheet['PID'] == pid_value, 'Vendor Name'].values[0]
            #print('vendor',vendor)
            # Find the Lead Time for the Vendor
            country = vendor_sheet.loc[vendor_sheet['Vendor Name'] == vendor, 'Country of Origin']
            #print('country',country)
            if not country.empty:
                country = country.values[0]  # Access the first value
            else:
                country = None  # Set to None if empty
            current_month_upper = current_month.upper()
            lead_time = vendor_sheet.loc[vendor_sheet['Vendor Name'] == vendor, 'Lead Time(weeks)']
            lead_time=lead_time.values[0] if not lead_time.empty else 8
            if current_month_upper=='NOV' or current_month_upper=='JAN' or current_month_upper=='DEC':
                if country=='China':
                    lead_time=10
            if current_month_upper=='MAY' or current_month_upper=='JUN' or current_month_upper=='JULY':
                if country=='Italy':
                    lead_time=14
            return lead_time 
        lead_time = get_lead_time_by_pid(pid_value)
        #print('lead_time',lead_time) 
            # Function to get the 3-letter month abbreviation based on the month number
        # Retail calendar month sequence
        retail_months = ["Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Jan"]

        def calculate_forecast_months(lead_time_weeks, current_month):
            # Handle NaN or None for lead_time_weeks by setting a default value of 8
            if lead_time_weeks is None or (isinstance(lead_time_weeks, float) and math.isnan(lead_time_weeks)):
                lead_time_weeks = 8

            # Calculate months from weeks (approx. 4 weeks in a month)
            lead_time_months = math.ceil(lead_time_weeks / 4)  # Use math.ceil to round up to the next full month

            # Get the index of the current month in the retail calendar
            current_index = retail_months.index(current_month)

            # List to store the forecast months
            forecast_months = []

            # Add months to the current month, starting from the next month
            for i in range(1, lead_time_months + 1):
                new_index = (current_index + i) % len(retail_months)  # Calculate the next month's index
                forecast_months.append(retail_months[new_index])

            return forecast_months


        forecast_months = calculate_forecast_months(lead_time,current_month)
        #print('forecast_months',forecast_months)
        retail_months = ["Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Jan"]
        #Step 2 :  Find STD period###################
        def get_std_months(current_month):
            # Retail calendar month sequence
            
            
            # Define Spring and Fall seasons
            spring_season = ["Feb", "Mar", "Apr", "May", "Jun", "Jul"]
            fall_season = ["Aug", "Sep", "Oct", "Nov", "Dec", "Jan"]
            
            # Get the index of the current month in the retail calendar
            current_index = retail_months.index(current_month)
            
            # Determine the season and the corresponding std period length
            if current_month in spring_season:
                std_period = 5  # Last 5 months for Spring
            elif current_month in fall_season:
                std_period = 4  # Last 3 months for Fall
            else:
                raise ValueError("Invalid month provided.")

            # Calculate the std months
            std_months = []
            for i in range(1, std_period + 1):
                std_index = (current_index - i) % len(retail_months)
                std_months.insert(0, retail_months[std_index])  # Insert at the beginning to maintain order
            
            return std_months

        def convert_month_to_abbr(full_month):
            # Define mapping of full month names to abbreviations
            month_mapping = {
                "January": "Jan",
                "February": "Feb",
                "March": "Mar",
                "April": "Apr",
                "May": "May",
                "June": "Jun",
                "July": "Jul",
                "August": "Aug",
                "September": "Sep",
                "October": "Oct",
                "November": "Nov",
                "December": "Dec"
            }
            
            # Return the corresponding abbreviation, or None if not found
            return month_mapping.get(full_month, None)
        
        months_new = ['FEB', 'MAR', 'APR', 'MAY', 'JUN', 'JUL', 'AUG', 'SEP', 'OCT', 'NOV', 'DEC', 'JAN']
        def get_month_range(months, month_start, endMonth):
            # Find the start and end indices in the months list
            start_index = months.index(month_start)
            end_index = months.index(endMonth)
        
            # Handle the circular nature of the list
            if end_index < start_index:
                # Rolling over to the next year
                selected_months = months[start_index:] + months[:end_index + 1]
            else:
                # Normal range
                selected_months = months[start_index:end_index + 1]
        
            return selected_months
        if month_from and month_to:
            first_month_std=convert_month_to_abbr(month_from).upper()
            last_mont_std=convert_month_to_abbr(month_to).upper()
            std_months = get_month_range(months_new, first_month_std, last_mont_std)
        else:
            std_months = get_std_months(current_month)

        #print(std_months)
        std_months_upper = [month.upper() for month in std_months]
        # TY_Unit_Sales_list=[TY_Unit_Sales['FEB'],TY_Unit_Sales['MAR'],TY_Unit_Sales['APR'],TY_Unit_Sales['MAY'],TY_Unit_Sales['JUN'],TY_Unit_Sales['JUL'],TY_Unit_Sales['AUG'],TY_Unit_Sales['SEP'],TY_Unit_Sales['OCT'],TY_Unit_Sales['NOV'],TY_Unit_Sales['DEC'],TY_Unit_Sales['JAN']]
        STD_TY_Unit_Sales_list = [TY_Unit_Sales[month] for month in std_months_upper]
        STD_LY_Unit_Sales_list=[LY_Unit_Sales[month] for month in std_months_upper]
        STD_index_value=[index_value[month] for month in std_months_upper]
        STD_index_value = sum(STD_index_value)

        if STD_index_value:
            month_12_fc_index = sum(STD_TY_Unit_Sales_list) / (STD_index_value )
        else:
            month_12_fc_index = 0
        STD_index_value=round(STD_index_value, 2)
        #print("STD Index:", STD_index_value)
        #print("Month 12 FC:", month_12_fc_index)

        if STD_LY_Unit_Sales_list:
            std_trend = (sum(STD_TY_Unit_Sales_list) - sum(STD_LY_Unit_Sales_list)) / sum(STD_LY_Unit_Sales_list)
        else:
            std_trend = 0
        std_trend=round(std_trend, 2)
        #print("std_trend:", std_trend)
        # Get the door count from cell C10
        retail_months_upper = [month.upper() for month in retail_months]
        def check_product_type( ):
            
            All_TY_Unit_Sales_list=[LY_Unit_Sales[month] for month in retail_months_upper]
            ALL_LY_MCOM_Unit_Sales=[LY_MCOM_Unit_Sales[month] for month in retail_months_upper]
            result_list = [
                (ly_mcom / ty_unit if ty_unit != 0 else 0)
                for ly_mcom, ty_unit in zip(ALL_LY_MCOM_Unit_Sales, All_TY_Unit_Sales_list)
            ]
            # Get the COM to TTL % Sales for this year (from I20 to T20)
    
            # Calculate the average COM to TTL % Sales for this year
            average = sum(result_list) / len(result_list) if result_list else 0 
            average_com_to_ttl_sales=average*100
    
            # Determine if the product is COM or Store based on door count and COM to TTL % Sales
            if KPI_Door_count is None or  KPI_Door_count <= 1 or average_com_to_ttl_sales > 65 or None:
                return True
            else:
                return False
    
        is_com_product = check_product_type()
                
            # #print the result
        #print("Is COM Product:", is_com_product)
        if not is_com_product:
            store_products.append(pid_value) 

            # #############################################
            # #############       step 5       ############      
            # #############################################
    
            STD_TY_OH_Units_list = [TY_OH_Units[month] for month in std_months_upper]
            STD_TY_OH_MCOM_Units_list=[TY_OH_MCOM_Units[month] for month in std_months_upper]

            this_year_std_store_eom_oh = [
                (ty_oh_unit - ty_oh_mcom )
                for ty_oh_unit,ty_oh_mcom  in zip(STD_TY_OH_Units_list, STD_TY_OH_MCOM_Units_list)
            ]
                
            STD_LY_OH_Units_list = [LY_OH_Units[month] for month in std_months_upper]
            STD_LY_OH_MCOM_Units_list=[LY_MCOM_OH_Units[month] for month in std_months_upper]

            last_year_std_store_eom_oh = [
                (ly_oh_unit - ly_oh_mcom )
                for ly_oh_unit,ly_oh_mcom  in zip(STD_LY_OH_Units_list, STD_LY_OH_MCOM_Units_list)
            ]

            spring_months = ['FEB', 'MAR', 'APR', 'MAY', 'JUN', 'JUL']
            season = "Spring" if forecast_months[0].upper() in spring_months else "Fall"
            def calculate_column_values(s1, k1,months,f8, row4_values, row17_values, row39_values):
                results = {}
                for month,row4, row17, row39 in zip(months,row4_values, row17_values, row39_values):
                    if s1 == 12 and k1 == 12:
                        result = round(row17 + row17 * f8, 0)
                    elif s1 == 12 and row4 < 7:
                        result = round(row39 + row39 * f8, 0)
                    elif s1 > 6 and row4 < 7:
                        result = round(row17 + row17 * f8, 0)
                    elif s1 < 7 and row4 > 6:
                        result = round(row39 + row39 * f8, 0)
                    elif s1 < 7 and row4 < 7:
                        result = round(row39 + row39 * f8, 0)
                    elif s1 > 6 and row4 > 6:
                        result = round(row39 + row39 * f8, 0)
                    else:
                        result = None  # Fallback case
            
                    results[month] = result
            
                return results
            s1=last_month_of_previous_month_numeric
            k1=current_month_number
            f8=std_trend
            row4_values=[i+1 for i in range(12)]
            row17_values=[TY_Unit_Sales[month] for month in retail_months_upper]
            row39_values=[LY_Unit_Sales[month] for month in retail_months_upper]
            fc_by_trend_all = calculate_column_values(s1, k1, retail_months_upper, f8, row4_values, row17_values, row39_values)
            spring_fc_by_trend_all = sum([fc_by_trend_all[month] for month in ['FEB', 'MAR', 'APR', 'MAY', 'JUN', 'JUL']])
            fall_fc_by_trend_all = sum([fc_by_trend_all[month] for month in ['AUG', 'SEP', 'OCT', 'NOV', 'DEC', 'JAN']])
            all_index_value=[index_value[month] for month in retail_months_upper]
            fc_by_index_all={}
            for i,month in enumerate(retail_months_upper):
                fc_by_index_all[month] = round((all_index_value[i]*month_12_fc_index),0)


            spring_fc_by_index_all = sum([fc_by_index_all[month] for month in ['FEB', 'MAR', 'APR', 'MAY', 'JUN', 'JUL']])
            fall_fc_by_index_all = sum([fc_by_index_all[month] for month in ['AUG', 'SEP', 'OCT', 'NOV', 'DEC', 'JAN']])
            spring_months = ["Feb", "Mar", "Apr", "May", "Jun", "Jul"]
            fall_months   = ["Aug", "Sep", "Oct", "Nov", "Dec", "Jan"]


            # 3. Count how many forecast months belong to Spring vs. Fall
            spring_count = sum(month in spring_months for month in forecast_months)
            fall_count   = sum(month in fall_months   for month in forecast_months)

            # 4. Determine which season is in the majority
            if spring_count > fall_count:
                selected_season = "Spring"
            elif fall_count > spring_count:
                selected_season = "Fall"
            else:
                selected_season = season

            fc_by_index = spring_fc_by_index_all if selected_season == "Spring" else fall_fc_by_index_all
            fc_by_trend = spring_fc_by_trend_all if selected_season == "Spring" else fall_fc_by_trend_all
            #print('selected_season',selected_season)
            #print('fc_by_index',fc_by_index)
            #print('fc_by_trend',fc_by_trend)

            if fc_by_index is not None and fc_by_trend is not None:
                difference = (abs(fc_by_trend - fc_by_index) / max(fc_by_index,fc_by_trend))* 100 
            #print('difference',difference)
            def find_forecasting_method_index(difference, door_count, this_year_std_store_eom_oh, month_12_fc_index):
                # Check if the difference is more than 15%
                if difference > 15:            
                    average_value = sum(this_year_std_store_eom_oh) / len(this_year_std_store_eom_oh)
                    loss = (door_count / average_value) - 1
                    #print("loss : ",loss)
                    rank =Item_Code

                    #print('rank',rank)
                    if rank == 'A' or rank == 'B':
                        loss_percent = min(loss, 0.45)  # Maximum of 45% for Rank A or B
                    elif rank == 'C':
                        loss_percent = min(loss, 0.15)  # Maximum of 15% for Rank C
                    else:
                        loss_percent = loss  # No loss for Rank D or E
                    #print('loss_percent',loss_percent)
                    month_12_fc_index = round(month_12_fc_index * (1 + loss_percent))
                    forecasting_method = "FC by Index"  # Write the value to the merged cell
                else:

                    forecasting_method = "FC by Index"  # Write the value to the merged cell
                return  forecasting_method  ,  month_12_fc_index



            def find_forecasting_method_trend(difference):
                    # Check if the difference is more than 15%
                    if difference > 15:            
                        forecasting_method = "FC by Trend"  # Write the value to the merged cell
            
                    else:
                        forecasting_method = "Average"  # Write the value to the merged cell
                    return  forecasting_method
            
            def find_method(difference,this_year_std_store_eom_oh, last_year_std_store_eom_oh, door_count, month_12_fc_index):
                # Define the threshold percentage for "maintained"
                threshold = 0.95  # 5% difference
                # Function to calculate the average of a list and check if it's maintained
                def is_maintained(eom_oh_list):
                    average_eom_oh = sum(eom_oh_list) / len(eom_oh_list) if eom_oh_list else 0
                    # Check if the average is within the 5% down range or more than the door count
                    return (average_eom_oh >= threshold * door_count) or (average_eom_oh > door_count) 
        
                # Check if this year's and last year's Store EOM OH are maintained
                last_year_maintained = is_maintained(last_year_std_store_eom_oh)
                this_year_maintained = is_maintained(this_year_std_store_eom_oh)
        
                #print("last year_maintained ", last_year_maintained)
                #print("this year maintained ", this_year_maintained)
    
                # Now check the three conditions:
                if last_year_maintained and this_year_maintained:
                    # "Last Year => Store EOM OH not maintained, This Year => Store EOM OH not maintained"
                    forecasting_method=find_forecasting_method_trend(difference)
                else:
                    forecasting_method , month_12_fc_index=find_forecasting_method_index(difference, door_count, this_year_std_store_eom_oh, month_12_fc_index)
                return forecasting_method , month_12_fc_index
            forecasting_method , month_12_fc_index=find_method(difference,this_year_std_store_eom_oh, last_year_std_store_eom_oh, KPI_Door_count, month_12_fc_index)
            #print('forecasting_method',forecasting_method)
            #print('month_12_fc_index',month_12_fc_index)
            # Preprocess the dictionaries to replace NaN or Infinity with 0
            fc_by_index_all = {key: (value if not (math.isnan(value) or math.isinf(value)) else 0) for key, value in fc_by_index_all.items()}
            fc_by_trend_all = {key: (value if not (math.isnan(value) or math.isinf(value)) else 0) for key, value in fc_by_trend_all.items()}

            # Compute the average
            fc_by_average_all = {
                key: round((fc_by_index_all[key] + fc_by_trend_all[key]) / 2) for key in fc_by_trend_all
            }
            #print('fc_by_trend_all',fc_by_trend_all)
            #print('fc_by_index_all',fc_by_index_all)
            #print('fc_by_average_all',fc_by_average_all)
            if forecasting_method == "FC by Index":
                recommended_fc = fc_by_index_all
            elif forecasting_method == "FC by Trend":
                recommended_fc = fc_by_trend_all
            else:
                recommended_fc = fc_by_average_all
            #print("recommended_fc ",recommended_fc)

            current_month_upper = current_month.upper()
                    # Define input data for each row
            row_4 = row4_values  # Example values for row 4
            row_9 = recommended_fc
            row_17 = TY_Unit_Sales
            row_43 = LY_OH_Units
            # Define variables
            V1 = rolling_method  
            #print('rolling_method',rolling_method) # Example: could be "YTD", "CURRENT MTH", "SPRING", etc.
            K1 = current_month_number       # A reference column value (e.g., month or index)
            # List of columns (iterable from I to T in your example)
            
            # Initialize results
            planned_fc = {}
            # Iterate through columns and calculate values
            for idx, col in enumerate(retail_months_upper):
                J4 = row_4[idx]  # Get value from row_4 list based on index
                J17 = row_17[col]
                J43 = row_43[col]
                J9 = row_9[col]
            
                if V1 == "YTD" and J4 < K1:
                    planned_fc[col] = J17
                elif V1 == "YTD" and J4 == K1:
                    planned_fc[col] = J17
                elif V1 == "CURRENT MTH" and J4 == K1:
                    planned_fc[col] = J17
                elif V1 == "SPRING" and J4 < 7:
                    planned_fc[col] = J17
                elif V1 == "FALL" and J4 > 6:
                    planned_fc[col] = J17
                elif V1 == "LY FALL" and J4 > 6:
                    planned_fc[col] = J43
                else:
                    planned_fc[col] = J9
            current_month_upper = current_month.upper()
            # #print the results dictionary
            #print("planned_fc:", planned_fc)
            #print("This Year sales",TY_Unit_Sales)
            current_month_sales_percentage = int(percentage)
            current_month_upper = current_month.upper()
            planned_fc[current_month_upper] =  round(planned_fc[current_month_upper] / (current_month_sales_percentage/100))
            #print("planned_fc_current%_re",planned_fc)
            #compare
            forecast_months_upper = [month.upper() for month in forecast_months]
            if forecasting_method=="FC by Trend" or forecasting_method=="Average":
                
                trend_std={key:fc_by_trend_all[key] for key in forecast_months_upper}
                avg_std={key:fc_by_average_all[key] for key in forecast_months_upper}
                last_year_std={key:LY_Unit_Sales[key] for key in forecast_months_upper}
                #print("last_year sales unit for forecast month",last_year_std)
                for key in last_year_std.keys():
                    if key in trend_std and key in avg_std:
                        # Calculate the difference from last_year_std
                        trend_diff = abs(trend_std[key] - last_year_std[key])
                        avg_diff = abs(avg_std[key] - last_year_std[key])

                        # Find the closest value
                        closest_value = trend_std[key] if trend_diff < avg_diff else avg_std[key]

                        # Update planned_fc
                        planned_fc[key] = closest_value

                    # #print updated planned_fc
            #print('planned_fc_comprewithlast',planned_fc)
            currnt_month=current_month_upper
            D13 = OO  # Replace with your value for D13
            E19 = nav_OO  # Replace with your value for E19
            in_transit=0 if D13 - E19 < 0 else D13 - E19
            #print("in_transit",in_transit)
            # Calculate using the logic in the Excel formula
            planned_shp=[Nav_Feb,Nav_Mar,Nav_Apr,Nav_May,Nav_Jun,Nav_Jul,Nav_Aug,Nav_Sep,Nav_Oct,Nav_Nov,Nav_Dec,Nav_Jan]
            planned_shp={key:planned_shp[i] for i,key in enumerate(retail_months_upper)}
            #print('planned_shp',planned_shp)
            planned_shp[current_month_upper] += in_transit
            #print('planned_shpadd_in_trasit',planned_shp)
            v1=rolling_method
            k1=current_month_number
            i4=row_4
            row_10=planned_fc
            row_11=planned_shp
            row17=TY_Unit_Sales
            TY_OH_Units['JAN'] = TY_OH_Units.pop('JAN')
            row_21=TY_OH_Units
            TY_Receipts['JAN'] = TY_Receipts.pop('JAN')
            row_37=TY_Receipts
            LY_OH_Units['JAN'] = LY_OH_Units.pop('JAN')
            row_43=LY_OH_Units

            row_17=TY_Unit_Sales
                    # Create a dictionary for Birthstones and corresponding months based on the Birthstone sheet
            # Initialize output dictionary for Planned EOH (Cal)

            # Loop through months
            def calculate_row_12(v1, k1, i4, row_10, row_11, row_21, row_37, row_43, row_17, current_month):
                # Define the fixed column-to-month mapping
                month_order = ['FEB', 'MAR', 'APR', 'MAY', 'JUN', 'JUL', 'AUG', 'SEP', 'OCT', 'NOV', 'DEC', 'JAN']
                month_to_index = {month: idx + 1 for idx, month in enumerate(month_order)}  # FEB=1, ..., JAN=12
            
                # Rearrange month_order to start with the current_month
                start_idx = month_order.index(current_month)
                reordered_months = month_order[start_idx:] + month_order[:start_idx]
            
                # Initialize the output dictionary for row_12
                plan_oh = {}
            
                for idx, month in enumerate(reordered_months):
                    # Determine the previous month (wrapping around if necessary)
                    prev_month = reordered_months[idx - 1] if idx > 0 else reordered_months[-1]
            
                    # Get the corresponding month number
                    col4 = month_to_index[month]  # Month number based on the fixed mapping
                    col10 = row_10.get(month, 0)
                    col11 = row_11.get(month, 0)
                    col21 = row_21.get(month, 0)
                    col37 = row_37.get(month, 0)
                    col43 = row_43.get(month, 0)
                    col17 = row_17.get(month, 0)
            
                    # Get the previous column's value
                    prev_col12 = plan_oh.get(prev_month, 0)
                    prev_col21 = row_21.get(prev_month, 0)
            
                    # Apply the logic for calculation
                    if v1 == "Current Mth" and k1 == col4 and col4 == 1:
                        val = row_43['JAN'] + col11 + col37 - col10
                    elif v1 == "Current Mth" and k1 == col4:
                        val = col21 + col11 - (col10 - col17)
                    elif v1 == "Current Mth" and col4 == 1 and k1 > 1:
                        val = plan_oh['JAN'] + col11 - col10
                    elif v1 == "YTD" and col4 < k1:
                        val = col21
                    elif v1 == "Spring" and col4 < 7:
                        val = col21
                    elif v1 == "Fall" and col4 > 6 and col4 < k1:
                        val = col21
                    elif v1 == "Fall" and col4 == 1 and k1 > 1:
                        val = plan_oh['JAN'] + col11 - col10
                    elif v1 == "Fall" and col4 > 6 and col4 == k1:
                        val = prev_col21 + col37 + col11 - col10
                    elif v1 == "LY Fall" and col4 > 6:
                        val = col43
                    elif col4 == 1 and v1 == "LY FALL":
                        val = row_43['JAN'] + col11 + col11 - col10
                    elif col4 == k1 and col4 > 1:
                        val = prev_col21 + col37 + col11 - col10
                    else:
                        val = prev_col12 + col11 - col10
            
                    # Store the calculated value in row_12 for the current month
                    plan_oh[month] = val
            
                return plan_oh  
            plan_oh = calculate_row_12(v1, k1, i4, row_10, row_11, row_21, row_37, row_43, row_17, current_month_upper)
            #print('plan_eoh',plan_oh)
            required_quantity_month_dict = {}

                # #print intermediate results for debugging


            # Check if the filtered DataFrame is not empty
            bsp_row = master_sheet[master_sheet['PID'] == pid_value]
            if not bsp_row.empty:
                bsp_or_not = bsp_row['BSP_or_not'].iloc[0]
                bsp_or_not = str(bsp_or_not).strip().upper() if bsp_or_not else None
                # Step 2: Determine BSP_status
                bsp_status = 'BSP' if bsp_or_not == 'BSP' else 'non_bsp'
                #print('bsp_status',bsp_status)
                # Step 3: Find birthstone_month if BSP
                birthstone_month = None
                if bsp_status == 'BSP':
                    category_bsp=bsp_row['category'].iloc[0]
                    birthstone = bsp_row['Birthstone'].iloc[0] if bsp_or_not == 'BSP' else None
                    #print('category_bsp',category_bsp)
                    #print('birthstone',birthstone)
                    birthstone = str(birthstone).strip() 
                    birthstone_sheet = birthstone_sheet.dropna(subset=['Birthstone'])
                    birthstone_sheet['Birthstone'] = birthstone_sheet['Birthstone'].astype(str)
                    birthstone_row = birthstone_sheet[birthstone_sheet['Birthstone'].str.upper() == birthstone.upper()]

                    if not birthstone_row.empty:
                        birthstone_month = birthstone_row['Month Name'].iloc[0]
                        #print('birthstone_month',birthstone_month)
                    if birthstone_month:
                        # Determine the previous month of the birthstone month
                        months_list = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
                        birthstone_month_abb = convert_month_to_abbr(birthstone_month)
                        previous_month = months_list[months_list.index(birthstone_month_abb) - 1]
                        # Check if the previous month is in the forecast months
                        for month in forecast_months:

                            #print(f"Processing forecast month: {month}")
                            if month == previous_month or month == 'Nov':  # Apply special condition for November
                                if category == 'Studs':
                                    required_quantity = 3 * KPI_Door_count  # For Studs, use 3 per door count
                                elif category == 'Pendants':
                                    required_quantity = 2 * KPI_Door_count  # For Pendants, use 2 per door count
                                birthstone_list.append(pid_value)
                            else:
                                required_quantity = KPI_Door_count  # If not the previous month or November, use door count only

                            # Store the result in the dictionary
                            required_quantity_month_dict[month] = required_quantity
                            #print(f"if  bsp Updated dictionary: {required_quantity_month_dict}")
                    # Fallback: Populate the dictionary with `forecast_months` and `door_count`
            if not required_quantity_month_dict:
                #print(f"PID not found or no BSP logic executed. Populating default values...")
                for month in forecast_months:
                    required_quantity_month_dict[month] = KPI_Door_count
            #print(f"Updatedf required_quantity_month_dict: {required_quantity_month_dict}")
            average_store_eom_oh = []
            #print(forecast_months)
            months_list = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
            # Find the last month in the forecast_months list
            last_forecast_month = forecast_months[-1]
            # Find the index of the last forecast month in months_list
            last_month_index = months_list.index(last_forecast_month)
            next_month_fldc = months_list[(last_month_index + 1) % len(months_list)]
            #print("Next month after the forecast months:", next_month_fldc)
            next_month_fldc=next_month_fldc.upper()
            Calculate_FLDC=round((planned_fc[next_month_fldc])/2)
            #print('Calculate_FLDC_for_next_month',Calculate_FLDC)
            std_ly_store_eom_oh=[LY_MCOM_OH_Units[month] for month in forecast_months_upper]
            #print('lastyear avg COM EOM OH for forecastmonth ',std_ly_store_eom_oh)
            value_to_add = round(((sum(std_ly_store_eom_oh)/len(std_ly_store_eom_oh))),0)
            #print("Extra value to add : ",value_to_add)
            required_quantity_month_dict = {key: val + value_to_add for key, val in required_quantity_month_dict.items()}
            required_quantity_month_dict[last_forecast_month]=required_quantity_month_dict[last_forecast_month]+Calculate_FLDC
            #print("Updated final required_quantity_month_dict",required_quantity_month_dict)
            #review
            required_quantity_month_dict = {key.upper(): value for key, value in required_quantity_month_dict.items()}
            for month, oh_value in plan_oh.items():
                if oh_value < KPI_Door_count:
                    pids_below_door_count.append(pid_value)  # Append the PID to the list if condition is met
                    break
                    # Loop through each month to check the conditions
            for i, month in enumerate(forecast_months_upper):
                # Step 1: Check if planned_oh for the month is less than the required quantity for that month
                required_quantity = required_quantity_month_dict.get(month, 0)  # Get the required quantity for the month
                #print(month,required_quantity)
                if plan_oh[month] < required_quantity:
                    # Calculate the difference between required quantity and planned OH
                    difference = required_quantity - plan_oh[month]            
                    # Update the gross_projection for that month with the calculated difference
                    planned_shp[month] += difference
                    plan_oh[month] += difference
            
                    #print(f"Updated gross projection for {month}: {planned_shp[month]}")
                else:
                    print(f"No update needed for {month} as planned_oh is greater than required quantity.")
        
                #print("i : ",i)
                if i+1 <= len(forecast_months_upper)-1:
                    next_month = forecast_months_upper[i+1]

                    # Update planned_oh for next month using the formula
                    #print("Next month:", next_month)
                    plan_oh[next_month] = (planned_shp[next_month] + plan_oh[month]) - planned_fc[next_month]
                    #print(f"Updated planned_oh for {next_month}: {plan_oh[next_month]}")
            # #print final values for verification
            #print('planned_fc',planned_fc)
            #print("Final planned_oh:", plan_oh)
            #print("Final gross_projection:", planned_shp)
            planned_shipments = planned_shp
            macys_proj_receipt=[Macys_Proj_Receipts_Feb,Macys_Proj_Receipts_Mar,Macys_Proj_Receipts_Apr,Macys_Proj_Receipts_May,Macys_Proj_Receipts_Jun,Macys_Proj_Receipts_Jul,Macys_Proj_Receipts_Aug,Macys_Proj_Receipts_Sep,Macys_Proj_Receipts_oct,Macys_Proj_Receipts_Nov,Macys_Proj_Receipts_Dec,Macys_Proj_Receipts_Jan]
            macys_proj_receipt={key:macys_proj_receipt[i] for i,key in enumerate(retail_months_upper)}
            omni_receipts=TY_Receipts
            
            # 3) Write a helper function to sum a dictionary's values by season
            def season_sum(data_dict, season_months):
                return sum(data_dict.get(month, 0) for month in season_months)
            
            # 4) Calculate sums for each season
            # --- Spring ---
            planned_spring = season_sum(planned_shipments, spring_months)
            omni_spring    = season_sum(omni_receipts, spring_months)
            macys_spring   = season_sum(macys_proj_receipt, spring_months)
            
            # --- Fall ---
            planned_fall = season_sum(planned_shipments, fall_months)
            omni_fall    = season_sum(omni_receipts, fall_months)
            macys_fall   = season_sum(macys_proj_receipt, fall_months)
            
            # 5) Calculate the combined (planned + omni) for each season
            combined_spring = planned_spring + omni_spring
            combined_fall   = planned_fall + omni_fall
            # 6) Compare Macys Projection vs. the Combined Total
            #    If Macys Projection is less than the combined total, #print a warning
            if macys_spring < combined_spring or macys_fall < combined_fall:
                notify_macys.appned[pid_value]
            
            # 7) Optional: #print out all the final values to see at a glance
            #print("\n--- Season Summaries ---")
            #print(f"Planned Shipments Spring: {planned_spring}, Fall: {planned_fall}")
            #print(f"Omni Receipts Spring:    {omni_spring},    Fall: {omni_fall}")
            #print(f"Macys Receipts Spring:   {macys_spring},   Fall: {macys_fall}")
            #print(f"Combined Spring:         {combined_spring}")
            #print(f"Combined Fall:           {combined_fall}")
        dynamic_formulas = {
            f"G{start_row + 34}": g_value,
            f"F{start_row + 1}": f"=C{start_row + 1}",
            f"C{start_row}": pid_value,
            f"D{start_row}": RLJ,
            f"F{start_row}":MKST,
            f"O1":std_months_upper[0].upper()+"-"+std_months_upper[-1],
            f"C{start_row + 1}":Current_FC_Index,
            f"C{start_row + 2}": sum(STD_TY_Unit_Sales_list),
            f"D{start_row + 2}": sum(STD_LY_Unit_Sales_list),
            f"C{start_row + 3}": f"=IFERROR(ROUND((C{start_row + 2}-D{start_row + 2})/D{start_row + 2},2),0)",
            f"F{start_row + 3}": std_trend,
            f"E{start_row + 3}": "Chg Trend",
            f"D{start_row + 1}": "Change Index",
            f"E{start_row + 2}": STD_index_value,
            f"F{start_row + 2}": month_12_fc_index,  # Added IFERROR to prevent divide-by-zero issues
            f"F{start_row + 4}": forecasting_method,
            f"C{start_row + 4}":  Item_Status,
            f"C{start_row + 5}": Door_Count,
            f"D{start_row + 5}": Last_Str_Cnt,
            f"E{start_row + 5}": f"=ROUND(C{start_row + 5}-D{start_row + 5},2)",
            f"F{start_row + 5}": Door_count_Updated,
            f"C{start_row + 6}": Store_Model,
            f"D{start_row + 6}": Com_Model,
            f"E{start_row + 6}": f"=ROUND(C{start_row + 6}+D{start_row + 6},2)",
            f"G{start_row + 7}": Holiday_Bld_FC,
            f"C{start_row + 8}":MCYOH,
            f"D{start_row + 8}": OO,
            f"E{start_row + 8}": f"=IF(D{start_row + 8}-E{start_row + 14}<0,0,D{start_row + 8}-E{start_row + 14})",
            f"E{start_row + 14}": nav_OO,
            f"F{start_row + 8}": MTD_SHIPMENTS,
            f"G{start_row + 8}":LW_Shipments,
            f"C{start_row + 9}": Wks_of_Stock_OH,
            f"D{start_row + 9}":Wks_of_on_Proj,
            f"F{start_row + 9}": Last_3Wks_Ships,
            f"C{start_row + 12}": "=0",
            f"C{start_row + 13}": Vendor_Name,
            f"G{start_row + 13}": Min_order,
            f"C{start_row + 14}": Proj,
            f"D{start_row + 14}": Net_Proj,
            f"F{start_row + 14}": Unalloc_Orders,
            f"G{start_row + 14}": f"=IF(C{start_row + 5}>G{start_row + 5},(C{start_row + 8}+D{start_row + 8})-C{start_row + 5},IF(OR(E{start_row + 6}>C{start_row + 5},E{start_row + 6}>G{start_row + 5}),(C{start_row + 8}+D{start_row + 8})-E{start_row + 6},(C{start_row + 8}+D{start_row + 8})-G{start_row + 5}))",
            f"C{start_row + 15}": RLJ_OH,
            f"D{start_row + 15}": FLDC,
            f"E{start_row + 15}": WIP,
            f"C{start_row + 16}": f"=C{start_row + 14}+F{start_row + 14}-C{start_row + 15}-E{start_row + 15}",
            f"D{start_row + 16}": f"=IF(AND(F{start_row + 14}>=C{start_row + 16},C{start_row + 16}>0),\"Demand due to Unalloc Sales Orders\",IF(C{start_row + 16}<0,\"Excess OH Units\",IF(AND(C{start_row + 16}>0,F{start_row + 14}<C{start_row + 16}),\"Demand\",\"\")))",
            f"C{start_row + 17}": MD_Status_MZ1,
            f"D{start_row + 17}":Repl_Flag,
            f"E{start_row + 17}": MCOM_RPL,
            f"F{start_row + 17}": Pool_stock,
            f"C{start_row + 18}": st_Rec_Date,
            f"D{start_row + 18}": Last_Rec_Date,
            f"E{start_row + 18}": Item_Age,
            f"F{start_row + 18}": f"=IFERROR((NOW()-VALUE(C{start_row + 18}))/30,0)",
            f"C{start_row + 19}": TY_Last_Cost,
            f"D{start_row + 19}": Own_Retail,
            f"E{start_row + 19}": AWR_1st_Tkt_Ret,
            f"F{start_row + 19}": f"=(D{start_row + 19}-C{start_row + 19})/D{start_row + 19}",
            f"C{start_row + 20}": Metal_Lock,
            f"D{start_row + 20}":MFG_Policy,
            f"C{start_row + 21}": KPI_Data_Updated,
            f"C{start_row + 22}": KPI_Door_count,
            f"C{start_row + 23}": f"=C{start_row + 22}-C{start_row + 5}",
            f"C{start_row + 24}": OOS_Locs,
            f"C{start_row + 25}": Suspended_Loc_Count,
            f"D{start_row + 21}": f"=SUBSTITUTE(D{start_row},\"/\",\"\")",
            f"B{start_row + 26}": f"=HYPERLINK(\"http://www.macys.com/shop/product/\"&C{start_row + 32}&\"?ID=\"&E{start_row + 31},\"Click to View online\")",
            f"C{start_row + 27}": Dpt_ID,
            f"D{start_row + 27}": Dpt_Desc,
            f"C{start_row + 28}":SC_ID,
            f"D{start_row + 28}": SC_Desc,
            f"C{start_row + 29}": MstrSt_ID,
            f"D{start_row + 29}":Masterstyle_Desc,
            f"C{start_row + 30}":PID_Desc,
            f"C{start_row + 31}": st_Live,
            f"D{start_row + 31}": f"{Live_Site}/{V2C}",
            f"E{start_row + 31}": Mktg_ID,
            f"F{start_row + 31}": STD_Store_Rtn,
            f"C{start_row + 32}":Prod_Desc,
            f"C{start_row + 33}":Last_Proj_Review_Date,
            f"D{start_row + 35}": Macys_Spring_Proj_Notes,
            f"B{start_row + 39}": "\"Past Review Comments\"",
            f"B{start_row + 40}": Planner_Response,
            f"I{start_row + 1}": index_value['FEB'],
            f"J{start_row + 1}": index_value['MAR'],
            f"K{start_row + 1}": index_value['APR'],
            f"L{start_row + 1}": index_value['MAY'],
            f"M{start_row + 1}": index_value['JUN'],
            f"N{start_row + 1}": index_value['JUL'],
            f"O{start_row + 1}": index_value['AUG'],
            f"P{start_row + 1}": index_value['SEP'],
            f"Q{start_row + 1}": index_value['OCT'],
            f"R{start_row + 1}": index_value['NOV'],
            f"S{start_row + 1}": index_value['DEC'],
            f"T{start_row + 1}": index_value['JAN'],
            f"U{start_row + 1}": f"=SUM(I{start_row + 1}:T{start_row + 1})",
            f"V{start_row + 1}": f"=IFERROR(SUM(I{start_row + 1}:N{start_row + 1}),0)",
            f"W{start_row + 1}": f"=IFERROR(SUM(O{start_row + 1}:T{start_row + 1}),0)",
            f"I{start_row + 2}": f"=ROUND($F{start_row + 2}*I{start_row + 1},0)",
            f"J{start_row + 2}": f"=ROUND($F{start_row + 2}*J{start_row + 1},0)",
            f"K{start_row + 2}": f"=ROUND($F{start_row + 2}*K{start_row + 1},0)",
            f"L{start_row + 2}": f"=ROUND($F{start_row + 2}*L{start_row + 1},0)",
            f"M{start_row + 2}": f"=ROUND($F{start_row + 2}*M{start_row + 1},0)",
            f"N{start_row + 2}": f"=ROUND($F{start_row + 2}*N{start_row + 1},0)",
            f"O{start_row + 2}": f"=ROUND($F{start_row + 2}*O{start_row + 1},0)",
            f"P{start_row + 2}": f"=ROUND($F{start_row + 2}*P{start_row + 1},0)",
            f"Q{start_row + 2}": f"=ROUND($F{start_row + 2}*Q{start_row + 1},0)",
            f"R{start_row + 2}": f"=ROUND($F{start_row + 2}*R{start_row + 1},0)",
            f"S{start_row + 2}": f"=ROUND($F{start_row + 2}*S{start_row + 1},0)",
            f"T{start_row + 2}": f"=ROUND($F{start_row + 2}*T{start_row + 1},0)",
            f"U{start_row + 2}": f"=SUM(I{start_row + 2}:T{start_row + 2})",
            f"V{start_row + 2}": f"=IFERROR(SUM(I{start_row + 2}:N{start_row + 2}),0)",
            f"W{start_row + 2}": f"=IFERROR(SUM(O{start_row + 2}:T{start_row + 2}),0)",

            f"I{start_row + 11}": "FEB",
            f"J{start_row + 11}": "MAR",
            f"K{start_row + 11}": "APR",
            f"L{start_row + 11}": "MAY",
            f"M{start_row + 11}": "JUN",
            f"N{start_row + 11}": "JUL",
            f"O{start_row + 11}": "AUG",
            f"P{start_row + 11}": "SEP",
            f"Q{start_row + 11}": "OCT",
            f"R{start_row + 11}": "NOV",
            f"S{start_row + 11}": "DEC",
            f"T{start_row + 11}": "JAN",
            f"U{start_row + 11}": "ANNUAL",
            f"V{start_row + 11}": "SPRING",
            f"W{start_row + 11}": "FALL",

            f"I{start_row + 33}": "FEB",
            f"J{start_row + 33}": "MAR",
            f"K{start_row + 33}": "APR",
            f"L{start_row + 33}": "MAY",
            f"M{start_row + 33}": "JUN",
            f"N{start_row + 33}": "JUL",
            f"O{start_row + 33}": "AUG",
            f"P{start_row + 33}": "SEP",
            f"Q{start_row + 33}": "OCT",
            f"R{start_row + 33}": "NOV",
            f"S{start_row + 33}": "DEC",
            f"T{start_row + 33}": "JAN",
            f"U{start_row + 33}": "ANNUAL",
            f"V{start_row + 33}": "SPRING",
            f"W{start_row + 33}": "FALL",


            f"F{start_row + 2}": f"=IFERROR(C{start_row + 2}/E{start_row + 2},0)",
            f"F{start_row + 7}": f"=C{start_row + 14}+E{start_row + 8}-E{start_row + 7}",
            f"F{start_row + 10}": f"=MAX(0,C{start_row + 24}-F{start_row + 9})",
            f"D{start_row + 12}": f"=C{start_row + 12}*C{start_row + 19}",
            f"I{start_row + 12}": TY_Unit_Sales['FEB'],
            f"J{start_row + 12}":TY_Unit_Sales['MAR'],
            f"K{start_row + 12}": TY_Unit_Sales['APR'],
            f"L{start_row + 12}":TY_Unit_Sales['MAY'],
            f"M{start_row + 12}":TY_Unit_Sales['JUN'],
            f"N{start_row + 12}": TY_Unit_Sales['JUL'],
            f"O{start_row + 12}":TY_Unit_Sales['AUG'],
            f"P{start_row + 12}": TY_Unit_Sales['SEP'],
            f"Q{start_row + 12}": TY_Unit_Sales['OCT'],
            f"R{start_row + 12}":TY_Unit_Sales['NOV'],
            f"S{start_row + 12}": TY_Unit_Sales['DEC'],
            f"T{start_row + 12}": TY_Unit_Sales['JAN'],
            f"U{start_row + 12}": f"=IFERROR(SUM(I{start_row + 12}:T{start_row + 12}),0)",
            f"V{start_row + 12}": f"=IFERROR(SUM(I{start_row + 12}:N{start_row + 12}),0)",
            f"W{start_row + 12}": f"=IFERROR(SUM(O{start_row + 12}:T{start_row + 12}),0)" ,
            f"I{start_row + 34}": LY_Unit_Sales['FEB'],
            f"J{start_row + 34}":LY_Unit_Sales['MAR'],
            f"K{start_row + 34}":LY_Unit_Sales['APR'],
            f"L{start_row + 34}":LY_Unit_Sales['MAY'],
            f"M{start_row + 34}":LY_Unit_Sales['JUN'],
            f"N{start_row + 34}":LY_Unit_Sales['JUL'],
            f"O{start_row + 34}":LY_Unit_Sales['AUG'],
            f"P{start_row + 34}":LY_Unit_Sales['SEP'],
            f"Q{start_row + 34}":LY_Unit_Sales['OCT'],
            f"R{start_row + 34}":LY_Unit_Sales['NOV'],
            f"S{start_row + 34}":LY_Unit_Sales['DEC'],
            f"T{start_row + 34}":LY_Unit_Sales['JAN'],
            f"U{start_row + 34}": f"=SUM(I{start_row + 34}:T{start_row + 34})",
            f"V{start_row + 34}": f"=SUM(I{start_row + 34}:N{start_row + 34})",
            f"W{start_row + 34}": f"=SUM(O{start_row + 34}:T{start_row + 34})",
            f"I{start_row + 3}": f"=IF(AND($S$1=12,$K$1=12),ROUND(I{start_row + 12}+I{start_row + 12}*$F{start_row + 3},0),"
                                f"IF(AND($S$1=12,I$4<7),ROUND(I{start_row + 34}+I{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1>6,I$4<7),ROUND(I{start_row + 12}+I{start_row + 12}*$F{start_row + 3},0),"
                                f"IF(AND($S$1<7,I$4>6),ROUND(I{start_row + 34}+I{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1<7,I$4<7),ROUND(I{start_row + 34}+I{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1>6,I$4>6),ROUND(I{start_row + 34}+I{start_row + 34}*$F{start_row + 3},0)))))))",

            f"J{start_row + 3}": f"=IF(AND($S$1=12,$K$1=12),ROUND(J{start_row + 12}+J{start_row + 12}*$F{start_row + 3},0),"
                                f"IF(AND($S$1=12,J$4<7),ROUND(J{start_row + 34}+J{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1>6,J$4<7),ROUND(J{start_row + 12}+J{start_row + 12}*$F{start_row + 3},0),"
                                f"IF(AND($S$1<7,J$4>6),ROUND(J{start_row + 34}+J{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1<7,J$4<7),ROUND(J{start_row + 34}+J{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1>6,J$4>6),ROUND(J{start_row + 34}+J{start_row + 34}*$F{start_row + 3},0)))))))",

            f"L{start_row + 3}": f"=IF(AND($S$1=12,$K$1=12),ROUND(L{start_row + 12}+L{start_row + 12}*$F{start_row + 3},0),"
                                f"IF(AND($S$1=12,L$4<7),ROUND(L{start_row + 34}+L{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1>6,L$4<7),ROUND(L{start_row + 12}+L{start_row + 12}*$F{start_row + 3},0),"
                                f"IF(AND($S$1<7,L$4>6),ROUND(L{start_row + 34}+L{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1<7,L$4<7),ROUND(L{start_row + 34}+L{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1>6,L$4>6),ROUND(L{start_row + 34}+L{start_row + 34}*$F{start_row + 3},0)))))))",

            f"M{start_row + 3}": f"=IF(AND($S$1=12,$K$1=12),ROUND(M{start_row + 12}+M{start_row + 12}*$F{start_row + 3},0),"
                                f"IF(AND($S$1=12,M$4<7),ROUND(M{start_row + 34}+M{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1>6,M$4<7),ROUND(M{start_row + 12}+M{start_row + 12}*$F{start_row + 3},0),"
                                f"IF(AND($S$1<7,M$4>6),ROUND(M{start_row + 34}+M{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1<7,M$4<7),ROUND(M{start_row + 34}+M{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1>6,M$4>6),ROUND(M{start_row + 34}+M{start_row + 34}*$F{start_row + 3},0)))))))",

            f"N{start_row + 3}": f"=IF(AND($S$1=12,$K$1=12),ROUND(N{start_row + 12}+N{start_row + 12}*$F{start_row + 3},0),"
                                f"IF(AND($S$1=12,N$4<7),ROUND(N{start_row + 34}+N{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1>6,N$4<7),ROUND(N{start_row + 12}+N{start_row + 12}*$F{start_row + 3},0),"
                                f"IF(AND($S$1<7,N$4>6),ROUND(N{start_row + 34}+N{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1<7,N$4<7),ROUND(N{start_row + 34}+N{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1>6,N$4>6),ROUND(N{start_row + 34}+N{start_row + 34}*$F{start_row + 3},0)))))))",

            f"O{start_row + 3}": f"=IF(AND($S$1=12,$K$1=12),ROUND(O{start_row + 12}+O{start_row + 12}*$F{start_row + 3},0),"
                                f"IF(AND($S$1=12,O$4<7),ROUND(O{start_row + 34}+O{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1>6,O$4<7),ROUND(O{start_row + 12}+O{start_row + 12}*$F{start_row + 3},0),"
                                f"IF(AND($S$1<7,O$4>6),ROUND(O{start_row + 34}+O{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1<7,O$4<7),ROUND(O{start_row + 34}+O{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1>6,O$4>6),ROUND(O{start_row + 34}+O{start_row + 34}*$F{start_row + 3},0)))))))",

            f"P{start_row + 3}": f"=IF(AND($S$1=12,$K$1=12),ROUND(P{start_row + 12}+P{start_row + 12}*$F{start_row + 3},0),"
                                f"IF(AND($S$1=12,P$4<7),ROUND(P{start_row + 34}+P{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1>6,P$4<7),ROUND(P{start_row + 12}+P{start_row + 12}*$F{start_row + 3},0),"
                                f"IF(AND($S$1<7,P$4>6),ROUND(P{start_row + 34}+P{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1<7,P$4<7),ROUND(P{start_row + 34}+P{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1>6,P$4>6),ROUND(P{start_row + 34}+P{start_row + 34}*$F{start_row + 3},0)))))))",

            f"Q{start_row + 3}": f"=IF(AND($S$1=12,$K$1=12),ROUND(Q{start_row + 12}+Q{start_row + 12}*$F{start_row + 3},0),"
                                f"IF(AND($S$1=12,Q$4<7),ROUND(Q{start_row + 34}+Q{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1>6,Q$4<7),ROUND(Q{start_row + 12}+Q{start_row + 12}*$F{start_row + 3},0),"
                                f"IF(AND($S$1<7,Q$4>6),ROUND(Q{start_row + 34}+Q{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1<7,Q$4<7),ROUND(Q{start_row + 34}+Q{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1>6,Q$4>6),ROUND(Q{start_row + 34}+Q{start_row + 34}*$F{start_row + 3},0)))))))",

            f"R{start_row + 3}": f"=IF(AND($S$1=12,$K$1=12),ROUND(R{start_row + 12}+R{start_row + 12}*$F{start_row + 3},0),"
                                f"IF(AND($S$1=12,R$4<7),ROUND(R{start_row + 34}+R{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1>6,R$4<7),ROUND(R{start_row + 12}+R{start_row + 12}*$F{start_row + 3},0),"
                                f"IF(AND($S$1<7,R$4>6),ROUND(R{start_row + 34}+R{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1<7,R$4<7),ROUND(R{start_row + 34}+R{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1>6,R$4>6),ROUND(R{start_row + 34}+R{start_row + 34}*$F{start_row + 3},0)))))))",

            f"S{start_row + 3}": f"=IF(AND($S$1=12,$K$1=12),ROUND(S{start_row + 12}+S{start_row + 12}*$F{start_row + 3},0),"
                                f"IF(AND($S$1=12,S$4<7),ROUND(S{start_row + 34}+S{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1>6,S$4<7),ROUND(S{start_row + 12}+S{start_row + 12}*$F{start_row + 3},0),"
                                f"IF(AND($S$1<7,S$4>6),ROUND(S{start_row + 34}+S{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1<7,S$4<7),ROUND(S{start_row + 34}+S{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1>6,S$4>6),ROUND(S{start_row + 34}+S{start_row + 34}*$F{start_row + 3},0)))))))",

            f"T{start_row + 3}": f"=IF(AND($S$1=12,$K$1=12),ROUND(T{start_row + 12}+T{start_row + 12}*$F{start_row + 3},0),"
                                f"IF(AND($S$1=12,T$4<7),ROUND(T{start_row + 34}+T{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1>6,T$4<7),ROUND(T{start_row + 12}+T{start_row + 12}*$F{start_row + 3},0),"
                                f"IF(AND($S$1<7,T$4>6),ROUND(T{start_row + 34}+T{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1<7,T$4<7),ROUND(T{start_row + 34}+T{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1>6,T$4>6),ROUND(T{start_row + 34}+T{start_row + 34}*$F{start_row + 3},0)))))))",
            f"K{start_row + 3}": f"=IF(AND($S$1=12,$K$1=12),ROUND(K{start_row + 12}+K{start_row + 12}*$F{start_row + 3},0),"
                                f"IF(AND($S$1=12,K$4<7),ROUND(K{start_row + 34}+K{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1>6,K$4<7),ROUND(K{start_row + 12}+K{start_row + 12}*$F{start_row + 3},0),"
                                f"IF(AND($S$1<7,K$4>6),ROUND(K{start_row + 34}+K{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1<7,K$4<7),ROUND(K{start_row + 34}+K{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1>6,K$4>6),ROUND(K{start_row + 34}+K{start_row + 34}*$F{start_row + 3},0)))))))",

            # Continue similarly for columns L through T
            f"U{start_row + 3}": f"=SUM(I{start_row + 3}:T{start_row + 3})",
            f"V{start_row + 3}": f"=IFERROR(SUM(I{start_row + 3}:N{start_row + 3}),0)",
            f"W{start_row + 3}": f"=IFERROR(SUM(O{start_row + 3}:T{start_row + 3}),0)",
            
            f"I{start_row + 4}": f"=IF($F{start_row + 4}=$H{start_row + 2},I{start_row + 2},"
                                f"IF($F{start_row + 4}=$H{start_row + 3},I{start_row + 3},"
                                f"IF($F{start_row + 4}=\"Current year\",I{start_row + 12},"
                                f"IF($F{start_row + 4}=\"Last year\",I{start_row + 34},"
                                f"ROUND(AVERAGE(I{start_row + 2}:I{start_row + 3}),0)))))",

            f"J{start_row + 4}": f"=IF($F{start_row + 4}=$H{start_row + 2},J{start_row + 2},"
                                f"IF($F{start_row + 4}=$H{start_row + 3},J{start_row + 3},"
                                f"IF($F{start_row + 4}=\"Current year\",J{start_row + 12},"
                                f"IF($F{start_row + 4}=\"Last year\",J{start_row + 34},"
                                f"ROUND(AVERAGE(J{start_row + 2}:J{start_row + 3}),0)))))",

            f"K{start_row + 4}": f"=IF($F{start_row + 4}=$H{start_row + 2},K{start_row + 2},"
                                f"IF($F{start_row + 4}=$H{start_row + 3},K{start_row + 3},"
                                f"IF($F{start_row + 4}=\"Current year\",K{start_row + 12},"
                                f"IF($F{start_row + 4}=\"Last year\",K{start_row + 34},"
                                f"ROUND(AVERAGE(K{start_row + 2}:K{start_row + 3}),0)))))",

            f"L{start_row + 4}": f"=IF($F{start_row + 4}=$H{start_row + 2},L{start_row + 2},"
                                f"IF($F{start_row + 4}=$H{start_row + 3},L{start_row + 3},"
                                f"IF($F{start_row + 4}=\"Current year\",L{start_row + 12},"
                                f"IF($F{start_row + 4}=\"Last year\",L{start_row + 34},"
                                f"ROUND(AVERAGE(L{start_row + 2}:L{start_row + 3}),0)))))",

            f"M{start_row + 4}": f"=IF($F{start_row + 4}=$H{start_row + 2},M{start_row + 2},"
                                f"IF($F{start_row + 4}=$H{start_row + 3},M{start_row + 3},"
                                f"IF($F{start_row + 4}=\"Current year\",M{start_row + 12},"
                                f"IF($F{start_row + 4}=\"Last year\",M{start_row + 34},"
                                f"ROUND(AVERAGE(M{start_row + 2}:M{start_row + 3}),0)))))",

            f"N{start_row + 4}": f"=IF($F{start_row + 4}=$H{start_row + 2},N{start_row + 2},"
                                f"IF($F{start_row + 4}=$H{start_row + 3},N{start_row + 3},"
                                f"IF($F{start_row + 4}=\"Current year\",N{start_row + 12},"
                                f"IF($F{start_row + 4}=\"Last year\",N{start_row + 34},"
                                f"ROUND(AVERAGE(N{start_row + 2}:N{start_row + 3}),0)))))",

            f"O{start_row + 4}": f"=IF($F{start_row + 4}=$H{start_row + 2},O{start_row + 2},"
                                f"IF($F{start_row + 4}=$H{start_row + 3},O{start_row + 3},"
                                f"IF($F{start_row + 4}=\"Current year\",O{start_row + 12},"
                                f"IF($F{start_row + 4}=\"Last year\",O{start_row + 34},"
                                f"ROUND(AVERAGE(O{start_row + 2}:O{start_row + 3}),0)))))",

            f"P{start_row + 4}": f"=IF($F{start_row + 4}=$H{start_row + 2},P{start_row + 2},"
                                f"IF($F{start_row + 4}=$H{start_row + 3},P{start_row + 3},"
                                f"IF($F{start_row + 4}=\"Current year\",P{start_row + 12},"
                                f"IF($F{start_row + 4}=\"Last year\",P{start_row + 34},"
                                f"ROUND(AVERAGE(P{start_row + 2}:P{start_row + 3}),0)))))",

            f"Q{start_row + 4}": f"=IF($F{start_row + 4}=$H{start_row + 2},Q{start_row + 2},"
                                f"IF($F{start_row + 4}=$H{start_row + 3},Q{start_row + 3},"
                                f"IF($F{start_row + 4}=\"Current year\",Q{start_row + 12},"
                                f"IF($F{start_row + 4}=\"Last year\",Q{start_row + 34},"
                                f"ROUND(AVERAGE(Q{start_row + 2}:Q{start_row + 3}),0)))))",

            f"R{start_row + 4}": f"=IF($F{start_row + 4}=$H{start_row + 2},R{start_row + 2},"
                                f"IF($F{start_row + 4}=$H{start_row + 3},R{start_row + 3},"
                                f"IF($F{start_row + 4}=\"Current year\",R{start_row + 12},"
                                f"IF($F{start_row + 4}=\"Last year\",R{start_row + 34},"
                                f"ROUND(AVERAGE(R{start_row + 2}:R{start_row + 3}),0)))))",

            f"S{start_row + 4}": f"=IF($F{start_row + 4}=$H{start_row + 2},S{start_row + 2},"
                                f"IF($F{start_row + 4}=$H{start_row + 3},S{start_row + 3},"
                                f"IF($F{start_row + 4}=\"Current year\",S{start_row + 12},"
                                f"IF($F{start_row + 4}=\"Last year\",S{start_row + 34},"
                                f"ROUND(AVERAGE(S{start_row + 2}:S{start_row + 3}),0)))))",

            f"T{start_row + 4}": f"=IF($F{start_row + 4}=$H{start_row + 2},T{start_row + 2},"
                                f"IF($F{start_row + 4}=$H{start_row + 3},T{start_row + 3},"
                                f"IF($F{start_row + 4}=\"Current year\",T{start_row + 12},"
                                f"IF($F{start_row + 4}=\"Last year\",T{start_row + 34},"
                                f"ROUND(AVERAGE(T{start_row + 2}:T{start_row + 3}),0)))))",

            f"U{start_row + 4}": f"=SUM(I{start_row + 4}:T{start_row + 4})",
            f"V{start_row + 4}": f"=IFERROR(SUM(I{start_row + 4}:N{start_row + 4}),0)",
            f"W{start_row + 4}": f"=IFERROR(SUM(O{start_row + 4}:T{start_row + 4}),0)",

            f"I{start_row + 38}": LY_OH_Units['FEB'],
            f"J{start_row + 38}": LY_OH_Units['MAR'],
            f"K{start_row + 38}": LY_OH_Units['APR'],
            f"L{start_row + 38}":LY_OH_Units['MAY'],
            f"M{start_row + 38}":LY_OH_Units['JUN'],
            f"N{start_row + 38}": LY_OH_Units['JUL'],
            f"O{start_row + 38}": LY_OH_Units['AUG'],
            f"P{start_row + 38}":LY_OH_Units['SEP'],
            f"Q{start_row + 38}":LY_OH_Units['OCT'],
            f"R{start_row + 38}": LY_OH_Units['NOV'],
            f"S{start_row + 38}":LY_OH_Units['DEC'],
            f"T{start_row + 38}": LY_OH_Units['JAN'],
            f"U{start_row + 38}": f"=AVERAGEIFS(I{start_row + 38}:T{start_row + 38}, $I$4:$T$4, \"<=\"&$T$4)",
            f"V{start_row + 38}": f"=AVERAGEIFS(I{start_row + 38}:N{start_row + 38}, $I$4:$N$4, \"<=\"&$N$4)",
            f"W{start_row + 38}": f"=AVERAGEIFS(O{start_row + 38}:T{start_row + 38}, $O$4:$T$4, \"<=\"&$T$4)",
            f"I{start_row + 5}": f"=IF(AND($V$1=\"YTD\",I$4<$K$1),I{start_row + 12},IF(AND($V$1=\"YTD\",I$4=$K$1),I{start_row + 12},"
                                f"IF(AND($V$1=\"CURRENT MTH\",I$4=$K$1),I{start_row + 12},IF(AND($V$1=\"SPRING\",I$4<7),I{start_row + 12},"
                                f"IF(AND($V$1=\"FALL\",I$4>6),I{start_row + 12},IF(AND($V$1=\"LY FALL\",I$4>6),I{start_row + 34},I{start_row + 4}))))))",

            f"J{start_row + 5}": f"=IF(AND($V$1=\"YTD\",J$4<$K$1),J{start_row + 12},IF(AND($V$1=\"YTD\",J$4=$K$1),J{start_row + 12},"
                                f"IF(AND($V$1=\"CURRENT MTH\",J$4=$K$1),J{start_row + 12},IF(AND($V$1=\"SPRING\",J$4<7),J{start_row + 12},"
                                f"IF(AND($V$1=\"FALL\",J$4>6),J{start_row + 12},IF(AND($V$1=\"LY FALL\",J$4>6),J{start_row + 34},J{start_row + 4}))))))",

            f"K{start_row + 5}": f"=IF(AND($V$1=\"YTD\",K$4<$K$1),K{start_row + 12},IF(AND($V$1=\"YTD\",K$4=$K$1),K{start_row + 12},"
                                f"IF(AND($V$1=\"CURRENT MTH\",K$4=$K$1),K{start_row + 12},IF(AND($V$1=\"SPRING\",K$4<7),K{start_row + 12},"
                                f"IF(AND($V$1=\"FALL\",K$4>6),K{start_row + 12},IF(AND($V$1=\"LY FALL\",K$4>6),K{start_row + 34},K{start_row + 4}))))))",

            f"L{start_row + 5}": f"=IF(AND($V$1=\"YTD\",L$4<$K$1),L{start_row + 12},IF(AND($V$1=\"YTD\",L$4=$K$1),L{start_row + 12},"
                                f"IF(AND($V$1=\"CURRENT MTH\",L$4=$K$1),L{start_row + 12},IF(AND($V$1=\"SPRING\",L$4<7),L{start_row + 12},"
                                f"IF(AND($V$1=\"FALL\",L$4>6),L{start_row + 12},IF(AND($V$1=\"LY FALL\",L$4>6),L{start_row + 34},L{start_row + 4}))))))",

            f"M{start_row + 5}": f"=IF(AND($V$1=\"YTD\",M$4<$K$1),M{start_row + 12},IF(AND($V$1=\"YTD\",M$4=$K$1),M{start_row + 12},"
                                f"IF(AND($V$1=\"CURRENT MTH\",M$4=$K$1),M{start_row + 12},IF(AND($V$1=\"SPRING\",M$4<7),M{start_row + 12},"
                                f"IF(AND($V$1=\"FALL\",M$4>6),M{start_row + 12},IF(AND($V$1=\"LY FALL\",M$4>6),M{start_row + 34},M{start_row + 4}))))))",

            f"N{start_row + 5}": f"=IF(AND($V$1=\"YTD\",N$4<$K$1),N{start_row + 12},IF(AND($V$1=\"YTD\",N$4=$K$1),N{start_row + 12},"
                                f"IF(AND($V$1=\"CURRENT MTH\",N$4=$K$1),N{start_row + 12},IF(AND($V$1=\"SPRING\",N$4<7),N{start_row + 12},"
                                f"IF(AND($V$1=\"FALL\",N$4>6),N{start_row + 12},IF(AND($V$1=\"LY FALL\",N$4>6),N{start_row + 34},N{start_row + 4}))))))",

            f"O{start_row + 5}": f"=IF(AND($V$1=\"YTD\",O$4<$K$1),O{start_row + 12},IF(AND($V$1=\"YTD\",O$4=$K$1),O{start_row + 12},"
                                f"IF(AND($V$1=\"CURRENT MTH\",O$4=$K$1),O{start_row + 12},IF(AND($V$1=\"SPRING\",O$4<7),O{start_row + 12},"
                                f"IF(AND($V$1=\"FALL\",O$4>6),O{start_row + 12},IF(AND($V$1=\"LY FALL\",O$4>6),O{start_row + 34},O{start_row + 4}))))))",
            f"P{start_row + 5}": f"=IF(AND($V$1=\"YTD\",P$4<$K$1),P{start_row + 12},IF(AND($V$1=\"YTD\",P$4=$K$1),P{start_row + 12},"
                                f"IF(AND($V$1=\"CURRENT MTH\",P$4=$K$1),P{start_row + 12},IF(AND($V$1=\"SPRING\",P$4<7),P{start_row + 12},"
                                f"IF(AND($V$1=\"FALL\",P$4>6),P{start_row + 12},IF(AND($V$1=\"LY FALL\",P$4>6),P{start_row + 34},P{start_row + 4}))))))",

            f"Q{start_row + 5}": f"=IF(AND($V$1=\"YTD\",Q$4<$K$1),Q{start_row + 12},IF(AND($V$1=\"YTD\",Q$4=$K$1),Q{start_row + 12},"
                                f"IF(AND($V$1=\"CURRENT MTH\",Q$4=$K$1),Q{start_row + 12},IF(AND($V$1=\"SPRING\",Q$4<7),Q{start_row + 12},"
                                f"IF(AND($V$1=\"FALL\",Q$4>6),Q{start_row + 12},IF(AND($V$1=\"LY FALL\",Q$4>6),Q{start_row + 34},Q{start_row + 4}))))))",
            
            f"R{start_row + 5}": f"=IF(AND($V$1=\"YTD\",R$4<$K$1),R{start_row + 12},IF(AND($V$1=\"YTD\",R$4=$K$1),R{start_row + 12},"
                                f"IF(AND($V$1=\"CURRENT MTH\",R$4=$K$1),R{start_row + 12},IF(AND($V$1=\"SPRING\",R$4<7),R{start_row + 12},"
                                f"IF(AND($V$1=\"FALL\",R$4>6),R{start_row + 12},IF(AND($V$1=\"LY FALL\",R$4>6),R{start_row + 34},R{start_row + 4}))))))",
            
            f"S{start_row + 5}": f"=IF(AND($V$1=\"YTD\",S$4<$K$1),S{start_row + 12},IF(AND($V$1=\"YTD\",S$4=$K$1),S{start_row + 12},"
                                f"IF(AND($V$1=\"CURRENT MTH\",S$4=$K$1),S{start_row + 12},IF(AND($V$1=\"SPRING\",S$4<7),S{start_row + 12},"
                                f"IF(AND($V$1=\"FALL\",S$4>6),S{start_row + 12},IF(AND($V$1=\"LY FALL\",S$4>6),S{start_row + 34},S{start_row + 4}))))))",
            
            f"T{start_row + 5}": f"=IF(AND($V$1=\"YTD\",T$4<$K$1),T{start_row + 12},IF(AND($V$1=\"YTD\",T$4=$K$1),T{start_row + 12},"
                                f"IF(AND($V$1=\"CURRENT MTH\",T$4=$K$1),T{start_row + 12},IF(AND($V$1=\"SPRING\",T$4<7),T{start_row + 12},"
                                f"IF(AND($V$1=\"FALL\",T$4>6),T{start_row + 12},IF(AND($V$1=\"LY FALL\",T$4>6),T{start_row + 34},T{start_row + 4}))))))",
            
            f"U{start_row + 5}": f"=SUM(I{start_row + 5}:T{start_row + 5})",
            
            f"V{start_row + 5}": f"=IFERROR(SUM(I{start_row + 5}:N{start_row + 5}),0)",
            
            f"W{start_row + 5}": f"=IFERROR(SUM(O{start_row + 5}:T{start_row + 5}),0)",

            f"U{start_row + 6}": f"=SUM(I{start_row + 6}:T{start_row + 6})",
            f"V{start_row + 6}": f"=IFERROR(SUM(I{start_row + 6}:N{start_row + 6}),0)",
            f"W{start_row + 6}": f"=IFERROR(SUM(O{start_row + 6}:T{start_row + 6}),0)",
            f"I{start_row + 16}": TY_OH_Units['FEB'],
            f"J{start_row + 16}": TY_OH_Units['MAR'],
            f"K{start_row + 16}":TY_OH_Units['APR'],
            f"L{start_row + 16}": TY_OH_Units['MAY'],
            f"M{start_row + 16}": TY_OH_Units['JUN'],
            f"N{start_row + 16}":TY_OH_Units['JUL'],
            f"O{start_row + 16}": TY_OH_Units['AUG'],
            f"P{start_row + 16}":TY_OH_Units['SEP'],
            f"Q{start_row + 16}":TY_OH_Units['OCT'],
            f"R{start_row + 16}": TY_OH_Units['NOV'],
            f"S{start_row + 16}": TY_OH_Units['DEC'],
            f"T{start_row + 16}": TY_OH_Units['JAN'],

            f"U{start_row + 16}": f"=AVERAGEIFS(I{start_row + 16}:T{start_row + 16}, $I$4:$T$4, \"<=\"&$K$1)",
            
            f"V{start_row + 16}": f"=AVERAGEIFS(I{start_row + 16}:N{start_row + 16}, $I$4:$N$4, \"<=\"&$K$2)",
            
            f"W{start_row + 16}": f"=AVERAGEIFS(O{start_row + 16}:T{start_row + 16}, $O$4:$T$4, \"<=\"&$M$2)",

            f"I{start_row + 32}":TY_Receipts['FEB'],
            
            f"J{start_row + 32}": TY_Receipts['MAR'],

            f"K{start_row + 32}":TY_Receipts['APR'],
            f"L{start_row + 32}":TY_Receipts['MAY'],
            f"M{start_row + 32}":TY_Receipts['JUN'],
            f"N{start_row + 32}":TY_Receipts['JUL'],
            f"O{start_row + 32}":TY_Receipts['AUG'],
            f"P{start_row + 32}":TY_Receipts['SEP'],

            f"Q{start_row + 32}":TY_Receipts['OCT'],

            f"R{start_row + 32}":TY_Receipts['NOV'],

            f"S{start_row + 32}":TY_Receipts['DEC'],

            f"T{start_row + 32}":TY_Receipts['JAN'],

            f"U{start_row + 32}": f"=IFERROR(SUM(I{start_row + 32}:T{start_row + 32}),0)",

            f"V{start_row + 32}": f"=IFERROR(SUM(I{start_row + 32}:N{start_row + 32}),0)",

            f"W{start_row + 32}": f"=IFERROR(SUM(O{start_row + 32}:T{start_row + 32}),0)",
            f"I{start_row + 7}": f"=IF(AND($V$1=\"Current Mth\",$K$1=I$4,I$4=1),$T{start_row + 38}+I{start_row + 6}+I{start_row + 32}-I{start_row + 5},IF(AND($V$1=\"Current Mth\",$K$1=I$4),I{start_row + 16}+I{start_row + 6}-(I{start_row + 5}-I{start_row + 12}),IF(AND($V$1=\"Current Mth\",I$4=1,$K$1>1),$T{start_row + 7}+I{start_row + 6}-I{start_row + 5},IF(AND($V$1=\"YTD\",I$4<$K$1),I{start_row + 16},IF(AND($V$1=\"Spring\",I$4<7),I{start_row + 16},IF(AND($V$1=\"Fall\",I$4>6,I$4<$K$1),I{start_row + 16},IF(AND($V$1=\"Fall\",I$4=1,$K$1>1),$T{start_row + 7}+I{start_row + 6}-I{start_row + 5},IF(AND($V$1=\"Fall\",I$4>6,I$4=$K$1),H{start_row + 16}+I{start_row + 32}+I{start_row + 6}-I{start_row + 5},IF(AND($V$1=\"LY Fall\",I$4>6),I{start_row + 38},IF(AND(I$4=1,$V$1=\"LY FALL\"),$T{start_row + 38}+I{start_row + 6}+I{start_row + 6}-I{start_row + 5},IF(AND(I$4=$K$1,I$4>1),H{start_row + 16}+I{start_row + 32}+I{start_row + 6}-I{start_row + 5},H{start_row + 7}+I{start_row + 6}-I{start_row + 5})))))))))))",

            f"J{start_row + 7}": f"=IF(AND($V$1=\"Current Mth\",$K$1=J$4,J$4=1),$T{start_row + 38}+J{start_row + 6}+J{start_row + 32}-J{start_row + 5},IF(AND($V$1=\"Current Mth\",$K$1=J$4),J{start_row + 16}+J{start_row + 6}-(J{start_row + 5}-J{start_row + 12}),IF(AND($V$1=\"Current Mth\",J$4=1,$K$1>1),$T{start_row + 7}+J{start_row + 6}-J{start_row + 5},IF(AND($V$1=\"YTD\",J$4<$K$1),J{start_row + 16},IF(AND($V$1=\"Spring\",J$4<7),J{start_row + 16},IF(AND($V$1=\"Fall\",J$4>6,J$4<$K$1),J{start_row + 16},IF(AND($V$1=\"Fall\",J$4=1,$K$1>1),$T{start_row + 7}+J{start_row + 6}-J{start_row + 5},IF(AND($V$1=\"Fall\",J$4>6,J$4=$K$1),I{start_row + 16}+J{start_row + 32}+J{start_row + 6}-J{start_row + 5},IF(AND($V$1=\"LY Fall\",J$4>6),J{start_row + 38},IF(AND(J$4=1,$V$1=\"LY FALL\"),$T{start_row + 38}+J{start_row + 6}+J{start_row + 6}-J{start_row + 5},IF(AND(J$4=$K$1,J$4>1),I{start_row + 16}+J{start_row + 32}+J{start_row + 6}-J{start_row + 5},I{start_row + 7}+J{start_row + 6}-J{start_row + 5})))))))))))",

            f"K{start_row + 7}": f"=IF(AND($V$1=\"Current Mth\",$K$1=K$4,K$4=1),$T{start_row + 38}+K{start_row + 6}+K{start_row + 32}-K{start_row + 5},IF(AND($V$1=\"Current Mth\",$K$1=K$4),K{start_row + 16}+K{start_row + 6}-(K{start_row + 5}-K{start_row + 12}),IF(AND($V$1=\"Current Mth\",K$4=1,$K$1>1),$T{start_row + 7}+K{start_row + 6}-K{start_row + 5},IF(AND($V$1=\"YTD\",K$4<$K$1),K{start_row + 16},IF(AND($V$1=\"Spring\",K$4<7),K{start_row + 16},IF(AND($V$1=\"Fall\",K$4>6,K$4<$K$1),K{start_row + 16},IF(AND($V$1=\"Fall\",K$4=1,$K$1>1),$T{start_row + 7}+K{start_row + 6}-K{start_row + 5},IF(AND($V$1=\"Fall\",K$4>6,K$4=$K$1),J{start_row + 16}+K{start_row + 32}+K{start_row + 6}-K{start_row + 5},IF(AND($V$1=\"LY Fall\",K$4>6),K{start_row + 38},IF(AND(K$4=1,$V$1=\"LY FALL\"),$T{start_row + 38}+K{start_row + 6}+K{start_row + 6}-K{start_row + 5},IF(AND(K$4=$K$1,K$4>1),J{start_row + 16}+K{start_row + 32}+K{start_row + 6}-K{start_row + 5},J{start_row + 7}+K{start_row + 6}-K{start_row + 5})))))))))))",

            f"L{start_row + 7}": f"=IF(AND($V$1=\"Current Mth\",$K$1=L$4,L$4=1),$T{start_row + 38}+L{start_row + 6}+L{start_row + 32}-L{start_row + 5},IF(AND($V$1=\"Current Mth\",$K$1=L$4),L{start_row + 16}+L{start_row + 6}-(L{start_row + 5}-L{start_row + 12}),IF(AND($V$1=\"Current Mth\",L$4=1,$K$1>1),$T{start_row + 7}+L{start_row + 6}-L{start_row + 5},IF(AND($V$1=\"YTD\",L$4<$K$1),L{start_row + 16},IF(AND($V$1=\"Spring\",L$4<7),L{start_row + 16},IF(AND($V$1=\"Fall\",L$4>6,L$4<$K$1),L{start_row + 16},IF(AND($V$1=\"Fall\",L$4=1,$K$1>1),$T{start_row + 7}+L{start_row + 6}-L{start_row + 5},IF(AND($V$1=\"Fall\",L$4>6,L$4=$K$1),K{start_row + 16}+L{start_row + 32}+L{start_row + 6}-L{start_row + 5},IF(AND($V$1=\"LY Fall\",L$4>6),L{start_row + 38},IF(AND(L$4=1,$V$1=\"LY FALL\"),$T{start_row + 38}+L{start_row + 6}+L{start_row + 6}-L{start_row + 5},IF(AND(L$4=$K$1,L$4>1),K{start_row + 16}+L{start_row + 32}+L{start_row + 6}-L{start_row + 5},K{start_row + 7}+L{start_row + 6}-L{start_row + 5})))))))))))",

            f"M{start_row + 7}": f"=IF(AND($V$1=\"Current Mth\",$K$1=M$4,M$4=1),$T{start_row + 38}+M{start_row + 6}+M{start_row + 32}-M{start_row + 5},IF(AND($V$1=\"Current Mth\",$K$1=M$4),M{start_row + 16}+M{start_row + 6}-(M{start_row + 5}-M{start_row + 12}),IF(AND($V$1=\"Current Mth\",M$4=1,$K$1>1),$T{start_row + 7}+M{start_row + 6}-M{start_row + 5},IF(AND($V$1=\"YTD\",M$4<$K$1),M{start_row + 16},IF(AND($V$1=\"Spring\",M$4<7),M{start_row + 16},IF(AND($V$1=\"Fall\",M$4>6,M$4<$K$1),M{start_row + 16},IF(AND($V$1=\"Fall\",M$4=1,$K$1>1),$T{start_row + 7}+M{start_row + 6}-M{start_row + 5},IF(AND($V$1=\"Fall\",M$4>6,M$4=$K$1),L{start_row + 16}+M{start_row + 32}+M{start_row + 6}-M{start_row + 5},IF(AND($V$1=\"LY Fall\",M$4>6),M{start_row + 38},IF(AND(M$4=1,$V$1=\"LY FALL\"),$T{start_row + 38}+M{start_row + 6}+M{start_row + 6}-M{start_row + 5},IF(AND(M$4=$K$1,M$4>1),L{start_row + 16}+M{start_row + 32}+M{start_row + 6}-M{start_row + 5},L{start_row + 7}+M{start_row + 6}-M{start_row + 5})))))))))))",

            f"O{start_row + 7}": f"=IF(AND($V$1=\"Current Mth\",$K$1=O$4,O$4=1),$T{start_row + 38}+O{start_row + 6}+O{start_row + 32}-O{start_row + 5},IF(AND($V$1=\"Current Mth\",$K$1=O$4),O{start_row + 16}+O{start_row + 6}-(O{start_row + 5}-O{start_row + 12}),IF(AND($V$1=\"Current Mth\",O$4=1,$K$1>1),$T{start_row + 7}+O{start_row + 6}-O{start_row + 5},IF(AND($V$1=\"YTD\",O$4<$K$1),O{start_row + 16},IF(AND($V$1=\"Spring\",O$4<7),O{start_row + 16},IF(AND($V$1=\"Fall\",O$4>6,O$4<$K$1),O{start_row + 16},IF(AND($V$1=\"Fall\",O$4=1,$K$1>1),$T{start_row + 7}+O{start_row + 6}-O{start_row + 5},IF(AND($V$1=\"Fall\",O$4>6,O$4=$K$1),N{start_row + 16}+O{start_row + 32}+O{start_row + 6}-O{start_row + 5},IF(AND($V$1=\"LY Fall\",O$4>6),O{start_row + 38},IF(AND(O$4=1,$V$1=\"LY FALL\"),$T{start_row + 38}+O{start_row + 6}+O{start_row + 6}-O{start_row + 5},IF(AND(O$4=$K$1,O$4>1),N{start_row + 16}+O{start_row + 32}+O{start_row + 6}-O{start_row + 5},N{start_row + 7}+O{start_row + 6}-O{start_row + 5})))))))))))",

            f"P{start_row + 7}": f"=IF(AND($V$1=\"Current Mth\",$K$1=P$4,P$4=1),$T{start_row + 38}+P{start_row + 6}+P{start_row + 32}-P{start_row + 5},IF(AND($V$1=\"Current Mth\",$K$1=P$4),P{start_row + 16}+P{start_row + 6}-(P{start_row + 5}-P{start_row + 12}),IF(AND($V$1=\"Current Mth\",P$4=1,$K$1>1),$T{start_row + 7}+P{start_row + 6}-P{start_row + 5},IF(AND($V$1=\"YTD\",P$4<$K$1),P{start_row + 16},IF(AND($V$1=\"Spring\",P$4<7),P{start_row + 16},IF(AND($V$1=\"Fall\",P$4>6,P$4<$K$1),P{start_row + 16},IF(AND($V$1=\"Fall\",P$4=1,$K$1>1),$T{start_row + 7}+P{start_row + 6}-P{start_row + 5},IF(AND($V$1=\"Fall\",P$4>6,P$4=$K$1),O{start_row + 16}+P{start_row + 32}+P{start_row + 6}-P{start_row + 5},IF(AND($V$1=\"LY Fall\",P$4>6),P{start_row + 38},IF(AND(P$4=1,$V$1=\"LY FALL\"),$T{start_row + 38}+P{start_row + 6}+P{start_row + 6}-P{start_row + 5},IF(AND(P$4=$K$1,P$4>1),O{start_row + 16}+P{start_row + 32}+P{start_row + 6}-P{start_row + 5},O{start_row + 7}+P{start_row + 6}-P{start_row + 5})))))))))))",

            f"Q{start_row + 7}": f"=IF(AND($V$1=\"Current Mth\",$K$1=Q$4,Q$4=1),$T{start_row + 38}+Q{start_row + 6}+Q{start_row + 32}-Q{start_row + 5},IF(AND($V$1=\"Current Mth\",$K$1=Q$4),Q{start_row + 16}+Q{start_row + 6}-(Q{start_row + 5}-Q{start_row + 12}),IF(AND($V$1=\"Current Mth\",Q$4=1,$K$1>1),$T{start_row + 7}+Q{start_row + 6}-Q{start_row + 5},IF(AND($V$1=\"YTD\",Q$4<$K$1),Q{start_row + 16},IF(AND($V$1=\"Spring\",Q$4<7),Q{start_row + 16},IF(AND($V$1=\"Fall\",Q$4>6,Q$4<$K$1),Q{start_row + 16},IF(AND($V$1=\"Fall\",Q$4=1,$K$1>1),$T{start_row + 7}+Q{start_row + 6}-Q{start_row + 5},IF(AND($V$1=\"Fall\",Q$4>6,Q$4=$K$1),P{start_row + 16}+Q{start_row + 32}+Q{start_row + 6}-Q{start_row + 5},IF(AND($V$1=\"LY Fall\",Q$4>6),Q{start_row + 38},IF(AND(Q$4=1,$V$1=\"LY FALL\"),$T{start_row + 38}+Q{start_row + 6}+Q{start_row + 6}-Q{start_row + 5},IF(AND(Q$4=$K$1,Q$4>1),P{start_row + 16}+Q{start_row + 32}+Q{start_row + 6}-Q{start_row + 5},P{start_row + 7}+Q{start_row + 6}-Q{start_row + 5})))))))))))",

            f"N{start_row + 7}": f"=IF(AND($V$1=\"Current Mth\",$K$1=N$4,N$4=1),$T{start_row + 38}+N{start_row + 6}+N{start_row + 32}-N{start_row + 5},IF(AND($V$1=\"Current Mth\",$K$1=N$4),N{start_row + 16}+N{start_row + 6}-(N{start_row + 5}-N{start_row + 12}),IF(AND($V$1=\"Current Mth\",N$4=1,$K$1>1),$T{start_row + 7}+N{start_row + 6}-N{start_row + 5},IF(AND($V$1=\"YTD\",N$4<$K$1),N{start_row + 16},IF(AND($V$1=\"Spring\",N$4<7),N{start_row + 16},IF(AND($V$1=\"Fall\",N$4>6,N$4<$K$1),N{start_row + 16},IF(AND($V$1=\"Fall\",N$4=1,$K$1>1),$T{start_row + 7}+N{start_row + 6}-N{start_row + 5},IF(AND($V$1=\"Fall\",N$4>6,N$4=$K$1),M{start_row + 16}+N{start_row + 32}+N{start_row + 6}-N{start_row + 5},IF(AND($V$1=\"LY Fall\",N$4>6),N{start_row + 38},IF(AND(N$4=1,$V$1=\"LY FALL\"),$T{start_row + 38}+N{start_row + 6}+N{start_row + 6}-N{start_row + 5},IF(AND(N$4=$K$1,N$4>1),M{start_row + 16}+N{start_row + 32}+N{start_row + 6}-N{start_row + 5},M{start_row + 7}+N{start_row + 6}-N{start_row + 5})))))))))))",

            f"R{start_row + 7}": f"=IF(AND($V$1=\"Current Mth\",$K$1=R$4,R$4=1),$T{start_row + 38}+R{start_row + 6}+R{start_row + 32}-R{start_row + 5},IF(AND($V$1=\"Current Mth\",$K$1=R$4),R{start_row + 16}+R{start_row + 6}-(R{start_row + 5}-R{start_row + 12}),IF(AND($V$1=\"Current Mth\",R$4=1,$K$1>1),$T{start_row + 7}+R{start_row + 6}-R{start_row + 5},IF(AND($V$1=\"YTD\",R$4<$K$1),R{start_row + 16},IF(AND($V$1=\"Spring\",R$4<7),R{start_row + 16},IF(AND($V$1=\"Fall\",R$4>6,R$4<$K$1),R{start_row + 16},IF(AND($V$1=\"Fall\",R$4=1,$K$1>1),$T{start_row + 7}+R{start_row + 6}-R{start_row + 5},IF(AND($V$1=\"Fall\",R$4>6,R$4=$K$1),Q{start_row + 16}+R{start_row + 32}+R{start_row + 6}-R{start_row + 5},IF(AND($V$1=\"LY Fall\",R$4>6),R{start_row + 38},IF(AND(R$4=1,$V$1=\"LY FALL\"),$T{start_row + 38}+R{start_row + 6}+R{start_row + 6}-R{start_row + 5},IF(AND(R$4=$K$1,R$4>1),Q{start_row + 16}+R{start_row + 32}+R{start_row + 6}-R{start_row + 5},Q{start_row + 7}+R{start_row + 6}-R{start_row + 5})))))))))))",

            f"S{start_row + 7}": f"=IF(AND($V$1=\"Current Mth\",$K$1=S$4,S$4=1),$T{start_row + 38}+S{start_row + 6}+S{start_row + 32}-S{start_row + 5},IF(AND($V$1=\"Current Mth\",$K$1=S$4),S{start_row + 16}+S{start_row + 6}-(S{start_row + 5}-S{start_row + 12}),IF(AND($V$1=\"Current Mth\",S$4=1,$K$1>1),$T{start_row + 7}+S{start_row + 6}-S{start_row + 5},IF(AND($V$1=\"YTD\",S$4<$K$1),S{start_row + 16},IF(AND($V$1=\"Spring\",S$4<7),S{start_row + 16},IF(AND($V$1=\"Fall\",S$4>6,S$4<$K$1),S{start_row + 16},IF(AND($V$1=\"Fall\",S$4=1,$K$1>1),$T{start_row + 7}+S{start_row + 6}-S{start_row + 5},IF(AND($V$1=\"Fall\",S$4>6,S$4=$K$1),R{start_row + 16}+S{start_row + 32}+S{start_row + 6}-S{start_row + 5},IF(AND($V$1=\"LY Fall\",S$4>6),S{start_row + 38},IF(AND(S$4=1,$V$1=\"LY FALL\"),$T{start_row + 38}+S{start_row + 6}+S{start_row + 6}-S{start_row + 5},IF(AND(S$4=$K$1,S$4>1),R{start_row + 16}+S{start_row + 32}+S{start_row + 6}-S{start_row + 5},R{start_row + 7}+S{start_row + 6}-S{start_row + 5})))))))))))",

            f"T{start_row + 7}": f"=IF(AND($V$1=\"Current Mth\",$K$1=T$4,T$4=1),$T{start_row + 38}+T{start_row + 6}+T{start_row + 32}-T{start_row + 5},IF(AND($V$1=\"Current Mth\",$K$1=T$4),T{start_row + 16}+T{start_row + 6}-(T{start_row + 5}-T{start_row + 12}),IF(AND($V$1=\"Current Mth\",T$4=1,$K$1>1),$T{start_row + 7}+T{start_row + 6}-T{start_row + 5},IF(AND($V$1=\"YTD\",T$4<$K$1),T{start_row + 16},IF(AND($V$1=\"Spring\",T$4<7),T{start_row + 16},IF(AND($V$1=\"Fall\",T$4>6,T$4<$K$1),T{start_row + 16},IF(AND($V$1=\"Fall\",T$4=1,$K$1>1),$T{start_row + 7}+T{start_row + 6}-T{start_row + 5},IF(AND($V$1=\"Fall\",T$4>6,T$4=$K$1),S{start_row + 16}+T{start_row + 32}+T{start_row + 6}-T{start_row + 5},IF(AND($V$1=\"LY Fall\",T$4>6),T{start_row + 38},IF(AND(T$4=1,$V$1=\"LY FALL\"),$T{start_row + 38}+T{start_row + 6}+T{start_row + 6}-T{start_row + 5},IF(AND(T$4=$K$1,T$4>1),S{start_row + 16}+T{start_row + 32}+T{start_row + 6}-T{start_row + 5},S{start_row + 7}+T{start_row + 6}-T{start_row + 5})))))))))))",

            f"U{start_row + 7}": f"=AVERAGEIFS(I{start_row + 7}:T{start_row + 7},$I$4:$T$4,\"<=\"&$K$1)",
            f"V{start_row + 7}": f"=AVERAGEIFS(I{start_row + 7}:N{start_row + 7},$I$4:$N$4,\"<=\"&$K$2)",
            f"W{start_row + 7}": f"=AVERAGEIFS(O{start_row + 7}:T{start_row + 7},$O$4:$T$4,\"<=\"&$M$2)",
            f"I{start_row + 8}": Nav_Feb,
            f"J{start_row + 8}": Nav_Mar,
            f"K{start_row + 8}": Nav_Apr,
            f"L{start_row + 8}": Nav_May,
            f"M{start_row + 8}": Nav_Jun,
            f"N{start_row + 8}": Nav_Jul,
            f"O{start_row + 8}": Nav_Aug,
            f"P{start_row + 8}":Nav_Sep,
            f"Q{start_row + 8}":Nav_Oct,
            f"R{start_row + 8}":Nav_Nov,
            f"S{start_row + 8}": Nav_Dec,
            f"T{start_row + 8}": Nav_Jan,
            f"I{start_row + 6}": planned_shp["FEB"],
            f"J{start_row + 6}": planned_shp["MAR"],
            f"K{start_row + 6}": planned_shp["APR"],
            f"L{start_row + 6}": planned_shp["MAY"],
            f"M{start_row + 6}": planned_shp["JUN"],
            f"N{start_row + 6}": planned_shp["JUL"],
            f"O{start_row + 6}": planned_shp["AUG"],
            f"P{start_row + 6}":planned_shp["SEP"],
            f"Q{start_row + 6}":planned_shp["OCT"],
            f"R{start_row + 6}":planned_shp["NOV"],
            f"S{start_row + 6}": planned_shp["DEC"],
            f"T{start_row + 6}": planned_shp["JAN"],
            f"U{start_row + 8}": f"=SUM(I{start_row + 8}:T{start_row + 8})",
            f"V{start_row + 8}": f"=IFERROR(SUM(I{start_row + 8}:N{start_row + 8}),0)",
            f"W{start_row + 8}": f"=IFERROR(SUM(O{start_row + 8}:T{start_row + 8}),0)",

            f"I{start_row + 9}": Macys_Proj_Receipts_Feb,
            f"J{start_row + 9}": Macys_Proj_Receipts_Mar,
            f"K{start_row + 9}":Macys_Proj_Receipts_Apr ,
            f"L{start_row + 9}":Macys_Proj_Receipts_May,
            f"M{start_row + 9}": Macys_Proj_Receipts_Jun,
            f"N{start_row + 9}": Macys_Proj_Receipts_Jul,
            f"O{start_row + 9}": Macys_Proj_Receipts_Aug,
            f"P{start_row + 9}":Macys_Proj_Receipts_Sep,
            f"Q{start_row + 9}": Macys_Proj_Receipts_oct,
            f"R{start_row + 9}": Macys_Proj_Receipts_Nov,
            f"S{start_row + 9}": Macys_Proj_Receipts_Dec,
            f"T{start_row + 9}": Macys_Proj_Receipts_Jan,
            f"U{start_row + 9}": f"=SUM(I{start_row + 9}:T{start_row + 9})",
            f"V{start_row + 9}": f"=IFERROR(SUM(I{start_row + 9}:N{start_row + 9}),0)",
            f"W{start_row + 9}": f"=IFERROR(SUM(O{start_row + 9}:T{start_row + 9}),0)",
            f"I{start_row + 10}": f"=I{start_row + 5}/(I{start_row + 5}+I{start_row + 7})",
            f"J{start_row + 10}": f"=J{start_row + 5}/(J{start_row + 5}+J{start_row + 7})",
            f"K{start_row + 10}": f"=K{start_row + 5}/(K{start_row + 5}+K{start_row + 7})",
            f"L{start_row + 10}": f"=L{start_row + 5}/(L{start_row + 5}+L{start_row + 7})",
            f"M{start_row + 10}": f"=M{start_row + 5}/(M{start_row + 5}+M{start_row + 7})",
            f"N{start_row + 10}": f"=N{start_row + 5}/(N{start_row + 5}+N{start_row + 7})",
            f"O{start_row + 10}": f"=O{start_row + 5}/(O{start_row + 5}+O{start_row + 7})",
            f"P{start_row + 10}": f"=P{start_row + 5}/(P{start_row + 5}+P{start_row + 7})",
            f"Q{start_row + 10}": f"=Q{start_row + 5}/(Q{start_row + 5}+Q{start_row + 7})",
            f"R{start_row + 10}": f"=R{start_row + 5}/(R{start_row + 5}+R{start_row + 7})",
            f"S{start_row + 10}": f"=S{start_row + 5}/(S{start_row + 5}+S{start_row + 7})",
            f"T{start_row + 10}": f"=T{start_row + 5}/(T{start_row + 5}+T{start_row + 7})",
            f"U{start_row + 10}": f"=U{start_row + 5}/(U{start_row + 5}+U{start_row + 7})",
            f"V{start_row + 10}": f"=V{start_row + 5}/(V{start_row + 5}+V{start_row + 7})",
            f"W{start_row + 10}": f"=W{start_row + 5}/(W{start_row + 5}+W{start_row + 7})",
            f"I{start_row + 14}": TY_MCOM_Unit_Sales['FEB'],

            f"J{start_row + 14}":TY_MCOM_Unit_Sales['MAR'],

            f"K{start_row + 14}":TY_MCOM_Unit_Sales['APR'],

            f"L{start_row + 14}":TY_MCOM_Unit_Sales['MAY'],

            f"M{start_row + 14}": TY_MCOM_Unit_Sales['JUN'],

            f"N{start_row + 14}": TY_MCOM_Unit_Sales['JUL'],

            f"O{start_row + 14}": TY_MCOM_Unit_Sales['AUG'],

            f"P{start_row + 14}":TY_MCOM_Unit_Sales['SEP'],

            f"Q{start_row + 14}": TY_MCOM_Unit_Sales['OCT'],

            f"R{start_row + 14}":TY_MCOM_Unit_Sales['NOV'],

            f"S{start_row + 14}": TY_MCOM_Unit_Sales['DEC'],

            f"T{start_row + 14}":TY_MCOM_Unit_Sales['JAN'],

            f"U{start_row + 14}": f"=IFERROR(SUM(I{start_row + 14}:T{start_row + 14}),0)",
            f"V{start_row + 14}": f"=IFERROR(SUM(I{start_row + 14}:N{start_row + 14}),0)",
            f"W{start_row + 14}": f"=IFERROR(SUM(O{start_row + 14}:T{start_row + 14}),0)",
            f"I{start_row + 13}": f"=I{start_row + 12}-I{start_row + 14}",
            f"J{start_row + 13}": f"=J{start_row + 12}-J{start_row + 14}",
            f"K{start_row + 13}": f"=K{start_row + 12}-K{start_row + 14}",
            f"L{start_row + 13}": f"=L{start_row + 12}-L{start_row + 14}",
            f"M{start_row + 13}": f"=M{start_row + 12}-M{start_row + 14}",
            f"N{start_row + 13}": f"=N{start_row + 12}-N{start_row + 14}",
            f"O{start_row + 13}": f"=O{start_row + 12}-O{start_row + 14}",
            f"P{start_row + 13}": f"=P{start_row + 12}-P{start_row + 14}",
            f"Q{start_row + 13}": f"=Q{start_row + 12}-Q{start_row + 14}",
            f"R{start_row + 13}": f"=R{start_row + 12}-R{start_row + 14}",
            f"S{start_row + 13}": f"=S{start_row + 12}-S{start_row + 14}",
            f"T{start_row + 13}": f"=T{start_row + 12}-T{start_row + 14}",

            f"U{start_row + 13}": f"=IFERROR(SUM(I{start_row + 13}:T{start_row + 13}),0)",
            f"V{start_row + 13}": f"=IFERROR(SUM(I{start_row + 13}:N{start_row + 13}),0)",
            f"W{start_row + 13}": f"=IFERROR(SUM(O{start_row + 13}:T{start_row + 13}),0)",

            f"I{start_row + 15}": f"=IFERROR(I{start_row + 14}/I{start_row + 12},0)",
            f"J{start_row + 15}": f"=IFERROR(J{start_row + 14}/J{start_row + 12},0)",
            f"K{start_row + 15}": f"=IFERROR(K{start_row + 14}/K{start_row + 12},0)",
            f"L{start_row + 15}": f"=IFERROR(L{start_row + 14}/L{start_row + 12},0)",
            f"M{start_row + 15}": f"=IFERROR(M{start_row + 14}/M{start_row + 12},0)",
            f"N{start_row + 15}": f"=IFERROR(N{start_row + 14}/N{start_row + 12},0)",
            f"O{start_row + 15}": f"=IFERROR(O{start_row + 14}/O{start_row + 12},0)",
            f"P{start_row + 15}": f"=IFERROR(P{start_row + 14}/P{start_row + 12},0)",
            f"Q{start_row + 15}": f"=IFERROR(Q{start_row + 14}/Q{start_row + 12},0)",
            f"R{start_row + 15}": f"=IFERROR(R{start_row + 14}/R{start_row + 12},0)",
            f"S{start_row + 15}": f"=IFERROR(S{start_row + 14}/S{start_row + 12},0)",
            f"T{start_row + 15}": f"=IFERROR(T{start_row + 14}/T{start_row + 12},0)",
            f"U{start_row + 15}": f"=IFERROR(U{start_row + 14}/U{start_row + 12},0)",
            f"V{start_row + 15}": f"=IFERROR(V{start_row + 14}/V{start_row + 12},0)",
            f"W{start_row + 15}": f"=IFERROR(W{start_row + 14}/W{start_row + 12},0)",
            f"I{start_row + 18}": TY_OH_MCOM_Units['FEB'],

            f"J{start_row + 18}": TY_OH_MCOM_Units['MAR'],

            f"K{start_row + 18}":TY_OH_MCOM_Units['APR'],

            f"L{start_row + 18}":TY_OH_MCOM_Units['MAY'],

            f"M{start_row + 18}": TY_OH_MCOM_Units['JUN'],

            f"N{start_row + 18}": TY_OH_MCOM_Units['JUL'],

            f"O{start_row + 18}":TY_OH_MCOM_Units['AUG'],

            f"P{start_row + 18}":TY_OH_MCOM_Units['SEP'],

            f"Q{start_row + 18}":TY_OH_MCOM_Units['OCT'],

            f"R{start_row + 18}": TY_OH_MCOM_Units['NOV'],

            f"S{start_row + 18}": TY_OH_MCOM_Units['DEC'],

            f"T{start_row + 18}":TY_OH_MCOM_Units['JAN'],

            f"U{start_row + 18}": f"=AVERAGEIFS(I{start_row + 18}:T{start_row + 18},$I$4:$T$4,\"<=\"&$K$1)",
            f"V{start_row + 18}": f"=AVERAGEIFS(I{start_row + 18}:N{start_row + 18},$I$4:$N$4,\"<=\"&$K$2)",
            f"W{start_row + 18}": f"=AVERAGEIFS(O{start_row + 18}:T{start_row + 18},$O$4:$T$4,\"<=\"&$M$2)",
            f"I{start_row + 17}": f"=I{start_row + 16}-I{start_row + 18}",
            f"J{start_row + 17}": f"=J{start_row + 16}-J{start_row + 18}",
            f"K{start_row + 17}": f"=K{start_row + 16}-K{start_row + 18}",
            f"L{start_row + 17}": f"=L{start_row + 16}-L{start_row + 18}",
            f"M{start_row + 17}": f"=M{start_row + 16}-M{start_row + 18}",
            f"N{start_row + 17}": f"=N{start_row + 16}-N{start_row + 18}",
            f"O{start_row + 17}": f"=O{start_row + 16}-O{start_row + 18}",
            f"P{start_row + 17}": f"=P{start_row + 16}-P{start_row + 18}",
            f"Q{start_row + 17}": f"=Q{start_row + 16}-Q{start_row + 18}",
            f"R{start_row + 17}": f"=R{start_row + 16}-R{start_row + 18}",
            f"S{start_row + 17}": f"=S{start_row + 16}-S{start_row + 18}",
            f"T{start_row + 17}": f"=T{start_row + 16}-T{start_row + 18}",

            f"U{start_row + 17}": f"=AVERAGEIFS(I{start_row + 17}:T{start_row + 17},$I$4:$T$4,\"<=\"&$K$1)",
            f"V{start_row + 17}": f"=AVERAGEIFS(I{start_row + 17}:N{start_row + 17},$I$4:$N$4,\"<=\"&$K$2)",
            f"W{start_row + 17}": f"=AVERAGEIFS(O{start_row + 17}:T{start_row + 17},$O$4:$T$4,\"<=\"&$M$2)",
            f"I{start_row + 19}": f"=IFERROR(I{start_row + 18}/I{start_row + 16},0)",
            f"J{start_row + 19}": f"=IFERROR(J{start_row + 18}/J{start_row + 16},0)",
            f"K{start_row + 19}": f"=IFERROR(K{start_row + 18}/K{start_row + 16},0)",
            f"L{start_row + 19}": f"=IFERROR(L{start_row + 18}/L{start_row + 16},0)",
            f"M{start_row + 19}": f"=IFERROR(M{start_row + 18}/M{start_row + 16},0)",
            f"N{start_row + 19}": f"=IFERROR(N{start_row + 18}/N{start_row + 16},0)",
            f"O{start_row + 19}": f"=IFERROR(O{start_row + 18}/O{start_row + 16},0)",
            f"P{start_row + 19}": f"=IFERROR(P{start_row + 18}/P{start_row + 16},0)",
            f"Q{start_row + 19}": f"=IFERROR(Q{start_row + 18}/Q{start_row + 16},0)",
            f"R{start_row + 19}": f"=IFERROR(R{start_row + 18}/R{start_row + 16},0)",
            f"S{start_row + 19}": f"=IFERROR(S{start_row + 18}/S{start_row + 16},0)",
            f"T{start_row + 19}": f"=IFERROR(T{start_row + 18}/T{start_row + 16},0)",
            f"U{start_row + 19}": f"=IFERROR(U{start_row + 18}/U{start_row + 16},0)",
            f"V{start_row + 19}": f"=IFERROR(V{start_row + 18}/V{start_row + 16},0)",
            f"W{start_row + 19}": f"=IFERROR(W{start_row + 18}/W{start_row + 16},0)",
            f"I{start_row + 20}":PTD_TY_Sales['FEB'] ,

            f"J{start_row + 20}": PTD_TY_Sales['MAR'],

            f"K{start_row + 20}": PTD_TY_Sales['APR'],

            f"L{start_row + 20}":PTD_TY_Sales['MAY'],

            f"M{start_row + 20}": PTD_TY_Sales['JUN'],

            f"N{start_row + 20}": PTD_TY_Sales['JUL'],

            f"O{start_row + 20}":PTD_TY_Sales['AUG'],

            f"P{start_row + 20}": PTD_TY_Sales['SEP'],

            f"Q{start_row + 20}": PTD_TY_Sales['OCT'],

            f"R{start_row + 20}": PTD_TY_Sales['NOV'],

            f"S{start_row + 20}":PTD_TY_Sales['DEC'],

            f"T{start_row + 20}": PTD_TY_Sales['JAN'],

            # Aggregates
            f"U{start_row + 20}": f"=IFERROR(SUM(I{start_row + 20}:T{start_row + 20}),0)",
            f"V{start_row + 20}": f"=IFERROR(SUM(I{start_row + 20}:N{start_row + 20}),0)",
            f"W{start_row + 20}": f"=IFERROR(SUM(O{start_row + 20}:T{start_row + 20}),0)",
            f"I{start_row + 21}":MCOM_PTD_TY_Sales['FEB'],

            f"J{start_row + 21}":MCOM_PTD_TY_Sales['MAR'],

            f"K{start_row + 21}": MCOM_PTD_TY_Sales['APR'],

            f"L{start_row + 21}":MCOM_PTD_TY_Sales['MAY'],

            f"M{start_row + 21}":MCOM_PTD_TY_Sales['JUN'],

            f"N{start_row + 21}": MCOM_PTD_TY_Sales['JUL'],

            f"O{start_row + 21}": MCOM_PTD_TY_Sales['AUG'],

            f"P{start_row + 21}":MCOM_PTD_TY_Sales['SEP'],

            f"Q{start_row + 21}": MCOM_PTD_TY_Sales['OCT'],

            f"R{start_row + 21}": MCOM_PTD_TY_Sales['NOV'],

            f"S{start_row + 21}": MCOM_PTD_TY_Sales['DEC'],

            f"T{start_row + 21}":MCOM_PTD_TY_Sales['JAN'],

            f"U{start_row + 21}": f"=IFERROR(SUM(I{start_row + 21}:T{start_row + 21}),0)",
            f"V{start_row + 21}": f"=IFERROR(SUM(I{start_row + 21}:N{start_row + 21}),0)",
            f"W{start_row + 21}": f"=IFERROR(SUM(O{start_row + 21}:T{start_row + 21}),0)",

            f"I{start_row + 22}": f"=IFERROR(TEXT(IFERROR(I{start_row + 20}/I{start_row + 16},0),\"$0\") & \"/ \" & TEXT(((I{start_row + 20}/I{start_row + 16})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"J{start_row + 22}": f"=IFERROR(TEXT(IFERROR(J{start_row + 20}/J{start_row + 16},0),\"$0\") & \"/ \" & TEXT(((J{start_row + 20}/J{start_row + 16})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"K{start_row + 22}": f"=IFERROR(TEXT(IFERROR(K{start_row + 20}/K{start_row + 16},0),\"$0\") & \"/ \" & TEXT(((K{start_row + 20}/K{start_row + 16})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"L{start_row + 22}": f"=IFERROR(TEXT(IFERROR(L{start_row + 20}/L{start_row + 16},0),\"$0\") & \"/ \" & TEXT(((L{start_row + 20}/L{start_row + 16})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"M{start_row + 22}": f"=IFERROR(TEXT(IFERROR(M{start_row + 20}/M{start_row + 16},0),\"$0\") & \"/ \" & TEXT(((M{start_row + 20}/M{start_row + 16})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"N{start_row + 22}": f"=IFERROR(TEXT(IFERROR(N{start_row + 20}/N{start_row + 16},0),\"$0\") & \"/ \" & TEXT(((N{start_row + 20}/N{start_row + 16})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"O{start_row + 22}": f"=IFERROR(TEXT(IFERROR(O{start_row + 20}/O{start_row + 16},0),\"$0\") & \"/ \" & TEXT(((O{start_row + 20}/O{start_row + 16})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"P{start_row + 22}": f"=IFERROR(TEXT(IFERROR(P{start_row + 20}/P{start_row + 16},0),\"$0\") & \"/ \" & TEXT(((P{start_row + 20}/P{start_row + 16})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"Q{start_row + 22}": f"=IFERROR(TEXT(IFERROR(Q{start_row + 20}/Q{start_row + 16},0),\"$0\") & \"/ \" & TEXT(((Q{start_row + 20}/Q{start_row + 16})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"R{start_row + 22}": f"=IFERROR(TEXT(IFERROR(R{start_row + 20}/R{start_row + 16},0),\"$0\") & \"/ \" & TEXT(((R{start_row + 20}/R{start_row + 16})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"S{start_row + 22}": f"=IFERROR(TEXT(IFERROR(S{start_row + 20}/S{start_row + 16},0),\"$0\") & \"/ \" & TEXT(((S{start_row + 20}/S{start_row + 16})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"T{start_row + 22}": f"=IFERROR(TEXT(IFERROR(T{start_row + 20}/T{start_row + 16},0),\"$0\") & \"/ \" & TEXT(((T{start_row + 20}/T{start_row + 16})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"U{start_row + 22}": f"=IFERROR(TEXT(IFERROR(U{start_row + 20}/U{start_row + 16},0),\"$0\") & \"/ \" & TEXT(((U{start_row + 20}/U{start_row + 16})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"V{start_row + 22}": f"=IFERROR(TEXT(IFERROR(V{start_row + 20}/V{start_row + 16},0),\"$0\") & \"/ \" & TEXT(((V{start_row + 20}/V{start_row + 16})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"W{start_row + 22}": f"=IFERROR(TEXT(IFERROR(W{start_row + 20}/W{start_row + 16},0),\"$0\") & \"/ \" & TEXT(((W{start_row + 20}/W{start_row + 16})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"I{start_row + 23}": f"=IFERROR(IFERROR(I{start_row + 12}/(I{start_row + 12}+I{start_row + 16}),0),0)",
            f"J{start_row + 23}": f"=IFERROR(IFERROR(J{start_row + 12}/(J{start_row + 12}+J{start_row + 16}),0),0)",
            f"K{start_row + 23}": f"=IFERROR(IFERROR(K{start_row + 12}/(K{start_row + 12}+K{start_row + 16}),0),0)",
            f"L{start_row + 23}": f"=IFERROR(IFERROR(L{start_row + 12}/(L{start_row + 12}+L{start_row + 16}),0),0)",
            f"M{start_row + 23}": f"=IFERROR(IFERROR(M{start_row + 12}/(M{start_row + 12}+M{start_row + 16}),0),0)",
            f"N{start_row + 23}": f"=IFERROR(IFERROR(N{start_row + 12}/(N{start_row + 12}+N{start_row + 16}),0),0)",
            f"O{start_row + 23}": f"=IFERROR(IFERROR(O{start_row + 12}/(O{start_row + 12}+O{start_row + 16}),0),0)",
            f"P{start_row + 23}": f"=IFERROR(IFERROR(P{start_row + 12}/(P{start_row + 12}+P{start_row + 16}),0),0)",
            f"Q{start_row + 23}": f"=IFERROR(IFERROR(Q{start_row + 12}/(Q{start_row + 12}+Q{start_row + 16}),0),0)",
            f"R{start_row + 23}": f"=IFERROR(IFERROR(R{start_row + 12}/(R{start_row + 12}+R{start_row + 16}),0),0)",
            f"S{start_row + 23}": f"=IFERROR(IFERROR(S{start_row + 12}/(S{start_row + 12}+S{start_row + 16}),0),0)",
            f"T{start_row + 23}": f"=IFERROR(IFERROR(T{start_row + 12}/(T{start_row + 12}+T{start_row + 16}),0),0)",

            f"U{start_row + 23}": f"=IFERROR(IFERROR(U{start_row + 12}/(U{start_row + 12}+U{start_row + 16}-U{start_row + 8}),0),0)",
            f"V{start_row + 23}": f"=IFERROR(IFERROR(V{start_row + 12}/(V{start_row + 12}+V{start_row + 16}-V{start_row + 8}),0),0)",
            f"W{start_row + 23}": f"=IFERROR(IFERROR(W{start_row + 12}/(W{start_row + 12}+W{start_row + 16}-W{start_row + 8}),0),0)",
            f"I{start_row + 24}": f"=IFERROR((I{start_row + 12}-I{start_row + 14})/((I{start_row + 12}-I{start_row + 14})+(I{start_row + 16}-I{start_row + 18})),0)",
            f"J{start_row + 24}": f"=IFERROR((J{start_row + 12}-J{start_row + 14})/((J{start_row + 12}-J{start_row + 14})+(J{start_row + 16}-J{start_row + 18})),0)",
            f"K{start_row + 24}": f"=IFERROR((K{start_row + 12}-K{start_row + 14})/((K{start_row + 12}-K{start_row + 14})+(K{start_row + 16}-K{start_row + 18})),0)",
            f"L{start_row + 24}": f"=IFERROR((L{start_row + 12}-L{start_row + 14})/((L{start_row + 12}-L{start_row + 14})+(L{start_row + 16}-L{start_row + 18})),0)",
            f"M{start_row + 24}": f"=IFERROR((M{start_row + 12}-M{start_row + 14})/((M{start_row + 12}-M{start_row + 14})+(M{start_row + 16}-M{start_row + 18})),0)",
            f"N{start_row + 24}": f"=IFERROR((N{start_row + 12}-N{start_row + 14})/((N{start_row + 12}-N{start_row + 14})+(N{start_row + 16}-N{start_row + 18})),0)",
            f"O{start_row + 24}": f"=IFERROR((O{start_row + 12}-O{start_row + 14})/((O{start_row + 12}-O{start_row + 14})+(O{start_row + 16}-O{start_row + 18})),0)",
            f"P{start_row + 24}": f"=IFERROR((P{start_row + 12}-P{start_row + 14})/((P{start_row + 12}-P{start_row + 14})+(P{start_row + 16}-P{start_row + 18})),0)",
            f"Q{start_row + 24}": f"=IFERROR((Q{start_row + 12}-Q{start_row + 14})/((Q{start_row + 12}-Q{start_row + 14})+(Q{start_row + 16}-Q{start_row + 18})),0)",
            f"R{start_row + 24}": f"=IFERROR((R{start_row + 12}-R{start_row + 14})/((R{start_row + 12}-R{start_row + 14})+(R{start_row + 16}-R{start_row + 18})),0)",
            f"S{start_row + 24}": f"=IFERROR((S{start_row + 12}-S{start_row + 14})/((S{start_row + 12}-S{start_row + 14})+(S{start_row + 16}-S{start_row + 18})),0)",
            f"T{start_row + 24}": f"=IFERROR((T{start_row + 12}-T{start_row + 14})/((T{start_row + 12}-T{start_row + 14})+(T{start_row + 16}-T{start_row + 18})),0)",

            f"U{start_row + 24}": f"=IFERROR((U{start_row + 12}-U{start_row + 14})/((U{start_row + 12}-U{start_row + 14})+(U{start_row + 16}-U{start_row + 18}-U{start_row + 8})),0)",
            f"V{start_row + 24}": f"=IFERROR((V{start_row + 12}-V{start_row + 14})/((V{start_row + 12}-V{start_row + 14})+(V{start_row + 16}-V{start_row + 18}-V{start_row + 8})),0)",
            f"W{start_row + 24}": f"=IFERROR((W{start_row + 12}-W{start_row + 14})/((W{start_row + 12}-W{start_row + 14})+(W{start_row + 16}-W{start_row + 18})),0)",
            f"I{start_row + 25}": f"=IFERROR(IFERROR(I{start_row + 12}/I{start_row + 16},0),0)",
            f"J{start_row + 25}": f"=IFERROR(IFERROR(SUM(I{start_row + 12}:J{start_row + 12})/AVERAGE(I{start_row + 16}:J{start_row + 16}),0),0)",
            f"K{start_row + 25}": f"=IFERROR(IFERROR(SUM(I{start_row + 12}:K{start_row + 12})/AVERAGE(I{start_row + 16}:K{start_row + 16}),0),0)",
            f"L{start_row + 25}": f"=IFERROR(IFERROR(SUM(I{start_row + 12}:L{start_row + 12})/AVERAGE(I{start_row + 16}:L{start_row + 16}),0),0)",
            f"M{start_row + 25}": f"=IFERROR(IFERROR(SUM(I{start_row + 12}:M{start_row + 12})/AVERAGE(I{start_row + 16}:M{start_row + 16}),0),0)",
            f"N{start_row + 25}": f"=IFERROR(IFERROR(SUM(I{start_row + 12}:N{start_row + 12})/AVERAGE(I{start_row + 16}:N{start_row + 16}),0),0)",
            f"O{start_row + 25}": f"=IFERROR(IFERROR(SUM(I{start_row + 12}:O{start_row + 12})/AVERAGE(I{start_row + 16}:O{start_row + 16}),0),0)",
            f"P{start_row + 25}": f"=IFERROR(IFERROR(SUM(I{start_row + 12}:P{start_row + 12})/AVERAGE(I{start_row + 16}:P{start_row + 16}),0),0)",
            f"Q{start_row + 25}": f"=IFERROR(IFERROR(SUM(I{start_row + 12}:Q{start_row + 12})/AVERAGE(I{start_row + 16}:Q{start_row + 16}),0),0)",
            f"R{start_row + 25}": f"=IFERROR(IFERROR(SUM(I{start_row + 12}:R{start_row + 12})/AVERAGE(I{start_row + 16}:R{start_row + 16}),0),0)",
            f"S{start_row + 25}": f"=IFERROR(IFERROR(SUM(I{start_row + 12}:S{start_row + 12})/AVERAGE(I{start_row + 16}:S{start_row + 16}),0),0)",
            f"T{start_row + 25}": f"=IFERROR(IFERROR(SUM(I{start_row + 12}:T{start_row + 12})/AVERAGE(I{start_row + 16}:T{start_row + 16}),0),0)",
            f"U{start_row + 25}": f"=IFERROR(U{start_row + 12}/U{start_row + 16},0)",
            f"V{start_row + 25}": f"=IFERROR(V{start_row + 12}/V{start_row + 16},0)",
            f"W{start_row + 25}": f"=IFERROR(W{start_row + 12}/W{start_row + 16},0)",
            f"I{start_row + 26}": f"=IFERROR(I{start_row + 13}/I{start_row + 17},0)",
            f"J{start_row + 26}": f"=IFERROR(SUM(I{start_row + 13}:J{start_row + 13})/AVERAGE(I{start_row + 17}:J{start_row + 17}),0)",
            f"K{start_row + 26}": f"=IFERROR(SUM(I{start_row + 13}:K{start_row + 13})/AVERAGE(I{start_row + 17}:K{start_row + 17}),0)",
            f"L{start_row + 26}": f"=IFERROR(SUM(I{start_row + 13}:L{start_row + 13})/AVERAGE(I{start_row + 17}:L{start_row + 17}),0)",
            f"M{start_row + 26}": f"=IFERROR(SUM(I{start_row + 13}:M{start_row + 13})/AVERAGE(I{start_row + 17}:M{start_row + 17}),0)",
            f"N{start_row + 26}": f"=IFERROR(SUM(I{start_row + 13}:N{start_row + 13})/AVERAGE(I{start_row + 17}:N{start_row + 17}),0)",
            f"O{start_row + 26}": f"=IFERROR(SUM(I{start_row + 13}:O{start_row + 13})/AVERAGE(I{start_row + 17}:O{start_row + 17}),0)",
            f"P{start_row + 26}": f"=IFERROR(SUM(I{start_row + 13}:P{start_row + 13})/AVERAGE(I{start_row + 17}:P{start_row + 17}),0)",
            f"Q{start_row + 26}": f"=IFERROR(SUM(I{start_row + 13}:Q{start_row + 13})/AVERAGE(I{start_row + 17}:Q{start_row + 17}),0)",
            f"R{start_row + 26}": f"=IFERROR(SUM(I{start_row + 13}:R{start_row + 13})/AVERAGE(I{start_row + 17}:R{start_row + 17}),0)",
            f"S{start_row + 26}": f"=IFERROR(SUM(I{start_row + 13}:S{start_row + 13})/AVERAGE(I{start_row + 17}:S{start_row + 17}),0)",
            f"T{start_row + 26}": f"=IFERROR(SUM(I{start_row + 13}:T{start_row + 13})/AVERAGE(I{start_row + 17}:T{start_row + 17}),0)",
            f"U{start_row + 26}": f"=IFERROR(U{start_row + 13}/U{start_row + 17},0)",
            f"V{start_row + 26}": f"=IFERROR(V{start_row + 13}/V{start_row + 17},0)",
            f"W{start_row + 26}": f"=IFERROR(W{start_row + 13}/W{start_row + 17},0)",
            f"I{start_row + 36}": LY_MCOM_Unit_Sales['FEB'],
            f"J{start_row + 36}":LY_MCOM_Unit_Sales['MAR'] ,
            f"K{start_row + 36}":LY_MCOM_Unit_Sales['APR'] ,
            f"L{start_row + 36}": LY_MCOM_Unit_Sales['MAY'],
            f"M{start_row + 36}":LY_MCOM_Unit_Sales['JUN'] ,
            f"N{start_row + 36}":LY_MCOM_Unit_Sales['JUL'] ,
            f"O{start_row + 36}":LY_MCOM_Unit_Sales['AUG'] ,
            f"P{start_row + 36}":LY_MCOM_Unit_Sales['SEP'],
            f"Q{start_row + 36}":LY_MCOM_Unit_Sales['OCT'] ,
            f"R{start_row + 36}":LY_MCOM_Unit_Sales['NOV'],
            f"S{start_row + 36}":LY_MCOM_Unit_Sales['DEC'] ,
            f"T{start_row + 36}":LY_MCOM_Unit_Sales['JAN'] ,

            f"U{start_row + 36}": f"=SUM(I{start_row + 36}:T{start_row + 36})",
            f"V{start_row + 36}": f"=SUM(I{start_row + 36}:N{start_row + 36})",
            f"W{start_row + 36}": f"=SUM(O{start_row + 36}:T{start_row + 36})",
            f"I{start_row + 35}": f"=I{start_row + 34}-I{start_row + 36}",
            f"J{start_row + 35}": f"=J{start_row + 34}-J{start_row + 36}",
            f"K{start_row + 35}": f"=K{start_row + 34}-K{start_row + 36}",
            f"L{start_row + 35}": f"=L{start_row + 34}-L{start_row + 36}",
            f"M{start_row + 35}": f"=M{start_row + 34}-M{start_row + 36}",
            f"N{start_row + 35}": f"=N{start_row + 34}-N{start_row + 36}",
            f"O{start_row + 35}": f"=O{start_row + 34}-O{start_row + 36}",
            f"P{start_row + 35}": f"=P{start_row + 34}-P{start_row + 36}",
            f"Q{start_row + 35}": f"=Q{start_row + 34}-Q{start_row + 36}",
            f"R{start_row + 35}": f"=R{start_row + 34}-R{start_row + 36}",
            f"S{start_row + 35}": f"=S{start_row + 34}-S{start_row + 36}",
            f"T{start_row + 35}": f"=T{start_row + 34}-T{start_row + 36}",

            f"U{start_row + 35}": f"=IFERROR(SUM(I{start_row + 35}:T{start_row + 35}),0)",
            f"V{start_row + 35}": f"=IFERROR(SUM(I{start_row + 35}:N{start_row + 35}),0)",
            f"W{start_row + 35}": f"=IFERROR(SUM(O{start_row + 35}:T{start_row + 35}),0)",
            f"I{start_row + 27}": f"=IFERROR(IF(I{start_row + 13}+I{start_row + 35}=0,0,(I{start_row + 13}-I{start_row + 35})/I{start_row + 35}),1)",
            f"J{start_row + 27}": f"=IFERROR(IF(J{start_row + 13}+J{start_row + 35}=0,0,(J{start_row + 13}-J{start_row + 35})/J{start_row + 35}),1)",
            f"K{start_row + 27}": f"=IFERROR(IF(K{start_row + 13}+K{start_row + 35}=0,0,(K{start_row + 13}-K{start_row + 35})/K{start_row + 35}),1)",
            f"L{start_row + 27}": f"=IFERROR(IF(L{start_row + 13}+L{start_row + 35}=0,0,(L{start_row + 13}-L{start_row + 35})/L{start_row + 35}),1)",
            f"M{start_row + 27}": f"=IFERROR(IF(M{start_row + 13}+M{start_row + 35}=0,0,(M{start_row + 13}-M{start_row + 35})/M{start_row + 35}),1)",
            f"N{start_row + 27}": f"=IFERROR(IF(N{start_row + 13}+N{start_row + 35}=0,0,(N{start_row + 13}-N{start_row + 35})/N{start_row + 35}),1)",
            f"O{start_row + 27}": f"=IFERROR(IF(O{start_row + 13}+O{start_row + 35}=0,0,(O{start_row + 13}-O{start_row + 35})/O{start_row + 35}),1)",
            f"P{start_row + 27}": f"=IFERROR(IF(P{start_row + 13}+P{start_row + 35}=0,0,(P{start_row + 13}-P{start_row + 35})/P{start_row + 35}),1)",
            f"Q{start_row + 27}": f"=IFERROR(IF(Q{start_row + 13}+Q{start_row + 35}=0,0,(Q{start_row + 13}-Q{start_row + 35})/Q{start_row + 35}),1)",
            f"R{start_row + 27}": f"=IFERROR(IF(R{start_row + 13}+R{start_row + 35}=0,0,(R{start_row + 13}-R{start_row + 35})/R{start_row + 35}),1)",
            f"S{start_row + 27}": f"=IFERROR(IF(S{start_row + 13}+S{start_row + 35}=0,0,(S{start_row + 13}-S{start_row + 35})/S{start_row + 35}),1)",
            f"T{start_row + 27}": f"=IFERROR(IF(T{start_row + 13}+T{start_row + 35}=0,0,(T{start_row + 13}-T{start_row + 35})/T{start_row + 35}),1)",
            f"U{start_row + 27}": f"=IFERROR(IF(U{start_row + 13}+U{start_row + 35}=0,0,(U{start_row + 13}-U{start_row + 35})/U{start_row + 35}),1)",
            f"V{start_row + 27}": f"=IFERROR(IF(V{start_row + 13}+V{start_row + 35}=0,0,(V{start_row + 13}-V{start_row + 35})/V{start_row + 35}),1)",
            f"W{start_row + 27}": f"=IFERROR(IF(W{start_row + 13}+W{start_row + 35}=0,0,(W{start_row + 13}-W{start_row + 35})/W{start_row + 35}),1)",
            f"I{start_row + 28}": f"=IFERROR(IF(I{start_row + 36}+I{start_row + 14}=0,0,(I{start_row + 14}-I{start_row + 36})/I{start_row + 36}),1)",
            f"J{start_row + 28}": f"=IFERROR(IF(J{start_row + 36}+J{start_row + 14}=0,0,(J{start_row + 14}-J{start_row + 36})/J{start_row + 36}),1)",
            f"K{start_row + 28}": f"=IFERROR(IF(K{start_row + 36}+K{start_row + 14}=0,0,(K{start_row + 14}-K{start_row + 36})/K{start_row + 36}),1)",
            f"L{start_row + 28}": f"=IFERROR(IF(L{start_row + 36}+L{start_row + 14}=0,0,(L{start_row + 14}-L{start_row + 36})/L{start_row + 36}),1)",
            f"M{start_row + 28}": f"=IFERROR(IF(M{start_row + 36}+M{start_row + 14}=0,0,(M{start_row + 14}-M{start_row + 36})/M{start_row + 36}),1)",
            f"N{start_row + 28}": f"=IFERROR(IF(N{start_row + 36}+N{start_row + 14}=0,0,(N{start_row + 14}-N{start_row + 36})/N{start_row + 36}),1)",
            f"O{start_row + 28}": f"=IFERROR(IF(O{start_row + 36}+O{start_row + 14}=0,0,(O{start_row + 14}-O{start_row + 36})/O{start_row + 36}),1)",
            f"P{start_row + 28}": f"=IFERROR(IF(P{start_row + 36}+P{start_row + 14}=0,0,(P{start_row + 14}-P{start_row + 36})/P{start_row + 36}),1)",
            f"Q{start_row + 28}": f"=IFERROR(IF(Q{start_row + 36}+Q{start_row + 14}=0,0,(Q{start_row + 14}-Q{start_row + 36})/Q{start_row + 36}),1)",
            f"R{start_row + 28}": f"=IFERROR(IF(R{start_row + 36}+R{start_row + 14}=0,0,(R{start_row + 14}-R{start_row + 36})/R{start_row + 36}),1)",
            f"S{start_row + 28}": f"=IFERROR(IF(S{start_row + 36}+S{start_row + 14}=0,0,(S{start_row + 14}-S{start_row + 36})/S{start_row + 36}),1)",
            f"T{start_row + 28}": f"=IFERROR(IF(T{start_row + 36}+T{start_row + 14}=0,0,(T{start_row + 14}-T{start_row + 36})/T{start_row + 36}),1)",
            f"U{start_row + 28}": f"=IFERROR(IF(U{start_row + 36}+U{start_row + 14}=0,0,(U{start_row + 14}-U{start_row + 36})/U{start_row + 36}),1)",
            f"V{start_row + 28}": f"=IFERROR(IF(V{start_row + 36}+V{start_row + 14}=0,0,(V{start_row + 14}-V{start_row + 36})/V{start_row + 36}),1)",
            f"W{start_row + 28}": f"=IFERROR(IF(W{start_row + 36}+W{start_row + 14}=0,0,(W{start_row + 14}-W{start_row + 36})/W{start_row + 36}),1)",    
            f"I{start_row + 40}":LY_MCOM_OH_Units['FEB'] ,
            f"J{start_row + 40}":LY_MCOM_OH_Units['MAR'],
            f"K{start_row + 40}":LY_MCOM_OH_Units['APR'],
            f"L{start_row + 40}":LY_MCOM_OH_Units['MAY'],
            f"M{start_row + 40}":LY_MCOM_OH_Units['JUN'] ,
            f"N{start_row + 40}":LY_MCOM_OH_Units['JUL'],
            f"O{start_row + 40}":LY_MCOM_OH_Units['AUG'],
            f"P{start_row + 40}":LY_MCOM_OH_Units['SEP'] ,
            f"Q{start_row + 40}":LY_MCOM_OH_Units['OCT'] ,
            f"R{start_row + 40}":LY_MCOM_OH_Units['NOV'],
            f"S{start_row + 40}":LY_MCOM_OH_Units['DEC'],
            f"T{start_row + 40}":LY_MCOM_OH_Units['JAN'] ,

            f"U{start_row + 40}": f"=AVERAGEIFS(I{start_row + 40}:T{start_row + 40},$I$4:$T$4,\"<=\"&$T$4)",
            f"V{start_row + 40}": f"=AVERAGEIFS(I{start_row + 40}:N{start_row + 40},$I$4:$N$4,\"<=\"&$N$4)",
            f"W{start_row + 40}": f"=AVERAGEIFS(O{start_row + 40}:T{start_row + 40},$O$4:$T$4,\"<=\"&$T$4)",  
            f"I{start_row + 39}": f"=I{start_row + 38}-I{start_row + 40}",
            f"J{start_row + 39}": f"=J{start_row + 38}-J{start_row + 40}",
            f"K{start_row + 39}": f"=K{start_row + 38}-K{start_row + 40}",
            f"L{start_row + 39}": f"=L{start_row + 38}-L{start_row + 40}",
            f"M{start_row + 39}": f"=M{start_row + 38}-M{start_row + 40}",
            f"N{start_row + 39}": f"=N{start_row + 38}-N{start_row + 40}",
            f"O{start_row + 39}": f"=O{start_row + 38}-O{start_row + 40}",
            f"P{start_row + 39}": f"=P{start_row + 38}-P{start_row + 40}",
            f"Q{start_row + 39}": f"=Q{start_row + 38}-Q{start_row + 40}",
            f"R{start_row + 39}": f"=R{start_row + 38}-R{start_row + 40}",
            f"S{start_row + 39}": f"=S{start_row + 38}-S{start_row + 40}",
            f"T{start_row + 39}": f"=T{start_row + 38}-T{start_row + 40}",

            f"U{start_row + 39}": f"=AVERAGEIFS(I{start_row + 39}:T{start_row + 39},$I$4:$T$4,\"<=\"&$T$4)",
            f"V{start_row + 39}": f"=AVERAGEIFS(I{start_row + 39}:N{start_row + 39},$I$4:$N$4,\"<=\"&$N$4)",
            f"W{start_row + 39}": f"=AVERAGEIFS(O{start_row + 39}:T{start_row + 39},$O$4:$T$4,\"<=\"&$T$4)",
            f"I{start_row + 29}": f"=IFERROR(IF(I{start_row + 17}+I{start_row + 39}=0,0,(I{start_row + 17}-I{start_row + 39})/I{start_row + 39}),1)",
            f"J{start_row + 29}": f"=IFERROR(IF(J{start_row + 17}+J{start_row + 39}=0,0,(J{start_row + 17}-J{start_row + 39})/J{start_row + 39}),1)",
            f"K{start_row + 29}": f"=IFERROR(IF(K{start_row + 17}+K{start_row + 39}=0,0,(K{start_row + 17}-K{start_row + 39})/K{start_row + 39}),1)",
            f"L{start_row + 29}": f"=IFERROR(IF(L{start_row + 17}+L{start_row + 39}=0,0,(L{start_row + 17}-L{start_row + 39})/L{start_row + 39}),1)",
            f"M{start_row + 29}": f"=IFERROR(IF(M{start_row + 17}+M{start_row + 39}=0,0,(M{start_row + 17}-M{start_row + 39})/M{start_row + 39}),1)",
            f"N{start_row + 29}": f"=IFERROR(IF(N{start_row + 17}+N{start_row + 39}=0,0,(N{start_row + 17}-N{start_row + 39})/N{start_row + 39}),1)",
            f"O{start_row + 29}": f"=IFERROR(IF(O{start_row + 17}+O{start_row + 39}=0,0,(O{start_row + 17}-O{start_row + 39})/O{start_row + 39}),1)",
            f"P{start_row + 29}": f"=IFERROR(IF(P{start_row + 17}+P{start_row + 39}=0,0,(P{start_row + 17}-P{start_row + 39})/P{start_row + 39}),1)",
            f"Q{start_row + 29}": f"=IFERROR(IF(Q{start_row + 17}+Q{start_row + 39}=0,0,(Q{start_row + 17}-Q{start_row + 39})/Q{start_row + 39}),1)",
            f"R{start_row + 29}": f"=IFERROR(IF(R{start_row + 17}+R{start_row + 39}=0,0,(R{start_row + 17}-R{start_row + 39})/R{start_row + 39}),1)",
            f"S{start_row + 29}": f"=IFERROR(IF(S{start_row + 17}+S{start_row + 39}=0,0,(S{start_row + 17}-S{start_row + 39})/S{start_row + 39}),1)",
            f"T{start_row + 29}": f"=IFERROR(IF(T{start_row + 17}+T{start_row + 39}=0,0,(T{start_row + 17}-T{start_row + 39})/T{start_row + 39}),1)",
            f"U{start_row + 29}": f"=IFERROR(IF(U{start_row + 17}+U{start_row + 39}=0,0,(U{start_row + 17}-U{start_row + 39})/U{start_row + 39}),1)",
            f"V{start_row + 29}": f"=IFERROR(IF(V{start_row + 17}+V{start_row + 39}=0,0,(V{start_row + 17}-V{start_row + 39})/V{start_row + 39}),1)",
            f"W{start_row + 29}": f"=IFERROR(IF(W{start_row + 17}+W{start_row + 39}=0,0,(W{start_row + 17}-W{start_row + 39})/W{start_row + 39}),1)",
            f"I{start_row + 30}": OO_Total_Units['FEB'],
            f"J{start_row + 30}": OO_Total_Units['MAR'],
            f"K{start_row + 30}": OO_Total_Units['APR'],
            f"L{start_row + 30}": OO_Total_Units['MAY'],
            f"M{start_row + 30}":OO_Total_Units['JUN'],
            f"N{start_row + 30}":OO_Total_Units['JUL'],
            f"O{start_row + 30}": OO_Total_Units['AUG'],
            f"P{start_row + 30}": OO_Total_Units['SEP'],
            f"Q{start_row + 30}": OO_Total_Units['OCT'],
            f"R{start_row + 30}": OO_Total_Units['NOV'],
            f"S{start_row + 30}": OO_Total_Units['DEC'],
            f"T{start_row + 30}": OO_Total_Units['JAN'],

            f"U{start_row + 30}": f"=IFERROR(LOOKUP(2,1/($I${start_row}:$T${start_row}=$J$1),I{start_row + 30}:T{start_row + 30}),0)",
            f"V{start_row + 30}": f"=IFERROR(IFERROR(IFERROR(LOOKUP(2,1/($I${start_row}:$N${start_row}=$J$2),I{start_row + 30}:N{start_row + 30}),0),0),0)",
            f"W{start_row + 30}": f"=IFERROR(IFERROR(IFERROR(LOOKUP(2,1/($O${start_row}:$T${start_row}=$J$1),O{start_row + 30}:T{start_row + 30}),0),0),0)",
            f"I{start_row + 31}": OO_MCOM_Total_Units['FEB'],
            f"J{start_row + 31}": OO_MCOM_Total_Units['MAR'],
            f"K{start_row + 31}": OO_MCOM_Total_Units['APR'],
            f"L{start_row + 31}": OO_MCOM_Total_Units['MAY'],
            f"M{start_row + 31}": OO_MCOM_Total_Units['JUN'],
            f"N{start_row + 31}": OO_MCOM_Total_Units['JUL'],
            f"O{start_row + 31}": OO_MCOM_Total_Units['AUG'],
            f"P{start_row + 31}":OO_MCOM_Total_Units['SEP'],
            f"Q{start_row + 31}": OO_MCOM_Total_Units['OCT'],
            f"R{start_row + 31}": OO_MCOM_Total_Units['NOV'],
            f"S{start_row + 31}": OO_MCOM_Total_Units['DEC'],
            f"T{start_row + 31}":OO_MCOM_Total_Units['JAN'],

            f"U{start_row + 31}": f"=IFERROR(LOOKUP(2,1/($I${start_row}:$T${start_row}=$J$1),I{start_row + 31}:T{start_row + 31}),0)",
            f"V{start_row + 31}": f"=IFERROR(IFERROR(IFERROR(LOOKUP(2,1/($I${start_row}:$N${start_row}=$J$2),I{start_row + 31}:N{start_row + 31}),0),0),0)",
            f"W{start_row + 31}": f"=IFERROR(IFERROR(IFERROR(LOOKUP(2,1/($O${start_row}:$T${start_row}=$J$1),O{start_row + 31}:T{start_row + 31}),0),0),0)",
            f"I{start_row + 37}": f"=IFERROR(I{start_row + 36}/I{start_row + 34},0)",
            f"J{start_row + 37}": f"=IFERROR(J{start_row + 36}/J{start_row + 34},0)",
            f"K{start_row + 37}": f"=IFERROR(K{start_row + 36}/K{start_row + 34},0)",
            f"L{start_row + 37}": f"=IFERROR(L{start_row + 36}/L{start_row + 34},0)",
            f"M{start_row + 37}": f"=IFERROR(M{start_row + 36}/M{start_row + 34},0)",
            f"N{start_row + 37}": f"=IFERROR(N{start_row + 36}/N{start_row + 34},0)",
            f"O{start_row + 37}": f"=IFERROR(O{start_row + 36}/O{start_row + 34},0)",
            f"P{start_row + 37}": f"=IFERROR(P{start_row + 36}/P{start_row + 34},0)",
            f"Q{start_row + 37}": f"=IFERROR(Q{start_row + 36}/Q{start_row + 34},0)",
            f"R{start_row + 37}": f"=IFERROR(R{start_row + 36}/R{start_row + 34},0)",
            f"S{start_row + 37}": f"=IFERROR(S{start_row + 36}/S{start_row + 34},0)",
            f"T{start_row + 37}": f"=IFERROR(T{start_row + 36}/T{start_row + 34},0)",
            f"U{start_row + 37}": f"=IFERROR(U{start_row + 36}/U{start_row + 34},0)",
            f"V{start_row + 37}": f"=IFERROR(V{start_row + 36}/V{start_row + 34},0)",
            f"W{start_row + 37}": f"=IFERROR(W{start_row + 36}/W{start_row + 34},0)",
            f"I{start_row + 41}": f"=IFERROR(I{start_row + 40}/I{start_row + 38},0)",
            f"J{start_row + 41}": f"=IFERROR(J{start_row + 40}/J{start_row + 38},0)",
            f"K{start_row + 41}": f"=IFERROR(K{start_row + 40}/K{start_row + 38},0)",
            f"L{start_row + 41}": f"=IFERROR(L{start_row + 40}/L{start_row + 38},0)",
            f"M{start_row + 41}": f"=IFERROR(M{start_row + 40}/M{start_row + 38},0)",
            f"N{start_row + 41}": f"=IFERROR(N{start_row + 40}/N{start_row + 38},0)",
            f"O{start_row + 41}": f"=IFERROR(O{start_row + 40}/O{start_row + 38},0)",
            f"P{start_row + 41}": f"=IFERROR(P{start_row + 40}/P{start_row + 38},0)",
            f"Q{start_row + 41}": f"=IFERROR(Q{start_row + 40}/Q{start_row + 38},0)",
            f"R{start_row + 41}": f"=IFERROR(R{start_row + 40}/R{start_row + 38},0)",
            f"S{start_row + 41}": f"=IFERROR(S{start_row + 40}/S{start_row + 38},0)",
            f"T{start_row + 41}": f"=IFERROR(T{start_row + 40}/T{start_row + 38},0)",
            f"U{start_row + 41}": f"=IFERROR(U{start_row + 40}/U{start_row + 38},0)",
            f"V{start_row + 41}": f"=IFERROR(V{start_row + 40}/V{start_row + 38},0)",
            f"W{start_row + 41}": f"=IFERROR(W{start_row + 40}/W{start_row + 38},0)",
            f"I{start_row + 42}":LY_Receipts['FEB'] ,
            f"J{start_row + 42}":LY_Receipts['MAR'],
            f"K{start_row + 42}":LY_Receipts['APR'],
            f"L{start_row + 42}": LY_Receipts['MAY'],
            f"M{start_row + 42}":LY_Receipts['JUN'],
            f"N{start_row + 42}": LY_Receipts['JUL'],
            f"O{start_row + 42}": LY_Receipts['AUG'],
            f"P{start_row + 42}": LY_Receipts['SEP'],
            f"Q{start_row + 42}": LY_Receipts['OCT'],
            f"R{start_row + 42}":LY_Receipts['NOV'],
            f"S{start_row + 42}":LY_Receipts['DEC'],
            f"T{start_row + 42}": LY_Receipts['JAN'],

            f"U{start_row + 42}": f"=SUM(I{start_row + 42}:T{start_row + 42})",
            f"V{start_row + 42}": f"=SUM(I{start_row + 42}:N{start_row + 42})",
            f"W{start_row + 42}": f"=SUM(O{start_row + 42}:T{start_row + 42})",
            f"I{start_row + 43}": f"=IFERROR(IFERROR(I{start_row + 34}/(I{start_row + 34}+I{start_row + 38}),0),0)",
            f"J{start_row + 43}": f"=IFERROR(IFERROR(J{start_row + 34}/(J{start_row + 34}+J{start_row + 38}),0),0)",
            f"K{start_row + 43}": f"=IFERROR(IFERROR(K{start_row + 34}/(K{start_row + 34}+K{start_row + 38}),0),0)",
            f"L{start_row + 43}": f"=IFERROR(IFERROR(L{start_row + 34}/(L{start_row + 34}+L{start_row + 38}),0),0)",
            f"M{start_row + 43}": f"=IFERROR(IFERROR(M{start_row + 34}/(M{start_row + 34}+M{start_row + 38}),0),0)",
            f"N{start_row + 43}": f"=IFERROR(IFERROR(N{start_row + 34}/(N{start_row + 34}+N{start_row + 38}),0),0)",
            f"O{start_row + 43}": f"=IFERROR(IFERROR(O{start_row + 34}/(O{start_row + 34}+O{start_row + 38}),0),0)",
            f"P{start_row + 43}": f"=IFERROR(IFERROR(P{start_row + 34}/(P{start_row + 34}+P{start_row + 38}),0),0)",
            f"Q{start_row + 43}": f"=IFERROR(IFERROR(Q{start_row + 34}/(Q{start_row + 34}+Q{start_row + 38}),0),0)",
            f"R{start_row + 43}": f"=IFERROR(IFERROR(R{start_row + 34}/(R{start_row + 34}+R{start_row + 38}),0),0)",
            f"S{start_row + 43}": f"=IFERROR(IFERROR(S{start_row + 34}/(S{start_row + 34}+S{start_row + 38}),0),0)",
            f"T{start_row + 43}": f"=IFERROR(IFERROR(T{start_row + 34}/(T{start_row + 34}+T{start_row + 38}),0),0)",
            f"U{start_row + 43}": f"=IFERROR(IFERROR(U{start_row + 34}/(U{start_row + 34}+U{start_row + 38}),0),0)",
            f"V{start_row + 43}": f"=IFERROR(IFERROR(V{start_row + 34}/(V{start_row + 34}+V{start_row + 38}),0),0)",
            f"W{start_row + 43}": f"=IFERROR(IFERROR(W{start_row + 34}/(W{start_row + 34}+W{start_row + 38}),0),0)",
            f"I{start_row + 44}": f"=IFERROR((I{start_row + 34}-I{start_row + 36})/((I{start_row + 34}-I{start_row + 36})+(I{start_row + 38}-I{start_row + 40})),0)",
            f"J{start_row + 44}": f"=IFERROR((J{start_row + 34}-J{start_row + 36})/((J{start_row + 34}-J{start_row + 36})+(J{start_row + 38}-J{start_row + 40})),0)",
            f"K{start_row + 44}": f"=IFERROR((K{start_row + 34}-K{start_row + 36})/((K{start_row + 34}-K{start_row + 36})+(K{start_row + 38}-K{start_row + 40})),0)",
            f"L{start_row + 44}": f"=IFERROR((L{start_row + 34}-L{start_row + 36})/((L{start_row + 34}-L{start_row + 36})+(L{start_row + 38}-L{start_row + 40})),0)",
            f"M{start_row + 44}": f"=IFERROR((M{start_row + 34}-M{start_row + 36})/((M{start_row + 34}-M{start_row + 36})+(M{start_row + 38}-M{start_row + 40})),0)",
            f"N{start_row + 44}": f"=IFERROR((N{start_row + 34}-N{start_row + 36})/((N{start_row + 34}-N{start_row + 36})+(N{start_row + 38}-N{start_row + 40})),0)",
            f"O{start_row + 44}": f"=IFERROR((O{start_row + 34}-O{start_row + 36})/((O{start_row + 34}-O{start_row + 36})+(O{start_row + 38}-O{start_row + 40})),0)",
            f"P{start_row + 44}": f"=IFERROR((P{start_row + 34}-P{start_row + 36})/((P{start_row + 34}-P{start_row + 36})+(P{start_row + 38}-P{start_row + 40})),0)",
            f"Q{start_row + 44}": f"=IFERROR((Q{start_row + 34}-Q{start_row + 36})/((Q{start_row + 34}-Q{start_row + 36})+(Q{start_row + 38}-Q{start_row + 40})),0)",
            f"R{start_row + 44}": f"=IFERROR((R{start_row + 34}-R{start_row + 36})/((R{start_row + 34}-R{start_row + 36})+(R{start_row + 38}-R{start_row + 40})),0)",
            f"S{start_row + 44}": f"=IFERROR((S{start_row + 34}-S{start_row + 36})/((S{start_row + 34}-S{start_row + 36})+(S{start_row + 38}-S{start_row + 40})),0)",
            f"T{start_row + 44}": f"=IFERROR((T{start_row + 34}-T{start_row + 36})/((T{start_row + 34}-T{start_row + 36})+(T{start_row + 38}-T{start_row + 40})),0)",
            f"U{start_row + 44}": f"=IFERROR((U{start_row + 34}-U{start_row + 36})/((U{start_row + 34}-U{start_row + 36})+(U{start_row + 38}-U{start_row + 40})),0)",
            f"V{start_row + 44}": f"=IFERROR((V{start_row + 34}-V{start_row + 36})/((V{start_row + 34}-V{start_row + 36})+(V{start_row + 38}-V{start_row + 40})),0)",
            f"W{start_row + 44}": f"=IFERROR((W{start_row + 34}-W{start_row + 36})/((W{start_row + 34}-W{start_row + 36})+(W{start_row + 38}-W{start_row + 40})),0)",
            f"I{start_row + 45}": f"=IFERROR(IFERROR(I{start_row + 34}/I{start_row + 38},0),0)",
            f"J{start_row + 45}": f"=IFERROR(IFERROR(SUM(I{start_row + 34}:J{start_row + 34})/AVERAGE(I{start_row + 38}:J{start_row + 38}),0),0)",
            f"K{start_row + 45}": f"=IFERROR(IFERROR(SUM(I{start_row + 34}:K{start_row + 34})/AVERAGE(I{start_row + 38}:K{start_row + 38}),0),0)",
            f"L{start_row + 45}": f"=IFERROR(IFERROR(SUM(I{start_row + 34}:L{start_row + 34})/AVERAGE(I{start_row + 38}:L{start_row + 38}),0),0)",
            f"M{start_row + 45}": f"=IFERROR(IFERROR(SUM(I{start_row + 34}:M{start_row + 34})/AVERAGE(I{start_row + 38}:M{start_row + 38}),0),0)",
            f"N{start_row + 45}": f"=IFERROR(IFERROR(SUM(I{start_row + 34}:N{start_row + 34})/AVERAGE(I{start_row + 38}:N{start_row + 38}),0),0)",
            f"O{start_row + 45}": f"=IFERROR(IFERROR(SUM(I{start_row + 34}:O{start_row + 34})/AVERAGE(I{start_row + 38}:O{start_row + 38}),0),0)",
            f"P{start_row + 45}": f"=IFERROR(IFERROR(SUM(I{start_row + 34}:P{start_row + 34})/AVERAGE(I{start_row + 38}:P{start_row + 38}),0),0)",
            f"Q{start_row + 45}": f"=IFERROR(IFERROR(SUM(I{start_row + 34}:Q{start_row + 34})/AVERAGE(I{start_row + 38}:Q{start_row + 38}),0),0)",
            f"R{start_row + 45}": f"=IFERROR(IFERROR(SUM(I{start_row + 34}:R{start_row + 34})/AVERAGE(I{start_row + 38}:R{start_row + 38}),0),0)",
            f"S{start_row + 45}": f"=IFERROR(IFERROR(SUM(I{start_row + 34}:S{start_row + 34})/AVERAGE(I{start_row + 38}:S{start_row + 38}),0),0)",
            f"T{start_row + 45}": f"=IFERROR(IFERROR(SUM(I{start_row + 34}:T{start_row + 34})/AVERAGE(I{start_row + 38}:T{start_row + 38}),0),0)",
            f"U{start_row + 45}": f"=IFERROR(U{start_row + 34}/U{start_row + 38},0)",
            f"V{start_row + 45}": f"=IFERROR(V{start_row + 34}/V{start_row + 38},0)",
            f"W{start_row + 45}": f"=IFERROR(W{start_row + 34}/W{start_row + 38},0)",
            f"I{start_row + 46}": f"=IFERROR(IFERROR(I{start_row + 35}/I{start_row + 39},0),0)",
            f"J{start_row + 46}": f"=IFERROR(IFERROR(SUM(I{start_row + 35}:J{start_row + 35})/AVERAGE(I{start_row + 39}:J{start_row + 39}),0),0)",
            f"K{start_row + 46}": f"=IFERROR(IFERROR(SUM(I{start_row + 35}:K{start_row + 35})/AVERAGE(I{start_row + 39}:K{start_row + 39}),0),0)",
            f"L{start_row + 46}": f"=IFERROR(IFERROR(SUM(I{start_row + 35}:L{start_row + 35})/AVERAGE(I{start_row + 39}:L{start_row + 39}),0),0)",
            f"M{start_row + 46}": f"=IFERROR(IFERROR(SUM(I{start_row + 35}:M{start_row + 35})/AVERAGE(I{start_row + 39}:M{start_row + 39}),0),0)",
            f"N{start_row + 46}": f"=IFERROR(IFERROR(SUM(I{start_row + 35}:N{start_row + 35})/AVERAGE(I{start_row + 39}:N{start_row + 39}),0),0)",
            f"O{start_row + 46}": f"=IFERROR(IFERROR(SUM(I{start_row + 35}:O{start_row + 35})/AVERAGE(I{start_row + 39}:O{start_row + 39}),0),0)",
            f"P{start_row + 46}": f"=IFERROR(IFERROR(SUM(I{start_row + 35}:P{start_row + 35})/AVERAGE(I{start_row + 39}:P{start_row + 39}),0),0)",
            f"Q{start_row + 46}": f"=IFERROR(IFERROR(SUM(I{start_row + 35}:Q{start_row + 35})/AVERAGE(I{start_row + 39}:Q{start_row + 39}),0),0)",
            f"R{start_row + 46}": f"=IFERROR(IFERROR(SUM(I{start_row + 35}:R{start_row + 35})/AVERAGE(I{start_row + 39}:R{start_row + 39}),0),0)",
            f"S{start_row + 46}": f"=IFERROR(IFERROR(SUM(I{start_row + 35}:S{start_row + 35})/AVERAGE(I{start_row + 39}:S{start_row + 39}),0),0)",
            f"T{start_row + 46}": f"=IFERROR(IFERROR(SUM(I{start_row + 35}:T{start_row + 35})/AVERAGE(I{start_row + 39}:T{start_row + 39}),0),0)",
            f"U{start_row + 46}": f"=IFERROR(U{start_row + 35}/U{start_row + 39},0)",
            f"V{start_row + 46}": f"=IFERROR(V{start_row + 35}/V{start_row + 39},0)",
            f"W{start_row + 46}": f"=IFERROR(W{start_row + 35}/W{start_row + 39},0)",
            f"I{start_row + 47}":LY_PTD_Sales['FEB']  ,
            f"J{start_row + 47}": LY_PTD_Sales['MAR'] ,
            f"K{start_row + 47}": LY_PTD_Sales['APR'] ,
            f"L{start_row + 47}":LY_PTD_Sales['MAY'] ,
            f"M{start_row + 47}":LY_PTD_Sales['JUN'] ,
            f"N{start_row + 47}":LY_PTD_Sales['JUL'] ,
            f"O{start_row + 47}": LY_PTD_Sales['AUG'] ,
            f"P{start_row + 47}": LY_PTD_Sales['SEP'] ,
            f"Q{start_row + 47}":LY_PTD_Sales['OCT'] ,
            f"R{start_row + 47}": LY_PTD_Sales['NOV'] ,
            f"S{start_row + 47}": LY_PTD_Sales['DEC'] ,
            f"T{start_row + 47}": LY_PTD_Sales['JAN'] ,

            f"U{start_row + 47}": f"=SUM(I{start_row + 47}:T{start_row + 47})",
            f"V{start_row + 47}": f"=SUM(I{start_row + 47}:N{start_row + 47})",
            f"W{start_row + 47}": f"=SUM(O{start_row + 47}:T{start_row + 47})",
            f"I{start_row + 48}":MCOM_PTD_LY_Sales['FEB'],
            f"J{start_row + 48}":MCOM_PTD_LY_Sales['MAR'],
            f"K{start_row + 48}": MCOM_PTD_LY_Sales['APR'],
            f"L{start_row + 48}":MCOM_PTD_LY_Sales['MAY'],
            f"M{start_row + 48}": MCOM_PTD_LY_Sales['JUN'],
            f"N{start_row + 48}": MCOM_PTD_LY_Sales['JUL'],
            f"O{start_row + 48}":MCOM_PTD_LY_Sales['AUG'],
            f"P{start_row + 48}": MCOM_PTD_LY_Sales['SEP'],
            f"Q{start_row + 48}":MCOM_PTD_LY_Sales['OCT'],
            f"R{start_row + 48}":MCOM_PTD_LY_Sales['NOV'],
            f"S{start_row + 48}":MCOM_PTD_LY_Sales['DEC'],
            f"T{start_row + 48}":MCOM_PTD_LY_Sales['JAN'],

            f"U{start_row + 48}": f"=SUM(I{start_row + 48}:T{start_row + 48})",
            f"V{start_row + 48}": f"=SUM(I{start_row + 48}:N{start_row + 48})",
            f"W{start_row + 48}": f"=SUM(O{start_row + 48}:T{start_row + 48})",
            f"I{start_row + 49}": f"=IFERROR(TEXT(IFERROR(I{start_row + 47}/I{start_row + 34},0),\"$0\") &\"/ \"&TEXT(((I{start_row + 47}/I{start_row + 34})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"J{start_row + 49}": f"=IFERROR(TEXT(IFERROR(J{start_row + 47}/J{start_row + 34},0),\"$0\") &\"/ \"&TEXT(((J{start_row + 47}/J{start_row + 34})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"K{start_row + 49}": f"=IFERROR(TEXT(IFERROR(K{start_row + 47}/K{start_row + 34},0),\"$0\") &\"/ \"&TEXT(((K{start_row + 47}/K{start_row + 34})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"L{start_row + 49}": f"=IFERROR(TEXT(IFERROR(L{start_row + 47}/L{start_row + 34},0),\"$0\") &\"/ \"&TEXT(((L{start_row + 47}/L{start_row + 34})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"M{start_row + 49}": f"=IFERROR(TEXT(IFERROR(M{start_row + 47}/M{start_row + 34},0),\"$0\") &\"/ \"&TEXT(((M{start_row + 47}/M{start_row + 34})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"N{start_row + 49}": f"=IFERROR(TEXT(IFERROR(N{start_row + 47}/N{start_row + 34},0),\"$0\") &\"/ \"&TEXT(((N{start_row + 47}/N{start_row + 34})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"O{start_row + 49}": f"=IFERROR(TEXT(IFERROR(O{start_row + 47}/O{start_row + 34},0),\"$0\") &\"/ \"&TEXT(((O{start_row + 47}/O{start_row + 34})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"P{start_row + 49}": f"=IFERROR(TEXT(IFERROR(P{start_row + 47}/P{start_row + 34},0),\"$0\") &\"/ \"&TEXT(((P{start_row + 47}/P{start_row + 34})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"Q{start_row + 49}": f"=IFERROR(TEXT(IFERROR(Q{start_row + 47}/Q{start_row + 34},0),\"$0\") &\"/ \"&TEXT(((Q{start_row + 47}/Q{start_row + 34})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"R{start_row + 49}": f"=IFERROR(TEXT(IFERROR(R{start_row + 47}/R{start_row + 34},0),\"$0\") &\"/ \"&TEXT(((R{start_row + 47}/R{start_row + 34})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"S{start_row + 49}": f"=IFERROR(TEXT(IFERROR(S{start_row + 47}/S{start_row + 34},0),\"$0\") &\"/ \"&TEXT(((S{start_row + 47}/S{start_row + 34})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"T{start_row + 49}": f"=IFERROR(TEXT(IFERROR(T{start_row + 47}/T{start_row + 34},0),\"$0\") &\"/ \"&TEXT(((T{start_row + 47}/T{start_row + 34})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"U{start_row + 49}": f"=IFERROR(TEXT(IFERROR(U{start_row + 47}/U{start_row + 34},0),\"$0\") &\"/ \"&TEXT(((U{start_row + 47}/U{start_row + 34})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"V{start_row + 49}": f"=IFERROR(TEXT(IFERROR(V{start_row + 47}/V{start_row + 34},0),\"$0\") &\"/ \"&TEXT(((V{start_row + 47}/V{start_row + 34})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"W{start_row + 49}": f"=IFERROR(TEXT(IFERROR(W{start_row + 47}/W{start_row + 34},0),\"$0\") &\"/ \"&TEXT(((W{start_row + 47}/W{start_row + 34})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"G{start_row + 3}": f"=IFERROR(U{start_row + 34}+U{start_row + 34}*F{start_row + 3},0)",  # Dynamic references from row 5 onward
            f"C{start_row + 7}": f"=IF(G1=\"SP\",V{start_row + 9},W{start_row + 9})",
            f"D{start_row + 7}": f"=IF($G$1=\"SP\",C{start_row + 7}-V{start_row + 32},C{start_row + 7}-W{start_row + 32})",
            f"E{start_row + 7}": f"=IF($G$1=\"SP\",C{start_row + 7}-V{start_row + 32},C{start_row + 7}-W{start_row + 32})",
            f"E{start_row + 9}": f"=SUMIFS(I{start_row + 5}:T{start_row + 5},$I$4:$T$4,\">=\"&$F$4,$I$4:$T$4,\"<=\"&$G$4)/SUMIFS($I$3:$T$3,$I$4:$T$4,\">=\"&$F$4,$I$4:$T$4,\"<=\"&$G$4)",  # Static references in rows 1 to 4
            f"C{start_row + 10}": f"=IFERROR(C{start_row + 15}/E{start_row + 9},0)",
            f"D{start_row + 10}": f"=IFERROR(C{start_row + 14}/E{start_row + 9},0)",
            f"C{start_row + 11}": f"=IFERROR((D{start_row + 10}-D{start_row + 9})*E{start_row + 9},0)",  # Dynamic D and E references from row 5 onward
            f"D{start_row + 11}": f"=C{start_row + 11}*C{start_row + 19}",  # C24 kept static because it was a specific reference
            f"F{start_row + 11}": f"=IFERROR(IF(U{start_row + 6}>C{start_row + 14},(U{start_row + 6}-C{start_row + 14}+C{start_row + 14}-F{start_row + 10}-C{start_row + 12})/E{start_row + 9},(C{start_row + 14}-F{start_row + 10}-C{start_row + 12})/E{start_row + 9}),0)",
            f"G{start_row + 11}": f"=F{start_row + 11}*E{start_row + 9}",
            f"G{start_row + 19}": f"=((U{start_row + 20}/U{start_row + 12})-C{start_row + 19})/(U{start_row + 20}/U{start_row + 12})",
            f"E{start_row + 33}": f"=U{start_row + 6}-U{start_row + 8}-E{start_row + 8}",
            f"A{start_row + 1}": f"=$C{start_row}&H{start_row + 1}",
            f"A{start_row + 2}": f"=$C{start_row}&H{start_row + 2}",
            f"A{start_row + 3}": f"=$C{start_row}&H{start_row + 3}",
            f"A{start_row + 4}": f"=$C{start_row}&H{start_row + 4}",
            f"A{start_row + 5}": f"=$C{start_row}&H{start_row + 5}",
            f"A{start_row + 6}": f"=$C{start_row}&H{start_row + 6}",
            f"A{start_row + 7}": f"=$C{start_row}&H{start_row + 7}",
            f"A{start_row + 8}": f"=$C{start_row}&H{start_row + 8}",
            f"A{start_row + 9}": f"=$C{start_row}&H{start_row + 9}",
            f"A{start_row + 10}": f"=$C{start_row}&H{start_row +10}",
            f"A{start_row + 11}": f"=$C{start_row + 11}&\"Excess Proj\"",
            f"A{start_row + 12}": f"=$C{start_row + 12}&\"Qty to Release\"",
            f"A{start_row + 13}": f"=$C{start_row}&H{start_row +12}",
            f"A{start_row + 14}": f"=$C{start_row}&H{start_row +13}",
            f"A{start_row + 15}": f"=$C{start_row}&H{start_row +14}",
            f"A{start_row + 16}": f"=$C{start_row}&H{start_row +16}",
            f"A{start_row + 17}": f"=$C{start_row}&H{start_row +17}",
            f"A{start_row + 18}": f"=$C{start_row}&H{start_row +18}",

            f"A{start_row + 20}": f"=$C{start_row}&H{start_row +20}",
            f"A{start_row + 21}": f"=$C{start_row}&H{start_row +21}",
            f"A{start_row + 22}": f"=$C{start_row}&H{start_row +22}",
            f"A{start_row + 23}": f"=$C{start_row}&H{start_row +23}",
            f"A{start_row + 24}": f"=$C{start_row}&H{start_row +24}",
            f"A{start_row + 25}": f"=$C{start_row}&H{start_row +25}",
            f"A{start_row + 26}": f"=$C{start_row}&H{start_row +26}",
            f"A{start_row + 27}": f"=$C{start_row}&H{start_row +27}",
            f"A{start_row + 28}": f"=$C{start_row}&H{start_row +28}",
            f"A{start_row + 29}": f"=$C{start_row}&H{start_row +29}",
            f"A{start_row + 33}": f"=$C{start_row + 33}&\"Qty to Enter\"",
            f"A{start_row + 35}": f"=$C{start_row}&B{start_row +34}",
            f"A{start_row + 40}": f"=$C{start_row}&B{start_row +39}",
            f"I{start_row + 5}":planned_fc["FEB"],
            f"J{start_row + 5}":planned_fc["MAR"],
            f"K{start_row + 5}":planned_fc["APR"],
            f"L{start_row + 5}":planned_fc["MAY"],
            f"M{start_row + 5}":planned_fc["JUN"],
            f"N{start_row + 5}":planned_fc["JUL"],
            f"O{start_row + 5}":planned_fc["AUG"],
            f"P{start_row + 5}":planned_fc["SEP"],
            f"Q{start_row + 5}":planned_fc["OCT"],
            f"R{start_row + 5}":planned_fc["NOV"],
            f"S{start_row + 5}":planned_fc["DEC"],
            f"T{start_row + 5}":planned_fc["JAN"],

                }
                # Add dropdown validation to the specific cell (e.g., F column) for each loop

        # Apply formulas
        percentage_ranges = [
        f"I{start_row + 1}:W{start_row + 1}",
        f"I{start_row + 10}:W{start_row + 10}",
        f"I{start_row + 15}:W{start_row + 15}",
        f"I{start_row + 19}:W{start_row + 19}",
        f"I{start_row + 23}:W{start_row + 23}",
        f"I{start_row + 24}:W{start_row + 24}",
        f"I{start_row + 27}:W{start_row + 27}",
        f"I{start_row +28}:W{start_row + 28}",
        f"I{start_row +29}:W{start_row + 29}",
        f"I{start_row + 37}:W{start_row + 37}",
        f"I{start_row + 41}:W{start_row + 41}",
        f"I{start_row + 43}:W{start_row + 43}",
        f"I{start_row + 44}:W{start_row + 44}",
        f"E{start_row + 2}",
        f"C{start_row + 3}",
        f"F{start_row + 19}",
        f"G{start_row + 19}",
        f"F{start_row + 3}",
    ]
        rounded_ranges = [
            f"F{start_row + 2}",
            f"G{start_row + 3}",
            f"E{start_row + 9}",
            f"C{start_row + 10}",
            f"D{start_row + 10}",
            f"C{start_row + 11}",
            f"F{start_row + 11}",
            f"F{start_row + 18}",
            f"U{start_row + 7}:W{start_row + 7}",
            f"U{start_row + 16}:W{start_row + 16}",
            f"U{start_row + 17}:W{start_row + 17}",
            f"U{start_row + 18}:W{start_row + 18}",
            f"I{start_row + 20}:W{start_row + 20}",
            f"I{start_row + 21}:W{start_row + 21}",
            f"U{start_row + 38}:W{start_row + 38}",
            f"U{start_row + 39}:W{start_row + 39}",
            f"U{start_row + 40}:W{start_row + 40}",


        ]
        rounded_ranges_one = [

            f"I{start_row + 25}:T{start_row + 25}",
            f"I{start_row + 26}:T{start_row + 26}",
            f"I{start_row + 45}:T{start_row + 45}",
            f"I{start_row + 46}:T{start_row + 46}",

        ]
        rounded_ranges_two = [
            f"U{start_row + 25}:W{start_row + 25}",
            f"U{start_row + 26}:W{start_row + 26}",
            f"U{start_row + 45}:W{start_row + 45}",
            f"U{start_row + 46}:W{start_row + 46}",




        ]
        currency_ranges = [
        f"D{start_row + 11}",
        f"C{start_row + 19}",
        f"D{start_row + 19}",
        f"C{start_row + 20}",
        f"D{start_row + 12}",
        f"E{start_row + 19}",
        f"I{start_row + 20}:W{start_row + 20}",
        f"I{start_row + 21}:W{start_row + 21}",
        f"I{start_row + 47}:W{start_row + 47}",
        f"I{start_row + 48}:W{start_row + 48}",
        
        # Add more ranges as needed
    ]
        sheet = ws
                # Add values to column B from ALL_VALUES and apply alignment
        for i, value in enumerate(ALL_VALUES, start=start_row):
            cell = ws.cell(row=i, column=2, value=value)  # Column B is the 2nd column
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Add values to column H from H_VALUES and apply alignment
        for i, value in enumerate(H_VALUES, start=start_row):
            cell = ws.cell(row=i, column=8, value=value)  # Column H is the 8th column
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Add monthly values starting from column I (9th column) in the specified row
        for col, value in enumerate(MONTHLY_VALUES, start=9):
            ws.cell(row=start_row, column=col, value=value)
        for cell, formula in dynamic_formulas.items():
            try:
                ws[cell] = formula  # Set the formula directly in the specified cell
            except Exception as e:
                print(f"Error setting formula in {cell}: {e}")

        # Apply ROUND formatting with different decimal places
        apply_round_format(ws, rounded_ranges, decimal_places=0)
        apply_round_format(ws, rounded_ranges_one, decimal_places=1)
        apply_round_format(ws, rounded_ranges_two, decimal_places=2)
    # Apply formats
        apply_format(ws, percentage_ranges, "0%")        # Apply percentage format
        apply_format(ws, currency_ranges, "$#,##0") 
        # add_dropdown(ws, f"F{start_row + 1}", category_options)
        add_dropdown(ws, f"F{start_row + 4}", forecast_method_options)




    #######loop formatinf
        red_font= Font(color="FF0000") 
        blue_font=Font(color="0000FF") 
        white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        dark_pink = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        gray = PatternFill(start_color="EDEDED", end_color="EDEDED", fill_type="solid")
        pink_font = Font(color="9C0006") 
        for row in ws[f"C{start_row}:G{start_row + 8}"]:
            for cell in row:
                cell.fill = gray

        for row in ws[f"C{start_row + 13}:G{start_row + 20}"]:  # Use the dynamically constructed range
            for cell in row:
                cell.fill = gray
        for row in ws[f"C{start_row + 27}:G{start_row + 32}"]:  # Offset the range dynamically
            for cell in row:
                cell.fill = gray  # Apply the gray fill 

        ws[f"D{start_row + 1}"].fill = white  # D6 becomes D(start_row + 1)
        ws[f"E{start_row + 3}"].fill = white  # E8 becomes E(start_row + 3)

        red_font_list = [f"B{start_row + 40}", f"B{start_row}", f"D{start_row + 16}", f"F{start_row + 7}"]
        for cell_address in red_font_list:
            ws[cell_address].font = red_font

        blue_font_list = [
            f"B{start_row + 9}",  # B14 -> start_row + 9
            f"B{start_row + 10}",  # B15 -> start_row + 10
            f"B{start_row + 33}",  # B38 -> start_row + 33
            f"D{start_row + 9}",  # D14 -> start_row + 9
            f"E{start_row + 9}",  # E14 -> start_row + 9
            f"D{start_row + 35}",  # D40 -> start_row + 35
            f"C{start_row + 10}",  # C15 -> start_row + 10
            f"D{start_row + 10}"   # D15 -> start_row + 10
        ]

        pink_fill_font = [
            f"C{start_row + 3}",  # C8 -> start_row + 3
            f"F{start_row + 3}",  # F8 -> start_row + 3
            f"F{start_row + 14}",  # F19 -> start_row + 14
            f"G{start_row + 14}",  # G19 -> start_row + 14
            f"C{start_row + 16}",  # C21 -> start_row + 16
            f"C{start_row + 24}"   # C29 -> start_row + 24
        ]
        for cell_address in pink_fill_font:
            ws[cell_address].fill = dark_pink
            ws[cell_address].font = pink_font


        # For the range H28:W31
        for row in ws[f"H{start_row + 23}:W{start_row + 26}"]:  # H28:W31 -> start_row + 23 to start_row + 26
            for cell in row:
                cell.font = blue_font

        # For the range H48:W51
        for row in ws[f"H{start_row + 43}:W{start_row + 46}"]:  # H48:W51 -> start_row + 43 to start_row + 46
            for cell in row:
                cell.font = blue_font

        # For the range I14:W14
        for row in ws[f"I{start_row + 9}:W{start_row + 9}"]:  # I14:W14 -> start_row + 9
            for cell in row:
                cell.font = red_font  
        yelow_fill_font_list=['E13','F13','Q13','R13','S13','G39',]
        for i in yelow_fill_font_list:
            ws[i].fill = light_yellow
            ws[i].font = yellow_font
        #Hyperlink
        ws[f"B{start_row + 26}"].font = Font(color="0563C1", underline="single", bold=False)  # B31 -> start_row + 26

        # Dynamically merge cells
        ws.merge_cells(f"D{start_row + 21}:G{start_row + 26}")  # D26:G31
        ws.merge_cells(f"D{start_row + 11}:E{start_row + 11}")  # D16:E16
        ws.merge_cells(f"D{start_row + 1}:E{start_row + 1}")    # D6:E6
        ws.merge_cells(f"D{start_row}:E{start_row}")            # D5:E5
        ws.merge_cells(f"D{start_row + 20}:F{start_row + 20}")  # D25:F25
        ws.merge_cells(f"B{start_row + 35}:C{start_row + 38}")  # B40:C43
        ws.merge_cells(f"D{start_row + 35}:G{start_row + 38}")  # D40:G43
        ws.merge_cells(f"C{start_row + 39}:G{start_row + 39}")  # C44:G44
        ws.merge_cells(f"B{start_row + 40}:G{start_row + 49}")  # B45:G54
        ws.merge_cells(f"C{start_row + 13}:F{start_row + 13}")  # C18:F18
        ws.merge_cells(f"D{start_row + 16}:G{start_row + 16}")  # D21:G21
        ws.merge_cells(f"C{start_row + 30}:G{start_row + 30}")  # C35:G35
        # Dynamically set alignment
        ws[f"B{start_row + 40}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)  # B45
        ws[f"D{start_row + 21}"].alignment = Alignment(horizontal='center', vertical='center')  # D26
        ws[f"D{start_row + 11}"].alignment = Alignment(horizontal='center', vertical='center')  # D16
        ws[f"D{start_row}"].alignment = Alignment(horizontal='center', vertical='center')       # D5
        ws[f"D{start_row + 1}"].alignment = Alignment(horizontal='right', vertical='center')    # D6
        ws[f"C{start_row + 3}"].alignment = Alignment(horizontal='center', vertical='center')   # C8
        ws[f"D{start_row + 16}"].alignment = Alignment(horizontal='left', vertical='center')    # D21
        ws[f"D{start_row + 20}"].alignment = Alignment(horizontal='center', vertical='center')  # D25
        ws[f"C{start_row + 30}"].alignment = Alignment(horizontal='center', vertical='top')     # C35
        ws[f"D{start_row + 35}"].alignment = Alignment(horizontal='left', vertical='top')       # D40

        #gradient bg
        two_yellow_gradient_fill = GradientFill(type="linear",degree=90,stop=("FFFFFF", "FFFF99", "FFFFFF")) 
        top_green_gradient_fill = GradientFill(type="linear", degree=90, stop=("E2EFDA", "FFFFFF"))
        full_green_gradient_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        bott_green_gradient_fill = GradientFill(type="linear", degree=90, stop=("FFFFFF","E2EFDA"))
        bott_light_gray_gradient_fill = GradientFill(type="linear", degree=90, stop=( "FFFFFF","EDEDED"))
        top_gray_gradient_fill = GradientFill(type="linear", degree=90, stop=("D9D9D9", "FFFFFF"))
        dark_gray_gradient_fill = GradientFill(type="linear", degree=90, stop=( "FFFFFF","DBDBDB"))
        top_orange_gradient_fill = GradientFill(type="linear", degree=90, stop=("FFE699", "FFFFFF"))
        bott_orange_gradient_fill = GradientFill(type="linear", degree=90, stop=( "FFFFFF","FFE699"))
        top_yellow_gradient_fill = GradientFill(type="linear", degree=90, stop=("FFFFCC", "FFFFFF"))
        full_yellow_gradient_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        bott_yellow_gradient_fill = GradientFill(type="linear", degree=90, stop=("FFFFFF","FFFFCC"))
        bott_blue_gradient_fill = GradientFill(type="linear", degree=90, stop=("FFFFFF","DFECF7"))
        full_blue_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        top_pink_gradient_fill = GradientFill(type="linear", degree=90, stop=("FCE4D6", "FFFFFF"))
        full_pink_gradient_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        bott_pink_gradient_fill = GradientFill(type="linear", degree=90, stop=("FFFFFF","FCE4D6"))
        # Pink
        for row in ws[f"C{start_row + 9}:G{start_row + 9}"]:  # C14:G14 -> start_row + 9
            for cell in row:
                cell.fill = bott_pink_gradient_fill  

        for row in ws[f"C{start_row + 10}:G{start_row + 11}"]:  # C15:G16 -> start_row + 10 to start_row + 11
            for cell in row:
                cell.fill = full_pink_gradient_fill  

        for row in ws[f"C{start_row + 12}:G{start_row + 12}"]:  # C17:G17 -> start_row + 12
            for cell in row:
                cell.fill = top_pink_gradient_fill 

        # Yellow
        for row in ws[f"I{start_row + 4}:W{start_row + 4}"]:  # I9:W9 -> start_row + 4
            for cell in row:
                cell.fill = two_yellow_gradient_fill  

        # Green
        for row in ws[f"I{start_row + 5}:W{start_row + 5}"]:  # I10:W10 -> start_row + 5
            for cell in row:
                cell.fill = bott_green_gradient_fill  

        for row in ws[f"I{start_row + 6}:W{start_row + 6}"]:  # I11:W11 -> start_row + 6
            for cell in row:
                cell.fill = full_green_gradient_fill  

        for row in ws[f"I{start_row + 7}:W{start_row + 7}"]:  # I12:W12 -> start_row + 7
            for cell in row:
                cell.fill = top_green_gradient_fill  

        # Blue
        bott_blue_fill_list = [
            f"I{start_row + 12}:W{start_row + 12}",  # I17:W17 -> start_row + 12
            f"I{start_row + 20}:W{start_row + 20}",  # I25:W25 -> start_row + 20
            f"I{start_row + 34}:W{start_row + 34}",  # I39:W39 -> start_row + 34
            f"I{start_row + 47}:W{start_row + 47}"   # I52:W52 -> start_row + 47
        ]
        for i in bott_blue_fill_list:
            for row in ws[i]:
                for cell in row:
                    cell.fill = bott_blue_gradient_fill  
                    cell.font = Font(bold=True)  

        full_blue_fill_list = [
            f"I{start_row + 13}:W{start_row + 14}",  # I18:W19 -> start_row + 13 to start_row + 14
            f"I{start_row + 21}:W{start_row + 22}",  # I26:W27 -> start_row + 21 to start_row + 22
            f"I{start_row + 35}:W{start_row + 36}",  # I40:W41 -> start_row + 35 to start_row + 36
            f"I{start_row + 48}:W{start_row + 49}"   # I53:W54 -> start_row + 48 to start_row + 49
        ]
        for i in full_blue_fill_list:
            for row in ws[i]:
                for cell in row:
                    cell.fill = full_blue_fill  
        # Gray
        for row in ws[f"I{start_row + 8}:W{start_row + 8}"]:  # I13:W13 -> start_row + 8
            for cell in row:
                cell.fill = bott_light_gray_gradient_fill 

        for row in ws[f"I{start_row + 9}:W{start_row + 9}"]:  # I14:W14 -> start_row + 9
            for cell in row:
                cell.fill = top_gray_gradient_fill 

        dark_gray_fill_list = [
            f"H{start_row}:W{start_row}",       # H5:W5 -> start_row
            f"H{start_row + 11}:W{start_row + 11}",  # H16:W16 -> start_row + 11
            f"H{start_row + 33}:W{start_row + 33}"   # H38:W38 -> start_row + 33
        ]
        for i in dark_gray_fill_list:
            for row in ws[i]:
                for cell in row:
                    cell.fill = dark_gray_gradient_fill 
                    cell.font=Font(bold=True)

        # Individual cells
        ws[f"F{start_row + 1}"].fill = dark_gray_gradient_fill  # F6 -> start_row + 1
        ws[f"F{start_row + 4}"].fill = dark_gray_gradient_fill  # F9 -> start_row + 4
        ws[f"B{start_row + 26}"].fill = dark_gray_gradient_fill  # B31 -> start_row + 26

        # Yellow
        for row in ws[f"I{start_row + 27}:W{start_row + 27}"]:  # I32:W32 -> start_row + 27
            for cell in row:
                cell.fill = bott_yellow_gradient_fill  

        for row in ws[f"I{start_row + 28}:W{start_row + 28}"]:  # I33:W33 -> start_row + 28
            for cell in row:
                cell.fill = full_yellow_gradient_fill  

        for row in ws[f"I{start_row + 29}:W{start_row + 29}"]:  # I34:W34 -> start_row + 29
            for cell in row:
                cell.fill = top_yellow_gradient_fill  

        # Orange
        top_orange_fill_list = [
            f"I{start_row + 15}:W{start_row + 15}",  # I20:W20 -> start_row + 15
            f"I{start_row + 37}:W{start_row + 37}"  # I42:W42 -> start_row + 37
        ]
        bott_orange_fill_list = [
            f"I{start_row + 19}:W{start_row + 19}",  # I24:W24 -> start_row + 19
            f"I{start_row + 41}:W{start_row + 41}"  # I46:W46 -> start_row + 41
        ]

        # Apply top orange gradient fill
        for i in top_orange_fill_list:
            for row in ws[i]:
                for cell in row:
                    cell.fill = top_orange_gradient_fill 

        # Apply bottom orange gradient fill
        for i in bott_orange_fill_list:
            for row in ws[i]:
                for cell in row:
                    cell.fill = bott_orange_gradient_fill
        bold_list=['C5','D5','F5','C6','B26','B39','B44','C38','E38','G39','H20','H21','H24','H32','H33','H34','H46','H43','H42']
        for i in bold_list:
            ws[i].font = Font(bold=True)


        for row in ws[f"C{start_row}:W{start_row + 49}"]:  # Use the dynamically calculated range
            for cell in row:
                cell.border = gridline1  # Apply the border
                # Hide column A
        # Define border style for top and bottom only

        gridline_top_bottom = Border(
        left=Side(style="thin", color='FFFFFF'),
        right=Side(style="thin", color='FFFFFF'),
        top=Side(style="thin", color='D9D9D9'),
        bottom=Side(style="thin", color='D9D9D9')
    )

        for row in ws[f"B{start_row+50}:W{start_row+50}"]:
            for cell in row:
                cell.border = gridline_top_bottom
    
    # Define border style for left and right only

        
    ws.column_dimensions.group("A", "A", outline_level=1, hidden=True)

    wb.save(output_file_path)  # Save the workbook as an .xlsx file
    print(f"Workbook '{output_file}' saved successfully.")
    result = {}
    # result[output_file] = {"Pid_to_review": pids_below_door_count, "Pid_to_birthstone": birthstone_list ,  "Pid_to_notify_to_macy":notify_macys,   "Pid_to_Store_product": store_products}
    #print("JSON structure to be written:\n", final_data)
    result = {
        output_file: {
            "Pid_to_review": pids_below_door_count,
            "Pid_to_birthstone": birthstone_list,
            "Pid_to_notify_to_macy": notify_macys,
            "Pid_to_Store_product": store_products,
        }
    }
    return result

def convert_to_tuples(categories):
    return [(category['name'], category['value']) for category in categories]

def process_data(input_path,file_path,month_from,month_to,percentage,input_tuple):

    print("input_path------>",input_path)

    sheet_config = {
        "Index": {"usecols": "A:P", "nrows": 41, "header": 2},
        "report grouping": {"header": None},
        "Repln Items": {"header": 2},
        "Setup Sales -L3M & Future": {"header": 9},
        "Macys Recpts": {"header": 1},
        "All_DATA": {"header": 0},
        "MCOM_Data": {"header": 0},
    }

    # Create a multiprocessing pool
    with mp.Pool(processes=mp.cpu_count()) as pool:
        # Prepare arguments for multiprocessing
        pool_args = [(sheet_name, config, input_path) for sheet_name, config in sheet_config.items()]
        
        # Process each sheet in parallel
        output = pool.starmap(process_sheet, pool_args)

    # Collect results into a dictionary
    dataframes = {sheet_name: data for sheet_name, data in output if data is not None}
    print("All sheets processed successfully!")

    # Store the results into specified variables
    index_df = dataframes.get("Index")
    report_grouping_df = dataframes.get("report grouping")
    planning_df = dataframes.get("Repln Items")
    TBL_Planning_VerticalReport__3 = dataframes.get("Setup Sales -L3M & Future")
    Macys_Recpts = dataframes.get("Macys Recpts")
    All_DATA = dataframes.get("All_DATA")
    MCOM_Data = dataframes.get("MCOM_Data")

    # Debugging: Print the results for verification
    if index_df is not None:
        print(f"Index sheet processed with {len(index_df)} rows.")
    if report_grouping_df is not None:
        print(f"Report Grouping sheet processed with {len(report_grouping_df)} rows.")
    if planning_df is not None:
        print(f"Planning sheet processed with {len(planning_df)} rows.")
    if TBL_Planning_VerticalReport__3 is not None:
        print(f"TBL_Planning_VerticalReport__3 sheet processed with {len(TBL_Planning_VerticalReport__3)} rows.")
    if Macys_Recpts is not None:
        print(f"Macys Recpts sheet processed with {len(Macys_Recpts)} rows.")
    if All_DATA is not None:
        print(f"All_DATA sheet processed with {len(All_DATA)} rows.")
    if MCOM_Data is not None:
        print(f"MCOM_Data sheet processed with {len(MCOM_Data)} rows.")

    #make master sheet# Specify the columns you want to extract
    columns_to_extract = [
        'PID', 'Cross ref', 'Sort', 'Dpt ID', 'Dpt Desc', 'Mkst', 'PID Desc',
        'CL ID', 'Class Desc', 'SC ID', 'SC Desc', 'MstrSt ID', 'Masterstyle Desc',
        'TY Last Cost', 'Own Retail', 'AWR 1st Tkt Ret', 'Prod Desc', 'Grouping',
        'Vendor', 'Vendor Name', 'FC Index', 'FLDC','Safe/Non-Safe', 'Item Code'
    ]

    # Extract the specified columns
    df_filtered = planning_df[columns_to_extract]

    # Create a new 'Gender' column based on the conditions
    df_filtered['Gender'] = 'Women'  # Default value
    df_filtered.loc[df_filtered['Dpt ID'].isin([768, 771]), 'Gender'] = 'Men'
    df_filtered.loc[df_filtered['CL ID'] == 86, 'Gender'] = 'Children'


    # List of birthstone names
    birthstones = [
        'GARNET', 'AMETHYST', 'AQUAMARINE', 'DIAMOND', 'EMERALD', 'PEARL', 
        'RUBY', 'PERIDOT', 'SAPPHIRE', 'OPAL', 'CITRINE', 'TANZANITE'
    ]

    # Function to check if 'Class Desc' contains any birthstone name
    def find_birthstone(class_desc):
        if isinstance(class_desc, str):  # Ensure the value is a string
            for stone in birthstones:
                if stone in class_desc.upper():  # Check if the birthstone is part of the string
                    return stone
        return ''  # Return blank if no birthstone is found

    # Apply the function to create the 'Birthstone' column
    df_filtered['Birthstone'] = df_filtered['Class Desc'].apply(find_birthstone)

    # Create the 'BSP_or_not' column based on the 'MstrSt ID' condition
    df_filtered['BSP_or_not'] = df_filtered['MstrSt ID'].apply(lambda x: 'BSP' if x in [26481, 74692] else '')


    def categorize_product(class_desc):
        if isinstance(class_desc, str):  # Ensure the value is a string
            class_desc = class_desc.upper()
            if 'EARRINGS' in class_desc or 'EARS' in class_desc:
                return 'Earrings'
            elif 'PENDANTS' in class_desc or 'PENDS' in class_desc:
                return 'Pendants'
            elif ' RING' in class_desc:  # Notice the space before "RINGS"
                return 'Rings'
            elif 'BRACELETS' in class_desc:
                return 'Bracelets'
            elif 'CHAIN' in class_desc:
                return 'Necklace'
            elif 'NECKS' in class_desc or 'NECKLACE' in class_desc:
                return 'Necklace'
            elif 'SET' in class_desc:
                return 'Set'
            elif 'BAND' in class_desc:
                return 'BAND'
        return ''  # Return blank if no match found

    # Apply the function to create the 'category' column
    df_filtered['category'] = df_filtered['Class Desc'].apply(categorize_product)


    def update_category_from_prod_desc(row):
        if not row['category']:  # If category is blank
            prod_desc = str(row['Prod Desc']).lower()
            categories_found = []

            # Search for keywords in Prod desc
            if 'earring' in prod_desc or 'stud' in prod_desc:
                categories_found.append('Earrings')
            if ' ring' in prod_desc:
                categories_found.append('Rings')
            if 'pendant' in prod_desc or 'necklace' in prod_desc:
                categories_found.append('Pendants' if 'pendant' in prod_desc else 'Necklace')
            if 'bracelet' in prod_desc:
                categories_found.append('Bracelets')
            if 'band' in prod_desc:
                categories_found.append('Band')
            if 'locket' in prod_desc:
                categories_found.append('Locket')
            # If multiple categories are found, set as 'Set'
            if len(categories_found) > 1:
                return 'Set'
            elif categories_found:
                return categories_found[0]
        return row['category']  # Return the original category if not blank

    # Apply the update function
    df_filtered['category'] = df_filtered.apply(update_category_from_prod_desc, axis=1)


    # Create the 'type' column based on conditions in 'Prod desc'
    def determine_type(prod_desc):
        if isinstance(prod_desc, str):  # Ensure the value is a string
            prod_desc = prod_desc.lower()
            if 'heart' in prod_desc:
                return 'Heart'
            elif 'stud' in prod_desc:
                return 'Studs'
            elif 'drop' in prod_desc:
                return 'Drop'
        return ''  # Return blank if no match found

    # Apply the function to create the 'type' column
    df_filtered['type'] = df_filtered['Prod Desc'].apply(determine_type)

    birthstone_data = [
        (1, 'January', 'Garnet'),
        (2, 'February', 'Amethyst'),
        (3, 'March', 'Aquamarine'),
        (4, 'April', 'Diamond'),
        (5, 'May', 'Emerald'),
        (6, 'June', 'Pearl'),
        (7, 'July', 'Ruby'),
        (8, 'August', 'Peridot'),
        (9, 'September', 'Sapphire'),
        (10, 'October', 'Opal'),
        (11, 'November', 'Citrine'),
        (12, 'December', 'Tanzanite')
    ]

    # Create a DataFrame from the data
    birthstone_sheet = pd.DataFrame(birthstone_data, columns=['Month', 'Month Name', 'Birthstone'])


    master_sheet = df_filtered[['PID','category','Birthstone','BSP_or_not','type','Gender','Vendor','Vendor Name','Own Retail','FC Index', 'FLDC','Safe/Non-Safe', 'Item Code']]

    # Create the unique sheets

    df_unique_vendor = df_filtered[['Vendor', 'Vendor Name']].drop_duplicates().dropna()


    data = [
        {"Vendor Name": "ALMOND (THAILAND) LIMITED", "Country of Origin": "Thailand", "Lead Time(weeks)": 8},
        {"Vendor Name": "AMMANTE JEWELLS LLP", "Country of Origin": "India", "Lead Time(weeks)": 8},
        {"Vendor Name": "ARIN", "Country of Origin": "Peru", "Lead Time(weeks)": 9},
        {"Vendor Name": "ARPAS INTERNATIONAL LTD", "Country of Origin": "US", "Lead Time(weeks)": 9},
        {"Vendor Name": "ASIAN STAR COMPANY LTD", "Country of Origin": "India", "Lead Time(weeks)": 8},
        {"Vendor Name": "BEAUTY GEMS FACTORY CO LTD", "Country of Origin": "Thailand", "Lead Time(weeks)": 8},
        {"Vendor Name": "CHARM AMERICA", "Country of Origin": "US", "Lead Time(weeks)": 9},
        {"Vendor Name": "CREATIONS JEWELLERY MFG. PVT. LTD.", "Country of Origin": "India", "Lead Time(weeks)": 8},
        {"Vendor Name": "DAISY'S COLLECTION, INC", "Country of Origin": "China", "Lead Time(weeks)": 8},
        {"Vendor Name": "DEORO - WIRE", "Country of Origin": "Peru", "Lead Time(weeks)": 9},
        {"Vendor Name": "DIA GOLD CREATION PVT. LTD.", "Country of Origin": "India", "Lead Time(weeks)": 8},
        {"Vendor Name": "DRL Manufacturing - LG", "Country of Origin": "DRL", "Lead Time(weeks)": 10},
        {"Vendor Name": "DRL MANUFACTURING S.A.", "Country of Origin": "DRL", "Lead Time(weeks)": 10},
        {"Vendor Name": "DRL/SARDELLI", "Country of Origin": "DRL", "Lead Time(weeks)": 10},
        {"Vendor Name": "GDL JEWELLERY LTD", "Country of Origin": "Thailand", "Lead Time(weeks)": 8},
        {"Vendor Name": "GOLDENLINE C/O BOGAZICI HEDIYELIK A.S.", "Country of Origin": "US", "Lead Time(weeks)": 9},
        {"Vendor Name": "IJM-INTERCONTINENTAL JEWELLERY MFG", "Country of Origin": "Thailand", "Lead Time(weeks)": 8},
        {"Vendor Name": "JOHN C NORDT COMPANY", "Country of Origin": "US", "Lead Time(weeks)": 9},
        {"Vendor Name": "LARA CORPORATION LIMITED", "Country of Origin": "China", "Lead Time(weeks)": 8},
        {"Vendor Name": "Leach & Garner (HK) Limited", "Country of Origin": "China", "Lead Time(weeks)": 8},
        {"Vendor Name": "LINEA NUOVA S.A.", "Country of Origin": "Peru", "Lead Time(weeks)": 9},
        {"Vendor Name": "Milor S.P.A", "Country of Origin": "Italy", "Lead Time(weeks)": 9},
        {"Vendor Name": "MIORO GOLD, LLC", "Country of Origin": "US", "Lead Time(weeks)": 9},
        {"Vendor Name": "NEW GOLD", "Country of Origin": "US", "Lead Time(weeks)": 9},
        {"Vendor Name": "PD&P LTD.", "Country of Origin": "China", "Lead Time(weeks)": 8},
        {"Vendor Name": "RICHLINE ITALY SRL", "Country of Origin": "Italy", "Lead Time(weeks)": 12},
        # Note: The country for "RICHLINE SA PTY" is unclear in the provided data. 
        # If "Zales" is not a country, you can adjust as needed or remove. 
        {"Vendor Name": "RICHLINE SA PTY", "Country of Origin": "Zales", "Lead Time(weeks)": None},
        {"Vendor Name": "RLG STANDARD MANUFACTURING VENDOR", "Country of Origin": "US", "Lead Time(weeks)": 9},
        {"Vendor Name": "S&S JEWELRY", "Country of Origin": "US", "Lead Time(weeks)": 9},
        {"Vendor Name": "SABELLI, S.A. DE C.V.", "Country of Origin": "Mexico", "Lead Time(weeks)": 9},
        {"Vendor Name": "SERENITY JEWELS PVT LTD", "Country of Origin": "India", "Lead Time(weeks)": 8},
        {"Vendor Name": "SHANGOLD INDIA LTD", "Country of Origin": "India", "Lead Time(weeks)": 8},
        {"Vendor Name": "SILO SPA-WIRES", "Country of Origin": "Italy", "Lead Time(weeks)": 12},
        {"Vendor Name": "STRONG JEWELRY(HK) CO LTD CHONG", "Country of Origin": "China", "Lead Time(weeks)": 8},
        {"Vendor Name": "T.C.I. LTD", "Country of Origin": "China", "Lead Time(weeks)": 8},
        {"Vendor Name": "National Chain", "Country of Origin": "US", "Lead Time(weeks)": 9},
        # Repeated entries (if you want to keep them exactly as listed):
        {"Vendor Name": "Leach & Garner (HK) Limited", "Country of Origin": "China", "Lead Time(weeks)": 8},
        {"Vendor Name": "LEACH & GARNER(HK) LTD", "Country of Origin": "China", "Lead Time(weeks)": 8},
        {"Vendor Name": "MANJUSAKA JEWELERS CO.,LTD", "Country of Origin": "China", "Lead Time(weeks)": 8},
        {"Vendor Name": "MEICHONG JEWELRY (HK) COMPANY LIMITED", "Country of Origin": "China", "Lead Time(weeks)": 8},
        {"Vendor Name": "NATIONAL CHAIN", "Country of Origin": "US", "Lead Time(weeks)": 9},
        {"Vendor Name": "PRIME DIRECT LIMITED", "Country of Origin": "China", "Lead Time(weeks)": 8},
    ]
    
    # Create the DataFrame
    df_coo = pd.DataFrame(data)

    # Merge the data on Vendor Name to add the 'Country of Origin' and 'Lead Time'
    vendor_sheet = pd.merge(df_unique_vendor, df_coo[['Vendor Name', 'Country of Origin', 'Lead Time(weeks)']], 
                        on='Vendor Name', how='left')



    def get_previous_retail_week():
        """
        Get the previous week's month, year of the previous month, 
        last year's occurrence of that month, last month before the previous month in numeric format,
        determine SP (Spring) or FA (Fall) based on the previous month,
        and calculate the number of retail weeks for each month individually.
        """
        # Use the current date as input
        current_date = datetime.now()

        # Find the current week's Sunday
        current_sunday = current_date - timedelta(days=current_date.weekday() + 1)

        # Calculate the previous week's Sunday
        previous_week_sunday = current_sunday - timedelta(days=7)

        # Determine the previous week number
        previous_week_number = (previous_week_sunday.day - 1) // 7 + 1

        # Get the month and year of the previous week
        current_month = previous_week_sunday.strftime('%b')
        year_of_previous_month = previous_week_sunday.year

        # Determine last year's occurrence of the same month
        last_year_of_previous_month = year_of_previous_month - 1

        # Determine the last month before the previous month
        last_month_of_previous_month_date = previous_week_sunday.replace(day=1) - timedelta(days=1)
        last_month = last_month_of_previous_month_date.strftime('%b').upper()

        # Custom mapping for months
        month_mapping = {
            'FEB': 1, 'MAR': 2, 'APR': 3, 'MAY': 4,
            'JUN': 5, 'JUL': 6, 'AUG': 7, 'SEP': 8,
            'OCT': 9, 'NOV': 10, 'DEC': 11, 'JAN': 12
        }
        last_month_of_previous_month_numeric = month_mapping[last_month]

        # Determine SP (Spring) or FA (Fall/Winter) based on the previous month
        spring_months = ['FEB', 'MAR', 'APR', 'MAY', 'JUN', 'JUL']
        fall_months = ['AUG', 'SEP', 'OCT', 'NOV', 'DEC', 'JAN']

        season = "SP" if current_month in spring_months else "FA"

        # Calculate the number of retail weeks for each month of the current year
        current_year = year_of_previous_month
        # Individual variables for retail weeks of each month
    

        def get_retail_weeks(year, month):
            """
            Calculate the number of retail weeks in a given month.
            Retail weeks follow the Sunday-to-Saturday structure, 
            and all days in a week belong to the month in which the week starts.
            
            Args:
                year (int): The year of the month.
                month (int): The month (1 for January, 12 for December).

            Returns:
                int: Number of retail weeks in the month.
            """
            # Get the first day and last day of the month
            first_day = datetime(year, month, 1)
            last_day = datetime(year, month, monthrange(year, month)[1])

            # Find the first Sunday of the month
            first_sunday = first_day + timedelta(days=(6 - first_day.weekday()) % 7)

            # Find the last Saturday of the month
            last_saturday = last_day - timedelta(days=last_day.weekday() + 1)

            # Count retail weeks
            current_week_start = first_sunday
            week_count = 0

            while current_week_start <= last_saturday:
                week_count += 1
                current_week_start += timedelta(days=7)  # Move to the next Sunday

            # Check if the final week starts in the current month (partial week rule)
            if current_week_start <= last_day:
                week_count += 1

            return week_count

        feb_weeks = get_retail_weeks(current_year,2)
        mar_weeks = get_retail_weeks(current_year,3)
        apr_weeks = get_retail_weeks(current_year,4)
        may_weeks = get_retail_weeks(current_year,5)
        jun_weeks = get_retail_weeks(current_year,6)
        jul_weeks = get_retail_weeks(current_year,7)
        aug_weeks = get_retail_weeks(current_year,8)
        sep_weeks = get_retail_weeks(current_year,9)
        oct_weeks = get_retail_weeks(current_year,10)
        nov_weeks = get_retail_weeks(current_year,11)
        dec_weeks = get_retail_weeks(current_year,12)
        jan_weeks = get_retail_weeks(current_year + 1, 1)  # January belongs to the next year

        return current_month, previous_week_number, year_of_previous_month,last_year_of_previous_month, last_month_of_previous_month_numeric,season, feb_weeks, mar_weeks, apr_weeks, may_weeks,jun_weeks, jul_weeks, aug_weeks, sep_weeks, oct_weeks,nov_weeks, dec_weeks, jan_weeks
        # return current_month, previous_week_number, year_of_previous_month, last_year_of_previous_month, last_month_of_previous_month_numeric, season

    # Call the function and #print the results
    current_month, previous_week_number, year_of_previous_month,last_year_of_previous_month, last_month_of_previous_month_numeric,season, feb_weeks, mar_weeks, apr_weeks, may_weeks,jun_weeks, jul_weeks, aug_weeks, sep_weeks, oct_weeks,nov_weeks, dec_weeks, jan_weeks = get_previous_retail_week()


   
#     category_tuples=[

#     ('Bridge Gem', '742'),
#     ('Gold', '746'),
#     ('Gold', '262&270'),
#     ('Womens Silver', '260&404'),
#     ('Precious', '264&268'),
#     ('Fine Pearl', '265&271'),
#     ('Semi', '272&733'),
#     ('Diamond', '734&737&748'),
#     ('Bridal', '739&267&263'),
#     ("Men's", '768&771')

# ]
    category_tuples = input_tuple
    
    
    dynamic_categories = [
    (
        category,
        code,
        report_grouping_df.loc[
            report_grouping_df[0].str.upper() == f"{category.upper()}{code}".upper(), 3
        ].iloc[0] if not report_grouping_df.loc[
            report_grouping_df[0].str.upper() == f"{category.upper()}{code}".upper()
        ].empty else None  # Handle missing values
    )
    for category, code in category_tuples]
    
    month_dict = {
        "Feb": 1,
        "Mar": 2,
        "Apr": 3,
        "May": 4,
        "Jun": 5,
        "Jul": 6,
        "Aug": 7,
        "Sep": 8,
        "Oct": 9,
        "Nov": 10,
        "Dec": 11,
        "Jan": 12
    }
    current_month_number = month_dict.get(current_month, "Month not found")
    if current_month in [ "Oct","Nov" "Dec","Jan"]:
        rolling_method="Current MTH"
    else:
        rolling_method="YTD"
    year_of_previous_month=2024
    last_year_of_previous_month=2023
    
    other_params = (
        year_of_previous_month,
        last_year_of_previous_month,
        season,
        current_month,
        current_month_number,
        previous_week_number,
        last_month_of_previous_month_numeric,
        rolling_method,
        feb_weeks,
        mar_weeks,
        apr_weeks,
        may_weeks,
        jun_weeks,
        jul_weeks,
        aug_weeks,
        sep_weeks,
        oct_weeks,
        nov_weeks,
        dec_weeks,
        jan_weeks,
        index_df,
        report_grouping_df,
        planning_df,
        TBL_Planning_VerticalReport__3,
        Macys_Recpts,
        All_DATA,
        MCOM_Data,
        percentage,
        month_from,
        month_to,
        master_sheet,
        vendor_sheet,
        birthstone_sheet

    ) 
    
    
    #main
    # args = [(category, (code, num_products), dynamic_categories, file_path, other_params) for category, (code, num_products) in dynamic_categories.items()]
    args = [(category, (code, num_products), dynamic_categories, file_path, other_params)    for category, code, num_products in dynamic_categories ]
    # Use multiprocessing pool to process the categories
    with Pool(processes=os.cpu_count()) as pool:
        results = pool.map(process_category, args)


    final_data = {}
    for result in results:
        final_data.update(result)

    # Write the final JSON to a file
    with open("media/output.json", "w") as f:
        json.dump(final_data, f, indent=2)

    print("Final JSON saved successfully.")

   



