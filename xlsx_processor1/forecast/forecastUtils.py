import os 
import zipfile 
import shutil
from calendar import monthrange
from datetime import datetime,timedelta
from openpyxl.worksheet.datavalidation import DataValidation
import pandas as pd
import openpyxl
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side,GradientFill
import math
import multiprocessing as mp
from multiprocessing import Manager, Pool
import json 
from .config import *
import numpy as np
from .adddatabase import save_macys_projection_receipts, save_monthly_forecasts
import logging
from .models import MonthlyForecast,ProductDetail
# Set up logging configuration
logging.basicConfig(filename=r'all_file.log', level=logging.INFO,
                    format='%(asctime)s:%(levelname)s:%(message)s',
                    filemode='w')
def make_zip_and_delete(folder_path):
    folder_path = os.path.normpath(folder_path)
    zip_file_path = os.path.normpath(f'{folder_path}.zip')
    
    try:
        # Create a ZIP file
        with zipfile.ZipFile(zip_file_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for root, dirs, files in os.walk(folder_path):
                for file in files:
                    file_path = os.path.join(root, file)
                    arcname = os.path.relpath(file_path, folder_path)  # Preserve folder structure
                    zipf.write(file_path, arcname)
        
        print(f"Folder '{folder_path}' has been compressed into '{zip_file_path}'")

        # # Delete the folder after zipping
        # shutil.rmtree(folder_path)
        # print(f"Folder '{folder_path}' has been deleted successfully.")
    
    except PermissionError:
        print(f"Permission denied: Cannot access '{folder_path}'. Please check folder permissions.")
    except FileNotFoundError:
        print(f"File not found: '{folder_path}' does not exist.")
    except Exception as e:
        print(f"An error occurred: {e}")
 

def add_dropdown(ws, cell, options):
    # Create a data validation object with the list of options
    dropdown = DataValidation(type="list", formula1=f'"{",".join(options)}"', allow_blank=True)
    dropdown.prompt = "Please select an option"
    dropdown.promptTitle = "Dropdown List"
    # Apply the data validation dropdown to the specified cell
    ws.add_data_validation(dropdown)
    dropdown.add(ws[cell])

def apply_round_format(ws, cell_ranges, decimal_places):
    for cell_range in cell_ranges:
        # Check if the cell range is a single cell
        if ":" in cell_range:
            # This is a range, so we can iterate through it
            for row in ws[cell_range]:
                for cell in row:
                    # Check if cell contains a formula by verifying if the value starts with "="
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                        # Wrap formula in ROUND with the specified decimal places
                        cell.value = f"=ROUND({cell.value[1:]}, {decimal_places})"
        else:
            # This is a single cell reference
            cell = ws[cell_range]
            if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                # Wrap formula in ROUND with the specified decimal places
                cell.value = f"=ROUND({cell.value[1:]}, {decimal_places})"

def apply_format(ws, cell_ranges, number_format):
    for cell_range in cell_ranges:
        # Check if the cell range is a single cell or a range
        if ":" in cell_range:
            # This is a range, so we can iterate through it
            for row in ws[cell_range]:
                for cell in row:
                    cell.number_format = number_format
        else:
            # This is a single cell
            cell = ws[cell_range]
            cell.number_format = number_format

def read_excel_sheet(file_details):
    file_path, sheet_name, kwargs = file_details
    return sheet_name, pd.read_excel(file_path, sheet_name=sheet_name, **kwargs)


def parse_date(date_value):
    """
    Parses a date value (string or pandas Timestamp) into a datetime.date object.
    Handles multiple formats and NaT (Not a Time) values.
    """
    if pd.isna(date_value) or date_value in ["NaT", None, ""]:
        return None  # Handle missing or NaT values

    # If the value is already a pandas Timestamp, convert it to date
    if isinstance(date_value, pd.Timestamp):
        return date_value.date()

    # Ensure the value is a string for parsing
    date_str = str(date_value).strip()

    try:
        # Try to parse "M/D/YYYY" or "MM/DD/YYYY" format
        return datetime.strptime(date_str.split()[0], "%m/%d/%Y").date()
    except ValueError:
        try:
            # Try to parse "M/D/YYYY H:MM:SS AM/PM" format
            return datetime.strptime(date_str.split(' ', 1)[0], "%m/%d/%Y").date()
        except ValueError:
            try:
                # Try parsing "YYYY-MM-DD HH:MM:SS" format (e.g., "2022-02-26 00:00:00")
                return datetime.strptime(date_str.split()[0], "%Y-%m-%d").date()
            except ValueError:
                # Try various other possible formats
                date_formats = ["%Y-%m-%d", "%d-%m-%Y", "%m-%d-%Y"]
                for fmt in date_formats:
                    try:
                        return datetime.strptime(date_str.split()[0], fmt).date()
                    except ValueError:
                        continue
                
                # If all parsing attempts fail, print a warning and return None
                print(f"Could not parse date: {date_value}")
                return None
            except Exception as e:
                print(f"Error parsing date {date_value}: {e}")
                return None
            

# multiprocess pool for reading excel 
def process_sheet(sheet_name, config, input_path):
    """
    Function to process a single sheet with specified configuration.
    """
    print(f"Processing sheet: {sheet_name}")
    try:
        excel_file = pd.ExcelFile(input_path)
        data = excel_file.parse(sheet_name=sheet_name, **config)
        print(f"Finished processing sheet: {sheet_name}")
        return sheet_name, data
    except Exception as e:
        print(f"Error processing sheet {sheet_name}: {e}")
        return sheet_name, None

def safe_int(value):
    """Convert value to int or return None if conversion fails"""
    if pd.isna(value):
        return None
    try:
        return int(value)
    except (ValueError, TypeError):
        return None
            
def safe_float(value):
    """Convert value to float or return None if conversion fails"""
    if pd.isna(value):
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None
        
def safe_str(value, max_length=None):
    """Convert value to string, respecting max_length if provided"""
    if pd.isna(value):
        return None
    try:
        result = str(value).strip()
        if max_length and len(result) > max_length:
            result = result[:max_length]
        return result
    except:
        return None
            
# multiprocess pool for generating excel 
def process_category(args):
    """
    Process a single category with its associated data.
    """
    category, (code, num_products), dynamic_categories, file_path, other_params = args

    # Extract required variables from other_params
    (
        year_of_previous_month,
        last_year_of_previous_month,
        season,
        current_month,
        current_month_number,
        previous_week_number,
        last_month_of_previous_month_numeric,
        rolling_method,
        feb_weeks,
        mar_weeks,
        apr_weeks,
        may_weeks,
        jun_weeks,
        jul_weeks,
        aug_weeks,
        sep_weeks,
        oct_weeks,
        nov_weeks,
        dec_weeks,
        jan_weeks,
        index_df,
        report_grouping_df,
        planning_df,
        TBL_Planning_VerticalReport__3,
        Macys_Recpts,
        All_DATA,
        MCOM_Data,
        percentage,
        month_from,
        month_to,
        master_sheet,
        vendor_sheet,
        birthstone_sheet

    ) = other_params

    # Placeholder for the processing logic of each category
    # Use the original code logic for processing categories here.
    print("Madhaveeeeeeeee ",year_of_previous_month,last_year_of_previous_month)
    print(f"Processing Category: {category} with Code: {code} and Num Products: {num_products}")
    data = [
        ["", "TY", year_of_previous_month, "LY", last_year_of_previous_month, "Season", season, "Current Year", "Month", current_month, current_month_number, "Week", previous_week_number, "", "MAY-SEP", "", "", "Last Completed Month", last_month_of_previous_month_numeric, "", "Use EOM Actual?", rolling_method],
        ["", "Count of Items", "", 8215, "", "", "", "Last SP / FA Months", "Month", "Jul", "", "Jan", 12, "Sorted by:", "Dept Grouping >Class ID", "", "", ""],
        ["", "", "", "", "", "", "", "# of Wks in Mth", feb_weeks, mar_weeks, apr_weeks, may_weeks, jun_weeks, jul_weeks, aug_weeks, sep_weeks, oct_weeks, nov_weeks, dec_weeks, jan_weeks],
        ["", category.upper(), code, "", "Avg Sales 1st & last Mth", 8, 11, "Month #", 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, "", "", ""]
    ]
    output_file_path = f"{category}{code}.xlsx"
    output_file = os.path.join(file_path, f'{output_file_path}.xlsx')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = category
    ws_index = wb.create_sheet(title="Index")
    ws_month = wb.create_sheet(title="Month")
    ws_dropdown = wb.create_sheet(title="DropdownData")
    all_products=[]
    com_products=[]
    store_products=[]
    pids_below_door_count_alert = []
    added_macys_proj_receipts_alert =[]
    notify_macys_alert=[]
    upcoming_birthstone_products = []
    all_birthstone_products = []
    min_order_alert = []
    


    # Loop through products to generate PIDs

    # Step 4: Populate Worksheet with Data
    for row_num, row_data in enumerate( data, 1):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")
    # Step 4: Freeze Top 4 Rows
    ws.freeze_panes = ws["A5"]
    # Define options for dropdowns
    lookup_key = f"{category.upper()}{code}"
# Perform a lookup in 'report grouping'
    lookup_value_c2 = report_grouping_df.loc[
    report_grouping_df[0].str.upper() == lookup_key.upper(), 3
].iloc[0]
    ws['C2'] = lookup_value_c2
    ws['D2'] = "=C2*51+4"
    ws['K1'] = "=VLOOKUP(J1,Month!A:B,2,0)"
    ws['K2'] = "=VLOOKUP(J2,Month!A:B,2,0)"
    ws['M2'] = "=VLOOKUP(L2,Month!A:B,2,0)"
    # Define additional dropdown options

    # Define your dropdown options (38 items)
    dropdown_options = [
        "BT", "Citrine", "Cross", "CZ", "Dia", "Ear", "EMER", "Garnet", "Gem",
        "GEM EAR", "Gold Chain", "GOLD EAR", "Amy", "Anklet", "Aqua", "Bridal",
        "Heart", "Heavy Gold Chain", "Jade", "KIDS", "Locket", "Mens Gold Bracelet",
        "Mens Misc", "Mens Silver chain", "Mom", "MOP", "Neck", "Onyx", "Opal",
        "Pearl", "Peridot", "Religious", "Ring", "Ruby", "Saph", "Womens Silver Chain",
        "Wrist", "Grand Total"
    ]

    # Write the dropdown options to the "DropdownData" sheet in column A
    for i, option in enumerate(dropdown_options, start=1):
        ws_dropdown.cell(row=i, column=1, value=option)

    # Define the named range for the dropdown options
    # Make sure the range covers exactly the 38 items (A1 to A38)
    named_range = DefinedName(name="DropdownOptions", attr_text="DropdownData!$A$1:$A$38")
    wb.defined_names.add(named_range)

    # Create a data validation that references the named range
    dropdown = DataValidation(type="list", formula1="DropdownOptions", allow_blank=True)
    dropdown.prompt = "Please select an option"
    dropdown.promptTitle = "Dropdown List"



    for loop in range(num_products):
        start_row = 5 + (51 * loop)  # Adjust the range as needed for your loop
        ws.add_data_validation(dropdown)
        dropdown.add(ws[f"F{start_row + 1}"])  # Apply to column F as an example

    forecast_method_options = ["FC by Index", "FC by Trend", "Average", "Current Year", "Last Year"]


    season_option = ["FA", "SP"]
    month_option = ["Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Jan"]
    year_option = ["Current MTH", "YTD", "SPRING", "FALL", "LY FALL"]


    # Function to add a dropdown to a specific cell
    def add_dropdown(ws, cell, options):
        # Create a data validation object with the list of options
        dropdown = DataValidation(type="list", formula1=f'"{",".join(options)}"', allow_blank=True)
        dropdown.prompt = "Please select an option"
        dropdown.promptTitle = "Dropdown List"
        # Apply the data validation dropdown to the specified cell
        ws.add_data_validation(dropdown)
        dropdown.add(ws[cell])
    # Add dropdowns to specified cells
    add_dropdown(ws, "G1", season_option)     # G1 for Season
    add_dropdown(ws, "V1", year_option)       # V1 for Year options

    # Add dropdowns to multiple cells for Months
    for cell in ["J1", "J2", "L2"]:
        add_dropdown(ws, cell, month_option)




    ALL_VALUES = [
        "PID/BLU/MKST", "Current FC Index", "(TY/LY) STD Sales Index/12M FC", "STD Trend / 12M FC",
        "Item Status/Forecasting Method/Safe", "Current Str Cnt/Last Str Cnt/Last Updated",
        "Store Model/Com Model/TTL Model", "MA Proj/Proj Ball/Holiday Bd FC",
        "MCY/OH/OH in Transit/MTD Ships/LW Ships", "Planned WOS/WOP/Wkly Avg Sale/Last 4 Wks Ships",
        "Actual WOS/WOP/Real OOS Loc", "Excess Proj Qty & $ / Recall Wks of Poj & Qty",
        "Proj Qty & $ to Release / Note", "Vendor/Min order", "RL TTL/Net Proj/ORD Unalloc/+/− to Model",
        "MA Bin/FLDC/WIP QTY/REPLN HOLD DATE", "WIP Demand", "MD Status/Store & Mcom Repl",
        "TTL Last Repl/Age/Mths Active", "Last Cost/Owned/TKT Ret/GM/Actual GM", "Metal Lock/MFG Policy",
        "KPI DATA", "Last KPI Door count", "Diff to Current Door",
        "Out of Stock Locations", "Suspended Location count",
        "Click to View online", "DEPT #", "Sub Class", "Masterstyle", "PID Desc",
        "COM 1st Live/Live Site/V2C/WebID/STD Rtn", "Web ID Description",
        "Last Reviewed Date/Code/Qty to Enter", "Current Review Comments"
    ]

    H_VALUES = [
        'ROLLING 12M FC', 'Index', 'FC by Index', 'FC by Trend', 'Recommended FC', 'Planned FC',
        'Planned Shipments', 'Planned EOH (Cal)', 'Gross Projection (Nav)', 'Macys Proj Receipts',
        'Planned Sell thru %', f"TOTAL {year_of_previous_month}",'Total Sales Units', 'Store Sales Units', 'Com Sales Units',
        'COM % to TTL (Sales)', 'TOTAL EOM OH', 'Store EOM OH', 'COM EOM OH',
        'COM % to TTL (EOH)', 'Omni Sales $', 'COM Sales $', 'Omni AUR/% Diff Own',
        'Omni Sell Thru %', 'Store SellThru %', 'Omni Turn', 'Store turn',
        'TY Store Sales U vs LY', 'TY COM sales U vs LY', 'TY Store EOH vs LY',
        'Omni OO Units', 'COM OO Units', 'Omni Receipts',    f"TOTAL {last_year_of_previous_month}",
        "Total Sales Units",
        "Store Sales Units",
        "Com Sales Units",
        "COM % to TTL (Sales)",
        "TOTAL EOM OH",
        "Store EOM OH",
        "COM EOM OH",
        "COM % to TTL (EOH)",
        "Omni Receipts",
        "Omni Sell Thru %",
        "Store SellThru %",
        "Omni Turn",
        "Store Turn",
        "Omni Sales $",
        "COM Sales $",
        "Omni AUR/% Diff Own"
    ]

    MONTHLY_VALUES = ['FEB', 'MAR', 'APR', 'MAY', 'JUN', 'JUL', 'AUG', 'SEP', 'OCT',
                    'NOV', 'DEC', 'JAN', 'ANNUAL', 'SPRING', 'FALL']
    def apply_round_format(ws, cell_ranges, decimal_places):
        for cell_range in cell_ranges:
            # Check if the cell range is a single cell
            if ":" in cell_range:
                # This is a range, so we can iterate through it
                for row in ws[cell_range]:
                    for cell in row:
                        # Check if cell contains a formula by verifying if the value starts with "="
                        if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                            # Wrap formula in ROUND with the specified decimal places
                            cell.value = f"=ROUND({cell.value[1:]}, {decimal_places})"
            else:
                # This is a single cell reference
                cell = ws[cell_range]
                if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                    # Wrap formula in ROUND with the specified decimal places
                    cell.value = f"=ROUND({cell.value[1:]}, {decimal_places})"

    def apply_format(ws, cell_ranges, number_format):
        for cell_range in cell_ranges:
            # Check if the cell range is a single cell or a range
            if ":" in cell_range:
                # This is a range, so we can iterate through it
                for row in ws[cell_range]:
                    for cell in row:
                        cell.number_format = number_format
            else:
                # This is a single cell
                cell = ws[cell_range]
                cell.number_format = number_format


    # Open the source workbook
    
    for row_num, row_data in enumerate(index_df.values, start=1):
        for col_num, value in enumerate(row_data, start=1):
            ws_index.cell(row=row_num, column=col_num, value=value)

    # Step 5: Apply Percentage Formatting to B4:P41
    for row in ws_index.iter_rows(min_row=4, max_row=41, min_col=2, max_col=16):
        for cell in row:
            if isinstance(cell.value, (int, float)) and 0 <= cell.value <= 1:
                cell.number_format = '0.00%'  # Format as percentage


    # Step 6: Apply Filters to A3:P3
    ws_index.auto_filter.ref = "A3:P3"


    month_data = [
        ["All", ""],
        ["Feb", 1],
        ["Mar", 2],
        ["Apr", 3],
        ["May", 4],
        ["Jun", 5],
        ["Jul", 6],
        ["Aug", 7],
        ["Sep", 8],
        ["Oct", 9],
        ["Nov", 10],
        ["Dec", 11],
        ["Jan", 12],
    ]

    # Write the data to the "Month" sheet
    for row_num, (month, number) in enumerate(month_data, start=1):
        ws_month.cell(row=row_num, column=1, value=month)  # Column A
        ws_month.cell(row=row_num, column=2, value=number)  # Column B

    border_color = "D9D9D9"  # Gridline color
    gridline1 = Border(
        left=Side(style="thin", color=border_color),
        right=Side(style="thin", color=border_color),
        top=Side(style="thin", color=border_color),
        bottom=Side(style="thin", color=border_color)
    )

    # Create PatternFill objects for the colors
    light_gray = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    light_yellow = PatternFill(start_color="FFEB9C" , end_color="FFEB9C" , fill_type="solid")
    light_pink = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    # Apply the fills to specific cells
    ws['C1'].fill = light_gray
    ws['E1'].fill = light_gray
    ws['G1'].fill = light_gray
    ws['J1'].fill = light_gray
    ws['K1'].fill = light_gray
    ws['M1'].fill = light_gray
    ws['S1'].fill = light_gray
    ws['J2'].fill = light_gray
    ws['L2'].fill = light_gray

    ws['K2'].fill = light_yellow
    ws['O1'].fill = light_yellow
    ws['M2'].fill = light_yellow

    ws['B2'].fill = yellow_fill
    ws['C2'].fill = yellow_fill
    ws['V1'].fill = light_pink

    #no boader
    # Define white border style
    white_border = Border(
        left=Side(style='thin', color="FFFFFF"),
        right=Side(style='thin', color="FFFFFF"),
        top=Side(style='thin', color="FFFFFF"),
        bottom=Side(style='thin', color="FFFFFF")
    )

    # Apply no border to the range B1:W2
    for row in range(1, 3):  # Rows 1 to 2
        for col in range(2, 24):  # Columns B (2) to W (23)
            ws.cell(row=row, column=col).border = white_border

    # Remove borders for specific cells to show gridlines
    # Define the border style
    border = Border(
        left=Side(style='thin', color='7F7F7F'),  # Thin black border on the left
        right=Side(style='thin', color='7F7F7F'),  # Thin black border on the right
        top=Side(style='thin', color='7F7F7F'),  # Thin black border on the top
        bottom=Side(style='thin', color='7F7F7F')  # Thin black border on the bottom
    )
    #no_boader list
    no_boader_list=['C1','E1','G1','J1','K1','L1','M1','O1','S1','V1','J2','K2','L2','M2']
    for i in no_boader_list:
        ws[i].border = border

    #font color
    red_font_italic= Font(color="FF0000", italic=True) 
    yellow_font = Font(color="9C5700") 
    blue_bold_font = Font(color="0563C1",bold=True) 
    ws['B4'].font = blue_bold_font
    ws['C4'].font = blue_bold_font
    ws['B4'].font = Font(bold=True)
    ws['C4'].font = Font(bold=True)
    ws['K2'].font = yellow_font
    ws['O1'].font = yellow_font
    ws['M2'].font = yellow_font

    ws['H1'].font = red_font_italic
    ws['H2'].font = red_font_italic
    ws['O2'].font = red_font_italic
    # Define the alignment (right-aligned)
    right_alignment = Alignment(horizontal='right')
    ws['H1'].alignment = right_alignment
    ws['H2'].alignment = right_alignment
    # Apply no border to the cells in the no_boader list
    bold_font_list=['C1','E1','G1','J1','K1','M1','S1','V1','L2','J2','C2','E4','F4','G4','H4',]
    for i in bold_font_list:
        ws[i].font = Font(bold=True)
    for row in ws["H3:W4"]:  # Use the dynamically calculated range
        for cell in row:
            cell.border = gridline1 

    ws.column_dimensions['c'].width = 20
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['H'].width = 25
    
    for loop in range(num_products):
        start_row = 5 + (51 * loop)
        g_value = loop + 1
        cross_ref = f"{g_value}{category.upper()}{code}"  # Ensure no spaces in category
        #Find the matching row
        matching_row = planning_df.loc[planning_df['Cross ref'].str.upper() == cross_ref]
        #matching_row = planning_df.loc[planning_df['PID'].str.upper() == pid]
        pid_value = matching_row['PID'].iloc[0]
        logging.info('PID : %s', pid_value)


        all_products.append(pid_value)


        RLJ = matching_row['Adjusted RLJ Item'].iloc[0] 
        MKST = matching_row['Mkst'].iloc[0] # Get the first matching PID
        Current_FC_Index = matching_row['FC Index'].iloc[0] # Get the first matching PID
        Safe_Non_Safe=matching_row['Safe/Non-Safe'].iloc[0]
        Item_Code=matching_row['Item Code'].iloc[0]
        Item_Status=f"{Safe_Non_Safe}/{Item_Code}"
        Door_Count=matching_row['Door Count'].iloc[0]
        Last_Str_Cnt=matching_row['Old Door count'].iloc[0]
        Door_count_Updated=matching_row['Door count Updated'].iloc[0]
        Store_Model=matching_row['Model'].iloc[0]
        Com_Model=matching_row['Com Model'].iloc[0]
        Holiday_Bld_FC=matching_row['HolidayBuildFC'].iloc[0]
        MCYOH=matching_row['OH Units'].iloc[0]
        OO=matching_row['OO Units'].iloc[0]
        nav_OO=matching_row['nav OO'].iloc[0]
        MTD_SHIPMENTS=matching_row['MTD SHIPMENTS'].iloc[0]
        LW_Shipments=matching_row['LW Shipments'].iloc[0]
        Wks_of_Stock_OH=matching_row['Wks of Stock OH'].iloc[0]
        Wks_of_on_Proj=matching_row['Wks of on Proj'].iloc[0]
        Last_3Wks_Ships=matching_row['Last 3Wks Ships'].iloc[0]
        Vendor_Name=matching_row['Vendor Name'].iloc[0]
        Min_order=matching_row['Min order'].iloc[0]
        Proj=matching_row['Proj'].iloc[0]
        Net_Proj=matching_row['Net Proj'].iloc[0]
        Unalloc_Orders=matching_row['Unalloc Orders'].iloc[0]
        RLJ_OH=matching_row['RLJ OH'].iloc[0]
        FLDC=matching_row['FLDC'].iloc[0]
        WIP=matching_row['WIP'].iloc[0]
        MD_Status_MZ1=matching_row['MD Status MZ1'].iloc[0]
        Repl_Flag=matching_row['Repl Flag'].iloc[0]
        MCOM_RPL=matching_row['MCOM RPL'].iloc[0]
        Pool_stock=matching_row['Pool stock'].iloc[0]
        st_Rec_Date=matching_row['1st Rec Date'].iloc[0]
        Last_Rec_Date=matching_row['Last Rec Date'].iloc[0]
        Item_Age=matching_row['Item Age'].iloc[0]
        TY_Last_Cost=matching_row['TY Last Cost'].iloc[0]
        Own_Retail=matching_row['Own Retail'].iloc[0]
        AWR_1st_Tkt_Ret=matching_row['AWR 1st Tkt Ret'].iloc[0]
        Metal_Lock=matching_row['Metal Lock'].iloc[0]
        MFG_Policy=matching_row['MFG Policy'].iloc[0]
        KPI_Data_Updated=matching_row['KPI Data Updated'].iloc[0]
        KPI_Door_count=matching_row['KPI Door count'].iloc[0]
        TBL_Planning_VerticalReport__3_matching_row=TBL_Planning_VerticalReport__3.loc[TBL_Planning_VerticalReport__3['PID'].str.upper() == pid_value]
        #STD_Sales = TBL_Planning_VerticalReport__3_matching_row['STD SALES'].iloc[0]
        #LY_STD_SALES = TBL_Planning_VerticalReport__3_matching_row['LY STD SALES'].iloc[0] 
        OOS_Locs=matching_row['OOS Locs'].iloc[0]
        Suspended_Loc_Count=matching_row['Suspended Loc Count'].iloc[0]
        Masterstyle_Desc=matching_row['Masterstyle Desc'].iloc[0]
        Dpt_ID=matching_row['Dpt ID'].iloc[0]  
        Dpt_Desc=matching_row['Dpt Desc'].iloc[0]
        SC_ID=matching_row['SC ID'].iloc[0]
        SC_Desc=matching_row['SC Desc'].iloc[0]
        MstrSt_ID=matching_row['MstrSt ID'].iloc[0]
        Masterstyle_Desc=matching_row['Masterstyle Desc'].iloc[0]
        PID_Desc=matching_row['PID Desc'].iloc[0]
        st_Live=matching_row['1st Live'].iloc[0]
        Live_Site=matching_row['Live Site'].iloc[0]
        V2C=matching_row['V2C'].iloc[0]
        Mktg_ID=matching_row['Mktg ID'].iloc[0]
        STD_Store_Rtn=matching_row['STD Store Rtn %'].iloc[0]
        Prod_Desc=matching_row['Prod Desc'].iloc[0]
        Last_Proj_Review_Date=matching_row['Last Proj Review Date'].iloc[0]
        Macys_Recpts_matching_row=Macys_Recpts.loc[Macys_Recpts['PID'].str.upper() == pid_value]
        Macys_Spring_Proj_Notes =  f"Macy's Spring Proj Notes: {Macys_Recpts_matching_row['ACTION'].iloc[0]}" if not Macys_Recpts_matching_row.empty else "Macy's Spring Proj Notes: "
        Planner_Response=matching_row['Planner Response'].iloc[0] 

        Nav_Feb=matching_row['Feb'].iloc[0]
        Nav_Mar=matching_row['Mar'].iloc[0]
        Nav_Apr=matching_row['Apr'].iloc[0]
        Nav_May=matching_row['May'].iloc[0]
        Nav_Jun=matching_row['Jun'].iloc[0]
        Nav_Jul=matching_row['Jul'].iloc[0]
        Nav_Aug=matching_row['Aug'].iloc[0]
        Nav_Sep=matching_row['Sep'].iloc[0]
        Nav_Oct=matching_row['Oct'].iloc[0]
        Nav_Nov=matching_row['Nov'].iloc[0]
        Nav_Dec=matching_row['Dec'].iloc[0]
        Nav_Jan=matching_row['Jan'].iloc[0]

        Macys_Proj_Receipts_Feb=matching_row['FEB RECPT'].iloc[0]
        Macys_Proj_Receipts_Mar=matching_row['MAR RECPT'].iloc[0]
        Macys_Proj_Receipts_Apr=matching_row['APR RECPT'].iloc[0]
        Macys_Proj_Receipts_May=matching_row['May RECPT'].iloc[0]
        Macys_Proj_Receipts_Jun=matching_row['JUN RECPT'].iloc[0]
        Macys_Proj_Receipts_Jul=matching_row['JUL RECPT'].iloc[0]
        Macys_Proj_Receipts_Aug=matching_row['AUG RECPT'].iloc[0]
        Macys_Proj_Receipts_Sep=matching_row['SEP RECPT'].iloc[0]
        Macys_Proj_Receipts_oct=matching_row['OCT RECPT'].iloc[0]
        Macys_Proj_Receipts_Nov=matching_row['NOV RECPT'].iloc[0]
        Macys_Proj_Receipts_Dec=matching_row['DEC RECPT'].iloc[0]
        Macys_Proj_Receipts_Jan=matching_row['JAN RECPT'].iloc[0]
        MCOM_Data_matching_row=MCOM_Data.loc[MCOM_Data['PID'].str.upper() == pid_value]
        if current_month  in ['Feb','Mar','Apr'] :
            this_year_value=year_of_previous_month-1
            last_year_value=last_year_of_previous_month-1
        else:
            this_year_value=year_of_previous_month
            last_year_value=last_year_of_previous_month
        this_year_data = All_DATA.loc[(All_DATA['PID'] == pid_value) & (All_DATA['Year'] == this_year_value)]
        last_year_data = All_DATA.loc[(All_DATA['PID'] == pid_value) & (All_DATA['Year'] == last_year_value)]
        this_year_MCOM=MCOM_Data.loc[(MCOM_Data['PID'] == pid_value) & (MCOM_Data['Year'] == this_year_value)]
        last_year_MCOM=MCOM_Data.loc[(MCOM_Data['PID'] == pid_value) & (MCOM_Data['Year'] == last_year_value)]
        # Define months in order
        months = ['JAN', 'FEB', 'MAR', 'APR', 'MAY', 'JUN', 'JUL', 'AUG', 'SEP', 'OCT', 'NOV', 'DEC']
        # Initialize dictionaries to store results
        TY_Unit_Sales = {month: 0 for month in months}
        LY_Unit_Sales = {month: 0 for month in months}
        LY_OH_Units = {month: 0 for month in months}
        TY_OH_Units = {month: 0 for month in months}
        TY_Receipts = {month: 0 for month in months}
        TY_MCOM_Unit_Sales = {month: 0 for month in months}
        TY_OH_MCOM_Units={month: 0 for month in months}
        PTD_TY_Sales={month: 0 for month in months}
        MCOM_PTD_TY_Sales={month: 0 for month in months}
        LY_MCOM_Unit_Sales={month: 0 for month in months}
        LY_MCOM_OH_Units = {month: 0 for month in months}
        OO_Total_Units={month: 0 for month in months}
        OO_MCOM_Total_Units={month: 0 for month in months}
        LY_Receipts={month: 0 for month in months}
        LY_PTD_Sales={month: 0 for month in months}
        MCOM_PTD_LY_Sales={month: 0 for month in months}
        # Sum data for each month
        for month in months:
            TY_Unit_Sales[month] = this_year_data.loc[this_year_data['Month'].str.upper() == month, 'PTD TY Unit Sales'].sum()
            LY_Unit_Sales[month] = last_year_data.loc[last_year_data['Month'].str.upper() == month, 'PTD TY Unit Sales'].sum()
            LY_OH_Units[month] = last_year_data.loc[last_year_data['Month'].str.upper() == month, 'OH TY Units'].sum()
            TY_OH_Units[month] = this_year_data.loc[this_year_data['Month'].str.upper() == month, 'OH TY Units'].sum()
            TY_Receipts[month] = this_year_data.loc[this_year_data['Month'].str.upper() == month, 'PTD TY RCVD Unit'].sum()
            TY_MCOM_Unit_Sales[month] = this_year_MCOM.loc[this_year_MCOM['Month'].str.upper() == month, 'PTD TY Unit Sales'].sum()
            LY_MCOM_Unit_Sales[month] = last_year_MCOM.loc[last_year_MCOM['Month'].str.upper() == month, 'PTD TY Unit Sales'].sum()
            TY_OH_MCOM_Units[month] = this_year_MCOM.loc[this_year_MCOM['Month'].str.upper() == month, 'OH TY Units'].sum()
            PTD_TY_Sales[month] = this_year_data.loc[this_year_data['Month'].str.upper() == month, 'PTD TY $ Sales'].sum()
            LY_PTD_Sales[month] = last_year_data.loc[last_year_data['Month'].str.upper() == month, 'PTD TY $ Sales'].sum()
            MCOM_PTD_TY_Sales[month] = this_year_MCOM.loc[this_year_MCOM['Month'].str.upper() == month, 'PTD TY $ Sales'].sum()
            MCOM_PTD_LY_Sales[month] = last_year_MCOM.loc[last_year_MCOM['Month'].str.upper() == month, 'PTD TY $ Sales'].sum()
            LY_MCOM_OH_Units[month] = last_year_MCOM.loc[last_year_MCOM['Month'].str.upper() == month, 'OH TY Units'].sum()
            OO_Total_Units[month] = this_year_data.loc[this_year_data['Month'].str.upper() == month, 'OO Total Units'].sum()
            OO_MCOM_Total_Units[month] = this_year_MCOM.loc[this_year_MCOM['Month'].str.upper() == month, 'OO Total Units'].sum()
            LY_Receipts[month] = last_year_data.loc[last_year_data['Month'].str.upper() == month, 'PTD TY RCVD Unit'].sum()
        index_df.columns = index_df.columns.str.strip().str.upper()
        logging.info("Current_FC_Index : %s ",Current_FC_Index)
        if pd.isna(Current_FC_Index):
            Current_FC_Index = "Dia"

        index_row_data = index_df.loc[index_df['INDEX'].astype(str).str.lower() == Current_FC_Index.lower()]
        print("index_row_data",index_row_data)	
        index_value = {}
        # Loop through each month and fetch its value
        for month in months:
            index_value[month] = index_row_data[month].iloc[0] if not index_row_data.empty else 0
        print("index_value",index_value)	



                        #algorithm
        #######step1####################
        # Function to get Lead Time by PID
        def get_lead_time_by_pid(pid_value,current_date):
            # Find the Vendor for the given PID
            vendor = master_sheet.loc[master_sheet['PID'] == pid_value, 'Vendor Name'].values[0]
            logging.info('Current Date  : %s ',current_date)
            logging.info('Vendor : %s',vendor)
            # Find the Lead Time for the Vendor
            country = vendor_sheet.loc[vendor_sheet['Vendor Name'] == vendor, 'Country of Origin']
            if not country.empty:
                country = country.values[0]  # Access the first value
            else:
                country = None  # Set to None if empty
            logging.info('Country : %s',country)
        

            lead_time = vendor_sheet.loc[vendor_sheet['Vendor Name'] == vendor, 'Lead Time(weeks)']
            lead_time = lead_time.values[0] if not lead_time.empty else 8


            if country == "South Africa" or country == "DRL":
                logging.info("Adding 2 weeks grace in Lead time as it is DRL and South africa")
                lead_time = lead_time + 2

            if np.isnan(lead_time):
                logging.info("Lead time is None So considering default lead time = 8")
                lead_time=8 

            logging.info("Lead time : %s",lead_time)

            lead_time_days = lead_time * 7
        
            # Calculate the forecast date
            forecast_date = current_date + timedelta(days=lead_time_days)
            logging.info("Forecast date : %s",forecast_date)
            currentdate = current_date.strftime("%Y-%m-%d")
            forecast_lead_time_date = forecast_date.strftime("%Y-%m-%d")
            holiday_start_date = "2025-01-22"
            holiday_end_date = "2025-02-05"
        
            if country=="China" and currentdate <= "2025-02-05" and forecast_lead_time_date >= "2025-01-22":
                logging.info("Considering china holiday for leadtime(Take lead time = 11)")
                lead_time = 11
            if country=="Italy" and currentdate <= "2025-08-31" and forecast_lead_time_date >= "2025-08-01":
                logging.info("Considering italy holiday for leadtime(Take lead time = 14)")
                lead_time = 14
            
            
            # lead_time = dynamicLead(currentdate,forecast_lead_time_date,lead_time,country)

            return lead_time
            
        current_date = datetime.today()
        


        # current_date = datetime(2025,3,14)
        # print("current_date",current_date)

        lead_time = get_lead_time_by_pid(pid_value,current_date)
        logging.info('Final lead_time : %s',lead_time) 

            # Function to get the 3-letter month abbreviation based on the month number
        # Retail calendar month sequence
        retail_months = ["Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Jan"]
        
        # logging.info intermediate results for debugging
        def convert_month_to_abbr(full_month):
            # Define mapping of full month names to abbreviations
            month_mapping = {
                "January": "Jan",
                "February": "Feb",
                "March": "Mar",
                "April": "Apr",
                "May": "May",
                "June": "Jun",
                "July": "Jul",
                "August": "Aug",
                "September": "Sep",
                "October": "Oct",
                "November": "Nov",
                "December": "Dec"
            }
                
            # Return the corresponding abbreviation, or None if not found
            return month_mapping.get(full_month, None)
        

        def calculate_forecast_months(lead_time_weeks, current_date,current_month):
            # Calculate the lead time in days
            lead_time_days = lead_time_weeks * 7
            
            # Calculate the forecast date
            forecast_date = current_date + timedelta(days=lead_time_days)
            logging.info("Final forecast date : %s ",forecast_date)
            
            # Extract the forecast month and year
            forecast_month = forecast_date.strftime("%B")  # Full month name
            forecast_year = forecast_date.year
            forecast_month_abbr = convert_month_to_abbr(forecast_month)

            check = False
            # Calculate the week of the forecast month
            first_day_of_month = forecast_date.replace(day=1)
            days_diff = (forecast_date - first_day_of_month).days
            week_of_forecast_month = math.ceil((days_diff + 1) / 7)  # +1 to include the current day in the week count

            if week_of_forecast_month > 2 :
                check = True
            
            # Calculate the week of the current month
            first_day_of_current_month = current_date.replace(day=1)
            days_diff_current = (current_date - first_day_of_current_month).days
            week_of_current_month = math.ceil((days_diff_current + 1) / 7)

            # Define the month names in 3-letter format
            #month_names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
            
            # # Create the list of months
            # start = month_names.index(current_month) + 1
            # end = month_names.index(forecast_month_abbr) + 1
            forecast_month_list =  [forecast_month_abbr]

        
            # Display the result
            logging.info(f"Forecast month is {forecast_month} {forecast_year}") 
            logging.info(f"Week of the forecast month: Week {week_of_forecast_month}")
            logging.info(f"Current month is {current_date.strftime('%B')} {current_date.year}")
            logging.info(f"Week of the current month: Week {week_of_current_month}")
            logging.info(f"Forecast months list: {forecast_month_list}")

            return forecast_month_list , week_of_forecast_month , forecast_month_abbr , check




        forecast_months ,week_of_forecast_month ,forecast_month, check = calculate_forecast_months(lead_time,current_date,current_month)
        #forecast_months = calculate_forecast_months(lead_time,current_date)
        #logging.info('forecast_months',forecast_months)
        
        #############################logging.info(madhavee)
        retail_months = ["Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Jan"]
        #Step 2 :  Find STD period###################

        def get_std_months(current_month):            
            # Define Spring and Fall seasons
            spring_season = ["Feb", "Mar", "Apr", "May", "Jun", "Jul"]
            fall_season = ["Aug", "Sep", "Oct", "Nov", "Dec", "Jan"]
            
            # Get the index of the current month in the retail calendar
            current_index = retail_months.index(current_month)
            
            # Determine the season and the corresponding std period length
            if current_month in spring_season:
                logging.info("Current season is : Spring , Taking previous 4 months as STD Period")
                std_period = 4  # Last 5 months for Spring
            elif current_month in fall_season:
                logging.info("Current season is : Fall , Taking previous 4 months as STD Period")
                std_period = 4  # Last 3 months for Fall
            else:
                raise ValueError("Invalid month provided.")

            # Calculate the std months
            std_months = []
            for i in range(1, std_period + 1):
                std_index = (current_index - i) % len(retail_months)
                std_months.insert(0, retail_months[std_index])  # Insert at the beginning to maintain order
            
            return std_months
        
        def get_month_range(months, month_start, endMonth):
            # Find the start and end indices in the months list
            start_index = months.index(month_start)
            end_index = months.index(endMonth)
        
            # Handle the circular nature of the list
            if end_index < start_index:
                # Rolling over to the next year
                selected_months = months[start_index:] + months[:end_index + 1]
            else:
                # Normal range
                selected_months = months[start_index:end_index + 1]
        
            return selected_months
        
        months_new = ['FEB', 'MAR', 'APR', 'MAY', 'JUN', 'JUL', 'AUG', 'SEP', 'OCT', 'NOV', 'DEC', 'JAN']


        if month_from and month_to:
            first_month_std=convert_month_to_abbr(month_from).upper()
            last_mont_std=convert_month_to_abbr(month_to).upper()
            std_months = get_month_range(months_new, first_month_std, last_mont_std)
        else:
            std_months = get_std_months(current_month)
        
       
        
        # std_months = get_std_months(current_month)
        # std_months = ['Nov','Dec','Jan','Feb']
        logging.info(std_months)

        std_months_upper = [month.upper() for month in std_months]
        STD_TY_Unit_Sales_list = [TY_Unit_Sales[month] for month in std_months_upper]
        logging.info(f"This year unit sales for STD period : {STD_TY_Unit_Sales_list}")

        STD_LY_Unit_Sales_list=[LY_Unit_Sales[month] for month in std_months_upper]
        logging.info(f"Last  year unit sales for STD period : {STD_LY_Unit_Sales_list}")
        target_months={'FEB', 'MAR', 'APR'}
        if any(month in std_months_upper for month in target_months):
            ac_this_year_value=year_of_previous_month

            L_this_year_value=year_of_previous_month-1
            ac_last_year_value=last_year_of_previous_month
            L_last_year_value=last_year_of_previous_month-1
            ac_this_year_data = All_DATA.loc[(All_DATA['PID'] == pid_value) & (All_DATA['Year'] == ac_this_year_value)]
            L_this_year_data = All_DATA.loc[(All_DATA['PID'] == pid_value) & (All_DATA['Year'] == L_this_year_value)]
            ac_last_year_data = All_DATA.loc[(All_DATA['PID'] == pid_value) & (All_DATA['Year'] == ac_last_year_value)]
            L_last_year_data = All_DATA.loc[(All_DATA['PID'] == pid_value) & (All_DATA['Year'] == L_last_year_value)]
            Ac_TY_Unit_Sales = {month: 0 for month in months}
            L_TY_Unit_Sales = {month: 0 for month in months}
            Ac_LY_Unit_Sales = {month: 0 for month in months}
            L_LY_Unit_Sales = {month: 0 for month in months}
            for month in months:
                TY_Unit_Sales[month] = this_year_data.loc[this_year_data['Month'].str.upper() == month, 'PTD TY Unit Sales'].sum()
                Ac_TY_Unit_Sales[month] = ac_this_year_data.loc[ac_this_year_data['Month'].str.upper() == month, 'PTD TY Unit Sales'].sum()
                L_TY_Unit_Sales[month] = L_this_year_data.loc[L_this_year_data['Month'].str.upper() == month, 'PTD TY Unit Sales'].sum()
                Ac_LY_Unit_Sales[month] = ac_last_year_data.loc[ac_last_year_data['Month'].str.upper() == month, 'PTD TY Unit Sales'].sum()
                L_LY_Unit_Sales[month] = L_last_year_data.loc[L_last_year_data['Month'].str.upper() == month, 'PTD TY Unit Sales'].sum()


            STD_TY_Unit_Sales_list = [
            Ac_TY_Unit_Sales.get(month, L_TY_Unit_Sales.get(month, 0)) if month in ['FEB', 'MAR', 'APR']
            else L_TY_Unit_Sales.get(month, 0)
            for month in std_months_upper
        ]
            STD_LY_Unit_Sales_list=[
            Ac_LY_Unit_Sales.get(month, L_LY_Unit_Sales.get(month, 0)) if month in ['FEB', 'MAR', 'APR']
            else L_LY_Unit_Sales.get(month, 0)
            for month in std_months_upper
        ]

        # TY_Unit_Sales_list=[TY_Unit_Sales['FEB'],TY_Unit_Sales['MAR'],TY_Unit_Sales['APR'],TY_Unit_Sales['MAY'],TY_Unit_Sales['JUN'],TY_Unit_Sales['JUL'],TY_Unit_Sales['AUG'],TY_Unit_Sales['SEP'],TY_Unit_Sales['OCT'],TY_Unit_Sales['NOV'],TY_Unit_Sales['DEC'],TY_Unit_Sales['JAN']]
        STD_index_value=[round(index_value[month],2) for month in std_months_upper]
        logging.info(f"Index percentage for STD period : {STD_index_value}")

        STD_index_value = sum(STD_index_value)
        logging.info(f"Sum of index percentage for STD period  : {STD_index_value}")


        if STD_index_value:
            month_12_fc_index = round((sum(STD_TY_Unit_Sales_list) / (STD_index_value)),0)
            logging.info(f"Sum of index percentage is not 0 that's why 12th month forecast by index =  {month_12_fc_index}")
        else:
            month_12_fc_index = 0
            logging.info(f"Sum of index percentage is  0 that's why 12th month forecast by index =  {month_12_fc_index}")
            

        if sum(STD_LY_Unit_Sales_list) and sum(STD_TY_Unit_Sales_list):            
            std_trend = round(((sum(STD_TY_Unit_Sales_list) - sum(STD_LY_Unit_Sales_list)) / sum(STD_LY_Unit_Sales_list)),2)
            logging.info(f"Sum of this year unit sales and last year unit sales is not 0 that's why std trend =  {std_trend}")
        else:
            std_trend = 0
            logging.info(f"Sum of this year unit sales and last year unit sales is 0 that's why std trend =  {std_trend}")
            

        # Get the door count from cell C10
        retail_months_upper = [month.upper() for month in retail_months]

        def check_product_type( ):
            
            All_LY_Unit_Sales_list=[LY_Unit_Sales[month] for month in retail_months_upper]
            ALL_LY_MCOM_Unit_Sales=[LY_MCOM_Unit_Sales[month] for month in retail_months_upper]
            result_list = [
                (ly_mcom / ty_unit if ty_unit != 0 else 0)
                for ly_mcom, ty_unit in zip(ALL_LY_MCOM_Unit_Sales, All_LY_Unit_Sales_list)
            ]
            # Get the COM to TTL % Sales for this year (from I20 to T20)

            # Calculate the average COM to TTL % Sales for this year
            average = sum(result_list) / len(result_list) if result_list else 0 
            average_com_to_ttl_sales=average*100
            logging.info(f"Average COM to TTL % Last Year : {average_com_to_ttl_sales}")
            logging.info(f"KPI Door Count: {KPI_Door_count}")

            # Determine if the product is COM or Store based on door count and COM to TTL % Sales
            if KPI_Door_count is None or  KPI_Door_count <= 1 or average_com_to_ttl_sales > 65 or None:
                return True
            else:
                return False

        is_com_product = check_product_type()
                
        logging.info(f"Is COM Product : {is_com_product}")

        if is_com_product:
            com_products.append(pid_value)
        else:
            store_products.append(pid_value) 

            # #############################################
            # #############       step 5       ############      
            # #############################################

            STD_TY_OH_Units_list = [TY_OH_Units[month] for month in std_months_upper]
            STD_TY_OH_MCOM_Units_list=[TY_OH_MCOM_Units[month] for month in std_months_upper]

            this_year_std_store_eom_oh = [
                (ty_oh_unit - ty_oh_mcom )
                for ty_oh_unit,ty_oh_mcom  in zip(STD_TY_OH_Units_list, STD_TY_OH_MCOM_Units_list)
            ]
                


            STD_LY_OH_Units_list = [LY_OH_Units[month] for month in std_months_upper]
            STD_LY_OH_MCOM_Units_list=[LY_MCOM_OH_Units[month] for month in std_months_upper]

            last_year_std_store_eom_oh = [
                (ly_oh_unit - ly_oh_mcom )
                for ly_oh_unit,ly_oh_mcom  in zip(STD_LY_OH_Units_list, STD_LY_OH_MCOM_Units_list)
            ]

            if current_month  in ['Feb','Mar','Apr'] :
                this_year_value=year_of_previous_month
                last_year_value=last_year_of_previous_month
                this_year_data = All_DATA.loc[(All_DATA['PID'] == pid_value) & (All_DATA['Year'] == this_year_value)]
                last_year_data = All_DATA.loc[(All_DATA['PID'] == pid_value) & (All_DATA['Year'] == last_year_value)]
                this_year_MCOM=MCOM_Data.loc[(MCOM_Data['PID'] == pid_value) & (MCOM_Data['Year'] == this_year_value)]
                last_year_MCOM=MCOM_Data.loc[(MCOM_Data['PID'] == pid_value) & (MCOM_Data['Year'] == last_year_value)]
                # Define months in order
                months = ['JAN', 'FEB', 'MAR', 'APR', 'MAY', 'JUN', 'JUL', 'AUG', 'SEP', 'OCT', 'NOV', 'DEC']
                # Initialize dictionaries to store results
                TY_Unit_Sales = {month: 0 for month in months}
                LY_Unit_Sales = {month: 0 for month in months}
                LY_OH_Units = {month: 0 for month in months}
                TY_OH_Units = {month: 0 for month in months}
                TY_Receipts = {month: 0 for month in months}
                TY_MCOM_Unit_Sales = {month: 0 for month in months}
                TY_OH_MCOM_Units={month: 0 for month in months}
                PTD_TY_Sales={month: 0 for month in months}
                MCOM_PTD_TY_Sales={month: 0 for month in months}
                LY_MCOM_Unit_Sales={month: 0 for month in months}
                LY_MCOM_OH_Units = {month: 0 for month in months}
                OO_Total_Units={month: 0 for month in months}
                OO_MCOM_Total_Units={month: 0 for month in months}
                LY_Receipts={month: 0 for month in months}
                LY_PTD_Sales={month: 0 for month in months}
                MCOM_PTD_LY_Sales={month: 0 for month in months}
                # Sum data for each month
                for month in months:
                    TY_Unit_Sales[month] = this_year_data.loc[this_year_data['Month'].str.upper() == month, 'PTD TY Unit Sales'].sum()
                    LY_Unit_Sales[month] = last_year_data.loc[last_year_data['Month'].str.upper() == month, 'PTD TY Unit Sales'].sum()
                    LY_OH_Units[month] = last_year_data.loc[last_year_data['Month'].str.upper() == month, 'OH TY Units'].sum()
                    TY_OH_Units[month] = this_year_data.loc[this_year_data['Month'].str.upper() == month, 'OH TY Units'].sum()
                    TY_Receipts[month] = this_year_data.loc[this_year_data['Month'].str.upper() == month, 'PTD TY RCVD Unit'].sum()
                    TY_MCOM_Unit_Sales[month] = this_year_MCOM.loc[this_year_MCOM['Month'].str.upper() == month, 'PTD TY Unit Sales'].sum()
                    LY_MCOM_Unit_Sales[month] = last_year_MCOM.loc[last_year_MCOM['Month'].str.upper() == month, 'PTD TY Unit Sales'].sum()
                    TY_OH_MCOM_Units[month] = this_year_MCOM.loc[this_year_MCOM['Month'].str.upper() == month, 'OH TY Units'].sum()
                    PTD_TY_Sales[month] = this_year_data.loc[this_year_data['Month'].str.upper() == month, 'PTD TY $ Sales'].sum()
                    LY_PTD_Sales[month] = last_year_data.loc[last_year_data['Month'].str.upper() == month, 'PTD TY $ Sales'].sum()
                    MCOM_PTD_TY_Sales[month] = this_year_MCOM.loc[this_year_MCOM['Month'].str.upper() == month, 'PTD TY $ Sales'].sum()
                    MCOM_PTD_LY_Sales[month] = last_year_MCOM.loc[last_year_MCOM['Month'].str.upper() == month, 'PTD TY $ Sales'].sum()
                    LY_MCOM_OH_Units[month] = last_year_MCOM.loc[last_year_MCOM['Month'].str.upper() == month, 'OH TY Units'].sum()
                    OO_Total_Units[month] = this_year_data.loc[this_year_data['Month'].str.upper() == month, 'OO Total Units'].sum()
                    OO_MCOM_Total_Units[month] = this_year_MCOM.loc[this_year_MCOM['Month'].str.upper() == month, 'OO Total Units'].sum()
                    LY_Receipts[month] = last_year_data.loc[last_year_data['Month'].str.upper() == month, 'PTD TY RCVD Unit'].sum()
            
            
            spring_months = ['FEB', 'MAR', 'APR', 'MAY', 'JUN', 'JUL']
            season = "Spring" if forecast_months[0].upper() in spring_months else "Fall"
            def calculate_column_values(s1, k1,months,f8, row4_values, row17_values, row39_values):
                results = {}
                for month,row4, row17, row39 in zip(months,row4_values, row17_values, row39_values):
                    if s1 == 12 and k1 == 12:
                        result = round(row17 + row17 * f8, 0)
                    elif s1 == 12 and row4 < 7:
                        result = round(row39 + row39 * f8, 0)
                    elif s1 > 6 and row4 < 7:
                        result = round(row17 + row17 * f8, 0)
                    elif s1 < 7 and row4 > 6:
                        result = round(row39 + row39 * f8, 0)
                    elif s1 < 7 and row4 < 7:
                        result = round(row39 + row39 * f8, 0)
                    elif s1 > 6 and row4 > 6:
                        result = round(row39 + row39 * f8, 0)
                    else:
                        result = None  # Fallback case
            
                    results[month] = result
            
                return results
            s1=last_month_of_previous_month_numeric
            k1=current_month_number
            f8=std_trend
            row4_values=[i+1 for i in range(12)]
            row17_values=[TY_Unit_Sales[month] for month in retail_months_upper]
            logging.info(f"This year unit sales : {row17_values}")


            row39_values=[LY_Unit_Sales[month] for month in retail_months_upper]
            logging.info(f"Last year unit sales : {row39_values}")


            fc_by_trend_all = calculate_column_values(s1, k1, retail_months_upper, f8, row4_values, row17_values, row39_values)
            logging.info(f"Forecast by trend all month: {fc_by_trend_all}")

            spring_fc_by_trend_all = sum([fc_by_trend_all[month] for month in ['FEB', 'MAR', 'APR', 'MAY', 'JUN', 'JUL']])
            logging.info(f"Forecast by trend spring sum: {spring_fc_by_trend_all}")

            fall_fc_by_trend_all = sum([fc_by_trend_all[month] for month in ['AUG', 'SEP', 'OCT', 'NOV', 'DEC', 'JAN']])
            logging.info(f"Forecast by trend fall sum: {fall_fc_by_trend_all}")
            
            all_index_value=[index_value[month] for month in retail_months_upper]
            logging.info(f"ALl month index percentage : {all_index_value}")

            fc_by_index_all={}

            for i,month in enumerate(retail_months_upper):
                fc_by_index_all[month] = round((all_index_value[i]*month_12_fc_index),0)

            logging.info(f"Forecast by index all month: {fc_by_index_all}")

            spring_fc_by_index_all = sum([fc_by_index_all[month] for month in ['FEB', 'MAR', 'APR', 'MAY', 'JUN', 'JUL']])
            logging.info(f"Forecast by index spring sum: {spring_fc_by_index_all}")

            fall_fc_by_index_all = sum([fc_by_index_all[month] for month in ['AUG', 'SEP', 'OCT', 'NOV', 'DEC', 'JAN']])
            logging.info(f"Forecast by index fall sum: {fall_fc_by_index_all}")

            spring_months = ["Feb", "Mar", "Apr", "May", "Jun", "Jul"]
            fall_months   = ["Aug", "Sep", "Oct", "Nov", "Dec", "Jan"]


            # 3. Count how many forecast months belong to Spring vs. Fall
            spring_count = sum(month in spring_months for month in forecast_months)
            fall_count   = sum(month in fall_months   for month in forecast_months)

            # 4. Determine which season is in the majority
            if spring_count > fall_count:
                selected_season = "Spring"
            elif fall_count > spring_count:
                selected_season = "Fall"
            else:
                selected_season = season

            fc_by_index = spring_fc_by_index_all if selected_season == "Spring" else fall_fc_by_index_all
            fc_by_trend = spring_fc_by_trend_all if selected_season == "Spring" else fall_fc_by_trend_all

            logging.info(f'Selected_season : {selected_season}')
            logging.info(f'{selected_season} fc_by_index : {fc_by_index}')
            logging.info(f'{selected_season} fc_by_trend : {fc_by_trend}')

            if fc_by_index is not None and fc_by_trend is not None:
                difference = (abs(fc_by_trend - fc_by_index) / max(fc_by_index,fc_by_trend))* 100 
            
            logging.info(f'FC by index and trend Difference : {difference}')
            


            def find_forecasting_method_index(difference, door_count, this_year_std_store_eom_oh, month_12_fc_index):
                # Check if the difference is more than 15%
                if difference > 15:            
                    average_value = sum(this_year_std_store_eom_oh) / len(this_year_std_store_eom_oh)
                    logging.info(f'Average value for store EOM OH for STD period : {average_value}')
                    loss = (door_count / average_value) - 1
                    logging.info(f"Loss business percentage : {loss}")

                    rank =Item_Code
                    logging.info(f'Rank : {rank}')
                    logging.info(f'Own Retail : {Own_Retail}')

                    if Own_Retail < 1000:
                        if rank == 'A' or rank == 'B':
                            loss_percent = min(loss, 0.45)  # Maximum of 45% for Rank A or B
                            logging.info(f'Final loss percentage for rank A and B and price is less than 1000 : {loss_percent}')
                        elif rank == 'C':
                            loss_percent = min(loss, 0.15)  # Maximum of 15% for Rank C
                            logging.info(f'Final loss percentage for rank C and price is less than 1000 : {loss_percent}')
                        else:
                            loss_percent = min(loss,0.10)  # No loss for Rank D or E
                            logging.info(f'Final loss percentage for other rank and price is less than 1000 : {loss_percent}')
                    else:
                        loss_percent = min(loss, 0.15)
                        logging.info(f'Final loss percentage for price is greater than 1000 : {loss_percent}')



                    month_12_fc_index = round(month_12_fc_index * (1 + loss_percent))
                    logging.info(f'Updated 12 month FC by index : {month_12_fc_index}')

                    forecasting_method = "FC by Index"  # Write the value to the merged cell
                    logging.info(f'Forecasting method : {forecasting_method} as index and trend difference is higher than 15')
                else:

                    forecasting_method = "Average"  # Write the value to the merged cell
                    logging.info(f'Forecasting method : {forecasting_method} as index and trend difference is less than 15')
                return  forecasting_method  ,  month_12_fc_index



            def find_forecasting_method_trend(difference,month_12_fc_index):
                # Check if the difference is more than 15%
                if difference > 15:                
                    forecasting_method = "FC by Trend"  # Write the value to the merged cell
                    logging.info(f'Forecasting method : {forecasting_method} as index and trend difference > 15')
        
                else:                        
                    forecasting_method = "Average"  # Write the value to the merged cell
                    logging.info(f'Forecasting method : {forecasting_method} as index and trend difference > 15')
                return  forecasting_method,month_12_fc_index
            
            def find_method(difference,this_year_std_store_eom_oh, last_year_std_store_eom_oh, door_count, month_12_fc_index):
                # Define the threshold percentage for "maintained"
                threshold = 0.95  # 5% difference
                # Function to calculate the average of a list and check if it's maintained
                def is_maintained(eom_oh_list):
                    average_eom_oh = sum(eom_oh_list) / len(eom_oh_list) if eom_oh_list else 0
                    # Check if the average is within the 5% down range or more than the door count
                    return (average_eom_oh >= threshold * door_count) or (average_eom_oh > door_count) 
        
                # Check if this year's and last year's Store EOM OH are maintained
                logging.info(f"Last year store EOM OH for STD period : {last_year_std_store_eom_oh}")
                logging.info(f"This year store EOM OH for STD period : {this_year_std_store_eom_oh}")

                last_year_maintained = is_maintained(last_year_std_store_eom_oh)
                this_year_maintained = is_maintained(this_year_std_store_eom_oh)
        
                logging.info(f"Last year inventory is maintained ?  {last_year_maintained}")
                logging.info(f"This year inventory is maintained ?  {this_year_maintained}")

                # Now check the three conditions:
                if last_year_maintained and this_year_maintained:
                    # "Last Year => Store EOM OH not maintained, This Year => Store EOM OH not maintained"
                    logging.info(f"Both year inventory is maintained so Taking Trend or Average ")
                    forecasting_method,month_12_fc_index=find_forecasting_method_trend(difference,month_12_fc_index)
                else:
                    logging.info(f"Last year or this year inventory is not maintained so Taking Index or Average ")
                    forecasting_method , month_12_fc_index=find_forecasting_method_index(difference, door_count, this_year_std_store_eom_oh, month_12_fc_index)
                return forecasting_method , month_12_fc_index

            forecasting_method , month_12_fc_index=find_method(difference,this_year_std_store_eom_oh, last_year_std_store_eom_oh, KPI_Door_count, month_12_fc_index)

            if forecasting_method == "FC by Trend":
                if std_trend < -40 :
                    std_trend = 0
                    f8 = 0
                    logging.info(f"Both year inventory is maintained and also trend is very less so std trend = 0")
                    fc_by_trend_all = calculate_column_values(s1, k1, retail_months_upper, f8, row4_values, row17_values, row39_values)
                    logging.info(f"Updated FC by trend value : {fc_by_trend_all}")
                if difference > 20 :
                    if fc_by_trend > fc_by_index and Own_Retail > 1000:
                        std_trend = 0
                        f8 = 0
                        logging.info(f"Both year inventory is maintained and also trend and index difference > 20 , SO considering std trend = 0")
                        fc_by_trend_all = calculate_column_values(s1, k1, retail_months_upper, f8, row4_values, row17_values, row39_values)
                        logging.info(f"Updated FC by trend value : {fc_by_trend_all}")



            # Preprocess the dictionaries to replace NaN or Infinity with 0
            fc_by_index_all = {key: (value if not (math.isnan(value) or math.isinf(value)) else 0) for key, value in fc_by_index_all.items()}
            fc_by_trend_all = {key: (value if not (math.isnan(value) or math.isinf(value)) else 0) for key, value in fc_by_trend_all.items()}
            
            # Compute the average
            fc_by_average_all = {
                key: round((fc_by_index_all[key] + fc_by_trend_all[key]) / 2) for key in fc_by_trend_all
            }
            logging.info(f"Final FC by trend value for all month : {fc_by_trend_all}")
            logging.info(f"Final FC by index value for all month : {fc_by_index_all}")
            logging.info(f"Final FC by average value for all month : {fc_by_average_all}")

            if forecasting_method == "FC by Index":
                recommended_fc = fc_by_index_all
            elif forecasting_method == "FC by Trend":
                recommended_fc = fc_by_trend_all
            else:
                recommended_fc = fc_by_average_all
            logging.info(f"Recommended_fc : {recommended_fc}")


            current_month_upper = current_month.upper()
                    # Define input data for each row
            row_4 = row4_values  # Example values for row 4
            row_9 = recommended_fc
            row_17 = TY_Unit_Sales
            row_43 = LY_OH_Units
            # Define variables
            V1 = rolling_method  
            logging.info(f'Rolling_method : {rolling_method}') # Example: could be "YTD", "CURRENT MTH", "SPRING", etc.
            K1 = current_month_number       # A reference column value (e.g., month or index)
            # List of columns (iterable from I to T in your example)
            
            # Initialize results
            planned_fc = {}

            # Iterate through columns and calculate values
            for idx, col in enumerate(retail_months_upper):
                J4 = row_4[idx]  # Get value from row_4 list based on index
                J17 = row_17[col]
                J43 = row_43[col]
                J9 = row_9[col]
            
                if V1 == "YTD" and J4 < K1:
                    planned_fc[col] = J17
                elif V1 == "YTD" and J4 == K1:
                    planned_fc[col] = J17
                elif V1 == "CURRENT MTH" and J4 == K1:
                    planned_fc[col] = J17
                elif V1 == "SPRING" and J4 < 7:
                    planned_fc[col] = J17
                elif V1 == "FALL" and J4 > 6:
                    planned_fc[col] = J17
                elif V1 == "LY FALL" and J4 > 6:
                    planned_fc[col] = J43
                else:
                    planned_fc[col] = J9


            current_month_upper = current_month.upper()
            # logging.info the results dictionary
            logging.info(f"planned_fc: {planned_fc}")
            logging.info(f"This Year unit sales : {TY_Unit_Sales}")

            current_month_sales_percentage = 19
            logging.info(f"Current month sales percentage : {current_month_sales_percentage}")

            current_month_upper = current_month.upper()
            current_month_fc = round(TY_Unit_Sales[current_month_upper] / (current_month_sales_percentage/100))
            logging.info(f"Current month forecast : {current_month_fc}")

            planned_fc[current_month_upper] =  current_month_fc if current_month_fc > 0 else 0
            logging.info(f"planned_fc with current%_re : {planned_fc}")

            #compare
            forecast_months_upper = [month.upper() for month in forecast_months]

            currnt_month=current_month_upper
            D13 = OO  # Replace with your value for D13
            E19 = nav_OO  # Replace with your value for E19
            in_transit=0 if D13 - E19 < 0 else D13 - E19
            logging.info(f"in_transit qty : {in_transit}")

            # Calculate using the logic in the Excel formula
            planned_shp=[Nav_Feb,Nav_Mar,Nav_Apr,Nav_May,Nav_Jun,Nav_Jul,Nav_Aug,Nav_Sep,Nav_Oct,Nav_Nov,Nav_Dec,Nav_Jan]
            planned_shp={key:abs(planned_shp[i]) for i,key in enumerate(retail_months_upper)}

            logging.info(f"Planned shipments : {planned_shp}")

            if np.isnan(in_transit):
                in_transit = 0
            planned_shp[current_month_upper] += in_transit
            logging.info(f"Planned shipments after adding in transit : {planned_shp}")

            v1=rolling_method
            k1=current_month_number
            i4=row_4
            row_10=planned_fc
            row_11=planned_shp
            row17=TY_Unit_Sales
            TY_OH_Units['JAN'] = TY_OH_Units.pop('JAN')
            row_21=TY_OH_Units
            TY_Receipts['JAN'] = TY_Receipts.pop('JAN')
            row_37=TY_Receipts
            LY_OH_Units['JAN'] = LY_OH_Units.pop('JAN')
            row_43=LY_OH_Units

            row_17=TY_Unit_Sales
            
            # Create a dictionary for Birthstones and corresponding months based on the Birthstone sheet
            # Initialize output dictionary for Planned EOH (Cal)

            # Loop through months
            def calculate_row_12(v1, k1, i4, row_10, row_11, row_21, row_37, row_43, row_17, current_month):
                # Define the fixed column-to-month mapping
                month_order = ['FEB', 'MAR', 'APR', 'MAY', 'JUN', 'JUL', 'AUG', 'SEP', 'OCT', 'NOV', 'DEC', 'JAN']
                month_to_index = {month: idx + 1 for idx, month in enumerate(month_order)}  # FEB=1, ..., JAN=12
            
                # Rearrange month_order to start with the current_month
                start_idx = month_order.index(current_month)
                reordered_months = month_order[start_idx:] + month_order[:start_idx]
                logging.info(f"reordered_months: {reordered_months}")
                # Initialize the output dictionary for row_12
                plan_oh = {}
            
                for idx, month in enumerate(reordered_months):
                    # Determine the previous month (wrapping around if necessary)
                    prev_month = reordered_months[idx - 1] if idx > 0 else reordered_months[-1]
            
                    # Get the corresponding month number
                    col4 = month_to_index[month]  # Month number based on the fixed mapping
                    col10 = row_10.get(month, 0)
                    col11 = row_11.get(month, 0)
                    col21 = row_21.get(month, 0)
                    col37 = row_37.get(month, 0)
                    col43 = row_43.get(month, 0)
                    col17 = row_17.get(month, 0)
            
                    # Get the previous column's value
                    prev_col12 = plan_oh.get(prev_month, 0)
                    prev_col21 = row_21.get(prev_month, 0)
                    # logging.info("v1",v1)
                    # logging.info("k1",k1)
                    # logging.info("col4",col4)
                    # Apply the logic for calculation
                    if v1 == "Current MTH" and k1 == col4 and col4 == 1:
                        val = row_43['JAN'] + col11 + col37 - col10
                        # logging.info("condition1")
                    elif v1 == "Current MTH" and k1 == col4:
                        # logging.info("col21",col21)
                        # logging.info("col11",col11)
                        # logging.info("col10",col10)
                        # logging.info("col17",col17)
                        val = col21 + col11 - (col10 - col17)
                        # logging.info("condition2")
                    elif v1 == "Current MTH" and col4 == 1 and k1 > 1:
                        # logging.info("condition3")
                        val = plan_oh['JAN'] + col11 - col10
                    elif v1 == "YTD" and col4 < k1:
                        # logging.info("condition3")
                        val = col21
                    elif v1 == "Spring" and col4 < 7:
                        # logging.info("condition4")
                        val = col21
                    elif v1 == "Fall" and col4 > 6 and col4 < k1:
                        # logging.info("condition5")
                        val = col21
                    elif v1 == "Fall" and col4 == 1 and k1 > 1:
                        # logging.info("condition6")
                        val = plan_oh['JAN'] + col11 - col10
                    elif v1 == "Fall" and col4 > 6 and col4 == k1:
                        # logging.info("condition7")
                        val = prev_col21 + col37 + col11 - col10
                    elif v1 == "LY Fall" and col4 > 6:
                        # logging.info("condition8")
                        val = col43
                    elif col4 == 1 and v1 == "LY FALL":
                        # logging.info("condition9")
                        val = row_43['JAN'] + col11 + col11 - col10
                    elif col4 == k1 and col4 > 1:
                        # logging.info("condition10")
                        # logging.info("prev_col21",prev_col21)
                        # logging.info("col37",col37)
                        # logging.info("col11",col11)
                        # logging.info("col10",col10)
                        val = prev_col21 + col37 + col11 - col10
                    else:
                        # logging.info("condition11")
                        # logging.info("prev_col12",prev_col12)
                        # logging.info("col11",col11)
                        # logging.info("col10",col10)
                        val = prev_col12 + col11 - col10
            
                    # Store the calculated value in row_12 for the current month
                    plan_oh[month] = val
            
                return plan_oh  
            
            plan_oh = calculate_row_12(v1, k1, i4, row_10, row_11, row_21, row_37, row_43, row_17, current_month_upper)
            logging.info(f"Planned EOH : {plan_oh}")

            required_quantity_month_dict = {}

            # Check if the filtered DataFrame is not empty
            bsp_row = master_sheet[master_sheet['PID'] == pid_value]
            if not bsp_row.empty:
                bsp_or_not = bsp_row['BSP_or_not'].iloc[0]
                bsp_or_not = str(bsp_or_not).strip().upper() if bsp_or_not else None
                # Step 2: Determine BSP_status
                bsp_status = 'BSP' if bsp_or_not == 'BSP' else 'non_bsp'
                logging.info(f'Bsp_status : {bsp_status}')
                # Step 3: Find birthstone_month if BSP
                birthstone_month = None
                if bsp_status == 'BSP':
                    all_birthstone_products.append(pid_value)

                    category_bsp=bsp_row['category'].iloc[0]
                    birthstone = bsp_row['Birthstone'].iloc[0] if bsp_or_not == 'BSP' else None
                    logging.info(f'category_bsp : {category_bsp}')
                    logging.info(f'birthstone : {birthstone}')
                    birthstone = str(birthstone).strip() 
                    birthstone_sheet = birthstone_sheet.dropna(subset=['Birthstone'])
                    birthstone_sheet['Birthstone'] = birthstone_sheet['Birthstone'].astype(str)
                    birthstone_row = birthstone_sheet[birthstone_sheet['Birthstone'].str.upper() == birthstone.upper()]

                    if not birthstone_row.empty:
                        birthstone_month = birthstone_row['Month Name'].iloc[0]
                        logging.info(f'birthstone_month : {birthstone_month}')
                    if birthstone_month:
                        # Determine the previous month of the birthstone month
                        months_list = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
                        birthstone_month_abb = convert_month_to_abbr(birthstone_month)
                        previous_month = months_list[months_list.index(birthstone_month_abb) - 1]
                        # Check if the previous month is in the forecast months
                        for month in forecast_months:

                            logging.info(f"Processing forecast month: {month}")
                            if month == previous_month or month == 'Nov' :  # Apply special condition for November
                                if category == 'Studs':
                                    required_quantity = 3 * KPI_Door_count  # For Studs, use 3 per door count
                                elif category == 'Pendants':
                                    required_quantity = 2 * KPI_Door_count  # For Pendants, use 2 per door count
                                upcoming_birthstone_products.append(pid_value)
                            else:
                                required_quantity = KPI_Door_count  # If not the previous month or November, use door count only

                            # Store the result in the dictionary
                            required_quantity_month_dict[month] = required_quantity
                            logging.info(f"if bsp ,Updated dictionary: {required_quantity_month_dict}")
                    # Fallback: Populate the dictionary with `forecast_months` and `door_count`
            if not required_quantity_month_dict:
                logging.info(f"PID not found or no BSP logic executed. No adding extra required quantity")
                for month in forecast_months:
                    required_quantity_month_dict[month] = KPI_Door_count
            
            
            required_quantity_month_dict = {key.upper(): value for key, value in required_quantity_month_dict.items()}
            logging.info(f"Updated required_quantity_month_dict: {required_quantity_month_dict}")

            ly_com_eom_oh=[LY_MCOM_OH_Units[month] for month in ['FEB', 'MAR', 'APR', 'MAY', 'JUN', 'JUL','AUG', 'SEP', 'OCT', 'NOV', 'DEC', 'JAN']]
            logging.info(f'Last year avg COM EOM OH for forecast month : {ly_com_eom_oh}')
            average_com_eom_oh = round(((sum(ly_com_eom_oh)/len(ly_com_eom_oh))),0)
            logging.info(f"Extra value to add for average_eom_oh : {average_com_eom_oh}")
            if not average_com_eom_oh :
                logging.info(f"There is no average COM EOM OH so taking 2 as default value ")
                average_com_eom_oh = 2

            months_list = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

            # Convert the months in `forecast_months` to uppercase to match keys in `planned_fc`
            forecast_months_upper = [month.upper() for month in forecast_months]

            # Loop through each forecast month, calculate FLDC, and update the dictionary
            for month in forecast_months:    
                index = months_list.index(month)
                next_month_fldc = months_list[(index + 1) % len(months_list)]
                next_month_fldc=next_month_fldc.upper()
                # Calculate FLDC for the current month
                Calculate_FLDC = round((planned_fc[next_month_fldc]) / 2)
                logging.info(f"Adding FLDC for {month}: {Calculate_FLDC}")
                month=month.upper()
                # Update the required_quantity_month_dict
                required_quantity_month_dict[month] = (
                    required_quantity_month_dict.get(month, 0) + Calculate_FLDC + average_com_eom_oh
                )          
            logging.info(f"Updated final required_quantity_month_dict : {required_quantity_month_dict}")



            for i, month in enumerate(forecast_months_upper):
                # Step 1: Check if planned_oh for the month is less than the required quantity for that month
                required_quantity = required_quantity_month_dict.get(month, 0)  # Get the required quantity for the month

                logging.info(f"plan_oh for {month} : {plan_oh[month]}")
                logging.info(f"required_quantity for {month} : {required_quantity}")
                if plan_oh[month] < required_quantity:
                    # Calculate the difference between required quantity and planned OH
                    difference = required_quantity - plan_oh[month]     
                    # Update the gross_projection for that month with the calculated difference
                    planned_shp[month] += difference
                    plan_oh[month] += difference
                    pids_below_door_count_alert.append(pid_value)
                    logging.info(f"Updated gross projection for {month}: {planned_shp[month]}")
                else:
                    logging.info(f"No update needed for {month} as planned_oh is greater than required quantity.")
        
            plan_oh = calculate_row_12(v1, k1, i4, row_10, planned_shp, row_21, row_37, row_43, row_17, current_month_upper)
            logging.info(f'Updated plan_eoh : {plan_oh}')
            logging.info(f'week_of_forecast_month : {week_of_forecast_month}')

            if check == True:                    
                logging.info(f"forecast_month list: {forecast_months}")
                last_forecast_month = forecast_months[-1] 
                # Find the index of the last month
                last_index = months_list.index(last_forecast_month)                
                # Get the next month (wrap around using modulo)
                next_forecast_month = months_list[(last_index + 1) % len(months_list)]
                logging.info(f"last forecast month : {last_forecast_month}")
                last_forecast_month = last_forecast_month.upper()
                next_forecast_month = next_forecast_month.upper()
                logging.info(f"Next forecast month : {next_forecast_month}")
                plan_oh[next_forecast_month] = (planned_shp[next_forecast_month] + plan_oh[last_forecast_month]) - planned_fc[next_forecast_month]
                required_quantity_for_next_month = KPI_Door_count + average_com_eom_oh
                logging.info(f"Required_quantity_for_next_month : {required_quantity_for_next_month}")
                if plan_oh[next_forecast_month] < required_quantity_for_next_month:
                    difference = required_quantity_for_next_month - plan_oh[next_forecast_month] 
                    planned_shp[next_forecast_month] += difference
                    logging.info(f"Updated gross projection for {next_forecast_month}: {planned_shp[next_forecast_month] }")
                    pids_below_door_count_alert.append(pid_value)

            plan_oh = calculate_row_12(v1, k1, i4, row_10, planned_shp, row_21, row_37, row_43, row_17, current_month_upper)
            logging.info(f'Updated plan_eoh : {plan_oh}')

            def calculate_week_and_month(start_month_abbr, start_week, year, weeks_to_add):
                # Map abbreviated month names to their respective numbers
                month_map = {
                    "Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4,
                    "May": 5, "Jun": 6, "Jul": 7, "Aug": 8,
                    "Sep": 9, "Oct": 10, "Nov": 11, "Dec": 12
                }
                # Convert the abbreviated month to its numeric equivalent
                if start_month_abbr not in month_map:
                    return None,None
            
                start_month = month_map[start_month_abbr]
            
                # Define the start date as the first day of the given month and year
                start_date = datetime(year, start_month, 1)
            
                # Calculate the date corresponding to the given week in the start month
                days_to_start_week = (start_week - 1) * 7
                start_week_date = start_date + timedelta(days=days_to_start_week)
            
                # Add the specified number of weeks (convert weeks to days)
                target_date = start_week_date + timedelta(weeks=weeks_to_add)
            
                # Determine the target month and week
                target_month = target_date.month
                target_year = target_date.year
            
                # Find the week number relative to the month (using Sunday to Saturday week structure)
                first_day_of_target_month = datetime(target_year, target_month, 1)
                days_difference = (target_date - first_day_of_target_month).days
                target_week = days_difference // 7 + 1
            
                # Convert the month number to its abbreviation
                target_month_abbr = target_date.strftime("%b")
                
                return target_month_abbr,target_week
        
            year = 2025               # Year
            weeks_to_add = 4          # Add 4 weeks
            
            forecast_month_abbr = convert_month_to_abbr(forecast_month)
            # Calculate and logging.info the result
            target_month_abbr, target_week = calculate_week_and_month(forecast_month_abbr, week_of_forecast_month, year, weeks_to_add)
            #logging.info("next_month_holiday_check_week",target_month_abbr,target_week)

            def check_holiday(target_month_abbr, target_week):
                holidays = [
                    {'Holiday': 'valentine_day', 'Month': 'Feb', 'Day': 14, 'Week': 2},
                    {'Holiday': 'women_day', 'Month': 'Mar', 'Day': 8, 'Week': 2},
                    {'Holiday': 'father_day', 'Month': 'Jun', 'Day': 16, 'Week': 3},
                    {'Holiday': 'men_day', 'Month': 'Nov', 'Day': 19, 'Week': 3}
                ]

                df_holidays = pd.DataFrame(holidays)
                
                # Filter the dataframe based on the target month and week
                result = df_holidays[(df_holidays['Month'] == target_month_abbr) & (df_holidays['Week'] == target_week)]
                
                if not result.empty:
                    return result.iloc[0]['Holiday'],True
                else:
                    return None,False
                

            holiday_name,check_is_holiday = check_holiday(target_month_abbr, target_week)
            logging.info(f"The holiday in {target_month_abbr} week {target_week} is: {holiday_name}")
            first_forecast_month = forecast_months_upper[0]
            if check_is_holiday:                    
                if category == "Men's" and holiday_name in ["father_day", "men_day"]:
                    planned_shp[first_forecast_month] += (required_quantity_month_dict[first_forecast_month] * 1.15)
                elif category != "Men's" and holiday_name not in ["father_day", "men_day"]:
                    planned_shp[first_forecast_month] += (required_quantity_month_dict[first_forecast_month] * 1.15)
            


            plan_oh = calculate_row_12(v1, k1, i4, row_10, planned_shp, row_21, row_37, row_43, row_17, current_month_upper)


            logging.info(F"Min_order : {Min_order}")
            # logging.info final values for verification
            logging.info(f'Final planned_fc : {planned_fc}')
            logging.info(f'Final Planned shipments : {planned_shp}')
            logging.info(f'Final Planned EOH : {plan_oh}')


            def calculate_store_sale_thru(LY_Unit_Sales, LY_MCOM_Unit_Sales, LY_OH_Units, LY_MCOM_OH_Units):
                months = ['JAN', 'FEB', 'MAR', 'APR', 'MAY', 'JUN', 'JUL', 'AUG', 'SEP', 'OCT', 'NOV', 'DEC']
                store_sell_thru = {}
                for month in months:
                    numerator = LY_Unit_Sales[month] - LY_MCOM_Unit_Sales[month]
                    denominator = numerator + (LY_OH_Units[month] - LY_MCOM_OH_Units[month])
                    store_sell_thru[month] =numerator / denominator if denominator != 0 else 0
                return store_sell_thru

            store_sell_thru = calculate_store_sale_thru(LY_Unit_Sales, LY_MCOM_Unit_Sales, LY_OH_Units, LY_MCOM_OH_Units)
            # Compute the average
            average_store_sale_thru = round((sum(store_sell_thru.values()) / len(store_sell_thru)),2)

            logging.info(f"Store Sale Thru : {store_sell_thru}")
            logging.info(f"Average Store Sale Thru : {average_store_sale_thru}")

            threshold = 0.95  # 5% difference
            # Function to calculate the average of a list and check if it's maintained
            def is_maintained(eom_oh_list):
                average_eom_oh = sum(eom_oh_list) / len(eom_oh_list) if eom_oh_list else 0
                # Check if the average is within the 5% down range or more than the door count
                return (average_eom_oh >= threshold * KPI_Door_count) or (average_eom_oh > KPI_Door_count) 

            # Check if this year's and last year's Store EOM OH are maintained
            last_year_maintained = is_maintained(last_year_std_store_eom_oh)
            this_year_maintained = is_maintained(this_year_std_store_eom_oh)



            planned_shipments = planned_shp
            macys_proj_receipt=[Macys_Proj_Receipts_Feb,Macys_Proj_Receipts_Mar,Macys_Proj_Receipts_Apr,Macys_Proj_Receipts_May,Macys_Proj_Receipts_Jun,Macys_Proj_Receipts_Jul,Macys_Proj_Receipts_Aug,Macys_Proj_Receipts_Sep,Macys_Proj_Receipts_oct,Macys_Proj_Receipts_Nov,Macys_Proj_Receipts_Dec,Macys_Proj_Receipts_Jan]
            macys_proj_receipt={key:macys_proj_receipt[i] for i,key in enumerate(retail_months_upper)}
            omni_receipts=TY_Receipts
            spring_months_upper = ['FEB', 'MAR', 'APR', 'MAY', 'JUN', 'JUL']
            fall_months_upper = [ 'AUG', 'SEP', 'OCT', 'NOV', 'DEC','JAN']
            logging.info(f"Macys Proj Receipts : {macys_proj_receipt}")
            # 3) Write a helper function to sum a dictionary's values by season
            def season_sum(data_dict, season_months):
                return sum(data_dict.get(month, 0) for month in season_months)
            
            # 4) Calculate sums for each season
            # --- Spring ---
            planned_spring = season_sum(planned_shipments, spring_months_upper)
            macys_spring   = season_sum(macys_proj_receipt, spring_months_upper)
            
            # --- Fall ---
            planned_fall = season_sum(planned_shipments, fall_months_upper)
            macys_fall   = season_sum(macys_proj_receipt, fall_months_upper)

            fcm = forecast_months[0]
        fcm = fcm.upper()

        def sum_planned_shipments(months_upper, fcm, planned_shipments):
            if fcm not in months_upper:
                raise ValueError("FCM must be in the list of spring months")
            
            start_index = 0  # Always start from the first month in spring_months_upper
            end_index = months_upper.index(fcm) + 1  # Include the FCM month
            
            selected_months = months_upper[start_index:end_index]
            
            return sum(planned_shipments.get(month, 0) for month in selected_months)


        def distribute_units(total_demand, supplied_until, percentage, current_month, months):
            """
            Distributes the remaining units across the upcoming months to meet the target percentage.
        
            :param total_demand: Total demand from the customer
            :param supplied_until: Units already supplied until the current month
            :param percentage: Percentage of total demand to be met by the last month
            :param current_month: The current month from which the remaining units need to be allocated
            :param months: List of months in order
            :return: Dictionary with units allocated per month
            """
            target_supply = int(total_demand * (percentage / 100))  # Target supply based on percentage
            remaining_units = target_supply - supplied_until  # Units that need to be distributed
        
            # Identify the remaining months from the current month onward
            current_index = months.index(current_month)
            remaining_months = months[current_index:]  # Include the current month and onward
        
            # Initialize distribution
            distribution = {month: 0 for month in months}
        
            # Assign already supplied units to past months
            for month in months[:current_index]:
                distribution[month] = 0  # Past months remain unchanged
        
            # Distribute remaining units across upcoming months (starting from current month)
            if remaining_units > 0:
                units_per_month = remaining_units // len(remaining_months)
                extra_units = remaining_units % len(remaining_months)  # Handle remainder
        
                for i, month in enumerate(remaining_months):
                    distribution[month] = units_per_month + (1 if i < extra_units else 0)
        
            return distribution



        # Now check the three conditions:
        if average_store_sale_thru > 0.14 :
            percentage = 1.0
            logging.info(f"Average store sale thorugh > 0.14 So Percentage = 100 %")
        elif Own_Retail <= 1000 and average_store_sale_thru >= 0.10 :
            percentage = 1.0
            logging.info(f"Average store sale thorugh >= 0.10 and Price <= 1000 So percentage = 100 %")
        elif Own_Retail > 1000 and average_store_sale_thru >= 0.10:
            percentage = 0.75
            logging.info(f"Average store sale thorugh >= 0.10 and Price > 1000 So percentage = 75 %")
        elif Own_Retail <= 1000 and average_store_sale_thru >= 0.04 :
            percentage = 0.75
            logging.info(f"Average store sale thorugh > 0.04 and Price <= 1000 So percentage = 75 %")
        elif Own_Retail > 1000 and average_store_sale_thru >= 0.04:
            percentage = 0.65
            logging.info(f"Average store sale thorugh >= 0.04 and Price > 1000 So percentage = 65 %")
        else:
            percentage = 0.40
            logging.info(f"Other condition So percentage = 40 %")

        
        if season == "Spring":
            total_gross_projection_season_wise = macys_spring
            logging.info(f"Total_macys_proj_receipt_for_current_{season} : {total_gross_projection_season_wise}")
            month_lists = spring_months_upper
            planned_shp_upto_current_month = sum_planned_shipments(spring_months_upper, fcm, planned_shipments)
            logging.info(f"Sum of planned shipments up to {fcm} :  {planned_shp_upto_current_month}")
            logging.info(f"Remaining macys proj receipts")
            
        else:
            total_gross_projection_season_wise = macys_fall
            logging.info(f"Total_macys_proj_receipt_for_{season} : {total_gross_projection_season_wise}")
            month_lists = fall_months_upper
            planned_shp_upto_current_month = sum_planned_shipments(fall_months_upper, fcm, planned_shipments)
            logging.info(f"Sum of planned shipments up to {fcm} : {planned_shp_upto_current_month}")

        

        result = distribute_units(total_gross_projection_season_wise, planned_shp_upto_current_month, percentage*100, fcm, month_lists)
        logging.info(f"Spltiing macys proj receipt result : {result}")

        # planned_shipments={key: round(value) for key, value in planned_shipments.items()}
       
        if check==True:
            keys = list(planned_shipments.keys())  # Get all keys in a list
            if fcm in keys:  
                index = keys.index(fcm)  # Find the index of "fcm"
                if index + 1 < len(keys):  # Ensure there's a next key
                    next_key = keys[index + 1]  
                    if int(result[next_key]) > 0:
                        added_macys_proj_receipts_alert.append(pid_value)
                    logging.info(f"Added quantity for {next_key} according to macys proj receipts (Adding in Next month as forecast week is greater than 2 ): {planned_shipments[next_key]} + {result[next_key]}")
                    planned_shipments[next_key] += result[next_key]  # Update only the next key

        else:
            logging.info(f"Added quantity for {fcm} according to macys proj receipts : {planned_shipments[fcm]} +{result[fcm]}")
            planned_shipments[fcm] += result[fcm]
            if int(result[fcm]) > 0:
                added_macys_proj_receipts_alert.append(pid_value)


        plan_oh = calculate_row_12(v1, k1, i4, row_10, planned_shipments, row_21, row_37, row_43, row_17, current_month_upper)
        logging.info(f"Latest plan_eoh : {plan_oh}")
        logging.info(f"Latest planned_shp : {planned_shipments}")


        planned_spring = season_sum(planned_shipments, spring_months_upper)
        planned_fall = season_sum(planned_shipments, fall_months_upper)

        def round_to_nearest_five(value):
            return math.ceil(value / 5) * 5

        # 6) Compare Macys Projection vs. the Combined Total
        #    If Macys Projection is less than the combined total, logging.info a warning
        if macys_spring < planned_spring or macys_fall < planned_fall:
            notify_macys_alert.append(pid_value)

        total_planned_shipments = planned_spring + planned_fall
        total_gross_projection = sum([Nav_Feb,Nav_Mar,Nav_Apr,Nav_May,Nav_Jun,Nav_Jul,Nav_Aug,Nav_Sep,Nav_Oct,Nav_Nov,Nav_Dec,Nav_Jan])
        
        total_added_quantity = total_planned_shipments - total_gross_projection - in_transit

        logging.info(f"Total_added_quantity : {total_added_quantity}")

        if total_added_quantity < Min_order:
            logging.info(f"Min order : {Min_order}")
            logging.info(f"Total_added_quantity is less than min order")
            min_order_alert.append(pid_value)

            # Check if the product already exists
        ProductDetail.objects.update_or_create(
            product_id=pid_value,
            defaults={
                "product_description": safe_str(PID_Desc),

                # Main identifiers
                "blu": safe_str(RLJ),
                "mkst": safe_str(MKST),
                "currect_fc_index": safe_str(Current_FC_Index),

                # Classification fields
                "safe_non_safe": safe_str(Safe_Non_Safe),
                "item_code": safe_str(Item_Code),

                # Store information
                "current_door_count": safe_int(Door_Count),
                "last_store_count": safe_int(Last_Str_Cnt),
                "door_count_updated": parse_date(Door_count_Updated),
                "store_model": safe_int(Store_Model),
                "com_model": safe_int(Com_Model),

                # Inventory and forecast fields
                "holiday_build_fc": safe_int(Holiday_Bld_FC),
                "macys_onhand": safe_int(MCYOH),
                "oo": safe_int(OO),
                "in_transit": safe_int(nav_OO),
                "month_to_date_shipment": safe_int(MTD_SHIPMENTS),
                "lastweek_shipment": safe_int(LW_Shipments),
                "planned_weeks_of_stock": safe_int(Wks_of_Stock_OH),
                "weeks_of_projection": safe_int(Wks_of_on_Proj),
                "last_4weeks_shipment": safe_int(Last_3Wks_Ships),

                # Vendor information
                "vendor_name": safe_str(Vendor_Name),
                "min_order": safe_int(Min_order),

                # Projection fields
                "rl_total": safe_int(Proj),
                "net_projection": safe_int(Net_Proj),
                "unalloc_order": safe_int(Unalloc_Orders),

                # Distribution center fields
                "ma_bin": safe_int(RLJ_OH),
                "fldc": safe_int(FLDC),
                "wip_quantity": safe_int(WIP),

                # Status fields
                "md_status": safe_str(MD_Status_MZ1),
                "replanishment_flag": safe_str(Repl_Flag),
                "mcom_replanishment": safe_str(MCOM_RPL),
                "pool_stock": safe_int(Pool_stock),

                # Date fields
                "first_reciept_date": parse_date(st_Rec_Date),
                "last_reciept_date": parse_date(Last_Rec_Date),
                "item_age": safe_int(Item_Age),
                "first_live_date": parse_date(st_Live),

                # Cost and retail fields
                "this_year_last_cost": safe_float(TY_Last_Cost),
                "macys_owned_retail": safe_float(Own_Retail),
                "awr_first_ticket_retail": safe_float(AWR_1st_Tkt_Ret),

                # Policy and configuration fields
                "metal_lock": safe_float(Metal_Lock),
                "mfg_policy": safe_str(MFG_Policy),

                # KPI fields
                "kpi_data_updated": safe_str(KPI_Data_Updated),
                "kpi_door_count": safe_int(KPI_Door_count),

                # Location fields
                "out_of_stock_location": safe_int(OOS_Locs),
                "suspended_location_count": safe_int(Suspended_Loc_Count),
                "live_site": safe_str(Live_Site),

                # Product categorization fields
                "masterstyle_description": safe_str(Masterstyle_Desc),
                "masterstyle_id": safe_str(MstrSt_ID),

                "department_id": safe_int(Dpt_ID),
                "department_description": safe_str(Dpt_Desc),

                "subclass_id": safe_int(SC_ID),
                "subclass_decription": safe_str(SC_Desc),
                "webid_description": safe_str(Prod_Desc),

                # Marketing fields
                "v2c": safe_str(V2C),
                "marketing_id": safe_str(Mktg_ID),
                "std_store_return": safe_float(STD_Store_Rtn),

                # Planning fields
                "last_project_review_date": parse_date(Last_Proj_Review_Date),
                "macy_spring_projection_note": safe_str(Macys_Spring_Proj_Notes),
                "planner_response": safe_str(Planner_Response)
            }
        )

       

        productmain = ProductDetail.objects.get(product_id=pid_value)
        # Save data to DB

        MonthlyForecast.objects.update_or_create(
            product=productmain,
            year=current_year,
            defaults={}
            )
     
        current_year = datetime.now().year
        save_macys_projection_receipts(productmain, matching_row, current_year)
        save_monthly_forecasts(productmain, current_year, months, TY_Unit_Sales, LY_Unit_Sales, LY_OH_Units, TY_OH_Units, TY_Receipts, LY_Receipts, TY_MCOM_Unit_Sales, LY_MCOM_Unit_Sales, TY_OH_MCOM_Units, LY_MCOM_OH_Units, PTD_TY_Sales, LY_PTD_Sales, MCOM_PTD_TY_Sales, MCOM_PTD_LY_Sales, OO_Total_Units, OO_MCOM_Total_Units)

        print(f"Product {pid_value} saved successfully")	

        

        dynamic_formulas = {
            f"G{start_row + 34}": 1,
            f"F{start_row + 1}": f"=C{start_row + 1}",
            f"C{start_row}": pid_value,
            f"D{start_row}": RLJ,
            f"F{start_row}":MKST,
            f"O1":std_months_upper[0].upper()+"-"+std_months_upper[-1],
            f"C{start_row + 1}":Current_FC_Index,
            f"C{start_row + 2}": sum(STD_TY_Unit_Sales_list),
            f"D{start_row + 2}": sum(STD_LY_Unit_Sales_list),
            f"C{start_row + 3}": f"=IFERROR(ROUND((C{start_row + 2}-D{start_row + 2})/D{start_row + 2},2),0)",
            f"F{start_row + 3}": std_trend,
            f"E{start_row + 3}": "Chg Trend",
            f"D{start_row + 1}": "Change Index",
            f"E{start_row + 2}": STD_index_value,
            f"F{start_row + 2}": month_12_fc_index,  # Added IFERROR to prevent divide-by-zero issues
            f"F{start_row + 4}": forecasting_method,
            f"C{start_row + 4}":  Item_Status,
            f"C{start_row + 5}": Door_Count,
            f"D{start_row + 5}": Last_Str_Cnt,
            f"E{start_row + 5}": f"=ROUND(C{start_row + 5}-D{start_row + 5},2)",
            f"F{start_row + 5}": Door_count_Updated,
            f"C{start_row + 6}": Store_Model,
            f"D{start_row + 6}": Com_Model,
            f"E{start_row + 6}": f"=ROUND(C{start_row + 6}+D{start_row + 6},2)",
            f"G{start_row + 7}": Holiday_Bld_FC,
            f"C{start_row + 8}":MCYOH,
            f"D{start_row + 8}": OO,
            f"E{start_row + 8}": f"=IF(D{start_row + 8}-E{start_row + 14}<0,0,D{start_row + 8}-E{start_row + 14})",
            f"E{start_row + 14}": nav_OO,
            f"F{start_row + 8}": MTD_SHIPMENTS,
            f"G{start_row + 8}":LW_Shipments,
            f"C{start_row + 9}": Wks_of_Stock_OH,
            f"D{start_row + 9}":Wks_of_on_Proj,
            f"F{start_row + 9}": Last_3Wks_Ships,
            f"C{start_row + 12}": "=0",
            f"C{start_row + 13}": Vendor_Name,
            f"G{start_row + 13}": Min_order,
            f"C{start_row + 14}": Proj,
            f"D{start_row + 14}": Net_Proj,
            f"F{start_row + 14}": Unalloc_Orders,
            f"G{start_row + 14}": f"=IF(C{start_row + 5}>G{start_row + 5},(C{start_row + 8}+D{start_row + 8})-C{start_row + 5},IF(OR(E{start_row + 6}>C{start_row + 5},E{start_row + 6}>G{start_row + 5}),(C{start_row + 8}+D{start_row + 8})-E{start_row + 6},(C{start_row + 8}+D{start_row + 8})-G{start_row + 5}))",
            f"C{start_row + 15}": RLJ_OH,
            f"D{start_row + 15}": FLDC,
            f"E{start_row + 15}": WIP,
            f"C{start_row + 16}": f"=C{start_row + 14}+F{start_row + 14}-C{start_row + 15}-E{start_row + 15}",
            f"D{start_row + 16}": f"=IF(AND(F{start_row + 14}>=C{start_row + 16},C{start_row + 16}>0),\"Demand due to Unalloc Sales Orders\",IF(C{start_row + 16}<0,\"Excess OH Units\",IF(AND(C{start_row + 16}>0,F{start_row + 14}<C{start_row + 16}),\"Demand\",\"\")))",
            f"C{start_row + 17}": MD_Status_MZ1,
            f"D{start_row + 17}":Repl_Flag,
            f"E{start_row + 17}": MCOM_RPL,
            f"F{start_row + 17}": Pool_stock,
            f"C{start_row + 18}": st_Rec_Date,
            f"D{start_row + 18}": Last_Rec_Date,
            f"E{start_row + 18}": Item_Age,
            f"F{start_row + 18}": f"=IFERROR((NOW()-VALUE(C{start_row + 18}))/30,0)",
            f"C{start_row + 19}": TY_Last_Cost,
            f"D{start_row + 19}": Own_Retail,
            f"E{start_row + 19}": AWR_1st_Tkt_Ret,
            f"F{start_row + 19}": f"=(D{start_row + 19}-C{start_row + 19})/D{start_row + 19}",
            f"C{start_row + 20}": Metal_Lock,
            f"D{start_row + 20}":MFG_Policy,
            f"C{start_row + 21}": KPI_Data_Updated,
            f"C{start_row + 22}": KPI_Door_count,
            f"C{start_row + 23}": f"=C{start_row + 22}-C{start_row + 5}",
            f"C{start_row + 24}": OOS_Locs,
            f"C{start_row + 25}": Suspended_Loc_Count,
            f"D{start_row + 21}": f"=SUBSTITUTE(D{start_row},\"/\",\"\")",
            f"B{start_row + 26}": f"=HYPERLINK(\"http://www.macys.com/shop/product/\"&C{start_row + 32}&\"?ID=\"&E{start_row + 31},\"Click to View online\")",
            f"C{start_row + 27}": Dpt_ID,
            f"D{start_row + 27}": Dpt_Desc,
            f"C{start_row + 28}":SC_ID,
            f"D{start_row + 28}": SC_Desc,
            f"C{start_row + 29}": MstrSt_ID,
            f"D{start_row + 29}":Masterstyle_Desc,
            f"C{start_row + 30}":PID_Desc,
            f"C{start_row + 31}": st_Live,
            f"D{start_row + 31}": f"{Live_Site}/{V2C}",
            f"E{start_row + 31}": Mktg_ID,
            f"F{start_row + 31}": STD_Store_Rtn,
            f"C{start_row + 32}":Prod_Desc,
            f"C{start_row + 33}":Last_Proj_Review_Date,
            f"D{start_row + 35}": Macys_Spring_Proj_Notes,
            f"B{start_row + 39}": "\"Past Review Comments\"",
            f"B{start_row + 40}": Planner_Response,
            f"I{start_row + 1}": index_value['FEB'],
            f"J{start_row + 1}": index_value['MAR'],
            f"K{start_row + 1}": index_value['APR'],
            f"L{start_row + 1}": index_value['MAY'],
            f"M{start_row + 1}": index_value['JUN'],
            f"N{start_row + 1}": index_value['JUL'],
            f"O{start_row + 1}": index_value['AUG'],
            f"P{start_row + 1}": index_value['SEP'],
            f"Q{start_row + 1}": index_value['OCT'],
            f"R{start_row + 1}": index_value['NOV'],
            f"S{start_row + 1}": index_value['DEC'],
            f"T{start_row + 1}": index_value['JAN'],
            f"U{start_row + 1}": f"=SUM(I{start_row + 1}:T{start_row + 1})",
            f"V{start_row + 1}": f"=IFERROR(SUM(I{start_row + 1}:N{start_row + 1}),0)",
            f"W{start_row + 1}": f"=IFERROR(SUM(O{start_row + 1}:T{start_row + 1}),0)",
            f"I{start_row + 2}": f"=ROUND($F{start_row + 2}*I{start_row + 1},0)",
            f"J{start_row + 2}": f"=ROUND($F{start_row + 2}*J{start_row + 1},0)",
            f"K{start_row + 2}": f"=ROUND($F{start_row + 2}*K{start_row + 1},0)",
            f"L{start_row + 2}": f"=ROUND($F{start_row + 2}*L{start_row + 1},0)",
            f"M{start_row + 2}": f"=ROUND($F{start_row + 2}*M{start_row + 1},0)",
            f"N{start_row + 2}": f"=ROUND($F{start_row + 2}*N{start_row + 1},0)",
            f"O{start_row + 2}": f"=ROUND($F{start_row + 2}*O{start_row + 1},0)",
            f"P{start_row + 2}": f"=ROUND($F{start_row + 2}*P{start_row + 1},0)",
            f"Q{start_row + 2}": f"=ROUND($F{start_row + 2}*Q{start_row + 1},0)",
            f"R{start_row + 2}": f"=ROUND($F{start_row + 2}*R{start_row + 1},0)",
            f"S{start_row + 2}": f"=ROUND($F{start_row + 2}*S{start_row + 1},0)",
            f"T{start_row + 2}": f"=ROUND($F{start_row + 2}*T{start_row + 1},0)",
            f"U{start_row + 2}": f"=SUM(I{start_row + 2}:T{start_row + 2})",
            f"V{start_row + 2}": f"=IFERROR(SUM(I{start_row + 2}:N{start_row + 2}),0)",
            f"W{start_row + 2}": f"=IFERROR(SUM(O{start_row + 2}:T{start_row + 2}),0)",

            f"I{start_row + 11}": "FEB",
            f"J{start_row + 11}": "MAR",
            f"K{start_row + 11}": "APR",
            f"L{start_row + 11}": "MAY",
            f"M{start_row + 11}": "JUN",
            f"N{start_row + 11}": "JUL",
            f"O{start_row + 11}": "AUG",
            f"P{start_row + 11}": "SEP",
            f"Q{start_row + 11}": "OCT",
            f"R{start_row + 11}": "NOV",
            f"S{start_row + 11}": "DEC",
            f"T{start_row + 11}": "JAN",
            f"U{start_row + 11}": "ANNUAL",
            f"V{start_row + 11}": "SPRING",
            f"W{start_row + 11}": "FALL",

            f"I{start_row + 33}": "FEB",
            f"J{start_row + 33}": "MAR",
            f"K{start_row + 33}": "APR",
            f"L{start_row + 33}": "MAY",
            f"M{start_row + 33}": "JUN",
            f"N{start_row + 33}": "JUL",
            f"O{start_row + 33}": "AUG",
            f"P{start_row + 33}": "SEP",
            f"Q{start_row + 33}": "OCT",
            f"R{start_row + 33}": "NOV",
            f"S{start_row + 33}": "DEC",
            f"T{start_row + 33}": "JAN",
            f"U{start_row + 33}": "ANNUAL",
            f"V{start_row + 33}": "SPRING",
            f"W{start_row + 33}": "FALL",


            # f"F{start_row + 2}": f"=IFERROR(C{start_row + 2}/E{start_row + 2},0)",
            f"F{start_row + 7}": f"=C{start_row + 14}+E{start_row + 8}-E{start_row + 7}",
            f"F{start_row + 10}": f"=MAX(0,C{start_row + 24}-F{start_row + 9})",
            f"D{start_row + 12}": f"=C{start_row + 12}*C{start_row + 19}",
            f"I{start_row + 12}": TY_Unit_Sales['FEB'],
            f"J{start_row + 12}":TY_Unit_Sales['MAR'],
            f"K{start_row + 12}": TY_Unit_Sales['APR'],
            f"L{start_row + 12}":TY_Unit_Sales['MAY'],
            f"M{start_row + 12}":TY_Unit_Sales['JUN'],
            f"N{start_row + 12}": TY_Unit_Sales['JUL'],
            f"O{start_row + 12}":TY_Unit_Sales['AUG'],
            f"P{start_row + 12}": TY_Unit_Sales['SEP'],
            f"Q{start_row + 12}": TY_Unit_Sales['OCT'],
            f"R{start_row + 12}":TY_Unit_Sales['NOV'],
            f"S{start_row + 12}": TY_Unit_Sales['DEC'],
            f"T{start_row + 12}": TY_Unit_Sales['JAN'],
            f"U{start_row + 12}": f"=IFERROR(SUM(I{start_row + 12}:T{start_row + 12}),0)",
            f"V{start_row + 12}": f"=IFERROR(SUM(I{start_row + 12}:N{start_row + 12}),0)",
            f"W{start_row + 12}": f"=IFERROR(SUM(O{start_row + 12}:T{start_row + 12}),0)" ,
            f"I{start_row + 34}": LY_Unit_Sales['FEB'],
            f"J{start_row + 34}":LY_Unit_Sales['MAR'],
            f"K{start_row + 34}":LY_Unit_Sales['APR'],
            f"L{start_row + 34}":LY_Unit_Sales['MAY'],
            f"M{start_row + 34}":LY_Unit_Sales['JUN'],
            f"N{start_row + 34}":LY_Unit_Sales['JUL'],
            f"O{start_row + 34}":LY_Unit_Sales['AUG'],
            f"P{start_row + 34}":LY_Unit_Sales['SEP'],
            f"Q{start_row + 34}":LY_Unit_Sales['OCT'],
            f"R{start_row + 34}":LY_Unit_Sales['NOV'],
            f"S{start_row + 34}":LY_Unit_Sales['DEC'],
            f"T{start_row + 34}":LY_Unit_Sales['JAN'],
            f"U{start_row + 34}": f"=SUM(I{start_row + 34}:T{start_row + 34})",
            f"V{start_row + 34}": f"=SUM(I{start_row + 34}:N{start_row + 34})",
            f"W{start_row + 34}": f"=SUM(O{start_row + 34}:T{start_row + 34})",
            f"I{start_row + 3}": f"=IF(AND($S$1=12,$K$1=12),ROUND(I{start_row + 12}+I{start_row + 12}*$F{start_row + 3},0),"
                                f"IF(AND($S$1=12,I$4<7),ROUND(I{start_row + 34}+I{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1>6,I$4<7),ROUND(I{start_row + 12}+I{start_row + 12}*$F{start_row + 3},0),"
                                f"IF(AND($S$1<7,I$4>6),ROUND(I{start_row + 34}+I{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1<7,I$4<7),ROUND(I{start_row + 34}+I{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1>6,I$4>6),ROUND(I{start_row + 34}+I{start_row + 34}*$F{start_row + 3},0)))))))",

            f"J{start_row + 3}": f"=IF(AND($S$1=12,$K$1=12),ROUND(J{start_row + 12}+J{start_row + 12}*$F{start_row + 3},0),"
                                f"IF(AND($S$1=12,J$4<7),ROUND(J{start_row + 34}+J{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1>6,J$4<7),ROUND(J{start_row + 12}+J{start_row + 12}*$F{start_row + 3},0),"
                                f"IF(AND($S$1<7,J$4>6),ROUND(J{start_row + 34}+J{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1<7,J$4<7),ROUND(J{start_row + 34}+J{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1>6,J$4>6),ROUND(J{start_row + 34}+J{start_row + 34}*$F{start_row + 3},0)))))))",

            f"L{start_row + 3}": f"=IF(AND($S$1=12,$K$1=12),ROUND(L{start_row + 12}+L{start_row + 12}*$F{start_row + 3},0),"
                                f"IF(AND($S$1=12,L$4<7),ROUND(L{start_row + 34}+L{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1>6,L$4<7),ROUND(L{start_row + 12}+L{start_row + 12}*$F{start_row + 3},0),"
                                f"IF(AND($S$1<7,L$4>6),ROUND(L{start_row + 34}+L{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1<7,L$4<7),ROUND(L{start_row + 34}+L{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1>6,L$4>6),ROUND(L{start_row + 34}+L{start_row + 34}*$F{start_row + 3},0)))))))",

            f"M{start_row + 3}": f"=IF(AND($S$1=12,$K$1=12),ROUND(M{start_row + 12}+M{start_row + 12}*$F{start_row + 3},0),"
                                f"IF(AND($S$1=12,M$4<7),ROUND(M{start_row + 34}+M{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1>6,M$4<7),ROUND(M{start_row + 12}+M{start_row + 12}*$F{start_row + 3},0),"
                                f"IF(AND($S$1<7,M$4>6),ROUND(M{start_row + 34}+M{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1<7,M$4<7),ROUND(M{start_row + 34}+M{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1>6,M$4>6),ROUND(M{start_row + 34}+M{start_row + 34}*$F{start_row + 3},0)))))))",

            f"N{start_row + 3}": f"=IF(AND($S$1=12,$K$1=12),ROUND(N{start_row + 12}+N{start_row + 12}*$F{start_row + 3},0),"
                                f"IF(AND($S$1=12,N$4<7),ROUND(N{start_row + 34}+N{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1>6,N$4<7),ROUND(N{start_row + 12}+N{start_row + 12}*$F{start_row + 3},0),"
                                f"IF(AND($S$1<7,N$4>6),ROUND(N{start_row + 34}+N{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1<7,N$4<7),ROUND(N{start_row + 34}+N{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1>6,N$4>6),ROUND(N{start_row + 34}+N{start_row + 34}*$F{start_row + 3},0)))))))",

            f"O{start_row + 3}": f"=IF(AND($S$1=12,$K$1=12),ROUND(O{start_row + 12}+O{start_row + 12}*$F{start_row + 3},0),"
                                f"IF(AND($S$1=12,O$4<7),ROUND(O{start_row + 34}+O{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1>6,O$4<7),ROUND(O{start_row + 12}+O{start_row + 12}*$F{start_row + 3},0),"
                                f"IF(AND($S$1<7,O$4>6),ROUND(O{start_row + 34}+O{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1<7,O$4<7),ROUND(O{start_row + 34}+O{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1>6,O$4>6),ROUND(O{start_row + 34}+O{start_row + 34}*$F{start_row + 3},0)))))))",

            f"P{start_row + 3}": f"=IF(AND($S$1=12,$K$1=12),ROUND(P{start_row + 12}+P{start_row + 12}*$F{start_row + 3},0),"
                                f"IF(AND($S$1=12,P$4<7),ROUND(P{start_row + 34}+P{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1>6,P$4<7),ROUND(P{start_row + 12}+P{start_row + 12}*$F{start_row + 3},0),"
                                f"IF(AND($S$1<7,P$4>6),ROUND(P{start_row + 34}+P{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1<7,P$4<7),ROUND(P{start_row + 34}+P{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1>6,P$4>6),ROUND(P{start_row + 34}+P{start_row + 34}*$F{start_row + 3},0)))))))",

            f"Q{start_row + 3}": f"=IF(AND($S$1=12,$K$1=12),ROUND(Q{start_row + 12}+Q{start_row + 12}*$F{start_row + 3},0),"
                                f"IF(AND($S$1=12,Q$4<7),ROUND(Q{start_row + 34}+Q{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1>6,Q$4<7),ROUND(Q{start_row + 12}+Q{start_row + 12}*$F{start_row + 3},0),"
                                f"IF(AND($S$1<7,Q$4>6),ROUND(Q{start_row + 34}+Q{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1<7,Q$4<7),ROUND(Q{start_row + 34}+Q{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1>6,Q$4>6),ROUND(Q{start_row + 34}+Q{start_row + 34}*$F{start_row + 3},0)))))))",

            f"R{start_row + 3}": f"=IF(AND($S$1=12,$K$1=12),ROUND(R{start_row + 12}+R{start_row + 12}*$F{start_row + 3},0),"
                                f"IF(AND($S$1=12,R$4<7),ROUND(R{start_row + 34}+R{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1>6,R$4<7),ROUND(R{start_row + 12}+R{start_row + 12}*$F{start_row + 3},0),"
                                f"IF(AND($S$1<7,R$4>6),ROUND(R{start_row + 34}+R{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1<7,R$4<7),ROUND(R{start_row + 34}+R{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1>6,R$4>6),ROUND(R{start_row + 34}+R{start_row + 34}*$F{start_row + 3},0)))))))",

            f"S{start_row + 3}": f"=IF(AND($S$1=12,$K$1=12),ROUND(S{start_row + 12}+S{start_row + 12}*$F{start_row + 3},0),"
                                f"IF(AND($S$1=12,S$4<7),ROUND(S{start_row + 34}+S{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1>6,S$4<7),ROUND(S{start_row + 12}+S{start_row + 12}*$F{start_row + 3},0),"
                                f"IF(AND($S$1<7,S$4>6),ROUND(S{start_row + 34}+S{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1<7,S$4<7),ROUND(S{start_row + 34}+S{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1>6,S$4>6),ROUND(S{start_row + 34}+S{start_row + 34}*$F{start_row + 3},0)))))))",

            f"T{start_row + 3}": f"=IF(AND($S$1=12,$K$1=12),ROUND(T{start_row + 12}+T{start_row + 12}*$F{start_row + 3},0),"
                                f"IF(AND($S$1=12,T$4<7),ROUND(T{start_row + 34}+T{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1>6,T$4<7),ROUND(T{start_row + 12}+T{start_row + 12}*$F{start_row + 3},0),"
                                f"IF(AND($S$1<7,T$4>6),ROUND(T{start_row + 34}+T{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1<7,T$4<7),ROUND(T{start_row + 34}+T{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1>6,T$4>6),ROUND(T{start_row + 34}+T{start_row + 34}*$F{start_row + 3},0)))))))",
            f"K{start_row + 3}": f"=IF(AND($S$1=12,$K$1=12),ROUND(K{start_row + 12}+K{start_row + 12}*$F{start_row + 3},0),"
                                f"IF(AND($S$1=12,K$4<7),ROUND(K{start_row + 34}+K{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1>6,K$4<7),ROUND(K{start_row + 12}+K{start_row + 12}*$F{start_row + 3},0),"
                                f"IF(AND($S$1<7,K$4>6),ROUND(K{start_row + 34}+K{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1<7,K$4<7),ROUND(K{start_row + 34}+K{start_row + 34}*$F{start_row + 3},0),"
                                f"IF(AND($S$1>6,K$4>6),ROUND(K{start_row + 34}+K{start_row + 34}*$F{start_row + 3},0)))))))",

            # Continue similarly for columns L through T
            f"U{start_row + 3}": f"=SUM(I{start_row + 3}:T{start_row + 3})",
            f"V{start_row + 3}": f"=IFERROR(SUM(I{start_row + 3}:N{start_row + 3}),0)",
            f"W{start_row + 3}": f"=IFERROR(SUM(O{start_row + 3}:T{start_row + 3}),0)",
            
            f"I{start_row + 4}": f"=IF($F{start_row + 4}=$H{start_row + 2},I{start_row + 2},"
                                f"IF($F{start_row + 4}=$H{start_row + 3},I{start_row + 3},"
                                f"IF($F{start_row + 4}=\"Current year\",I{start_row + 12},"
                                f"IF($F{start_row + 4}=\"Last year\",I{start_row + 34},"
                                f"ROUND(AVERAGE(I{start_row + 2}:I{start_row + 3}),0)))))",

            f"J{start_row + 4}": f"=IF($F{start_row + 4}=$H{start_row + 2},J{start_row + 2},"
                                f"IF($F{start_row + 4}=$H{start_row + 3},J{start_row + 3},"
                                f"IF($F{start_row + 4}=\"Current year\",J{start_row + 12},"
                                f"IF($F{start_row + 4}=\"Last year\",J{start_row + 34},"
                                f"ROUND(AVERAGE(J{start_row + 2}:J{start_row + 3}),0)))))",

            f"K{start_row + 4}": f"=IF($F{start_row + 4}=$H{start_row + 2},K{start_row + 2},"
                                f"IF($F{start_row + 4}=$H{start_row + 3},K{start_row + 3},"
                                f"IF($F{start_row + 4}=\"Current year\",K{start_row + 12},"
                                f"IF($F{start_row + 4}=\"Last year\",K{start_row + 34},"
                                f"ROUND(AVERAGE(K{start_row + 2}:K{start_row + 3}),0)))))",

            f"L{start_row + 4}": f"=IF($F{start_row + 4}=$H{start_row + 2},L{start_row + 2},"
                                f"IF($F{start_row + 4}=$H{start_row + 3},L{start_row + 3},"
                                f"IF($F{start_row + 4}=\"Current year\",L{start_row + 12},"
                                f"IF($F{start_row + 4}=\"Last year\",L{start_row + 34},"
                                f"ROUND(AVERAGE(L{start_row + 2}:L{start_row + 3}),0)))))",

            f"M{start_row + 4}": f"=IF($F{start_row + 4}=$H{start_row + 2},M{start_row + 2},"
                                f"IF($F{start_row + 4}=$H{start_row + 3},M{start_row + 3},"
                                f"IF($F{start_row + 4}=\"Current year\",M{start_row + 12},"
                                f"IF($F{start_row + 4}=\"Last year\",M{start_row + 34},"
                                f"ROUND(AVERAGE(M{start_row + 2}:M{start_row + 3}),0)))))",

            f"N{start_row + 4}": f"=IF($F{start_row + 4}=$H{start_row + 2},N{start_row + 2},"
                                f"IF($F{start_row + 4}=$H{start_row + 3},N{start_row + 3},"
                                f"IF($F{start_row + 4}=\"Current year\",N{start_row + 12},"
                                f"IF($F{start_row + 4}=\"Last year\",N{start_row + 34},"
                                f"ROUND(AVERAGE(N{start_row + 2}:N{start_row + 3}),0)))))",

            f"O{start_row + 4}": f"=IF($F{start_row + 4}=$H{start_row + 2},O{start_row + 2},"
                                f"IF($F{start_row + 4}=$H{start_row + 3},O{start_row + 3},"
                                f"IF($F{start_row + 4}=\"Current year\",O{start_row + 12},"
                                f"IF($F{start_row + 4}=\"Last year\",O{start_row + 34},"
                                f"ROUND(AVERAGE(O{start_row + 2}:O{start_row + 3}),0)))))",

            f"P{start_row + 4}": f"=IF($F{start_row + 4}=$H{start_row + 2},P{start_row + 2},"
                                f"IF($F{start_row + 4}=$H{start_row + 3},P{start_row + 3},"
                                f"IF($F{start_row + 4}=\"Current year\",P{start_row + 12},"
                                f"IF($F{start_row + 4}=\"Last year\",P{start_row + 34},"
                                f"ROUND(AVERAGE(P{start_row + 2}:P{start_row + 3}),0)))))",

            f"Q{start_row + 4}": f"=IF($F{start_row + 4}=$H{start_row + 2},Q{start_row + 2},"
                                f"IF($F{start_row + 4}=$H{start_row + 3},Q{start_row + 3},"
                                f"IF($F{start_row + 4}=\"Current year\",Q{start_row + 12},"
                                f"IF($F{start_row + 4}=\"Last year\",Q{start_row + 34},"
                                f"ROUND(AVERAGE(Q{start_row + 2}:Q{start_row + 3}),0)))))",

            f"R{start_row + 4}": f"=IF($F{start_row + 4}=$H{start_row + 2},R{start_row + 2},"
                                f"IF($F{start_row + 4}=$H{start_row + 3},R{start_row + 3},"
                                f"IF($F{start_row + 4}=\"Current year\",R{start_row + 12},"
                                f"IF($F{start_row + 4}=\"Last year\",R{start_row + 34},"
                                f"ROUND(AVERAGE(R{start_row + 2}:R{start_row + 3}),0)))))",

            f"S{start_row + 4}": f"=IF($F{start_row + 4}=$H{start_row + 2},S{start_row + 2},"
                                f"IF($F{start_row + 4}=$H{start_row + 3},S{start_row + 3},"
                                f"IF($F{start_row + 4}=\"Current year\",S{start_row + 12},"
                                f"IF($F{start_row + 4}=\"Last year\",S{start_row + 34},"
                                f"ROUND(AVERAGE(S{start_row + 2}:S{start_row + 3}),0)))))",

            f"T{start_row + 4}": f"=IF($F{start_row + 4}=$H{start_row + 2},T{start_row + 2},"
                                f"IF($F{start_row + 4}=$H{start_row + 3},T{start_row + 3},"
                                f"IF($F{start_row + 4}=\"Current year\",T{start_row + 12},"
                                f"IF($F{start_row + 4}=\"Last year\",T{start_row + 34},"
                                f"ROUND(AVERAGE(T{start_row + 2}:T{start_row + 3}),0)))))",

            f"U{start_row + 4}": f"=SUM(I{start_row + 4}:T{start_row + 4})",
            f"V{start_row + 4}": f"=IFERROR(SUM(I{start_row + 4}:N{start_row + 4}),0)",
            f"W{start_row + 4}": f"=IFERROR(SUM(O{start_row + 4}:T{start_row + 4}),0)",

            f"I{start_row + 38}": LY_OH_Units['FEB'],
            f"J{start_row + 38}": LY_OH_Units['MAR'],
            f"K{start_row + 38}": LY_OH_Units['APR'],
            f"L{start_row + 38}":LY_OH_Units['MAY'],
            f"M{start_row + 38}":LY_OH_Units['JUN'],
            f"N{start_row + 38}": LY_OH_Units['JUL'],
            f"O{start_row + 38}": LY_OH_Units['AUG'],
            f"P{start_row + 38}":LY_OH_Units['SEP'],
            f"Q{start_row + 38}":LY_OH_Units['OCT'],
            f"R{start_row + 38}": LY_OH_Units['NOV'],
            f"S{start_row + 38}":LY_OH_Units['DEC'],
            f"T{start_row + 38}": LY_OH_Units['JAN'],
            f"U{start_row + 38}": f"=AVERAGEIFS(I{start_row + 38}:T{start_row + 38}, $I$4:$T$4, \"<=\"&$T$4)",
            f"V{start_row + 38}": f"=AVERAGEIFS(I{start_row + 38}:N{start_row + 38}, $I$4:$N$4, \"<=\"&$N$4)",
            f"W{start_row + 38}": f"=AVERAGEIFS(O{start_row + 38}:T{start_row + 38}, $O$4:$T$4, \"<=\"&$T$4)",
            f"I{start_row + 5}": f"=IF(AND($V$1=\"YTD\",I$4<$K$1),I{start_row + 12},IF(AND($V$1=\"YTD\",I$4=$K$1),I{start_row + 12},"
                                f"IF(AND($V$1=\"CURRENT MTH\",I$4=$K$1),I{start_row + 12},IF(AND($V$1=\"SPRING\",I$4<7),I{start_row + 12},"
                                f"IF(AND($V$1=\"FALL\",I$4>6),I{start_row + 12},IF(AND($V$1=\"LY FALL\",I$4>6),I{start_row + 34},I{start_row + 4}))))))",

            f"J{start_row + 5}": f"=IF(AND($V$1=\"YTD\",J$4<$K$1),J{start_row + 12},IF(AND($V$1=\"YTD\",J$4=$K$1),J{start_row + 12},"
                                f"IF(AND($V$1=\"CURRENT MTH\",J$4=$K$1),J{start_row + 12},IF(AND($V$1=\"SPRING\",J$4<7),J{start_row + 12},"
                                f"IF(AND($V$1=\"FALL\",J$4>6),J{start_row + 12},IF(AND($V$1=\"LY FALL\",J$4>6),J{start_row + 34},J{start_row + 4}))))))",

            f"K{start_row + 5}": f"=IF(AND($V$1=\"YTD\",K$4<$K$1),K{start_row + 12},IF(AND($V$1=\"YTD\",K$4=$K$1),K{start_row + 12},"
                                f"IF(AND($V$1=\"CURRENT MTH\",K$4=$K$1),K{start_row + 12},IF(AND($V$1=\"SPRING\",K$4<7),K{start_row + 12},"
                                f"IF(AND($V$1=\"FALL\",K$4>6),K{start_row + 12},IF(AND($V$1=\"LY FALL\",K$4>6),K{start_row + 34},K{start_row + 4}))))))",

            f"L{start_row + 5}": f"=IF(AND($V$1=\"YTD\",L$4<$K$1),L{start_row + 12},IF(AND($V$1=\"YTD\",L$4=$K$1),L{start_row + 12},"
                                f"IF(AND($V$1=\"CURRENT MTH\",L$4=$K$1),L{start_row + 12},IF(AND($V$1=\"SPRING\",L$4<7),L{start_row + 12},"
                                f"IF(AND($V$1=\"FALL\",L$4>6),L{start_row + 12},IF(AND($V$1=\"LY FALL\",L$4>6),L{start_row + 34},L{start_row + 4}))))))",

            f"M{start_row + 5}": f"=IF(AND($V$1=\"YTD\",M$4<$K$1),M{start_row + 12},IF(AND($V$1=\"YTD\",M$4=$K$1),M{start_row + 12},"
                                f"IF(AND($V$1=\"CURRENT MTH\",M$4=$K$1),M{start_row + 12},IF(AND($V$1=\"SPRING\",M$4<7),M{start_row + 12},"
                                f"IF(AND($V$1=\"FALL\",M$4>6),M{start_row + 12},IF(AND($V$1=\"LY FALL\",M$4>6),M{start_row + 34},M{start_row + 4}))))))",

            f"N{start_row + 5}": f"=IF(AND($V$1=\"YTD\",N$4<$K$1),N{start_row + 12},IF(AND($V$1=\"YTD\",N$4=$K$1),N{start_row + 12},"
                                f"IF(AND($V$1=\"CURRENT MTH\",N$4=$K$1),N{start_row + 12},IF(AND($V$1=\"SPRING\",N$4<7),N{start_row + 12},"
                                f"IF(AND($V$1=\"FALL\",N$4>6),N{start_row + 12},IF(AND($V$1=\"LY FALL\",N$4>6),N{start_row + 34},N{start_row + 4}))))))",

            f"O{start_row + 5}": f"=IF(AND($V$1=\"YTD\",O$4<$K$1),O{start_row + 12},IF(AND($V$1=\"YTD\",O$4=$K$1),O{start_row + 12},"
                                f"IF(AND($V$1=\"CURRENT MTH\",O$4=$K$1),O{start_row + 12},IF(AND($V$1=\"SPRING\",O$4<7),O{start_row + 12},"
                                f"IF(AND($V$1=\"FALL\",O$4>6),O{start_row + 12},IF(AND($V$1=\"LY FALL\",O$4>6),O{start_row + 34},O{start_row + 4}))))))",
            f"P{start_row + 5}": f"=IF(AND($V$1=\"YTD\",P$4<$K$1),P{start_row + 12},IF(AND($V$1=\"YTD\",P$4=$K$1),P{start_row + 12},"
                                f"IF(AND($V$1=\"CURRENT MTH\",P$4=$K$1),P{start_row + 12},IF(AND($V$1=\"SPRING\",P$4<7),P{start_row + 12},"
                                f"IF(AND($V$1=\"FALL\",P$4>6),P{start_row + 12},IF(AND($V$1=\"LY FALL\",P$4>6),P{start_row + 34},P{start_row + 4}))))))",

            f"Q{start_row + 5}": f"=IF(AND($V$1=\"YTD\",Q$4<$K$1),Q{start_row + 12},IF(AND($V$1=\"YTD\",Q$4=$K$1),Q{start_row + 12},"
                                f"IF(AND($V$1=\"CURRENT MTH\",Q$4=$K$1),Q{start_row + 12},IF(AND($V$1=\"SPRING\",Q$4<7),Q{start_row + 12},"
                                f"IF(AND($V$1=\"FALL\",Q$4>6),Q{start_row + 12},IF(AND($V$1=\"LY FALL\",Q$4>6),Q{start_row + 34},Q{start_row + 4}))))))",
            
            f"R{start_row + 5}": f"=IF(AND($V$1=\"YTD\",R$4<$K$1),R{start_row + 12},IF(AND($V$1=\"YTD\",R$4=$K$1),R{start_row + 12},"
                                f"IF(AND($V$1=\"CURRENT MTH\",R$4=$K$1),R{start_row + 12},IF(AND($V$1=\"SPRING\",R$4<7),R{start_row + 12},"
                                f"IF(AND($V$1=\"FALL\",R$4>6),R{start_row + 12},IF(AND($V$1=\"LY FALL\",R$4>6),R{start_row + 34},R{start_row + 4}))))))",
            
            f"S{start_row + 5}": f"=IF(AND($V$1=\"YTD\",S$4<$K$1),S{start_row + 12},IF(AND($V$1=\"YTD\",S$4=$K$1),S{start_row + 12},"
                                f"IF(AND($V$1=\"CURRENT MTH\",S$4=$K$1),S{start_row + 12},IF(AND($V$1=\"SPRING\",S$4<7),S{start_row + 12},"
                                f"IF(AND($V$1=\"FALL\",S$4>6),S{start_row + 12},IF(AND($V$1=\"LY FALL\",S$4>6),S{start_row + 34},S{start_row + 4}))))))",
            
            f"T{start_row + 5}": f"=IF(AND($V$1=\"YTD\",T$4<$K$1),T{start_row + 12},IF(AND($V$1=\"YTD\",T$4=$K$1),T{start_row + 12},"
                                f"IF(AND($V$1=\"CURRENT MTH\",T$4=$K$1),T{start_row + 12},IF(AND($V$1=\"SPRING\",T$4<7),T{start_row + 12},"
                                f"IF(AND($V$1=\"FALL\",T$4>6),T{start_row + 12},IF(AND($V$1=\"LY FALL\",T$4>6),T{start_row + 34},T{start_row + 4}))))))",
            
            f"U{start_row + 5}": f"=SUM(I{start_row + 5}:T{start_row + 5})",
            
            f"V{start_row + 5}": f"=IFERROR(SUM(I{start_row + 5}:N{start_row + 5}),0)",
            
            f"W{start_row + 5}": f"=IFERROR(SUM(O{start_row + 5}:T{start_row + 5}),0)",

            f"U{start_row + 6}": f"=SUM(I{start_row + 6}:T{start_row + 6})",
            f"V{start_row + 6}": f"=IFERROR(SUM(I{start_row + 6}:N{start_row + 6}),0)",
            f"W{start_row + 6}": f"=IFERROR(SUM(O{start_row + 6}:T{start_row + 6}),0)",
            f"I{start_row + 16}": TY_OH_Units['FEB'],
            f"J{start_row + 16}": TY_OH_Units['MAR'],
            f"K{start_row + 16}":TY_OH_Units['APR'],
            f"L{start_row + 16}": TY_OH_Units['MAY'],
            f"M{start_row + 16}": TY_OH_Units['JUN'],
            f"N{start_row + 16}":TY_OH_Units['JUL'],
            f"O{start_row + 16}": TY_OH_Units['AUG'],
            f"P{start_row + 16}":TY_OH_Units['SEP'],
            f"Q{start_row + 16}":TY_OH_Units['OCT'],
            f"R{start_row + 16}": TY_OH_Units['NOV'],
            f"S{start_row + 16}": TY_OH_Units['DEC'],
            f"T{start_row + 16}": TY_OH_Units['JAN'],

            f"U{start_row + 16}": f"=AVERAGEIFS(I{start_row + 16}:T{start_row + 16}, $I$4:$T$4, \"<=\"&$K$1)",
            
            f"V{start_row + 16}": f"=AVERAGEIFS(I{start_row + 16}:N{start_row + 16}, $I$4:$N$4, \"<=\"&$K$2)",
            
            f"W{start_row + 16}": f"=AVERAGEIFS(O{start_row + 16}:T{start_row + 16}, $O$4:$T$4, \"<=\"&$M$2)",

            f"I{start_row + 32}":TY_Receipts['FEB'],
            
            f"J{start_row + 32}": TY_Receipts['MAR'],

            f"K{start_row + 32}":TY_Receipts['APR'],
            f"L{start_row + 32}":TY_Receipts['MAY'],
            f"M{start_row + 32}":TY_Receipts['JUN'],
            f"N{start_row + 32}":TY_Receipts['JUL'],
            f"O{start_row + 32}":TY_Receipts['AUG'],
            f"P{start_row + 32}":TY_Receipts['SEP'],

            f"Q{start_row + 32}":TY_Receipts['OCT'],

            f"R{start_row + 32}":TY_Receipts['NOV'],

            f"S{start_row + 32}":TY_Receipts['DEC'],

            f"T{start_row + 32}":TY_Receipts['JAN'],

            f"U{start_row + 32}": f"=IFERROR(SUM(I{start_row + 32}:T{start_row + 32}),0)",

            f"V{start_row + 32}": f"=IFERROR(SUM(I{start_row + 32}:N{start_row + 32}),0)",

            f"W{start_row + 32}": f"=IFERROR(SUM(O{start_row + 32}:T{start_row + 32}),0)",
            f"I{start_row + 7}": f"=IF(AND($V$1=\"Current Mth\",$K$1=I$4,I$4=1),$T{start_row + 38}+I{start_row + 6}+I{start_row + 32}-I{start_row + 5},IF(AND($V$1=\"Current Mth\",$K$1=I$4),I{start_row + 16}+I{start_row + 6}-(I{start_row + 5}-I{start_row + 12}),IF(AND($V$1=\"Current Mth\",I$4=1,$K$1>1),$T{start_row + 7}+I{start_row + 6}-I{start_row + 5},IF(AND($V$1=\"YTD\",I$4<$K$1),I{start_row + 16},IF(AND($V$1=\"Spring\",I$4<7),I{start_row + 16},IF(AND($V$1=\"Fall\",I$4>6,I$4<$K$1),I{start_row + 16},IF(AND($V$1=\"Fall\",I$4=1,$K$1>1),$T{start_row + 7}+I{start_row + 6}-I{start_row + 5},IF(AND($V$1=\"Fall\",I$4>6,I$4=$K$1),H{start_row + 16}+I{start_row + 32}+I{start_row + 6}-I{start_row + 5},IF(AND($V$1=\"LY Fall\",I$4>6),I{start_row + 38},IF(AND(I$4=1,$V$1=\"LY FALL\"),$T{start_row + 38}+I{start_row + 6}+I{start_row + 6}-I{start_row + 5},IF(AND(I$4=$K$1,I$4>1),H{start_row + 16}+I{start_row + 32}+I{start_row + 6}-I{start_row + 5},H{start_row + 7}+I{start_row + 6}-I{start_row + 5})))))))))))",

            f"J{start_row + 7}": f"=IF(AND($V$1=\"Current Mth\",$K$1=J$4,J$4=1),$T{start_row + 38}+J{start_row + 6}+J{start_row + 32}-J{start_row + 5},IF(AND($V$1=\"Current Mth\",$K$1=J$4),J{start_row + 16}+J{start_row + 6}-(J{start_row + 5}-J{start_row + 12}),IF(AND($V$1=\"Current Mth\",J$4=1,$K$1>1),$T{start_row + 7}+J{start_row + 6}-J{start_row + 5},IF(AND($V$1=\"YTD\",J$4<$K$1),J{start_row + 16},IF(AND($V$1=\"Spring\",J$4<7),J{start_row + 16},IF(AND($V$1=\"Fall\",J$4>6,J$4<$K$1),J{start_row + 16},IF(AND($V$1=\"Fall\",J$4=1,$K$1>1),$T{start_row + 7}+J{start_row + 6}-J{start_row + 5},IF(AND($V$1=\"Fall\",J$4>6,J$4=$K$1),I{start_row + 16}+J{start_row + 32}+J{start_row + 6}-J{start_row + 5},IF(AND($V$1=\"LY Fall\",J$4>6),J{start_row + 38},IF(AND(J$4=1,$V$1=\"LY FALL\"),$T{start_row + 38}+J{start_row + 6}+J{start_row + 6}-J{start_row + 5},IF(AND(J$4=$K$1,J$4>1),I{start_row + 16}+J{start_row + 32}+J{start_row + 6}-J{start_row + 5},I{start_row + 7}+J{start_row + 6}-J{start_row + 5})))))))))))",

            f"K{start_row + 7}": f"=IF(AND($V$1=\"Current Mth\",$K$1=K$4,K$4=1),$T{start_row + 38}+K{start_row + 6}+K{start_row + 32}-K{start_row + 5},IF(AND($V$1=\"Current Mth\",$K$1=K$4),K{start_row + 16}+K{start_row + 6}-(K{start_row + 5}-K{start_row + 12}),IF(AND($V$1=\"Current Mth\",K$4=1,$K$1>1),$T{start_row + 7}+K{start_row + 6}-K{start_row + 5},IF(AND($V$1=\"YTD\",K$4<$K$1),K{start_row + 16},IF(AND($V$1=\"Spring\",K$4<7),K{start_row + 16},IF(AND($V$1=\"Fall\",K$4>6,K$4<$K$1),K{start_row + 16},IF(AND($V$1=\"Fall\",K$4=1,$K$1>1),$T{start_row + 7}+K{start_row + 6}-K{start_row + 5},IF(AND($V$1=\"Fall\",K$4>6,K$4=$K$1),J{start_row + 16}+K{start_row + 32}+K{start_row + 6}-K{start_row + 5},IF(AND($V$1=\"LY Fall\",K$4>6),K{start_row + 38},IF(AND(K$4=1,$V$1=\"LY FALL\"),$T{start_row + 38}+K{start_row + 6}+K{start_row + 6}-K{start_row + 5},IF(AND(K$4=$K$1,K$4>1),J{start_row + 16}+K{start_row + 32}+K{start_row + 6}-K{start_row + 5},J{start_row + 7}+K{start_row + 6}-K{start_row + 5})))))))))))",

            f"L{start_row + 7}": f"=IF(AND($V$1=\"Current Mth\",$K$1=L$4,L$4=1),$T{start_row + 38}+L{start_row + 6}+L{start_row + 32}-L{start_row + 5},IF(AND($V$1=\"Current Mth\",$K$1=L$4),L{start_row + 16}+L{start_row + 6}-(L{start_row + 5}-L{start_row + 12}),IF(AND($V$1=\"Current Mth\",L$4=1,$K$1>1),$T{start_row + 7}+L{start_row + 6}-L{start_row + 5},IF(AND($V$1=\"YTD\",L$4<$K$1),L{start_row + 16},IF(AND($V$1=\"Spring\",L$4<7),L{start_row + 16},IF(AND($V$1=\"Fall\",L$4>6,L$4<$K$1),L{start_row + 16},IF(AND($V$1=\"Fall\",L$4=1,$K$1>1),$T{start_row + 7}+L{start_row + 6}-L{start_row + 5},IF(AND($V$1=\"Fall\",L$4>6,L$4=$K$1),K{start_row + 16}+L{start_row + 32}+L{start_row + 6}-L{start_row + 5},IF(AND($V$1=\"LY Fall\",L$4>6),L{start_row + 38},IF(AND(L$4=1,$V$1=\"LY FALL\"),$T{start_row + 38}+L{start_row + 6}+L{start_row + 6}-L{start_row + 5},IF(AND(L$4=$K$1,L$4>1),K{start_row + 16}+L{start_row + 32}+L{start_row + 6}-L{start_row + 5},K{start_row + 7}+L{start_row + 6}-L{start_row + 5})))))))))))",

            f"M{start_row + 7}": f"=IF(AND($V$1=\"Current Mth\",$K$1=M$4,M$4=1),$T{start_row + 38}+M{start_row + 6}+M{start_row + 32}-M{start_row + 5},IF(AND($V$1=\"Current Mth\",$K$1=M$4),M{start_row + 16}+M{start_row + 6}-(M{start_row + 5}-M{start_row + 12}),IF(AND($V$1=\"Current Mth\",M$4=1,$K$1>1),$T{start_row + 7}+M{start_row + 6}-M{start_row + 5},IF(AND($V$1=\"YTD\",M$4<$K$1),M{start_row + 16},IF(AND($V$1=\"Spring\",M$4<7),M{start_row + 16},IF(AND($V$1=\"Fall\",M$4>6,M$4<$K$1),M{start_row + 16},IF(AND($V$1=\"Fall\",M$4=1,$K$1>1),$T{start_row + 7}+M{start_row + 6}-M{start_row + 5},IF(AND($V$1=\"Fall\",M$4>6,M$4=$K$1),L{start_row + 16}+M{start_row + 32}+M{start_row + 6}-M{start_row + 5},IF(AND($V$1=\"LY Fall\",M$4>6),M{start_row + 38},IF(AND(M$4=1,$V$1=\"LY FALL\"),$T{start_row + 38}+M{start_row + 6}+M{start_row + 6}-M{start_row + 5},IF(AND(M$4=$K$1,M$4>1),L{start_row + 16}+M{start_row + 32}+M{start_row + 6}-M{start_row + 5},L{start_row + 7}+M{start_row + 6}-M{start_row + 5})))))))))))",

            f"O{start_row + 7}": f"=IF(AND($V$1=\"Current Mth\",$K$1=O$4,O$4=1),$T{start_row + 38}+O{start_row + 6}+O{start_row + 32}-O{start_row + 5},IF(AND($V$1=\"Current Mth\",$K$1=O$4),O{start_row + 16}+O{start_row + 6}-(O{start_row + 5}-O{start_row + 12}),IF(AND($V$1=\"Current Mth\",O$4=1,$K$1>1),$T{start_row + 7}+O{start_row + 6}-O{start_row + 5},IF(AND($V$1=\"YTD\",O$4<$K$1),O{start_row + 16},IF(AND($V$1=\"Spring\",O$4<7),O{start_row + 16},IF(AND($V$1=\"Fall\",O$4>6,O$4<$K$1),O{start_row + 16},IF(AND($V$1=\"Fall\",O$4=1,$K$1>1),$T{start_row + 7}+O{start_row + 6}-O{start_row + 5},IF(AND($V$1=\"Fall\",O$4>6,O$4=$K$1),N{start_row + 16}+O{start_row + 32}+O{start_row + 6}-O{start_row + 5},IF(AND($V$1=\"LY Fall\",O$4>6),O{start_row + 38},IF(AND(O$4=1,$V$1=\"LY FALL\"),$T{start_row + 38}+O{start_row + 6}+O{start_row + 6}-O{start_row + 5},IF(AND(O$4=$K$1,O$4>1),N{start_row + 16}+O{start_row + 32}+O{start_row + 6}-O{start_row + 5},N{start_row + 7}+O{start_row + 6}-O{start_row + 5})))))))))))",

            f"P{start_row + 7}": f"=IF(AND($V$1=\"Current Mth\",$K$1=P$4,P$4=1),$T{start_row + 38}+P{start_row + 6}+P{start_row + 32}-P{start_row + 5},IF(AND($V$1=\"Current Mth\",$K$1=P$4),P{start_row + 16}+P{start_row + 6}-(P{start_row + 5}-P{start_row + 12}),IF(AND($V$1=\"Current Mth\",P$4=1,$K$1>1),$T{start_row + 7}+P{start_row + 6}-P{start_row + 5},IF(AND($V$1=\"YTD\",P$4<$K$1),P{start_row + 16},IF(AND($V$1=\"Spring\",P$4<7),P{start_row + 16},IF(AND($V$1=\"Fall\",P$4>6,P$4<$K$1),P{start_row + 16},IF(AND($V$1=\"Fall\",P$4=1,$K$1>1),$T{start_row + 7}+P{start_row + 6}-P{start_row + 5},IF(AND($V$1=\"Fall\",P$4>6,P$4=$K$1),O{start_row + 16}+P{start_row + 32}+P{start_row + 6}-P{start_row + 5},IF(AND($V$1=\"LY Fall\",P$4>6),P{start_row + 38},IF(AND(P$4=1,$V$1=\"LY FALL\"),$T{start_row + 38}+P{start_row + 6}+P{start_row + 6}-P{start_row + 5},IF(AND(P$4=$K$1,P$4>1),O{start_row + 16}+P{start_row + 32}+P{start_row + 6}-P{start_row + 5},O{start_row + 7}+P{start_row + 6}-P{start_row + 5})))))))))))",

            f"Q{start_row + 7}": f"=IF(AND($V$1=\"Current Mth\",$K$1=Q$4,Q$4=1),$T{start_row + 38}+Q{start_row + 6}+Q{start_row + 32}-Q{start_row + 5},IF(AND($V$1=\"Current Mth\",$K$1=Q$4),Q{start_row + 16}+Q{start_row + 6}-(Q{start_row + 5}-Q{start_row + 12}),IF(AND($V$1=\"Current Mth\",Q$4=1,$K$1>1),$T{start_row + 7}+Q{start_row + 6}-Q{start_row + 5},IF(AND($V$1=\"YTD\",Q$4<$K$1),Q{start_row + 16},IF(AND($V$1=\"Spring\",Q$4<7),Q{start_row + 16},IF(AND($V$1=\"Fall\",Q$4>6,Q$4<$K$1),Q{start_row + 16},IF(AND($V$1=\"Fall\",Q$4=1,$K$1>1),$T{start_row + 7}+Q{start_row + 6}-Q{start_row + 5},IF(AND($V$1=\"Fall\",Q$4>6,Q$4=$K$1),P{start_row + 16}+Q{start_row + 32}+Q{start_row + 6}-Q{start_row + 5},IF(AND($V$1=\"LY Fall\",Q$4>6),Q{start_row + 38},IF(AND(Q$4=1,$V$1=\"LY FALL\"),$T{start_row + 38}+Q{start_row + 6}+Q{start_row + 6}-Q{start_row + 5},IF(AND(Q$4=$K$1,Q$4>1),P{start_row + 16}+Q{start_row + 32}+Q{start_row + 6}-Q{start_row + 5},P{start_row + 7}+Q{start_row + 6}-Q{start_row + 5})))))))))))",

            f"N{start_row + 7}": f"=IF(AND($V$1=\"Current Mth\",$K$1=N$4,N$4=1),$T{start_row + 38}+N{start_row + 6}+N{start_row + 32}-N{start_row + 5},IF(AND($V$1=\"Current Mth\",$K$1=N$4),N{start_row + 16}+N{start_row + 6}-(N{start_row + 5}-N{start_row + 12}),IF(AND($V$1=\"Current Mth\",N$4=1,$K$1>1),$T{start_row + 7}+N{start_row + 6}-N{start_row + 5},IF(AND($V$1=\"YTD\",N$4<$K$1),N{start_row + 16},IF(AND($V$1=\"Spring\",N$4<7),N{start_row + 16},IF(AND($V$1=\"Fall\",N$4>6,N$4<$K$1),N{start_row + 16},IF(AND($V$1=\"Fall\",N$4=1,$K$1>1),$T{start_row + 7}+N{start_row + 6}-N{start_row + 5},IF(AND($V$1=\"Fall\",N$4>6,N$4=$K$1),M{start_row + 16}+N{start_row + 32}+N{start_row + 6}-N{start_row + 5},IF(AND($V$1=\"LY Fall\",N$4>6),N{start_row + 38},IF(AND(N$4=1,$V$1=\"LY FALL\"),$T{start_row + 38}+N{start_row + 6}+N{start_row + 6}-N{start_row + 5},IF(AND(N$4=$K$1,N$4>1),M{start_row + 16}+N{start_row + 32}+N{start_row + 6}-N{start_row + 5},M{start_row + 7}+N{start_row + 6}-N{start_row + 5})))))))))))",

            f"R{start_row + 7}": f"=IF(AND($V$1=\"Current Mth\",$K$1=R$4,R$4=1),$T{start_row + 38}+R{start_row + 6}+R{start_row + 32}-R{start_row + 5},IF(AND($V$1=\"Current Mth\",$K$1=R$4),R{start_row + 16}+R{start_row + 6}-(R{start_row + 5}-R{start_row + 12}),IF(AND($V$1=\"Current Mth\",R$4=1,$K$1>1),$T{start_row + 7}+R{start_row + 6}-R{start_row + 5},IF(AND($V$1=\"YTD\",R$4<$K$1),R{start_row + 16},IF(AND($V$1=\"Spring\",R$4<7),R{start_row + 16},IF(AND($V$1=\"Fall\",R$4>6,R$4<$K$1),R{start_row + 16},IF(AND($V$1=\"Fall\",R$4=1,$K$1>1),$T{start_row + 7}+R{start_row + 6}-R{start_row + 5},IF(AND($V$1=\"Fall\",R$4>6,R$4=$K$1),Q{start_row + 16}+R{start_row + 32}+R{start_row + 6}-R{start_row + 5},IF(AND($V$1=\"LY Fall\",R$4>6),R{start_row + 38},IF(AND(R$4=1,$V$1=\"LY FALL\"),$T{start_row + 38}+R{start_row + 6}+R{start_row + 6}-R{start_row + 5},IF(AND(R$4=$K$1,R$4>1),Q{start_row + 16}+R{start_row + 32}+R{start_row + 6}-R{start_row + 5},Q{start_row + 7}+R{start_row + 6}-R{start_row + 5})))))))))))",

            f"S{start_row + 7}": f"=IF(AND($V$1=\"Current Mth\",$K$1=S$4,S$4=1),$T{start_row + 38}+S{start_row + 6}+S{start_row + 32}-S{start_row + 5},IF(AND($V$1=\"Current Mth\",$K$1=S$4),S{start_row + 16}+S{start_row + 6}-(S{start_row + 5}-S{start_row + 12}),IF(AND($V$1=\"Current Mth\",S$4=1,$K$1>1),$T{start_row + 7}+S{start_row + 6}-S{start_row + 5},IF(AND($V$1=\"YTD\",S$4<$K$1),S{start_row + 16},IF(AND($V$1=\"Spring\",S$4<7),S{start_row + 16},IF(AND($V$1=\"Fall\",S$4>6,S$4<$K$1),S{start_row + 16},IF(AND($V$1=\"Fall\",S$4=1,$K$1>1),$T{start_row + 7}+S{start_row + 6}-S{start_row + 5},IF(AND($V$1=\"Fall\",S$4>6,S$4=$K$1),R{start_row + 16}+S{start_row + 32}+S{start_row + 6}-S{start_row + 5},IF(AND($V$1=\"LY Fall\",S$4>6),S{start_row + 38},IF(AND(S$4=1,$V$1=\"LY FALL\"),$T{start_row + 38}+S{start_row + 6}+S{start_row + 6}-S{start_row + 5},IF(AND(S$4=$K$1,S$4>1),R{start_row + 16}+S{start_row + 32}+S{start_row + 6}-S{start_row + 5},R{start_row + 7}+S{start_row + 6}-S{start_row + 5})))))))))))",

            f"T{start_row + 7}": f"=IF(AND($V$1=\"Current Mth\",$K$1=T$4,T$4=1),$T{start_row + 38}+T{start_row + 6}+T{start_row + 32}-T{start_row + 5},IF(AND($V$1=\"Current Mth\",$K$1=T$4),T{start_row + 16}+T{start_row + 6}-(T{start_row + 5}-T{start_row + 12}),IF(AND($V$1=\"Current Mth\",T$4=1,$K$1>1),$T{start_row + 7}+T{start_row + 6}-T{start_row + 5},IF(AND($V$1=\"YTD\",T$4<$K$1),T{start_row + 16},IF(AND($V$1=\"Spring\",T$4<7),T{start_row + 16},IF(AND($V$1=\"Fall\",T$4>6,T$4<$K$1),T{start_row + 16},IF(AND($V$1=\"Fall\",T$4=1,$K$1>1),$T{start_row + 7}+T{start_row + 6}-T{start_row + 5},IF(AND($V$1=\"Fall\",T$4>6,T$4=$K$1),S{start_row + 16}+T{start_row + 32}+T{start_row + 6}-T{start_row + 5},IF(AND($V$1=\"LY Fall\",T$4>6),T{start_row + 38},IF(AND(T$4=1,$V$1=\"LY FALL\"),$T{start_row + 38}+T{start_row + 6}+T{start_row + 6}-T{start_row + 5},IF(AND(T$4=$K$1,T$4>1),S{start_row + 16}+T{start_row + 32}+T{start_row + 6}-T{start_row + 5},S{start_row + 7}+T{start_row + 6}-T{start_row + 5})))))))))))",

            f"U{start_row + 7}": f"=AVERAGEIFS(I{start_row + 7}:T{start_row + 7},$I$4:$T$4,\"<=\"&$K$1)",
            f"V{start_row + 7}": f"=AVERAGEIFS(I{start_row + 7}:N{start_row + 7},$I$4:$N$4,\"<=\"&$K$2)",
            f"W{start_row + 7}": f"=AVERAGEIFS(O{start_row + 7}:T{start_row + 7},$O$4:$T$4,\"<=\"&$M$2)",
            f"I{start_row + 8}": Nav_Feb,
            f"J{start_row + 8}": Nav_Mar,
            f"K{start_row + 8}": Nav_Apr,
            f"L{start_row + 8}": Nav_May,
            f"M{start_row + 8}": Nav_Jun,
            f"N{start_row + 8}": Nav_Jul,
            f"O{start_row + 8}": Nav_Aug,
            f"P{start_row + 8}":Nav_Sep,
            f"Q{start_row + 8}":Nav_Oct,
            f"R{start_row + 8}":Nav_Nov,
            f"S{start_row + 8}": Nav_Dec,
            f"T{start_row + 8}": Nav_Jan,
            f"I{start_row + 6}": planned_shipments["FEB"],
            f"J{start_row + 6}": planned_shipments["MAR"],
            f"K{start_row + 6}": planned_shipments["APR"],
            f"L{start_row + 6}": planned_shipments["MAY"],
            f"M{start_row + 6}": planned_shipments["JUN"],
            f"N{start_row + 6}": planned_shipments["JUL"],
            f"O{start_row + 6}": planned_shipments["AUG"],
            f"P{start_row + 6}":planned_shipments["SEP"],
            f"Q{start_row + 6}":planned_shipments["OCT"],
            f"R{start_row + 6}":planned_shipments["NOV"],
            f"S{start_row + 6}": planned_shipments["DEC"],
            f"T{start_row + 6}": planned_shipments["JAN"],
            f"U{start_row + 8}": f"=SUM(I{start_row + 8}:T{start_row + 8})",
            f"V{start_row + 8}": f"=IFERROR(SUM(I{start_row + 8}:N{start_row + 8}),0)",
            f"W{start_row + 8}": f"=IFERROR(SUM(O{start_row + 8}:T{start_row + 8}),0)",

            f"I{start_row + 9}": Macys_Proj_Receipts_Feb,
            f"J{start_row + 9}": Macys_Proj_Receipts_Mar,
            f"K{start_row + 9}":Macys_Proj_Receipts_Apr ,
            f"L{start_row + 9}":Macys_Proj_Receipts_May,
            f"M{start_row + 9}": Macys_Proj_Receipts_Jun,
            f"N{start_row + 9}": Macys_Proj_Receipts_Jul,
            f"O{start_row + 9}": Macys_Proj_Receipts_Aug,
            f"P{start_row + 9}":Macys_Proj_Receipts_Sep,
            f"Q{start_row + 9}": Macys_Proj_Receipts_oct,
            f"R{start_row + 9}": Macys_Proj_Receipts_Nov,
            f"S{start_row + 9}": Macys_Proj_Receipts_Dec,
            f"T{start_row + 9}": Macys_Proj_Receipts_Jan,
            f"U{start_row + 9}": f"=SUM(I{start_row + 9}:T{start_row + 9})",
            f"V{start_row + 9}": f"=IFERROR(SUM(I{start_row + 9}:N{start_row + 9}),0)",
            f"W{start_row + 9}": f"=IFERROR(SUM(O{start_row + 9}:T{start_row + 9}),0)",
            f"I{start_row + 10}": f"=I{start_row + 5}/(I{start_row + 5}+I{start_row + 7})",
            f"J{start_row + 10}": f"=J{start_row + 5}/(J{start_row + 5}+J{start_row + 7})",
            f"K{start_row + 10}": f"=K{start_row + 5}/(K{start_row + 5}+K{start_row + 7})",
            f"L{start_row + 10}": f"=L{start_row + 5}/(L{start_row + 5}+L{start_row + 7})",
            f"M{start_row + 10}": f"=M{start_row + 5}/(M{start_row + 5}+M{start_row + 7})",
            f"N{start_row + 10}": f"=N{start_row + 5}/(N{start_row + 5}+N{start_row + 7})",
            f"O{start_row + 10}": f"=O{start_row + 5}/(O{start_row + 5}+O{start_row + 7})",
            f"P{start_row + 10}": f"=P{start_row + 5}/(P{start_row + 5}+P{start_row + 7})",
            f"Q{start_row + 10}": f"=Q{start_row + 5}/(Q{start_row + 5}+Q{start_row + 7})",
            f"R{start_row + 10}": f"=R{start_row + 5}/(R{start_row + 5}+R{start_row + 7})",
            f"S{start_row + 10}": f"=S{start_row + 5}/(S{start_row + 5}+S{start_row + 7})",
            f"T{start_row + 10}": f"=T{start_row + 5}/(T{start_row + 5}+T{start_row + 7})",
            f"U{start_row + 10}": f"=U{start_row + 5}/(U{start_row + 5}+U{start_row + 7})",
            f"V{start_row + 10}": f"=V{start_row + 5}/(V{start_row + 5}+V{start_row + 7})",
            f"W{start_row + 10}": f"=W{start_row + 5}/(W{start_row + 5}+W{start_row + 7})",
            f"I{start_row + 14}": TY_MCOM_Unit_Sales['FEB'],

            f"J{start_row + 14}":TY_MCOM_Unit_Sales['MAR'],

            f"K{start_row + 14}":TY_MCOM_Unit_Sales['APR'],

            f"L{start_row + 14}":TY_MCOM_Unit_Sales['MAY'],

            f"M{start_row + 14}": TY_MCOM_Unit_Sales['JUN'],

            f"N{start_row + 14}": TY_MCOM_Unit_Sales['JUL'],

            f"O{start_row + 14}": TY_MCOM_Unit_Sales['AUG'],

            f"P{start_row + 14}":TY_MCOM_Unit_Sales['SEP'],

            f"Q{start_row + 14}": TY_MCOM_Unit_Sales['OCT'],

            f"R{start_row + 14}":TY_MCOM_Unit_Sales['NOV'],

            f"S{start_row + 14}": TY_MCOM_Unit_Sales['DEC'],

            f"T{start_row + 14}":TY_MCOM_Unit_Sales['JAN'],

            f"U{start_row + 14}": f"=IFERROR(SUM(I{start_row + 14}:T{start_row + 14}),0)",
            f"V{start_row + 14}": f"=IFERROR(SUM(I{start_row + 14}:N{start_row + 14}),0)",
            f"W{start_row + 14}": f"=IFERROR(SUM(O{start_row + 14}:T{start_row + 14}),0)",
            f"I{start_row + 13}": f"=I{start_row + 12}-I{start_row + 14}",
            f"J{start_row + 13}": f"=J{start_row + 12}-J{start_row + 14}",
            f"K{start_row + 13}": f"=K{start_row + 12}-K{start_row + 14}",
            f"L{start_row + 13}": f"=L{start_row + 12}-L{start_row + 14}",
            f"M{start_row + 13}": f"=M{start_row + 12}-M{start_row + 14}",
            f"N{start_row + 13}": f"=N{start_row + 12}-N{start_row + 14}",
            f"O{start_row + 13}": f"=O{start_row + 12}-O{start_row + 14}",
            f"P{start_row + 13}": f"=P{start_row + 12}-P{start_row + 14}",
            f"Q{start_row + 13}": f"=Q{start_row + 12}-Q{start_row + 14}",
            f"R{start_row + 13}": f"=R{start_row + 12}-R{start_row + 14}",
            f"S{start_row + 13}": f"=S{start_row + 12}-S{start_row + 14}",
            f"T{start_row + 13}": f"=T{start_row + 12}-T{start_row + 14}",

            f"U{start_row + 13}": f"=IFERROR(SUM(I{start_row + 13}:T{start_row + 13}),0)",
            f"V{start_row + 13}": f"=IFERROR(SUM(I{start_row + 13}:N{start_row + 13}),0)",
            f"W{start_row + 13}": f"=IFERROR(SUM(O{start_row + 13}:T{start_row + 13}),0)",

            f"I{start_row + 15}": f"=IFERROR(I{start_row + 14}/I{start_row + 12},0)",
            f"J{start_row + 15}": f"=IFERROR(J{start_row + 14}/J{start_row + 12},0)",
            f"K{start_row + 15}": f"=IFERROR(K{start_row + 14}/K{start_row + 12},0)",
            f"L{start_row + 15}": f"=IFERROR(L{start_row + 14}/L{start_row + 12},0)",
            f"M{start_row + 15}": f"=IFERROR(M{start_row + 14}/M{start_row + 12},0)",
            f"N{start_row + 15}": f"=IFERROR(N{start_row + 14}/N{start_row + 12},0)",
            f"O{start_row + 15}": f"=IFERROR(O{start_row + 14}/O{start_row + 12},0)",
            f"P{start_row + 15}": f"=IFERROR(P{start_row + 14}/P{start_row + 12},0)",
            f"Q{start_row + 15}": f"=IFERROR(Q{start_row + 14}/Q{start_row + 12},0)",
            f"R{start_row + 15}": f"=IFERROR(R{start_row + 14}/R{start_row + 12},0)",
            f"S{start_row + 15}": f"=IFERROR(S{start_row + 14}/S{start_row + 12},0)",
            f"T{start_row + 15}": f"=IFERROR(T{start_row + 14}/T{start_row + 12},0)",
            f"U{start_row + 15}": f"=IFERROR(U{start_row + 14}/U{start_row + 12},0)",
            f"V{start_row + 15}": f"=IFERROR(V{start_row + 14}/V{start_row + 12},0)",
            f"W{start_row + 15}": f"=IFERROR(W{start_row + 14}/W{start_row + 12},0)",
            f"I{start_row + 18}": TY_OH_MCOM_Units['FEB'],

            f"J{start_row + 18}": TY_OH_MCOM_Units['MAR'],

            f"K{start_row + 18}":TY_OH_MCOM_Units['APR'],

            f"L{start_row + 18}":TY_OH_MCOM_Units['MAY'],

            f"M{start_row + 18}": TY_OH_MCOM_Units['JUN'],

            f"N{start_row + 18}": TY_OH_MCOM_Units['JUL'],

            f"O{start_row + 18}":TY_OH_MCOM_Units['AUG'],

            f"P{start_row + 18}":TY_OH_MCOM_Units['SEP'],

            f"Q{start_row + 18}":TY_OH_MCOM_Units['OCT'],

            f"R{start_row + 18}": TY_OH_MCOM_Units['NOV'],

            f"S{start_row + 18}": TY_OH_MCOM_Units['DEC'],

            f"T{start_row + 18}":TY_OH_MCOM_Units['JAN'],

            f"U{start_row + 18}": f"=AVERAGEIFS(I{start_row + 18}:T{start_row + 18},$I$4:$T$4,\"<=\"&$K$1)",
            f"V{start_row + 18}": f"=AVERAGEIFS(I{start_row + 18}:N{start_row + 18},$I$4:$N$4,\"<=\"&$K$2)",
            f"W{start_row + 18}": f"=AVERAGEIFS(O{start_row + 18}:T{start_row + 18},$O$4:$T$4,\"<=\"&$M$2)",
            f"I{start_row + 17}": f"=I{start_row + 16}-I{start_row + 18}",
            f"J{start_row + 17}": f"=J{start_row + 16}-J{start_row + 18}",
            f"K{start_row + 17}": f"=K{start_row + 16}-K{start_row + 18}",
            f"L{start_row + 17}": f"=L{start_row + 16}-L{start_row + 18}",
            f"M{start_row + 17}": f"=M{start_row + 16}-M{start_row + 18}",
            f"N{start_row + 17}": f"=N{start_row + 16}-N{start_row + 18}",
            f"O{start_row + 17}": f"=O{start_row + 16}-O{start_row + 18}",
            f"P{start_row + 17}": f"=P{start_row + 16}-P{start_row + 18}",
            f"Q{start_row + 17}": f"=Q{start_row + 16}-Q{start_row + 18}",
            f"R{start_row + 17}": f"=R{start_row + 16}-R{start_row + 18}",
            f"S{start_row + 17}": f"=S{start_row + 16}-S{start_row + 18}",
            f"T{start_row + 17}": f"=T{start_row + 16}-T{start_row + 18}",

            f"U{start_row + 17}": f"=AVERAGEIFS(I{start_row + 17}:T{start_row + 17},$I$4:$T$4,\"<=\"&$K$1)",
            f"V{start_row + 17}": f"=AVERAGEIFS(I{start_row + 17}:N{start_row + 17},$I$4:$N$4,\"<=\"&$K$2)",
            f"W{start_row + 17}": f"=AVERAGEIFS(O{start_row + 17}:T{start_row + 17},$O$4:$T$4,\"<=\"&$M$2)",
            f"I{start_row + 19}": f"=IFERROR(I{start_row + 18}/I{start_row + 16},0)",
            f"J{start_row + 19}": f"=IFERROR(J{start_row + 18}/J{start_row + 16},0)",
            f"K{start_row + 19}": f"=IFERROR(K{start_row + 18}/K{start_row + 16},0)",
            f"L{start_row + 19}": f"=IFERROR(L{start_row + 18}/L{start_row + 16},0)",
            f"M{start_row + 19}": f"=IFERROR(M{start_row + 18}/M{start_row + 16},0)",
            f"N{start_row + 19}": f"=IFERROR(N{start_row + 18}/N{start_row + 16},0)",
            f"O{start_row + 19}": f"=IFERROR(O{start_row + 18}/O{start_row + 16},0)",
            f"P{start_row + 19}": f"=IFERROR(P{start_row + 18}/P{start_row + 16},0)",
            f"Q{start_row + 19}": f"=IFERROR(Q{start_row + 18}/Q{start_row + 16},0)",
            f"R{start_row + 19}": f"=IFERROR(R{start_row + 18}/R{start_row + 16},0)",
            f"S{start_row + 19}": f"=IFERROR(S{start_row + 18}/S{start_row + 16},0)",
            f"T{start_row + 19}": f"=IFERROR(T{start_row + 18}/T{start_row + 16},0)",
            f"U{start_row + 19}": f"=IFERROR(U{start_row + 18}/U{start_row + 16},0)",
            f"V{start_row + 19}": f"=IFERROR(V{start_row + 18}/V{start_row + 16},0)",
            f"W{start_row + 19}": f"=IFERROR(W{start_row + 18}/W{start_row + 16},0)",
            f"I{start_row + 20}":PTD_TY_Sales['FEB'] ,

            f"J{start_row + 20}": PTD_TY_Sales['MAR'],

            f"K{start_row + 20}": PTD_TY_Sales['APR'],

            f"L{start_row + 20}":PTD_TY_Sales['MAY'],

            f"M{start_row + 20}": PTD_TY_Sales['JUN'],

            f"N{start_row + 20}": PTD_TY_Sales['JUL'],

            f"O{start_row + 20}":PTD_TY_Sales['AUG'],

            f"P{start_row + 20}": PTD_TY_Sales['SEP'],

            f"Q{start_row + 20}": PTD_TY_Sales['OCT'],

            f"R{start_row + 20}": PTD_TY_Sales['NOV'],

            f"S{start_row + 20}":PTD_TY_Sales['DEC'],

            f"T{start_row + 20}": PTD_TY_Sales['JAN'],

            # Aggregates
            f"U{start_row + 20}": f"=IFERROR(SUM(I{start_row + 20}:T{start_row + 20}),0)",
            f"V{start_row + 20}": f"=IFERROR(SUM(I{start_row + 20}:N{start_row + 20}),0)",
            f"W{start_row + 20}": f"=IFERROR(SUM(O{start_row + 20}:T{start_row + 20}),0)",
            f"I{start_row + 21}":MCOM_PTD_TY_Sales['FEB'],

            f"J{start_row + 21}":MCOM_PTD_TY_Sales['MAR'],

            f"K{start_row + 21}": MCOM_PTD_TY_Sales['APR'],

            f"L{start_row + 21}":MCOM_PTD_TY_Sales['MAY'],

            f"M{start_row + 21}":MCOM_PTD_TY_Sales['JUN'],

            f"N{start_row + 21}": MCOM_PTD_TY_Sales['JUL'],

            f"O{start_row + 21}": MCOM_PTD_TY_Sales['AUG'],

            f"P{start_row + 21}":MCOM_PTD_TY_Sales['SEP'],

            f"Q{start_row + 21}": MCOM_PTD_TY_Sales['OCT'],

            f"R{start_row + 21}": MCOM_PTD_TY_Sales['NOV'],

            f"S{start_row + 21}": MCOM_PTD_TY_Sales['DEC'],

            f"T{start_row + 21}":MCOM_PTD_TY_Sales['JAN'],

            f"U{start_row + 21}": f"=IFERROR(SUM(I{start_row + 21}:T{start_row + 21}),0)",
            f"V{start_row + 21}": f"=IFERROR(SUM(I{start_row + 21}:N{start_row + 21}),0)",
            f"W{start_row + 21}": f"=IFERROR(SUM(O{start_row + 21}:T{start_row + 21}),0)",

            f"I{start_row + 22}": f"=IFERROR(TEXT(IFERROR(I{start_row + 20}/I{start_row + 16},0),\"$0\") & \"/ \" & TEXT(((I{start_row + 20}/I{start_row + 16})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"J{start_row + 22}": f"=IFERROR(TEXT(IFERROR(J{start_row + 20}/J{start_row + 16},0),\"$0\") & \"/ \" & TEXT(((J{start_row + 20}/J{start_row + 16})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"K{start_row + 22}": f"=IFERROR(TEXT(IFERROR(K{start_row + 20}/K{start_row + 16},0),\"$0\") & \"/ \" & TEXT(((K{start_row + 20}/K{start_row + 16})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"L{start_row + 22}": f"=IFERROR(TEXT(IFERROR(L{start_row + 20}/L{start_row + 16},0),\"$0\") & \"/ \" & TEXT(((L{start_row + 20}/L{start_row + 16})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"M{start_row + 22}": f"=IFERROR(TEXT(IFERROR(M{start_row + 20}/M{start_row + 16},0),\"$0\") & \"/ \" & TEXT(((M{start_row + 20}/M{start_row + 16})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"N{start_row + 22}": f"=IFERROR(TEXT(IFERROR(N{start_row + 20}/N{start_row + 16},0),\"$0\") & \"/ \" & TEXT(((N{start_row + 20}/N{start_row + 16})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"O{start_row + 22}": f"=IFERROR(TEXT(IFERROR(O{start_row + 20}/O{start_row + 16},0),\"$0\") & \"/ \" & TEXT(((O{start_row + 20}/O{start_row + 16})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"P{start_row + 22}": f"=IFERROR(TEXT(IFERROR(P{start_row + 20}/P{start_row + 16},0),\"$0\") & \"/ \" & TEXT(((P{start_row + 20}/P{start_row + 16})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"Q{start_row + 22}": f"=IFERROR(TEXT(IFERROR(Q{start_row + 20}/Q{start_row + 16},0),\"$0\") & \"/ \" & TEXT(((Q{start_row + 20}/Q{start_row + 16})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"R{start_row + 22}": f"=IFERROR(TEXT(IFERROR(R{start_row + 20}/R{start_row + 16},0),\"$0\") & \"/ \" & TEXT(((R{start_row + 20}/R{start_row + 16})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"S{start_row + 22}": f"=IFERROR(TEXT(IFERROR(S{start_row + 20}/S{start_row + 16},0),\"$0\") & \"/ \" & TEXT(((S{start_row + 20}/S{start_row + 16})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"T{start_row + 22}": f"=IFERROR(TEXT(IFERROR(T{start_row + 20}/T{start_row + 16},0),\"$0\") & \"/ \" & TEXT(((T{start_row + 20}/T{start_row + 16})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"U{start_row + 22}": f"=IFERROR(TEXT(IFERROR(U{start_row + 20}/U{start_row + 16},0),\"$0\") & \"/ \" & TEXT(((U{start_row + 20}/U{start_row + 16})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"V{start_row + 22}": f"=IFERROR(TEXT(IFERROR(V{start_row + 20}/V{start_row + 16},0),\"$0\") & \"/ \" & TEXT(((V{start_row + 20}/V{start_row + 16})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"W{start_row + 22}": f"=IFERROR(TEXT(IFERROR(W{start_row + 20}/W{start_row + 16},0),\"$0\") & \"/ \" & TEXT(((W{start_row + 20}/W{start_row + 16})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"I{start_row + 23}": f"=IFERROR(IFERROR(I{start_row + 12}/(I{start_row + 12}+I{start_row + 16}),0),0)",
            f"J{start_row + 23}": f"=IFERROR(IFERROR(J{start_row + 12}/(J{start_row + 12}+J{start_row + 16}),0),0)",
            f"K{start_row + 23}": f"=IFERROR(IFERROR(K{start_row + 12}/(K{start_row + 12}+K{start_row + 16}),0),0)",
            f"L{start_row + 23}": f"=IFERROR(IFERROR(L{start_row + 12}/(L{start_row + 12}+L{start_row + 16}),0),0)",
            f"M{start_row + 23}": f"=IFERROR(IFERROR(M{start_row + 12}/(M{start_row + 12}+M{start_row + 16}),0),0)",
            f"N{start_row + 23}": f"=IFERROR(IFERROR(N{start_row + 12}/(N{start_row + 12}+N{start_row + 16}),0),0)",
            f"O{start_row + 23}": f"=IFERROR(IFERROR(O{start_row + 12}/(O{start_row + 12}+O{start_row + 16}),0),0)",
            f"P{start_row + 23}": f"=IFERROR(IFERROR(P{start_row + 12}/(P{start_row + 12}+P{start_row + 16}),0),0)",
            f"Q{start_row + 23}": f"=IFERROR(IFERROR(Q{start_row + 12}/(Q{start_row + 12}+Q{start_row + 16}),0),0)",
            f"R{start_row + 23}": f"=IFERROR(IFERROR(R{start_row + 12}/(R{start_row + 12}+R{start_row + 16}),0),0)",
            f"S{start_row + 23}": f"=IFERROR(IFERROR(S{start_row + 12}/(S{start_row + 12}+S{start_row + 16}),0),0)",
            f"T{start_row + 23}": f"=IFERROR(IFERROR(T{start_row + 12}/(T{start_row + 12}+T{start_row + 16}),0),0)",

            f"U{start_row + 23}": f"=IFERROR(IFERROR(U{start_row + 12}/(U{start_row + 12}+U{start_row + 16}-U{start_row + 8}),0),0)",
            f"V{start_row + 23}": f"=IFERROR(IFERROR(V{start_row + 12}/(V{start_row + 12}+V{start_row + 16}-V{start_row + 8}),0),0)",
            f"W{start_row + 23}": f"=IFERROR(IFERROR(W{start_row + 12}/(W{start_row + 12}+W{start_row + 16}-W{start_row + 8}),0),0)",
            f"I{start_row + 24}": f"=IFERROR((I{start_row + 12}-I{start_row + 14})/((I{start_row + 12}-I{start_row + 14})+(I{start_row + 16}-I{start_row + 18})),0)",
            f"J{start_row + 24}": f"=IFERROR((J{start_row + 12}-J{start_row + 14})/((J{start_row + 12}-J{start_row + 14})+(J{start_row + 16}-J{start_row + 18})),0)",
            f"K{start_row + 24}": f"=IFERROR((K{start_row + 12}-K{start_row + 14})/((K{start_row + 12}-K{start_row + 14})+(K{start_row + 16}-K{start_row + 18})),0)",
            f"L{start_row + 24}": f"=IFERROR((L{start_row + 12}-L{start_row + 14})/((L{start_row + 12}-L{start_row + 14})+(L{start_row + 16}-L{start_row + 18})),0)",
            f"M{start_row + 24}": f"=IFERROR((M{start_row + 12}-M{start_row + 14})/((M{start_row + 12}-M{start_row + 14})+(M{start_row + 16}-M{start_row + 18})),0)",
            f"N{start_row + 24}": f"=IFERROR((N{start_row + 12}-N{start_row + 14})/((N{start_row + 12}-N{start_row + 14})+(N{start_row + 16}-N{start_row + 18})),0)",
            f"O{start_row + 24}": f"=IFERROR((O{start_row + 12}-O{start_row + 14})/((O{start_row + 12}-O{start_row + 14})+(O{start_row + 16}-O{start_row + 18})),0)",
            f"P{start_row + 24}": f"=IFERROR((P{start_row + 12}-P{start_row + 14})/((P{start_row + 12}-P{start_row + 14})+(P{start_row + 16}-P{start_row + 18})),0)",
            f"Q{start_row + 24}": f"=IFERROR((Q{start_row + 12}-Q{start_row + 14})/((Q{start_row + 12}-Q{start_row + 14})+(Q{start_row + 16}-Q{start_row + 18})),0)",
            f"R{start_row + 24}": f"=IFERROR((R{start_row + 12}-R{start_row + 14})/((R{start_row + 12}-R{start_row + 14})+(R{start_row + 16}-R{start_row + 18})),0)",
            f"S{start_row + 24}": f"=IFERROR((S{start_row + 12}-S{start_row + 14})/((S{start_row + 12}-S{start_row + 14})+(S{start_row + 16}-S{start_row + 18})),0)",
            f"T{start_row + 24}": f"=IFERROR((T{start_row + 12}-T{start_row + 14})/((T{start_row + 12}-T{start_row + 14})+(T{start_row + 16}-T{start_row + 18})),0)",

            f"U{start_row + 24}": f"=IFERROR((U{start_row + 12}-U{start_row + 14})/((U{start_row + 12}-U{start_row + 14})+(U{start_row + 16}-U{start_row + 18}-U{start_row + 8})),0)",
            f"V{start_row + 24}": f"=IFERROR((V{start_row + 12}-V{start_row + 14})/((V{start_row + 12}-V{start_row + 14})+(V{start_row + 16}-V{start_row + 18}-V{start_row + 8})),0)",
            f"W{start_row + 24}": f"=IFERROR((W{start_row + 12}-W{start_row + 14})/((W{start_row + 12}-W{start_row + 14})+(W{start_row + 16}-W{start_row + 18})),0)",
            f"I{start_row + 25}": f"=IFERROR(IFERROR(I{start_row + 12}/I{start_row + 16},0),0)",
            f"J{start_row + 25}": f"=IFERROR(IFERROR(SUM(I{start_row + 12}:J{start_row + 12})/AVERAGE(I{start_row + 16}:J{start_row + 16}),0),0)",
            f"K{start_row + 25}": f"=IFERROR(IFERROR(SUM(I{start_row + 12}:K{start_row + 12})/AVERAGE(I{start_row + 16}:K{start_row + 16}),0),0)",
            f"L{start_row + 25}": f"=IFERROR(IFERROR(SUM(I{start_row + 12}:L{start_row + 12})/AVERAGE(I{start_row + 16}:L{start_row + 16}),0),0)",
            f"M{start_row + 25}": f"=IFERROR(IFERROR(SUM(I{start_row + 12}:M{start_row + 12})/AVERAGE(I{start_row + 16}:M{start_row + 16}),0),0)",
            f"N{start_row + 25}": f"=IFERROR(IFERROR(SUM(I{start_row + 12}:N{start_row + 12})/AVERAGE(I{start_row + 16}:N{start_row + 16}),0),0)",
            f"O{start_row + 25}": f"=IFERROR(IFERROR(SUM(I{start_row + 12}:O{start_row + 12})/AVERAGE(I{start_row + 16}:O{start_row + 16}),0),0)",
            f"P{start_row + 25}": f"=IFERROR(IFERROR(SUM(I{start_row + 12}:P{start_row + 12})/AVERAGE(I{start_row + 16}:P{start_row + 16}),0),0)",
            f"Q{start_row + 25}": f"=IFERROR(IFERROR(SUM(I{start_row + 12}:Q{start_row + 12})/AVERAGE(I{start_row + 16}:Q{start_row + 16}),0),0)",
            f"R{start_row + 25}": f"=IFERROR(IFERROR(SUM(I{start_row + 12}:R{start_row + 12})/AVERAGE(I{start_row + 16}:R{start_row + 16}),0),0)",
            f"S{start_row + 25}": f"=IFERROR(IFERROR(SUM(I{start_row + 12}:S{start_row + 12})/AVERAGE(I{start_row + 16}:S{start_row + 16}),0),0)",
            f"T{start_row + 25}": f"=IFERROR(IFERROR(SUM(I{start_row + 12}:T{start_row + 12})/AVERAGE(I{start_row + 16}:T{start_row + 16}),0),0)",
            f"U{start_row + 25}": f"=IFERROR(U{start_row + 12}/U{start_row + 16},0)",
            f"V{start_row + 25}": f"=IFERROR(V{start_row + 12}/V{start_row + 16},0)",
            f"W{start_row + 25}": f"=IFERROR(W{start_row + 12}/W{start_row + 16},0)",
            f"I{start_row + 26}": f"=IFERROR(I{start_row + 13}/I{start_row + 17},0)",
            f"J{start_row + 26}": f"=IFERROR(SUM(I{start_row + 13}:J{start_row + 13})/AVERAGE(I{start_row + 17}:J{start_row + 17}),0)",
            f"K{start_row + 26}": f"=IFERROR(SUM(I{start_row + 13}:K{start_row + 13})/AVERAGE(I{start_row + 17}:K{start_row + 17}),0)",
            f"L{start_row + 26}": f"=IFERROR(SUM(I{start_row + 13}:L{start_row + 13})/AVERAGE(I{start_row + 17}:L{start_row + 17}),0)",
            f"M{start_row + 26}": f"=IFERROR(SUM(I{start_row + 13}:M{start_row + 13})/AVERAGE(I{start_row + 17}:M{start_row + 17}),0)",
            f"N{start_row + 26}": f"=IFERROR(SUM(I{start_row + 13}:N{start_row + 13})/AVERAGE(I{start_row + 17}:N{start_row + 17}),0)",
            f"O{start_row + 26}": f"=IFERROR(SUM(I{start_row + 13}:O{start_row + 13})/AVERAGE(I{start_row + 17}:O{start_row + 17}),0)",
            f"P{start_row + 26}": f"=IFERROR(SUM(I{start_row + 13}:P{start_row + 13})/AVERAGE(I{start_row + 17}:P{start_row + 17}),0)",
            f"Q{start_row + 26}": f"=IFERROR(SUM(I{start_row + 13}:Q{start_row + 13})/AVERAGE(I{start_row + 17}:Q{start_row + 17}),0)",
            f"R{start_row + 26}": f"=IFERROR(SUM(I{start_row + 13}:R{start_row + 13})/AVERAGE(I{start_row + 17}:R{start_row + 17}),0)",
            f"S{start_row + 26}": f"=IFERROR(SUM(I{start_row + 13}:S{start_row + 13})/AVERAGE(I{start_row + 17}:S{start_row + 17}),0)",
            f"T{start_row + 26}": f"=IFERROR(SUM(I{start_row + 13}:T{start_row + 13})/AVERAGE(I{start_row + 17}:T{start_row + 17}),0)",
            f"U{start_row + 26}": f"=IFERROR(U{start_row + 13}/U{start_row + 17},0)",
            f"V{start_row + 26}": f"=IFERROR(V{start_row + 13}/V{start_row + 17},0)",
            f"W{start_row + 26}": f"=IFERROR(W{start_row + 13}/W{start_row + 17},0)",
            f"I{start_row + 36}": LY_MCOM_Unit_Sales['FEB'],
            f"J{start_row + 36}":LY_MCOM_Unit_Sales['MAR'] ,
            f"K{start_row + 36}":LY_MCOM_Unit_Sales['APR'] ,
            f"L{start_row + 36}": LY_MCOM_Unit_Sales['MAY'],
            f"M{start_row + 36}":LY_MCOM_Unit_Sales['JUN'] ,
            f"N{start_row + 36}":LY_MCOM_Unit_Sales['JUL'] ,
            f"O{start_row + 36}":LY_MCOM_Unit_Sales['AUG'] ,
            f"P{start_row + 36}":LY_MCOM_Unit_Sales['SEP'],
            f"Q{start_row + 36}":LY_MCOM_Unit_Sales['OCT'] ,
            f"R{start_row + 36}":LY_MCOM_Unit_Sales['NOV'],
            f"S{start_row + 36}":LY_MCOM_Unit_Sales['DEC'] ,
            f"T{start_row + 36}":LY_MCOM_Unit_Sales['JAN'] ,

            f"U{start_row + 36}": f"=SUM(I{start_row + 36}:T{start_row + 36})",
            f"V{start_row + 36}": f"=SUM(I{start_row + 36}:N{start_row + 36})",
            f"W{start_row + 36}": f"=SUM(O{start_row + 36}:T{start_row + 36})",
            f"I{start_row + 35}": f"=I{start_row + 34}-I{start_row + 36}",
            f"J{start_row + 35}": f"=J{start_row + 34}-J{start_row + 36}",
            f"K{start_row + 35}": f"=K{start_row + 34}-K{start_row + 36}",
            f"L{start_row + 35}": f"=L{start_row + 34}-L{start_row + 36}",
            f"M{start_row + 35}": f"=M{start_row + 34}-M{start_row + 36}",
            f"N{start_row + 35}": f"=N{start_row + 34}-N{start_row + 36}",
            f"O{start_row + 35}": f"=O{start_row + 34}-O{start_row + 36}",
            f"P{start_row + 35}": f"=P{start_row + 34}-P{start_row + 36}",
            f"Q{start_row + 35}": f"=Q{start_row + 34}-Q{start_row + 36}",
            f"R{start_row + 35}": f"=R{start_row + 34}-R{start_row + 36}",
            f"S{start_row + 35}": f"=S{start_row + 34}-S{start_row + 36}",
            f"T{start_row + 35}": f"=T{start_row + 34}-T{start_row + 36}",

            f"U{start_row + 35}": f"=IFERROR(SUM(I{start_row + 35}:T{start_row + 35}),0)",
            f"V{start_row + 35}": f"=IFERROR(SUM(I{start_row + 35}:N{start_row + 35}),0)",
            f"W{start_row + 35}": f"=IFERROR(SUM(O{start_row + 35}:T{start_row + 35}),0)",
            f"I{start_row + 27}": f"=IFERROR(IF(I{start_row + 13}+I{start_row + 35}=0,0,(I{start_row + 13}-I{start_row + 35})/I{start_row + 35}),1)",
            f"J{start_row + 27}": f"=IFERROR(IF(J{start_row + 13}+J{start_row + 35}=0,0,(J{start_row + 13}-J{start_row + 35})/J{start_row + 35}),1)",
            f"K{start_row + 27}": f"=IFERROR(IF(K{start_row + 13}+K{start_row + 35}=0,0,(K{start_row + 13}-K{start_row + 35})/K{start_row + 35}),1)",
            f"L{start_row + 27}": f"=IFERROR(IF(L{start_row + 13}+L{start_row + 35}=0,0,(L{start_row + 13}-L{start_row + 35})/L{start_row + 35}),1)",
            f"M{start_row + 27}": f"=IFERROR(IF(M{start_row + 13}+M{start_row + 35}=0,0,(M{start_row + 13}-M{start_row + 35})/M{start_row + 35}),1)",
            f"N{start_row + 27}": f"=IFERROR(IF(N{start_row + 13}+N{start_row + 35}=0,0,(N{start_row + 13}-N{start_row + 35})/N{start_row + 35}),1)",
            f"O{start_row + 27}": f"=IFERROR(IF(O{start_row + 13}+O{start_row + 35}=0,0,(O{start_row + 13}-O{start_row + 35})/O{start_row + 35}),1)",
            f"P{start_row + 27}": f"=IFERROR(IF(P{start_row + 13}+P{start_row + 35}=0,0,(P{start_row + 13}-P{start_row + 35})/P{start_row + 35}),1)",
            f"Q{start_row + 27}": f"=IFERROR(IF(Q{start_row + 13}+Q{start_row + 35}=0,0,(Q{start_row + 13}-Q{start_row + 35})/Q{start_row + 35}),1)",
            f"R{start_row + 27}": f"=IFERROR(IF(R{start_row + 13}+R{start_row + 35}=0,0,(R{start_row + 13}-R{start_row + 35})/R{start_row + 35}),1)",
            f"S{start_row + 27}": f"=IFERROR(IF(S{start_row + 13}+S{start_row + 35}=0,0,(S{start_row + 13}-S{start_row + 35})/S{start_row + 35}),1)",
            f"T{start_row + 27}": f"=IFERROR(IF(T{start_row + 13}+T{start_row + 35}=0,0,(T{start_row + 13}-T{start_row + 35})/T{start_row + 35}),1)",
            f"U{start_row + 27}": f"=IFERROR(IF(U{start_row + 13}+U{start_row + 35}=0,0,(U{start_row + 13}-U{start_row + 35})/U{start_row + 35}),1)",
            f"V{start_row + 27}": f"=IFERROR(IF(V{start_row + 13}+V{start_row + 35}=0,0,(V{start_row + 13}-V{start_row + 35})/V{start_row + 35}),1)",
            f"W{start_row + 27}": f"=IFERROR(IF(W{start_row + 13}+W{start_row + 35}=0,0,(W{start_row + 13}-W{start_row + 35})/W{start_row + 35}),1)",
            f"I{start_row + 28}": f"=IFERROR(IF(I{start_row + 36}+I{start_row + 14}=0,0,(I{start_row + 14}-I{start_row + 36})/I{start_row + 36}),1)",
            f"J{start_row + 28}": f"=IFERROR(IF(J{start_row + 36}+J{start_row + 14}=0,0,(J{start_row + 14}-J{start_row + 36})/J{start_row + 36}),1)",
            f"K{start_row + 28}": f"=IFERROR(IF(K{start_row + 36}+K{start_row + 14}=0,0,(K{start_row + 14}-K{start_row + 36})/K{start_row + 36}),1)",
            f"L{start_row + 28}": f"=IFERROR(IF(L{start_row + 36}+L{start_row + 14}=0,0,(L{start_row + 14}-L{start_row + 36})/L{start_row + 36}),1)",
            f"M{start_row + 28}": f"=IFERROR(IF(M{start_row + 36}+M{start_row + 14}=0,0,(M{start_row + 14}-M{start_row + 36})/M{start_row + 36}),1)",
            f"N{start_row + 28}": f"=IFERROR(IF(N{start_row + 36}+N{start_row + 14}=0,0,(N{start_row + 14}-N{start_row + 36})/N{start_row + 36}),1)",
            f"O{start_row + 28}": f"=IFERROR(IF(O{start_row + 36}+O{start_row + 14}=0,0,(O{start_row + 14}-O{start_row + 36})/O{start_row + 36}),1)",
            f"P{start_row + 28}": f"=IFERROR(IF(P{start_row + 36}+P{start_row + 14}=0,0,(P{start_row + 14}-P{start_row + 36})/P{start_row + 36}),1)",
            f"Q{start_row + 28}": f"=IFERROR(IF(Q{start_row + 36}+Q{start_row + 14}=0,0,(Q{start_row + 14}-Q{start_row + 36})/Q{start_row + 36}),1)",
            f"R{start_row + 28}": f"=IFERROR(IF(R{start_row + 36}+R{start_row + 14}=0,0,(R{start_row + 14}-R{start_row + 36})/R{start_row + 36}),1)",
            f"S{start_row + 28}": f"=IFERROR(IF(S{start_row + 36}+S{start_row + 14}=0,0,(S{start_row + 14}-S{start_row + 36})/S{start_row + 36}),1)",
            f"T{start_row + 28}": f"=IFERROR(IF(T{start_row + 36}+T{start_row + 14}=0,0,(T{start_row + 14}-T{start_row + 36})/T{start_row + 36}),1)",
            f"U{start_row + 28}": f"=IFERROR(IF(U{start_row + 36}+U{start_row + 14}=0,0,(U{start_row + 14}-U{start_row + 36})/U{start_row + 36}),1)",
            f"V{start_row + 28}": f"=IFERROR(IF(V{start_row + 36}+V{start_row + 14}=0,0,(V{start_row + 14}-V{start_row + 36})/V{start_row + 36}),1)",
            f"W{start_row + 28}": f"=IFERROR(IF(W{start_row + 36}+W{start_row + 14}=0,0,(W{start_row + 14}-W{start_row + 36})/W{start_row + 36}),1)",    
            f"I{start_row + 40}":LY_MCOM_OH_Units['FEB'] ,
            f"J{start_row + 40}":LY_MCOM_OH_Units['MAR'],
            f"K{start_row + 40}":LY_MCOM_OH_Units['APR'],
            f"L{start_row + 40}":LY_MCOM_OH_Units['MAY'],
            f"M{start_row + 40}":LY_MCOM_OH_Units['JUN'] ,
            f"N{start_row + 40}":LY_MCOM_OH_Units['JUL'],
            f"O{start_row + 40}":LY_MCOM_OH_Units['AUG'],
            f"P{start_row + 40}":LY_MCOM_OH_Units['SEP'] ,
            f"Q{start_row + 40}":LY_MCOM_OH_Units['OCT'] ,
            f"R{start_row + 40}":LY_MCOM_OH_Units['NOV'],
            f"S{start_row + 40}":LY_MCOM_OH_Units['DEC'],
            f"T{start_row + 40}":LY_MCOM_OH_Units['JAN'] ,

            f"U{start_row + 40}": f"=AVERAGEIFS(I{start_row + 40}:T{start_row + 40},$I$4:$T$4,\"<=\"&$T$4)",
            f"V{start_row + 40}": f"=AVERAGEIFS(I{start_row + 40}:N{start_row + 40},$I$4:$N$4,\"<=\"&$N$4)",
            f"W{start_row + 40}": f"=AVERAGEIFS(O{start_row + 40}:T{start_row + 40},$O$4:$T$4,\"<=\"&$T$4)",  
            f"I{start_row + 39}": f"=I{start_row + 38}-I{start_row + 40}",
            f"J{start_row + 39}": f"=J{start_row + 38}-J{start_row + 40}",
            f"K{start_row + 39}": f"=K{start_row + 38}-K{start_row + 40}",
            f"L{start_row + 39}": f"=L{start_row + 38}-L{start_row + 40}",
            f"M{start_row + 39}": f"=M{start_row + 38}-M{start_row + 40}",
            f"N{start_row + 39}": f"=N{start_row + 38}-N{start_row + 40}",
            f"O{start_row + 39}": f"=O{start_row + 38}-O{start_row + 40}",
            f"P{start_row + 39}": f"=P{start_row + 38}-P{start_row + 40}",
            f"Q{start_row + 39}": f"=Q{start_row + 38}-Q{start_row + 40}",
            f"R{start_row + 39}": f"=R{start_row + 38}-R{start_row + 40}",
            f"S{start_row + 39}": f"=S{start_row + 38}-S{start_row + 40}",
            f"T{start_row + 39}": f"=T{start_row + 38}-T{start_row + 40}",

            f"U{start_row + 39}": f"=AVERAGEIFS(I{start_row + 39}:T{start_row + 39},$I$4:$T$4,\"<=\"&$T$4)",
            f"V{start_row + 39}": f"=AVERAGEIFS(I{start_row + 39}:N{start_row + 39},$I$4:$N$4,\"<=\"&$N$4)",
            f"W{start_row + 39}": f"=AVERAGEIFS(O{start_row + 39}:T{start_row + 39},$O$4:$T$4,\"<=\"&$T$4)",
            f"I{start_row + 29}": f"=IFERROR(IF(I{start_row + 17}+I{start_row + 39}=0,0,(I{start_row + 17}-I{start_row + 39})/I{start_row + 39}),1)",
            f"J{start_row + 29}": f"=IFERROR(IF(J{start_row + 17}+J{start_row + 39}=0,0,(J{start_row + 17}-J{start_row + 39})/J{start_row + 39}),1)",
            f"K{start_row + 29}": f"=IFERROR(IF(K{start_row + 17}+K{start_row + 39}=0,0,(K{start_row + 17}-K{start_row + 39})/K{start_row + 39}),1)",
            f"L{start_row + 29}": f"=IFERROR(IF(L{start_row + 17}+L{start_row + 39}=0,0,(L{start_row + 17}-L{start_row + 39})/L{start_row + 39}),1)",
            f"M{start_row + 29}": f"=IFERROR(IF(M{start_row + 17}+M{start_row + 39}=0,0,(M{start_row + 17}-M{start_row + 39})/M{start_row + 39}),1)",
            f"N{start_row + 29}": f"=IFERROR(IF(N{start_row + 17}+N{start_row + 39}=0,0,(N{start_row + 17}-N{start_row + 39})/N{start_row + 39}),1)",
            f"O{start_row + 29}": f"=IFERROR(IF(O{start_row + 17}+O{start_row + 39}=0,0,(O{start_row + 17}-O{start_row + 39})/O{start_row + 39}),1)",
            f"P{start_row + 29}": f"=IFERROR(IF(P{start_row + 17}+P{start_row + 39}=0,0,(P{start_row + 17}-P{start_row + 39})/P{start_row + 39}),1)",
            f"Q{start_row + 29}": f"=IFERROR(IF(Q{start_row + 17}+Q{start_row + 39}=0,0,(Q{start_row + 17}-Q{start_row + 39})/Q{start_row + 39}),1)",
            f"R{start_row + 29}": f"=IFERROR(IF(R{start_row + 17}+R{start_row + 39}=0,0,(R{start_row + 17}-R{start_row + 39})/R{start_row + 39}),1)",
            f"S{start_row + 29}": f"=IFERROR(IF(S{start_row + 17}+S{start_row + 39}=0,0,(S{start_row + 17}-S{start_row + 39})/S{start_row + 39}),1)",
            f"T{start_row + 29}": f"=IFERROR(IF(T{start_row + 17}+T{start_row + 39}=0,0,(T{start_row + 17}-T{start_row + 39})/T{start_row + 39}),1)",
            f"U{start_row + 29}": f"=IFERROR(IF(U{start_row + 17}+U{start_row + 39}=0,0,(U{start_row + 17}-U{start_row + 39})/U{start_row + 39}),1)",
            f"V{start_row + 29}": f"=IFERROR(IF(V{start_row + 17}+V{start_row + 39}=0,0,(V{start_row + 17}-V{start_row + 39})/V{start_row + 39}),1)",
            f"W{start_row + 29}": f"=IFERROR(IF(W{start_row + 17}+W{start_row + 39}=0,0,(W{start_row + 17}-W{start_row + 39})/W{start_row + 39}),1)",
            f"I{start_row + 30}": OO_Total_Units['FEB'],
            f"J{start_row + 30}": OO_Total_Units['MAR'],
            f"K{start_row + 30}": OO_Total_Units['APR'],
            f"L{start_row + 30}": OO_Total_Units['MAY'],
            f"M{start_row + 30}":OO_Total_Units['JUN'],
            f"N{start_row + 30}":OO_Total_Units['JUL'],
            f"O{start_row + 30}": OO_Total_Units['AUG'],
            f"P{start_row + 30}": OO_Total_Units['SEP'],
            f"Q{start_row + 30}": OO_Total_Units['OCT'],
            f"R{start_row + 30}": OO_Total_Units['NOV'],
            f"S{start_row + 30}": OO_Total_Units['DEC'],
            f"T{start_row + 30}": OO_Total_Units['JAN'],

            f"U{start_row + 30}": f"=IFERROR(LOOKUP(2,1/($I${start_row}:$T${start_row}=$J$1),I{start_row + 30}:T{start_row + 30}),0)",
            f"V{start_row + 30}": f"=IFERROR(IFERROR(IFERROR(LOOKUP(2,1/($I${start_row}:$N${start_row}=$J$2),I{start_row + 30}:N{start_row + 30}),0),0),0)",
            f"W{start_row + 30}": f"=IFERROR(IFERROR(IFERROR(LOOKUP(2,1/($O${start_row}:$T${start_row}=$J$1),O{start_row + 30}:T{start_row + 30}),0),0),0)",
            f"I{start_row + 31}": OO_MCOM_Total_Units['FEB'],
            f"J{start_row + 31}": OO_MCOM_Total_Units['MAR'],
            f"K{start_row + 31}": OO_MCOM_Total_Units['APR'],
            f"L{start_row + 31}": OO_MCOM_Total_Units['MAY'],
            f"M{start_row + 31}": OO_MCOM_Total_Units['JUN'],
            f"N{start_row + 31}": OO_MCOM_Total_Units['JUL'],
            f"O{start_row + 31}": OO_MCOM_Total_Units['AUG'],
            f"P{start_row + 31}":OO_MCOM_Total_Units['SEP'],
            f"Q{start_row + 31}": OO_MCOM_Total_Units['OCT'],
            f"R{start_row + 31}": OO_MCOM_Total_Units['NOV'],
            f"S{start_row + 31}": OO_MCOM_Total_Units['DEC'],
            f"T{start_row + 31}":OO_MCOM_Total_Units['JAN'],

            f"U{start_row + 31}": f"=IFERROR(LOOKUP(2,1/($I${start_row}:$T${start_row}=$J$1),I{start_row + 31}:T{start_row + 31}),0)",
            f"V{start_row + 31}": f"=IFERROR(IFERROR(IFERROR(LOOKUP(2,1/($I${start_row}:$N${start_row}=$J$2),I{start_row + 31}:N{start_row + 31}),0),0),0)",
            f"W{start_row + 31}": f"=IFERROR(IFERROR(IFERROR(LOOKUP(2,1/($O${start_row}:$T${start_row}=$J$1),O{start_row + 31}:T{start_row + 31}),0),0),0)",
            f"I{start_row + 37}": f"=IFERROR(I{start_row + 36}/I{start_row + 34},0)",
            f"J{start_row + 37}": f"=IFERROR(J{start_row + 36}/J{start_row + 34},0)",
            f"K{start_row + 37}": f"=IFERROR(K{start_row + 36}/K{start_row + 34},0)",
            f"L{start_row + 37}": f"=IFERROR(L{start_row + 36}/L{start_row + 34},0)",
            f"M{start_row + 37}": f"=IFERROR(M{start_row + 36}/M{start_row + 34},0)",
            f"N{start_row + 37}": f"=IFERROR(N{start_row + 36}/N{start_row + 34},0)",
            f"O{start_row + 37}": f"=IFERROR(O{start_row + 36}/O{start_row + 34},0)",
            f"P{start_row + 37}": f"=IFERROR(P{start_row + 36}/P{start_row + 34},0)",
            f"Q{start_row + 37}": f"=IFERROR(Q{start_row + 36}/Q{start_row + 34},0)",
            f"R{start_row + 37}": f"=IFERROR(R{start_row + 36}/R{start_row + 34},0)",
            f"S{start_row + 37}": f"=IFERROR(S{start_row + 36}/S{start_row + 34},0)",
            f"T{start_row + 37}": f"=IFERROR(T{start_row + 36}/T{start_row + 34},0)",
            f"U{start_row + 37}": f"=IFERROR(U{start_row + 36}/U{start_row + 34},0)",
            f"V{start_row + 37}": f"=IFERROR(V{start_row + 36}/V{start_row + 34},0)",
            f"W{start_row + 37}": f"=IFERROR(W{start_row + 36}/W{start_row + 34},0)",
            f"I{start_row + 41}": f"=IFERROR(I{start_row + 40}/I{start_row + 38},0)",
            f"J{start_row + 41}": f"=IFERROR(J{start_row + 40}/J{start_row + 38},0)",
            f"K{start_row + 41}": f"=IFERROR(K{start_row + 40}/K{start_row + 38},0)",
            f"L{start_row + 41}": f"=IFERROR(L{start_row + 40}/L{start_row + 38},0)",
            f"M{start_row + 41}": f"=IFERROR(M{start_row + 40}/M{start_row + 38},0)",
            f"N{start_row + 41}": f"=IFERROR(N{start_row + 40}/N{start_row + 38},0)",
            f"O{start_row + 41}": f"=IFERROR(O{start_row + 40}/O{start_row + 38},0)",
            f"P{start_row + 41}": f"=IFERROR(P{start_row + 40}/P{start_row + 38},0)",
            f"Q{start_row + 41}": f"=IFERROR(Q{start_row + 40}/Q{start_row + 38},0)",
            f"R{start_row + 41}": f"=IFERROR(R{start_row + 40}/R{start_row + 38},0)",
            f"S{start_row + 41}": f"=IFERROR(S{start_row + 40}/S{start_row + 38},0)",
            f"T{start_row + 41}": f"=IFERROR(T{start_row + 40}/T{start_row + 38},0)",
            f"U{start_row + 41}": f"=IFERROR(U{start_row + 40}/U{start_row + 38},0)",
            f"V{start_row + 41}": f"=IFERROR(V{start_row + 40}/V{start_row + 38},0)",
            f"W{start_row + 41}": f"=IFERROR(W{start_row + 40}/W{start_row + 38},0)",
            f"I{start_row + 42}":LY_Receipts['FEB'] ,
            f"J{start_row + 42}":LY_Receipts['MAR'],
            f"K{start_row + 42}":LY_Receipts['APR'],
            f"L{start_row + 42}": LY_Receipts['MAY'],
            f"M{start_row + 42}":LY_Receipts['JUN'],
            f"N{start_row + 42}": LY_Receipts['JUL'],
            f"O{start_row + 42}": LY_Receipts['AUG'],
            f"P{start_row + 42}": LY_Receipts['SEP'],
            f"Q{start_row + 42}": LY_Receipts['OCT'],
            f"R{start_row + 42}":LY_Receipts['NOV'],
            f"S{start_row + 42}":LY_Receipts['DEC'],
            f"T{start_row + 42}": LY_Receipts['JAN'],

            f"U{start_row + 42}": f"=SUM(I{start_row + 42}:T{start_row + 42})",
            f"V{start_row + 42}": f"=SUM(I{start_row + 42}:N{start_row + 42})",
            f"W{start_row + 42}": f"=SUM(O{start_row + 42}:T{start_row + 42})",
            f"I{start_row + 43}": f"=IFERROR(IFERROR(I{start_row + 34}/(I{start_row + 34}+I{start_row + 38}),0),0)",
            f"J{start_row + 43}": f"=IFERROR(IFERROR(J{start_row + 34}/(J{start_row + 34}+J{start_row + 38}),0),0)",
            f"K{start_row + 43}": f"=IFERROR(IFERROR(K{start_row + 34}/(K{start_row + 34}+K{start_row + 38}),0),0)",
            f"L{start_row + 43}": f"=IFERROR(IFERROR(L{start_row + 34}/(L{start_row + 34}+L{start_row + 38}),0),0)",
            f"M{start_row + 43}": f"=IFERROR(IFERROR(M{start_row + 34}/(M{start_row + 34}+M{start_row + 38}),0),0)",
            f"N{start_row + 43}": f"=IFERROR(IFERROR(N{start_row + 34}/(N{start_row + 34}+N{start_row + 38}),0),0)",
            f"O{start_row + 43}": f"=IFERROR(IFERROR(O{start_row + 34}/(O{start_row + 34}+O{start_row + 38}),0),0)",
            f"P{start_row + 43}": f"=IFERROR(IFERROR(P{start_row + 34}/(P{start_row + 34}+P{start_row + 38}),0),0)",
            f"Q{start_row + 43}": f"=IFERROR(IFERROR(Q{start_row + 34}/(Q{start_row + 34}+Q{start_row + 38}),0),0)",
            f"R{start_row + 43}": f"=IFERROR(IFERROR(R{start_row + 34}/(R{start_row + 34}+R{start_row + 38}),0),0)",
            f"S{start_row + 43}": f"=IFERROR(IFERROR(S{start_row + 34}/(S{start_row + 34}+S{start_row + 38}),0),0)",
            f"T{start_row + 43}": f"=IFERROR(IFERROR(T{start_row + 34}/(T{start_row + 34}+T{start_row + 38}),0),0)",
            f"U{start_row + 43}": f"=IFERROR(IFERROR(U{start_row + 34}/(U{start_row + 34}+U{start_row + 38}),0),0)",
            f"V{start_row + 43}": f"=IFERROR(IFERROR(V{start_row + 34}/(V{start_row + 34}+V{start_row + 38}),0),0)",
            f"W{start_row + 43}": f"=IFERROR(IFERROR(W{start_row + 34}/(W{start_row + 34}+W{start_row + 38}),0),0)",
            f"I{start_row + 44}": f"=IFERROR((I{start_row + 34}-I{start_row + 36})/((I{start_row + 34}-I{start_row + 36})+(I{start_row + 38}-I{start_row + 40})),0)",
            f"J{start_row + 44}": f"=IFERROR((J{start_row + 34}-J{start_row + 36})/((J{start_row + 34}-J{start_row + 36})+(J{start_row + 38}-J{start_row + 40})),0)",
            f"K{start_row + 44}": f"=IFERROR((K{start_row + 34}-K{start_row + 36})/((K{start_row + 34}-K{start_row + 36})+(K{start_row + 38}-K{start_row + 40})),0)",
            f"L{start_row + 44}": f"=IFERROR((L{start_row + 34}-L{start_row + 36})/((L{start_row + 34}-L{start_row + 36})+(L{start_row + 38}-L{start_row + 40})),0)",
            f"M{start_row + 44}": f"=IFERROR((M{start_row + 34}-M{start_row + 36})/((M{start_row + 34}-M{start_row + 36})+(M{start_row + 38}-M{start_row + 40})),0)",
            f"N{start_row + 44}": f"=IFERROR((N{start_row + 34}-N{start_row + 36})/((N{start_row + 34}-N{start_row + 36})+(N{start_row + 38}-N{start_row + 40})),0)",
            f"O{start_row + 44}": f"=IFERROR((O{start_row + 34}-O{start_row + 36})/((O{start_row + 34}-O{start_row + 36})+(O{start_row + 38}-O{start_row + 40})),0)",
            f"P{start_row + 44}": f"=IFERROR((P{start_row + 34}-P{start_row + 36})/((P{start_row + 34}-P{start_row + 36})+(P{start_row + 38}-P{start_row + 40})),0)",
            f"Q{start_row + 44}": f"=IFERROR((Q{start_row + 34}-Q{start_row + 36})/((Q{start_row + 34}-Q{start_row + 36})+(Q{start_row + 38}-Q{start_row + 40})),0)",
            f"R{start_row + 44}": f"=IFERROR((R{start_row + 34}-R{start_row + 36})/((R{start_row + 34}-R{start_row + 36})+(R{start_row + 38}-R{start_row + 40})),0)",
            f"S{start_row + 44}": f"=IFERROR((S{start_row + 34}-S{start_row + 36})/((S{start_row + 34}-S{start_row + 36})+(S{start_row + 38}-S{start_row + 40})),0)",
            f"T{start_row + 44}": f"=IFERROR((T{start_row + 34}-T{start_row + 36})/((T{start_row + 34}-T{start_row + 36})+(T{start_row + 38}-T{start_row + 40})),0)",
            f"U{start_row + 44}": f"=IFERROR((U{start_row + 34}-U{start_row + 36})/((U{start_row + 34}-U{start_row + 36})+(U{start_row + 38}-U{start_row + 40})),0)",
            f"V{start_row + 44}": f"=IFERROR((V{start_row + 34}-V{start_row + 36})/((V{start_row + 34}-V{start_row + 36})+(V{start_row + 38}-V{start_row + 40})),0)",
            f"W{start_row + 44}": f"=IFERROR((W{start_row + 34}-W{start_row + 36})/((W{start_row + 34}-W{start_row + 36})+(W{start_row + 38}-W{start_row + 40})),0)",
            f"I{start_row + 45}": f"=IFERROR(IFERROR(I{start_row + 34}/I{start_row + 38},0),0)",
            f"J{start_row + 45}": f"=IFERROR(IFERROR(SUM(I{start_row + 34}:J{start_row + 34})/AVERAGE(I{start_row + 38}:J{start_row + 38}),0),0)",
            f"K{start_row + 45}": f"=IFERROR(IFERROR(SUM(I{start_row + 34}:K{start_row + 34})/AVERAGE(I{start_row + 38}:K{start_row + 38}),0),0)",
            f"L{start_row + 45}": f"=IFERROR(IFERROR(SUM(I{start_row + 34}:L{start_row + 34})/AVERAGE(I{start_row + 38}:L{start_row + 38}),0),0)",
            f"M{start_row + 45}": f"=IFERROR(IFERROR(SUM(I{start_row + 34}:M{start_row + 34})/AVERAGE(I{start_row + 38}:M{start_row + 38}),0),0)",
            f"N{start_row + 45}": f"=IFERROR(IFERROR(SUM(I{start_row + 34}:N{start_row + 34})/AVERAGE(I{start_row + 38}:N{start_row + 38}),0),0)",
            f"O{start_row + 45}": f"=IFERROR(IFERROR(SUM(I{start_row + 34}:O{start_row + 34})/AVERAGE(I{start_row + 38}:O{start_row + 38}),0),0)",
            f"P{start_row + 45}": f"=IFERROR(IFERROR(SUM(I{start_row + 34}:P{start_row + 34})/AVERAGE(I{start_row + 38}:P{start_row + 38}),0),0)",
            f"Q{start_row + 45}": f"=IFERROR(IFERROR(SUM(I{start_row + 34}:Q{start_row + 34})/AVERAGE(I{start_row + 38}:Q{start_row + 38}),0),0)",
            f"R{start_row + 45}": f"=IFERROR(IFERROR(SUM(I{start_row + 34}:R{start_row + 34})/AVERAGE(I{start_row + 38}:R{start_row + 38}),0),0)",
            f"S{start_row + 45}": f"=IFERROR(IFERROR(SUM(I{start_row + 34}:S{start_row + 34})/AVERAGE(I{start_row + 38}:S{start_row + 38}),0),0)",
            f"T{start_row + 45}": f"=IFERROR(IFERROR(SUM(I{start_row + 34}:T{start_row + 34})/AVERAGE(I{start_row + 38}:T{start_row + 38}),0),0)",
            f"U{start_row + 45}": f"=IFERROR(U{start_row + 34}/U{start_row + 38},0)",
            f"V{start_row + 45}": f"=IFERROR(V{start_row + 34}/V{start_row + 38},0)",
            f"W{start_row + 45}": f"=IFERROR(W{start_row + 34}/W{start_row + 38},0)",
            f"I{start_row + 46}": f"=IFERROR(IFERROR(I{start_row + 35}/I{start_row + 39},0),0)",
            f"J{start_row + 46}": f"=IFERROR(IFERROR(SUM(I{start_row + 35}:J{start_row + 35})/AVERAGE(I{start_row + 39}:J{start_row + 39}),0),0)",
            f"K{start_row + 46}": f"=IFERROR(IFERROR(SUM(I{start_row + 35}:K{start_row + 35})/AVERAGE(I{start_row + 39}:K{start_row + 39}),0),0)",
            f"L{start_row + 46}": f"=IFERROR(IFERROR(SUM(I{start_row + 35}:L{start_row + 35})/AVERAGE(I{start_row + 39}:L{start_row + 39}),0),0)",
            f"M{start_row + 46}": f"=IFERROR(IFERROR(SUM(I{start_row + 35}:M{start_row + 35})/AVERAGE(I{start_row + 39}:M{start_row + 39}),0),0)",
            f"N{start_row + 46}": f"=IFERROR(IFERROR(SUM(I{start_row + 35}:N{start_row + 35})/AVERAGE(I{start_row + 39}:N{start_row + 39}),0),0)",
            f"O{start_row + 46}": f"=IFERROR(IFERROR(SUM(I{start_row + 35}:O{start_row + 35})/AVERAGE(I{start_row + 39}:O{start_row + 39}),0),0)",
            f"P{start_row + 46}": f"=IFERROR(IFERROR(SUM(I{start_row + 35}:P{start_row + 35})/AVERAGE(I{start_row + 39}:P{start_row + 39}),0),0)",
            f"Q{start_row + 46}": f"=IFERROR(IFERROR(SUM(I{start_row + 35}:Q{start_row + 35})/AVERAGE(I{start_row + 39}:Q{start_row + 39}),0),0)",
            f"R{start_row + 46}": f"=IFERROR(IFERROR(SUM(I{start_row + 35}:R{start_row + 35})/AVERAGE(I{start_row + 39}:R{start_row + 39}),0),0)",
            f"S{start_row + 46}": f"=IFERROR(IFERROR(SUM(I{start_row + 35}:S{start_row + 35})/AVERAGE(I{start_row + 39}:S{start_row + 39}),0),0)",
            f"T{start_row + 46}": f"=IFERROR(IFERROR(SUM(I{start_row + 35}:T{start_row + 35})/AVERAGE(I{start_row + 39}:T{start_row + 39}),0),0)",
            f"U{start_row + 46}": f"=IFERROR(U{start_row + 35}/U{start_row + 39},0)",
            f"V{start_row + 46}": f"=IFERROR(V{start_row + 35}/V{start_row + 39},0)",
            f"W{start_row + 46}": f"=IFERROR(W{start_row + 35}/W{start_row + 39},0)",
            f"I{start_row + 47}":LY_PTD_Sales['FEB']  ,
            f"J{start_row + 47}": LY_PTD_Sales['MAR'] ,
            f"K{start_row + 47}": LY_PTD_Sales['APR'] ,
            f"L{start_row + 47}":LY_PTD_Sales['MAY'] ,
            f"M{start_row + 47}":LY_PTD_Sales['JUN'] ,
            f"N{start_row + 47}":LY_PTD_Sales['JUL'] ,
            f"O{start_row + 47}": LY_PTD_Sales['AUG'] ,
            f"P{start_row + 47}": LY_PTD_Sales['SEP'] ,
            f"Q{start_row + 47}":LY_PTD_Sales['OCT'] ,
            f"R{start_row + 47}": LY_PTD_Sales['NOV'] ,
            f"S{start_row + 47}": LY_PTD_Sales['DEC'] ,
            f"T{start_row + 47}": LY_PTD_Sales['JAN'] ,

            f"U{start_row + 47}": f"=SUM(I{start_row + 47}:T{start_row + 47})",
            f"V{start_row + 47}": f"=SUM(I{start_row + 47}:N{start_row + 47})",
            f"W{start_row + 47}": f"=SUM(O{start_row + 47}:T{start_row + 47})",
            f"I{start_row + 48}":MCOM_PTD_LY_Sales['FEB'],
            f"J{start_row + 48}":MCOM_PTD_LY_Sales['MAR'],
            f"K{start_row + 48}": MCOM_PTD_LY_Sales['APR'],
            f"L{start_row + 48}":MCOM_PTD_LY_Sales['MAY'],
            f"M{start_row + 48}": MCOM_PTD_LY_Sales['JUN'],
            f"N{start_row + 48}": MCOM_PTD_LY_Sales['JUL'],
            f"O{start_row + 48}":MCOM_PTD_LY_Sales['AUG'],
            f"P{start_row + 48}": MCOM_PTD_LY_Sales['SEP'],
            f"Q{start_row + 48}":MCOM_PTD_LY_Sales['OCT'],
            f"R{start_row + 48}":MCOM_PTD_LY_Sales['NOV'],
            f"S{start_row + 48}":MCOM_PTD_LY_Sales['DEC'],
            f"T{start_row + 48}":MCOM_PTD_LY_Sales['JAN'],

            f"U{start_row + 48}": f"=SUM(I{start_row + 48}:T{start_row + 48})",
            f"V{start_row + 48}": f"=SUM(I{start_row + 48}:N{start_row + 48})",
            f"W{start_row + 48}": f"=SUM(O{start_row + 48}:T{start_row + 48})",
            f"I{start_row + 49}": f"=IFERROR(TEXT(IFERROR(I{start_row + 47}/I{start_row + 34},0),\"$0\") &\"/ \"&TEXT(((I{start_row + 47}/I{start_row + 34})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"J{start_row + 49}": f"=IFERROR(TEXT(IFERROR(J{start_row + 47}/J{start_row + 34},0),\"$0\") &\"/ \"&TEXT(((J{start_row + 47}/J{start_row + 34})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"K{start_row + 49}": f"=IFERROR(TEXT(IFERROR(K{start_row + 47}/K{start_row + 34},0),\"$0\") &\"/ \"&TEXT(((K{start_row + 47}/K{start_row + 34})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"L{start_row + 49}": f"=IFERROR(TEXT(IFERROR(L{start_row + 47}/L{start_row + 34},0),\"$0\") &\"/ \"&TEXT(((L{start_row + 47}/L{start_row + 34})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"M{start_row + 49}": f"=IFERROR(TEXT(IFERROR(M{start_row + 47}/M{start_row + 34},0),\"$0\") &\"/ \"&TEXT(((M{start_row + 47}/M{start_row + 34})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"N{start_row + 49}": f"=IFERROR(TEXT(IFERROR(N{start_row + 47}/N{start_row + 34},0),\"$0\") &\"/ \"&TEXT(((N{start_row + 47}/N{start_row + 34})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"O{start_row + 49}": f"=IFERROR(TEXT(IFERROR(O{start_row + 47}/O{start_row + 34},0),\"$0\") &\"/ \"&TEXT(((O{start_row + 47}/O{start_row + 34})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"P{start_row + 49}": f"=IFERROR(TEXT(IFERROR(P{start_row + 47}/P{start_row + 34},0),\"$0\") &\"/ \"&TEXT(((P{start_row + 47}/P{start_row + 34})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"Q{start_row + 49}": f"=IFERROR(TEXT(IFERROR(Q{start_row + 47}/Q{start_row + 34},0),\"$0\") &\"/ \"&TEXT(((Q{start_row + 47}/Q{start_row + 34})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"R{start_row + 49}": f"=IFERROR(TEXT(IFERROR(R{start_row + 47}/R{start_row + 34},0),\"$0\") &\"/ \"&TEXT(((R{start_row + 47}/R{start_row + 34})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"S{start_row + 49}": f"=IFERROR(TEXT(IFERROR(S{start_row + 47}/S{start_row + 34},0),\"$0\") &\"/ \"&TEXT(((S{start_row + 47}/S{start_row + 34})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"T{start_row + 49}": f"=IFERROR(TEXT(IFERROR(T{start_row + 47}/T{start_row + 34},0),\"$0\") &\"/ \"&TEXT(((T{start_row + 47}/T{start_row + 34})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"U{start_row + 49}": f"=IFERROR(TEXT(IFERROR(U{start_row + 47}/U{start_row + 34},0),\"$0\") &\"/ \"&TEXT(((U{start_row + 47}/U{start_row + 34})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"V{start_row + 49}": f"=IFERROR(TEXT(IFERROR(V{start_row + 47}/V{start_row + 34},0),\"$0\") &\"/ \"&TEXT(((V{start_row + 47}/V{start_row + 34})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"W{start_row + 49}": f"=IFERROR(TEXT(IFERROR(W{start_row + 47}/W{start_row + 34},0),\"$0\") &\"/ \"&TEXT(((W{start_row + 47}/W{start_row + 34})-$D{start_row + 19})/$D{start_row + 19},\"%0\"),0)",
            f"G{start_row + 3}": f"=IFERROR(U{start_row + 34}+U{start_row + 34}*F{start_row + 3},0)",  # Dynamic references from row 5 onward
            f"C{start_row + 7}": f"=IF(G1=\"SP\",V{start_row + 9},W{start_row + 9})",
            f"D{start_row + 7}": f"=IF($G$1=\"SP\",C{start_row + 7}-V{start_row + 32},C{start_row + 7}-W{start_row + 32})",
            f"E{start_row + 7}": f"=IF($G$1=\"SP\",C{start_row + 7}-V{start_row + 32},C{start_row + 7}-W{start_row + 32})",
            f"E{start_row + 9}": f"=SUMIFS(I{start_row + 5}:T{start_row + 5},$I$4:$T$4,\">=\"&$F$4,$I$4:$T$4,\"<=\"&$G$4)/SUMIFS($I$3:$T$3,$I$4:$T$4,\">=\"&$F$4,$I$4:$T$4,\"<=\"&$G$4)",  # Static references in rows 1 to 4
            f"C{start_row + 10}": f"=IFERROR(C{start_row + 15}/E{start_row + 9},0)",
            f"D{start_row + 10}": f"=IFERROR(C{start_row + 14}/E{start_row + 9},0)",
            f"C{start_row + 11}": f"=IFERROR((D{start_row + 10}-D{start_row + 9})*E{start_row + 9},0)",  # Dynamic D and E references from row 5 onward
            f"D{start_row + 11}": f"=C{start_row + 11}*C{start_row + 19}",  # C24 kept static because it was a specific reference
            f"F{start_row + 11}": f"=IFERROR(IF(U{start_row + 6}>C{start_row + 14},(U{start_row + 6}-C{start_row + 14}+C{start_row + 14}-F{start_row + 10}-C{start_row + 12})/E{start_row + 9},(C{start_row + 14}-F{start_row + 10}-C{start_row + 12})/E{start_row + 9}),0)",
            f"G{start_row + 11}": f"=F{start_row + 11}*E{start_row + 9}",
            f"G{start_row + 19}": f"=((U{start_row + 20}/U{start_row + 12})-C{start_row + 19})/(U{start_row + 20}/U{start_row + 12})",
            f"E{start_row + 33}": f"=ROUND((U{start_row + 6}-U{start_row + 8}-E{start_row + 8})/5, 0) * 5",
            f"A{start_row + 1}": f"=$C{start_row}&H{start_row + 1}",
            f"A{start_row + 2}": f"=$C{start_row}&H{start_row + 2}",
            f"A{start_row + 3}": f"=$C{start_row}&H{start_row + 3}",
            f"A{start_row + 4}": f"=$C{start_row}&H{start_row + 4}",
            f"A{start_row + 5}": f"=$C{start_row}&H{start_row + 5}",
            f"A{start_row + 6}": f"=$C{start_row}&H{start_row + 6}",
            f"A{start_row + 7}": f"=$C{start_row}&H{start_row + 7}",
            f"A{start_row + 8}": f"=$C{start_row}&H{start_row + 8}",
            f"A{start_row + 9}": f"=$C{start_row}&H{start_row + 9}",
            f"A{start_row + 10}": f"=$C{start_row}&H{start_row +10}",
            f"A{start_row + 11}": f"=$C{start_row + 11}&\"Excess Proj\"",
            f"A{start_row + 12}": f"=$C{start_row + 12}&\"Qty to Release\"",
            f"A{start_row + 13}": f"=$C{start_row}&H{start_row +12}",
            f"A{start_row + 14}": f"=$C{start_row}&H{start_row +13}",
            f"A{start_row + 15}": f"=$C{start_row}&H{start_row +14}",
            f"A{start_row + 16}": f"=$C{start_row}&H{start_row +16}",
            f"A{start_row + 17}": f"=$C{start_row}&H{start_row +17}",
            f"A{start_row + 18}": f"=$C{start_row}&H{start_row +18}",

            f"A{start_row + 20}": f"=$C{start_row}&H{start_row +20}",
            f"A{start_row + 21}": f"=$C{start_row}&H{start_row +21}",
            f"A{start_row + 22}": f"=$C{start_row}&H{start_row +22}",
            f"A{start_row + 23}": f"=$C{start_row}&H{start_row +23}",
            f"A{start_row + 24}": f"=$C{start_row}&H{start_row +24}",
            f"A{start_row + 25}": f"=$C{start_row}&H{start_row +25}",
            f"A{start_row + 26}": f"=$C{start_row}&H{start_row +26}",
            f"A{start_row + 27}": f"=$C{start_row}&H{start_row +27}",
            f"A{start_row + 28}": f"=$C{start_row}&H{start_row +28}",
            f"A{start_row + 29}": f"=$C{start_row}&H{start_row +29}",
            f"A{start_row + 33}": f"=$C{start_row + 33}&\"Qty to Enter\"",
            f"A{start_row + 35}": f"=$C{start_row}&B{start_row +34}",
            f"A{start_row + 40}": f"=$C{start_row}&B{start_row +39}",
            f"I{start_row + 5}":planned_fc["FEB"],
            f"J{start_row + 5}":planned_fc["MAR"],
            f"K{start_row + 5}":planned_fc["APR"],
            f"L{start_row + 5}":planned_fc["MAY"],
            f"M{start_row + 5}":planned_fc["JUN"],
            f"N{start_row + 5}":planned_fc["JUL"],
            f"O{start_row + 5}":planned_fc["AUG"],
            f"P{start_row + 5}":planned_fc["SEP"],
            f"Q{start_row + 5}":planned_fc["OCT"],
            f"R{start_row + 5}":planned_fc["NOV"],
            f"S{start_row + 5}":planned_fc["DEC"],
            f"T{start_row + 5}":planned_fc["JAN"],

                }
                # Add dropdown validation to the specific cell (e.g., F column) for each loop

        # Apply formulas
        percentage_ranges = [
        f"I{start_row + 1}:W{start_row + 1}",
        f"I{start_row + 10}:W{start_row + 10}",
        f"I{start_row + 15}:W{start_row + 15}",
        f"I{start_row + 19}:W{start_row + 19}",
        f"I{start_row + 23}:W{start_row + 23}",
        f"I{start_row + 24}:W{start_row + 24}",
        f"I{start_row + 27}:W{start_row + 27}",
        f"I{start_row +28}:W{start_row + 28}",
        f"I{start_row +29}:W{start_row + 29}",
        f"I{start_row + 37}:W{start_row + 37}",
        f"I{start_row + 41}:W{start_row + 41}",
        f"I{start_row + 43}:W{start_row + 43}",
        f"I{start_row + 44}:W{start_row + 44}",
        f"E{start_row + 2}",
        f"C{start_row + 3}",
        f"F{start_row + 19}",
        f"G{start_row + 19}",
        f"F{start_row + 3}",
    ]
        rounded_ranges = [
            f"F{start_row + 2}",
            f"G{start_row + 3}",
            f"E{start_row + 9}",
            f"C{start_row + 10}",
            f"D{start_row + 10}",
            f"C{start_row + 11}",
            f"F{start_row + 11}",
            f"F{start_row + 18}",
            f"U{start_row + 7}:W{start_row + 7}",
            f"U{start_row + 16}:W{start_row + 16}",
            f"U{start_row + 17}:W{start_row + 17}",
            f"U{start_row + 18}:W{start_row + 18}",
            f"I{start_row + 20}:W{start_row + 20}",
            f"I{start_row + 21}:W{start_row + 21}",
            f"U{start_row + 38}:W{start_row + 38}",
            f"U{start_row + 39}:W{start_row + 39}",
            f"U{start_row + 40}:W{start_row + 40}",


        ]
        rounded_ranges_one = [

            f"I{start_row + 25}:T{start_row + 25}",
            f"I{start_row + 26}:T{start_row + 26}",
            f"I{start_row + 45}:T{start_row + 45}",
            f"I{start_row + 46}:T{start_row + 46}",

        ]
        rounded_ranges_two = [
            f"U{start_row + 25}:W{start_row + 25}",
            f"U{start_row + 26}:W{start_row + 26}",
            f"U{start_row + 45}:W{start_row + 45}",
            f"U{start_row + 46}:W{start_row + 46}",




        ]
        currency_ranges = [
        f"D{start_row + 11}",
        f"C{start_row + 19}",
        f"D{start_row + 19}",
        f"C{start_row + 20}",
        f"D{start_row + 12}",
        f"E{start_row + 19}",
        f"I{start_row + 20}:W{start_row + 20}",
        f"I{start_row + 21}:W{start_row + 21}",
        f"I{start_row + 47}:W{start_row + 47}",
        f"I{start_row + 48}:W{start_row + 48}",
        
        # Add more ranges as needed
    ]
        sheet = ws
                # Add values to column B from ALL_VALUES and apply alignment
        for i, value in enumerate(ALL_VALUES, start=start_row):
            cell = ws.cell(row=i, column=2, value=value)  # Column B is the 2nd column
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Add values to column H from H_VALUES and apply alignment
        for i, value in enumerate(H_VALUES, start=start_row):
            cell = ws.cell(row=i, column=8, value=value)  # Column H is the 8th column
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Add monthly values starting from column I (9th column) in the specified row
        for col, value in enumerate(MONTHLY_VALUES, start=9):
            ws.cell(row=start_row, column=col, value=value)
        for cell, formula in dynamic_formulas.items():
            try:
                ws[cell] = formula  # Set the formula directly in the specified cell
            except Exception as e:
                logging.info(f"Error setting formula in {cell}: {e}")

        # Apply ROUND formatting with different decimal places
        apply_round_format(ws, rounded_ranges, decimal_places=0)
        apply_round_format(ws, rounded_ranges_one, decimal_places=1)
        apply_round_format(ws, rounded_ranges_two, decimal_places=2)
    # Apply formats
        apply_format(ws, percentage_ranges, "0%")        # Apply percentage format
        apply_format(ws, currency_ranges, "$#,##0") 
        # add_dropdown(ws, f"F{start_row + 1}", category_options)
        add_dropdown(ws, f"F{start_row + 4}", forecast_method_options)




    #######loop formatinf
        red_font= Font(color="FF0000") 
        blue_font=Font(color="0000FF") 
        white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        dark_pink = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        gray = PatternFill(start_color="EDEDED", end_color="EDEDED", fill_type="solid")
        pink_font = Font(color="9C0006") 
        for row in ws[f"C{start_row}:G{start_row + 8}"]:
            for cell in row:
                cell.fill = gray

        for row in ws[f"C{start_row + 13}:G{start_row + 20}"]:  # Use the dynamically constructed range
            for cell in row:
                cell.fill = gray
        for row in ws[f"C{start_row + 27}:G{start_row + 32}"]:  # Offset the range dynamically
            for cell in row:
                cell.fill = gray  # Apply the gray fill 

        ws[f"D{start_row + 1}"].fill = white  # D6 becomes D(start_row + 1)
        ws[f"E{start_row + 3}"].fill = white  # E8 becomes E(start_row + 3)

        red_font_list = [f"B{start_row + 40}", f"B{start_row}", f"D{start_row + 16}", f"F{start_row + 7}"]
        for cell_address in red_font_list:
            ws[cell_address].font = red_font

        blue_font_list = [
            f"B{start_row + 9}",  # B14 -> start_row + 9
            f"B{start_row + 10}",  # B15 -> start_row + 10
            f"B{start_row + 33}",  # B38 -> start_row + 33
            f"D{start_row + 9}",  # D14 -> start_row + 9
            f"E{start_row + 9}",  # E14 -> start_row + 9
            f"D{start_row + 35}",  # D40 -> start_row + 35
            f"C{start_row + 10}",  # C15 -> start_row + 10
            f"D{start_row + 10}"   # D15 -> start_row + 10
        ]

        pink_fill_font = [
            f"C{start_row + 3}",  # C8 -> start_row + 3
            f"F{start_row + 3}",  # F8 -> start_row + 3
            f"F{start_row + 14}",  # F19 -> start_row + 14
            f"G{start_row + 14}",  # G19 -> start_row + 14
            f"C{start_row + 16}",  # C21 -> start_row + 16
            f"C{start_row + 24}"   # C29 -> start_row + 24
        ]
        for cell_address in pink_fill_font:
            ws[cell_address].fill = dark_pink
            ws[cell_address].font = pink_font


        # For the range H28:W31
        for row in ws[f"H{start_row + 23}:W{start_row + 26}"]:  # H28:W31 -> start_row + 23 to start_row + 26
            for cell in row:
                cell.font = blue_font

        # For the range H48:W51
        for row in ws[f"H{start_row + 43}:W{start_row + 46}"]:  # H48:W51 -> start_row + 43 to start_row + 46
            for cell in row:
                cell.font = blue_font

        # For the range I14:W14
        for row in ws[f"I{start_row + 9}:W{start_row + 9}"]:  # I14:W14 -> start_row + 9
            for cell in row:
                cell.font = red_font  
        yelow_fill_font_list=['E13','F13','Q13','R13','S13','G39',]
        for i in yelow_fill_font_list:
            ws[i].fill = light_yellow
            ws[i].font = yellow_font
        #Hyperlink
        ws[f"B{start_row + 26}"].font = Font(color="0563C1", underline="single", bold=False)  # B31 -> start_row + 26

        # Dynamically merge cells
        ws.merge_cells(f"D{start_row + 21}:G{start_row + 26}")  # D26:G31
        ws.merge_cells(f"D{start_row + 11}:E{start_row + 11}")  # D16:E16
        ws.merge_cells(f"D{start_row + 1}:E{start_row + 1}")    # D6:E6
        ws.merge_cells(f"D{start_row}:E{start_row}")            # D5:E5
        ws.merge_cells(f"D{start_row + 20}:F{start_row + 20}")  # D25:F25
        ws.merge_cells(f"B{start_row + 35}:C{start_row + 38}")  # B40:C43
        ws.merge_cells(f"D{start_row + 35}:G{start_row + 38}")  # D40:G43
        ws.merge_cells(f"C{start_row + 39}:G{start_row + 39}")  # C44:G44
        ws.merge_cells(f"B{start_row + 40}:G{start_row + 49}")  # B45:G54
        ws.merge_cells(f"C{start_row + 13}:F{start_row + 13}")  # C18:F18
        ws.merge_cells(f"D{start_row + 16}:G{start_row + 16}")  # D21:G21
        ws.merge_cells(f"C{start_row + 30}:G{start_row + 30}")  # C35:G35
        # Dynamically set alignment
        ws[f"B{start_row + 40}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)  # B45
        ws[f"D{start_row + 21}"].alignment = Alignment(horizontal='center', vertical='center')  # D26
        ws[f"D{start_row + 11}"].alignment = Alignment(horizontal='center', vertical='center')  # D16
        ws[f"D{start_row}"].alignment = Alignment(horizontal='center', vertical='center')       # D5
        ws[f"D{start_row + 1}"].alignment = Alignment(horizontal='right', vertical='center')    # D6
        ws[f"C{start_row + 3}"].alignment = Alignment(horizontal='center', vertical='center')   # C8
        ws[f"D{start_row + 16}"].alignment = Alignment(horizontal='left', vertical='center')    # D21
        ws[f"D{start_row + 20}"].alignment = Alignment(horizontal='center', vertical='center')  # D25
        ws[f"C{start_row + 30}"].alignment = Alignment(horizontal='center', vertical='top')     # C35
        ws[f"D{start_row + 35}"].alignment = Alignment(horizontal='left', vertical='top')       # D40

        #gradient bg
        two_yellow_gradient_fill = GradientFill(type="linear",degree=90,stop=("FFFFFF", "FFFF99", "FFFFFF")) 
        top_green_gradient_fill = GradientFill(type="linear", degree=90, stop=("E2EFDA", "FFFFFF"))
        full_green_gradient_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        bott_green_gradient_fill = GradientFill(type="linear", degree=90, stop=("FFFFFF","E2EFDA"))
        bott_light_gray_gradient_fill = GradientFill(type="linear", degree=90, stop=( "FFFFFF","EDEDED"))
        top_gray_gradient_fill = GradientFill(type="linear", degree=90, stop=("D9D9D9", "FFFFFF"))
        dark_gray_gradient_fill = GradientFill(type="linear", degree=90, stop=( "FFFFFF","DBDBDB"))
        top_orange_gradient_fill = GradientFill(type="linear", degree=90, stop=("FFE699", "FFFFFF"))
        bott_orange_gradient_fill = GradientFill(type="linear", degree=90, stop=( "FFFFFF","FFE699"))
        top_yellow_gradient_fill = GradientFill(type="linear", degree=90, stop=("FFFFCC", "FFFFFF"))
        full_yellow_gradient_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        bott_yellow_gradient_fill = GradientFill(type="linear", degree=90, stop=("FFFFFF","FFFFCC"))
        bott_blue_gradient_fill = GradientFill(type="linear", degree=90, stop=("FFFFFF","DFECF7"))
        full_blue_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        top_pink_gradient_fill = GradientFill(type="linear", degree=90, stop=("FCE4D6", "FFFFFF"))
        full_pink_gradient_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        bott_pink_gradient_fill = GradientFill(type="linear", degree=90, stop=("FFFFFF","FCE4D6"))
        # Pink
        for row in ws[f"C{start_row + 9}:G{start_row + 9}"]:  # C14:G14 -> start_row + 9
            for cell in row:
                cell.fill = bott_pink_gradient_fill  

        for row in ws[f"C{start_row + 10}:G{start_row + 11}"]:  # C15:G16 -> start_row + 10 to start_row + 11
            for cell in row:
                cell.fill = full_pink_gradient_fill  

        for row in ws[f"C{start_row + 12}:G{start_row + 12}"]:  # C17:G17 -> start_row + 12
            for cell in row:
                cell.fill = top_pink_gradient_fill 

        # Yellow
        for row in ws[f"I{start_row + 4}:W{start_row + 4}"]:  # I9:W9 -> start_row + 4
            for cell in row:
                cell.fill = two_yellow_gradient_fill  

        # Green
        for row in ws[f"I{start_row + 5}:W{start_row + 5}"]:  # I10:W10 -> start_row + 5
            for cell in row:
                cell.fill = bott_green_gradient_fill  

        for row in ws[f"I{start_row + 6}:W{start_row + 6}"]:  # I11:W11 -> start_row + 6
            for cell in row:
                cell.fill = full_green_gradient_fill  

        for row in ws[f"I{start_row + 7}:W{start_row + 7}"]:  # I12:W12 -> start_row + 7
            for cell in row:
                cell.fill = top_green_gradient_fill  

        # Blue
        bott_blue_fill_list = [
            f"I{start_row + 12}:W{start_row + 12}",  # I17:W17 -> start_row + 12
            f"I{start_row + 20}:W{start_row + 20}",  # I25:W25 -> start_row + 20
            f"I{start_row + 34}:W{start_row + 34}",  # I39:W39 -> start_row + 34
            f"I{start_row + 47}:W{start_row + 47}"   # I52:W52 -> start_row + 47
        ]
        for i in bott_blue_fill_list:
            for row in ws[i]:
                for cell in row:
                    cell.fill = bott_blue_gradient_fill  
                    cell.font = Font(bold=True)  

        full_blue_fill_list = [
            f"I{start_row + 13}:W{start_row + 14}",  # I18:W19 -> start_row + 13 to start_row + 14
            f"I{start_row + 21}:W{start_row + 22}",  # I26:W27 -> start_row + 21 to start_row + 22
            f"I{start_row + 35}:W{start_row + 36}",  # I40:W41 -> start_row + 35 to start_row + 36
            f"I{start_row + 48}:W{start_row + 49}"   # I53:W54 -> start_row + 48 to start_row + 49
        ]
        for i in full_blue_fill_list:
            for row in ws[i]:
                for cell in row:
                    cell.fill = full_blue_fill  
        # Gray
        for row in ws[f"I{start_row + 8}:W{start_row + 8}"]:  # I13:W13 -> start_row + 8
            for cell in row:
                cell.fill = bott_light_gray_gradient_fill 

        for row in ws[f"I{start_row + 9}:W{start_row + 9}"]:  # I14:W14 -> start_row + 9
            for cell in row:
                cell.fill = top_gray_gradient_fill 

        dark_gray_fill_list = [
            f"H{start_row}:W{start_row}",       # H5:W5 -> start_row
            f"H{start_row + 11}:W{start_row + 11}",  # H16:W16 -> start_row + 11
            f"H{start_row + 33}:W{start_row + 33}"   # H38:W38 -> start_row + 33
        ]
        for i in dark_gray_fill_list:
            for row in ws[i]:
                for cell in row:
                    cell.fill = dark_gray_gradient_fill 
                    cell.font=Font(bold=True)

        # Individual cells
        ws[f"F{start_row + 1}"].fill = dark_gray_gradient_fill  # F6 -> start_row + 1
        ws[f"F{start_row + 4}"].fill = dark_gray_gradient_fill  # F9 -> start_row + 4
        ws[f"B{start_row + 26}"].fill = dark_gray_gradient_fill  # B31 -> start_row + 26

        # Yellow
        for row in ws[f"I{start_row + 27}:W{start_row + 27}"]:  # I32:W32 -> start_row + 27
            for cell in row:
                cell.fill = bott_yellow_gradient_fill  

        for row in ws[f"I{start_row + 28}:W{start_row + 28}"]:  # I33:W33 -> start_row + 28
            for cell in row:
                cell.fill = full_yellow_gradient_fill  

        for row in ws[f"I{start_row + 29}:W{start_row + 29}"]:  # I34:W34 -> start_row + 29
            for cell in row:
                cell.fill = top_yellow_gradient_fill  

        # Orange
        top_orange_fill_list = [
            f"I{start_row + 15}:W{start_row + 15}",  # I20:W20 -> start_row + 15
            f"I{start_row + 37}:W{start_row + 37}"  # I42:W42 -> start_row + 37
        ]
        bott_orange_fill_list = [
            f"I{start_row + 19}:W{start_row + 19}",  # I24:W24 -> start_row + 19
            f"I{start_row + 41}:W{start_row + 41}"  # I46:W46 -> start_row + 41
        ]

        # Apply top orange gradient fill
        for i in top_orange_fill_list:
            for row in ws[i]:
                for cell in row:
                    cell.fill = top_orange_gradient_fill 

        # Apply bottom orange gradient fill
        for i in bott_orange_fill_list:
            for row in ws[i]:
                for cell in row:
                    cell.fill = bott_orange_gradient_fill
        bold_list=['C5','D5','F5','C6','B26','B39','B44','C38','E38','G39','H20','H21','H24','H32','H33','H34','H46','H43','H42']
        for i in bold_list:
            ws[i].font = Font(bold=True)


        for row in ws[f"C{start_row}:W{start_row + 49}"]:  # Use the dynamically calculated range
            for cell in row:
                cell.border = gridline1  # Apply the border
                # Hide column A
        # Define border style for top and bottom only

        gridline_top_bottom = Border(
        left=Side(style="thin", color='FFFFFF'),
        right=Side(style="thin", color='FFFFFF'),
        top=Side(style="thin", color='D9D9D9'),
        bottom=Side(style="thin", color='D9D9D9')
    )

        for row in ws[f"B{start_row+50}:W{start_row+50}"]:
            for cell in row:
                cell.border = gridline_top_bottom
    # Define border style for left and right only

        
    ws.column_dimensions.group("A", "A", outline_level=1, hidden=True)

    wb.save(output_file)  # Save the workbook as an .xlsx file
    logging.info(f"Workbook '{output_file}' saved successfully.")

    result[output_file] = {"All products":all_products,"Com products":com_products,"Store products":store_products,"Pid_to_review": pids_below_door_count_alert, "Birthstone products": all_birthstone_products,  "Pid_to_notify_to_macys": notify_macys_alert,"Upcoming birthstone products": upcoming_birthstone_products,"Added quatity using macys proj receipts":added_macys_proj_receipts_alert,"Quantity below min order":min_order_alert}
    

    return result


def convert_to_tuples(categories):
    return [(category['name'], category['value']) for category in categories]


def process_data(input_path,file_path,month_from,month_to,percentage,input_tuple):

    print("input_path------>",input_path)

    sheet_config = {
        "Index": {"usecols": "A:P", "nrows": 41, "header": 2},
        "report grouping": {"header": None},
        "Repln Items": {"header": 2},
        "Setup Sales -L3M & Future": {"header": 9},
        "Macys Recpts": {"header": 1},
        "All_DATA": {"header": 0},
        "MCOM_Data": {"header": 0},
    }

    # Create a multiprocessing pool
    with mp.Pool(processes=mp.cpu_count()) as pool:
        # Prepare arguments for multiprocessing
        pool_args = [(sheet_name, config, input_path) for sheet_name, config in sheet_config.items()]
        
        # Process each sheet in parallel
        output = pool.starmap(process_sheet, pool_args)

    # Collect results into a dictionary
    dataframes = {sheet_name: data for sheet_name, data in output if data is not None}
    print("All sheets processed successfully!")

    # Store the results into specified variables
    index_df = dataframes.get("Index")
    report_grouping_df = dataframes.get("report grouping")
    planning_df = dataframes.get("Repln Items")
    TBL_Planning_VerticalReport__3 = dataframes.get("Setup Sales -L3M & Future")
    Macys_Recpts = dataframes.get("Macys Recpts")
    All_DATA = dataframes.get("All_DATA")
    MCOM_Data = dataframes.get("MCOM_Data")

    # Debugging: Print the results for verification
    if index_df is not None:
        print(f"Index sheet processed with {len(index_df)} rows.")
    if report_grouping_df is not None:
        print(f"Report Grouping sheet processed with {len(report_grouping_df)} rows.")
    if planning_df is not None:
        print(f"Planning sheet processed with {len(planning_df)} rows.")
    if TBL_Planning_VerticalReport__3 is not None:
        print(f"TBL_Planning_VerticalReport__3 sheet processed with {len(TBL_Planning_VerticalReport__3)} rows.")
    if Macys_Recpts is not None:
        print(f"Macys Recpts sheet processed with {len(Macys_Recpts)} rows.")
    if All_DATA is not None:
        print(f"All_DATA sheet processed with {len(All_DATA)} rows.")
    if MCOM_Data is not None:
        print(f"MCOM_Data sheet processed with {len(MCOM_Data)} rows.")

    #make master sheet# Specify the columns you want to extract
    columns_to_extract = [
        'PID', 'Cross ref', 'Sort', 'Dpt ID', 'Dpt Desc', 'Mkst', 'PID Desc',
        'CL ID', 'Class Desc', 'SC ID', 'SC Desc', 'MstrSt ID', 'Masterstyle Desc',
        'TY Last Cost', 'Own Retail', 'AWR 1st Tkt Ret', 'Prod Desc', 'Grouping',
        'Vendor', 'Vendor Name', 'FC Index', 'FLDC','Safe/Non-Safe', 'Item Code'
    ]

    # Extract the specified columns
    df_filtered = planning_df[columns_to_extract]

    # Create a new 'Gender' column based on the conditions
    df_filtered['Gender'] = 'Women'  # Default value
    df_filtered.loc[df_filtered['Dpt ID'].isin([768, 771]), 'Gender'] = 'Men'
    df_filtered.loc[df_filtered['CL ID'] == 86, 'Gender'] = 'Children'


    # List of birthstone names
    birthstones = [
        'GARNET', 'AMETHYST', 'AQUAMARINE', 'DIAMOND', 'EMERALD', 'PEARL', 
        'RUBY', 'PERIDOT', 'SAPPHIRE', 'OPAL', 'CITRINE', 'TANZANITE'
    ]

    # Function to check if 'Class Desc' contains any birthstone name
    def find_birthstone(class_desc):
        if isinstance(class_desc, str):  # Ensure the value is a string
            for stone in birthstones:
                if stone in class_desc.upper():  # Check if the birthstone is part of the string
                    return stone
        return ''  # Return blank if no birthstone is found

    # Apply the function to create the 'Birthstone' column
    df_filtered['Birthstone'] = df_filtered['Class Desc'].apply(find_birthstone)

    # Create the 'BSP_or_not' column based on the 'MstrSt ID' condition
    df_filtered['BSP_or_not'] = df_filtered['MstrSt ID'].apply(lambda x: 'BSP' if x in [26481, 74692] else '')


    def categorize_product(class_desc):
        if isinstance(class_desc, str):  # Ensure the value is a string
            class_desc = class_desc.upper()
            if 'EARRINGS' in class_desc or 'EARS' in class_desc:
                return 'Earrings'
            elif 'PENDANTS' in class_desc or 'PENDS' in class_desc:
                return 'Pendants'
            elif ' RING' in class_desc:  # Notice the space before "RINGS"
                return 'Rings'
            elif 'BRACELETS' in class_desc:
                return 'Bracelets'
            elif 'CHAIN' in class_desc:
                return 'Necklace'
            elif 'NECKS' in class_desc or 'NECKLACE' in class_desc:
                return 'Necklace'
            elif 'SET' in class_desc:
                return 'Set'
            elif 'BAND' in class_desc:
                return 'BAND'
        return ''  # Return blank if no match found

    # Apply the function to create the 'category' column
    df_filtered['category'] = df_filtered['Class Desc'].apply(categorize_product)


    def update_category_from_prod_desc(row):
        if not row['category']:  # If category is blank
            prod_desc = str(row['Prod Desc']).lower()
            categories_found = []

            # Search for keywords in Prod desc
            if 'earring' in prod_desc or 'stud' in prod_desc:
                categories_found.append('Earrings')
            if ' ring' in prod_desc:
                categories_found.append('Rings')
            if 'pendant' in prod_desc or 'necklace' in prod_desc:
                categories_found.append('Pendants' if 'pendant' in prod_desc else 'Necklace')
            if 'bracelet' in prod_desc:
                categories_found.append('Bracelets')
            if 'band' in prod_desc:
                categories_found.append('Band')
            if 'locket' in prod_desc:
                categories_found.append('Locket')
            # If multiple categories are found, set as 'Set'
            if len(categories_found) > 1:
                return 'Set'
            elif categories_found:
                return categories_found[0]
        return row['category']  # Return the original category if not blank

    # Apply the update function
    df_filtered['category'] = df_filtered.apply(update_category_from_prod_desc, axis=1)


    # Create the 'type' column based on conditions in 'Prod desc'
    def determine_type(prod_desc):
        if isinstance(prod_desc, str):  # Ensure the value is a string
            prod_desc = prod_desc.lower()
            if 'heart' in prod_desc:
                return 'Heart'
            elif 'stud' in prod_desc:
                return 'Studs'
            elif 'drop' in prod_desc:
                return 'Drop'
        return ''  # Return blank if no match found

    # Apply the function to create the 'type' column
    df_filtered['type'] = df_filtered['Prod Desc'].apply(determine_type)

    birthstone_data = [
        (1, 'January', 'Garnet'),
        (2, 'February', 'Amethyst'),
        (3, 'March', 'Aquamarine'),
        (4, 'April', 'Diamond'),
        (5, 'May', 'Emerald'),
        (6, 'June', 'Pearl'),
        (7, 'July', 'Ruby'),
        (8, 'August', 'Peridot'),
        (9, 'September', 'Sapphire'),
        (10, 'October', 'Opal'),
        (11, 'November', 'Citrine'),
        (12, 'December', 'Tanzanite')
    ]

    # Create a DataFrame from the data
    birthstone_sheet = pd.DataFrame(birthstone_data, columns=['Month', 'Month Name', 'Birthstone'])


    master_sheet = df_filtered[['PID','category','Birthstone','BSP_or_not','type','Gender','Vendor','Vendor Name','Own Retail','FC Index', 'FLDC','Safe/Non-Safe', 'Item Code']]

    # Create the unique sheets

    df_unique_vendor = df_filtered[['Vendor', 'Vendor Name']].drop_duplicates().dropna()


    data = [
        {"Vendor Name": "ALMOND (THAILAND) LIMITED", "Country of Origin": "Thailand", "Lead Time(weeks)": 8},
        {"Vendor Name": "AMMANTE JEWELLS LLP", "Country of Origin": "India", "Lead Time(weeks)": 8},
        {"Vendor Name": "ARIN", "Country of Origin": "Peru", "Lead Time(weeks)": 9},
        {"Vendor Name": "ARPAS INTERNATIONAL LTD", "Country of Origin": "US", "Lead Time(weeks)": 9},
        {"Vendor Name": "ASIAN STAR COMPANY LTD", "Country of Origin": "India", "Lead Time(weeks)": 8},
        {"Vendor Name": "BEAUTY GEMS FACTORY CO LTD", "Country of Origin": "Thailand", "Lead Time(weeks)": 8},
        {"Vendor Name": "CHARM AMERICA", "Country of Origin": "US", "Lead Time(weeks)": 9},
        {"Vendor Name": "CREATIONS JEWELLERY MFG. PVT. LTD.", "Country of Origin": "India", "Lead Time(weeks)": 8},
        {"Vendor Name": "DAISY'S COLLECTION, INC", "Country of Origin": "China", "Lead Time(weeks)": 8},
        {"Vendor Name": "DEORO - WIRE", "Country of Origin": "Peru", "Lead Time(weeks)": 9},
        {"Vendor Name": "DIA GOLD CREATION PVT. LTD.", "Country of Origin": "India", "Lead Time(weeks)": 8},
        {"Vendor Name": "DRL Manufacturing - LG", "Country of Origin": "DRL", "Lead Time(weeks)": 10},
        {"Vendor Name": "DRL MANUFACTURING S.A.", "Country of Origin": "DRL", "Lead Time(weeks)": 10},
        {"Vendor Name": "DRL/SARDELLI", "Country of Origin": "DRL", "Lead Time(weeks)": 10},
        {"Vendor Name": "GDL JEWELLERY LTD", "Country of Origin": "Thailand", "Lead Time(weeks)": 8},
        {"Vendor Name": "GOLDENLINE C/O BOGAZICI HEDIYELIK A.S.", "Country of Origin": "US", "Lead Time(weeks)": 9},
        {"Vendor Name": "IJM-INTERCONTINENTAL JEWELLERY MFG", "Country of Origin": "Thailand", "Lead Time(weeks)": 8},
        {"Vendor Name": "JOHN C NORDT COMPANY", "Country of Origin": "US", "Lead Time(weeks)": 9},
        {"Vendor Name": "LARA CORPORATION LIMITED", "Country of Origin": "China", "Lead Time(weeks)": 8},
        {"Vendor Name": "Leach & Garner (HK) Limited", "Country of Origin": "China", "Lead Time(weeks)": 8},
        {"Vendor Name": "LINEA NUOVA S.A.", "Country of Origin": "Peru", "Lead Time(weeks)": 9},
        {"Vendor Name": "Milor S.P.A", "Country of Origin": "Italy", "Lead Time(weeks)": 9},
        {"Vendor Name": "MIORO GOLD, LLC", "Country of Origin": "US", "Lead Time(weeks)": 9},
        {"Vendor Name": "NEW GOLD", "Country of Origin": "US", "Lead Time(weeks)": 9},
        {"Vendor Name": "PD&P LTD.", "Country of Origin": "China", "Lead Time(weeks)": 8},
        {"Vendor Name": "RICHLINE ITALY SRL", "Country of Origin": "Italy", "Lead Time(weeks)": 12},
        # Note: The country for "RICHLINE SA PTY" is unclear in the provided data. 
        # If "Zales" is not a country, you can adjust as needed or remove. 
        {"Vendor Name": "RICHLINE SA PTY", "Country of Origin": "Zales", "Lead Time(weeks)": None},
        {"Vendor Name": "RLG STANDARD MANUFACTURING VENDOR", "Country of Origin": "US", "Lead Time(weeks)": 9},
        {"Vendor Name": "S&S JEWELRY", "Country of Origin": "US", "Lead Time(weeks)": 9},
        {"Vendor Name": "SABELLI, S.A. DE C.V.", "Country of Origin": "Mexico", "Lead Time(weeks)": 9},
        {"Vendor Name": "SERENITY JEWELS PVT LTD", "Country of Origin": "India", "Lead Time(weeks)": 8},
        {"Vendor Name": "SHANGOLD INDIA LTD", "Country of Origin": "India", "Lead Time(weeks)": 8},
        {"Vendor Name": "SILO SPA-WIRES", "Country of Origin": "Italy", "Lead Time(weeks)": 12},
        {"Vendor Name": "STRONG JEWELRY(HK) CO LTD CHONG", "Country of Origin": "China", "Lead Time(weeks)": 8},
        {"Vendor Name": "T.C.I. LTD", "Country of Origin": "China", "Lead Time(weeks)": 8},
        {"Vendor Name": "National Chain", "Country of Origin": "US", "Lead Time(weeks)": 9},
        # Repeated entries (if you want to keep them exactly as listed):
        {"Vendor Name": "Leach & Garner (HK) Limited", "Country of Origin": "China", "Lead Time(weeks)": 8},
        {"Vendor Name": "LEACH & GARNER(HK) LTD", "Country of Origin": "China", "Lead Time(weeks)": 8},
        {"Vendor Name": "MANJUSAKA JEWELERS CO.,LTD", "Country of Origin": "China", "Lead Time(weeks)": 8},
        {"Vendor Name": "MEICHONG JEWELRY (HK) COMPANY LIMITED", "Country of Origin": "China", "Lead Time(weeks)": 8},
        {"Vendor Name": "NATIONAL CHAIN", "Country of Origin": "US", "Lead Time(weeks)": 9},
        {"Vendor Name": "PRIME DIRECT LIMITED", "Country of Origin": "China", "Lead Time(weeks)": 8},
    ]
    
    # Create the DataFrame
    df_coo = pd.DataFrame(data)

    # Merge the data on Vendor Name to add the 'Country of Origin' and 'Lead Time'
    vendor_sheet = pd.merge(df_unique_vendor, df_coo[['Vendor Name', 'Country of Origin', 'Lead Time(weeks)']], 
                        on='Vendor Name', how='left')




   
#     category_tuples=[

#     ('Bridge Gem', '742'),
#     ('Gold', '746'),
#     ('Gold', '262&270'),
#     ('Womens Silver', '260&404'),
#     ('Precious', '264&268'),
#     ('Fine Pearl', '265&271'),
#     ('Semi', '272&733'),
#     ('Diamond', '734&737&748'),
#     ('Bridal', '739&267&263'),
#     ("Men's", '768&771')

# ]
    category_tuples = input_tuple
    
    
    dynamic_categories = [
    (
        category,
        code,
        report_grouping_df.loc[
            report_grouping_df[0].str.upper() == f"{category.upper()}{code}".upper(), 3
        ].iloc[0] if not report_grouping_df.loc[
            report_grouping_df[0].str.upper() == f"{category.upper()}{code}".upper()
        ].empty else None  # Handle missing values
    )
    for category, code in category_tuples]
    
    month_dict = {
        "Feb": 1,
        "Mar": 2,
        "Apr": 3,
        "May": 4,
        "Jun": 5,
        "Jul": 6,
        "Aug": 7,
        "Sep": 8,
        "Oct": 9,
        "Nov": 10,
        "Dec": 11,
        "Jan": 12
    }
    current_month_number = month_dict.get(current_month, "Month not found")
    if current_month in ["Oct","Nov" "Dec","Jan","Feb","Mar","Apr","May"]:
        rolling_method="Current MTH"
    else:
        rolling_method="YTD"
    # year_of_previous_month=2024
    # last_year_of_previous_month=2023
    
    other_params = (
        year_of_previous_month,
        last_year_of_previous_month,
        season,
        current_month,
        current_month_number,
        previous_week_number,
        last_month_of_previous_month_numeric,
        rolling_method,
        feb_weeks,
        mar_weeks,
        apr_weeks,
        may_weeks,
        jun_weeks,
        jul_weeks,
        aug_weeks,
        sep_weeks,
        oct_weeks,
        nov_weeks,
        dec_weeks,
        jan_weeks,
        index_df,
        report_grouping_df,
        planning_df,
        TBL_Planning_VerticalReport__3,
        Macys_Recpts,
        All_DATA,
        MCOM_Data,
        percentage,
        month_from,
        month_to,
        master_sheet,
        vendor_sheet,
        birthstone_sheet

    ) 
    
    
    #main
    # args = [(category, (code, num_products), dynamic_categories, file_path, other_params) for category, (code, num_products) in dynamic_categories.items()]
    args = [(category, (code, num_products), dynamic_categories, file_path, other_params)    for category, code, num_products in dynamic_categories ]
    # Use multiprocessing pool to process the categories
    with Pool(processes=os.cpu_count()) as pool:
        results = pool.map(process_category, args)


    final_data = {}
    for result in results:
        final_data.update(result)

    # Write the final JSON to a file
    with open("media/output.json", "w") as f:
        json.dump(final_data, f, indent=2)

    print("Final JSON saved successfully.")

   





















            